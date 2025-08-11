# app/tasks.py
from datetime import datetime
from io import BytesIO

from flask import current_app
from flask_mail import Message
from app.utils.db import get_db_connection
from app import mail

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Inclua ambos se em algum lugar usar "Atrasada" em vez de "Vencida"
STATUS_ALVOS = ("Não iniciada", "Em andamento", "Vencida", "Atrasada")

def _auto_fit(ws):
    # largura automática simples
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

def _build_excel(acoes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ações pendentes"

    headers = ["ID", "Descrição", "Origem", "Criado por", "Responsável", "Centro de Custos", "Prazo", "Status"]
    ws.append(headers)

    for a in acoes:
        prazo = a.get("prazo")
        if prazo and hasattr(prazo, "strftime"):
            prazo_fmt = prazo.strftime("%d/%m/%Y")
        else:
            try:
                # caso venha como string 'YYYY-MM-DD'
                prazo_fmt = datetime.strptime(str(prazo), "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                prazo_fmt = ""

        ws.append([
            a.get("id"),
            a.get("descricao"),
            a.get("descricao_origem"),
            a.get("nome_criador"),
            a.get("nome_responsavel"),
            a.get("descricao_cc"),
            prazo_fmt,
            a.get("status"),
        ])

    _auto_fit(ws)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def _send_report_to_responsavel(responsavel):
    """
    responsavel: dict com id, nome, email
    Retorna True se enviou, False se não havia ações.
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        # Monta IN (%s, %s, ...) dinamicamente
        in_placeholders = ", ".join(["%s"] * len(STATUS_ALVOS))

        query = f"""
            SELECT 
                a.id, a.descricao, a.prazo, a.status,
                o.descricao AS descricao_origem,
                u.nome AS nome_responsavel,
                uc.nome AS nome_criador,
                cc.descricao AS descricao_cc
            FROM acoes a
            JOIN usuarios u  ON a.responsavel_id = u.id          -- responsável
            JOIN usuarios uc ON a.criado_por    = uc.id          -- criador
            JOIN centros_custos cc ON u.centro_custos_id = cc.id
            JOIN origens o ON a.origem_id = o.id
            WHERE a.responsavel_id = %s
              AND a.status IN ({in_placeholders})
            ORDER BY a.prazo ASC, a.id ASC
        """

        params = (responsavel["id"], *STATUS_ALVOS)
        cursor.execute(query, params)
        acoes = cursor.fetchall()
    finally:
        conn.close()

    if not acoes:
        return False

    # Gera Excel
    excel_io = _build_excel(acoes)
    hoje = datetime.now().strftime("%Y-%m-%d")

    # Monta e-mail
    msg = Message(
        subject=f"[TrackPlan] Relatório semanal de ações — {hoje}",
        recipients=[responsavel["email"]],
    )
    msg.html = f"""
    <div style="font-family: Arial, sans-serif; font-size: 15px; color:#333;">
      <div style="text-align:center; margin-bottom:18px;">
        <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height:50px;">
      </div>
      <p>Olá <strong>{responsavel["nome"]}</strong>,</p>
      <p>Segue em anexo seu relatório semanal de ações que estão com status
         <em>Não iniciada</em>, <em>Em andamento</em> ou <em>Vencida/Atrasada</em>.</p>
      <p style="font-size:13px; color:#666;">Equipe TrackPlan</p>
    </div>
    """

    filename = f"acoes_{responsavel['id']}_{hoje}.xlsx"
    msg.attach(
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        excel_io.read()
    )
    mail.send(msg)
    return True

def send_weekly_reports():
    """
    Envia um xlsx para cada responsável com ações pendentes.
    Retorna um resumo: {"enviados": int, "sem_acoes": int, "erros": [(email, msg), ...]}
    """
    enviados = 0
    sem_acoes = 0
    erros = []

    with current_app.app_context():
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            # Busca todos os usuários ativos com e-mail
            cursor.execute("""
                SELECT u.id, u.nome, u.email
                FROM usuarios u
                WHERE u.ativo = 1
                  AND u.email IS NOT NULL AND u.email <> ''
            """)
            responsaveis = cursor.fetchall()
        finally:
            conn.close()

        for r in responsaveis:
            try:
                enviou = _send_report_to_responsavel(r)
                if enviou:
                    enviados += 1
                else:
                    sem_acoes += 1
            except Exception as e:
                erros.append((r.get("email"), str(e)))

    # Log no console pra depurar rápido
    print(f"[Relatórios] enviados={enviados}, sem_acoes={sem_acoes}, erros={len(erros)}")
    for em, err in erros[:5]:
        print("  erro:", em, err)

    return {"enviados": enviados, "sem_acoes": sem_acoes, "erros": erros}
