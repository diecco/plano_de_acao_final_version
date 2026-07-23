import os
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.decorators import login_required, module_required, perfil_required
from app.permissions import filtrar_usuario_ids_permitidos, resolver_escopo_usuarios_treinamento
from app.upload_security import UploadService
from app.utils.db import get_db_connection


def register_treinamentos_routes(blueprint):
    def allowed_pdf(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'pdf'

    def salvar_evidencia_treinamento(arquivo):
        if not arquivo or not arquivo.filename:
            return None

        if not allowed_pdf(arquivo.filename):
            raise ValueError('O arquivo de evidência deve estar em PDF.')

        return UploadService.salvar(
            arquivo,
            {"pdf"},
            prefixo="treinamento",
            diretorio=os.path.join('static', 'evidencias_treinamentos'),
        )

    @blueprint.route('/treinamentos_realizados', methods=['GET', 'POST'])
    @login_required
    @module_required('acesso_treinamentos')
    def treinamentos_realizados():

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')
        centro_custo_id = session.get('centro_custos_id')

        if not usuario_id or not centro_custo_id:
            conn.close()
            flash('Usuário logado inválido.', 'danger')
            return redirect(url_for('main.dashboard'))

        if request.method == 'POST':

            procedimento_id = request.form.get('procedimento_id') or None
            procedimento_revisao_id = request.form.get('procedimento_revisao_id') or None
            data_treinamento = request.form.get('data_treinamento') or None
            hora_inicio = request.form.get('hora_inicio') or None
            hora_fim = request.form.get('hora_fim') or None
            local_treinamento = (request.form.get('local_treinamento') or '').strip()
            tipo_instrutor = (request.form.get('tipo_instrutor') or '').strip()
            instrutor_usuario_id = request.form.get('instrutor_usuario_id') or None
            instrutor_externo_id = request.form.get('instrutor_externo_id') or None
            observacoes = (request.form.get('observacoes') or '').strip() or None
            participante_ids = request.form.getlist('participante_ids[]')
            evidencia_arquivo = request.files.get('evidencia_arquivo')

            if not procedimento_id:
                flash('Selecione o procedimento.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if not procedimento_revisao_id:
                flash('Selecione a revisão do procedimento.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if not data_treinamento:
                flash('Informe a data do treinamento.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if not hora_inicio or not hora_fim:
                flash('Informe a hora de início e a hora de fim.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if not local_treinamento:
                flash('Informe o local do treinamento.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if tipo_instrutor not in ('interno', 'externo'):
                flash('Selecione um tipo de instrutor válido.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if tipo_instrutor == 'interno' and not instrutor_usuario_id:
                flash('Selecione o instrutor interno.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if tipo_instrutor == 'externo' and not instrutor_externo_id:
                flash('Selecione o instrutor externo.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            if not participante_ids:
                flash('Adicione pelo menos um participante.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            try:

                # 🔒 VALIDA PARTICIPANTES CONFORME ESCOPO
                participantes_ids_processados = set()

                for participante_id in participante_ids:

                    participante_id = (participante_id or '').strip()

                    if not participante_id:
                        continue

                    if participante_id in participantes_ids_processados:
                        continue

                    if perfil in ['administrador', 'avancado']:

                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                        """, (participante_id,))

                    else:

                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                              AND centro_custos_id = %s
                        """, (participante_id, centro_custo_id))

                    participante_valido = cursor.fetchone()

                    if not participante_valido:
                        flash('Existe participante fora do seu escopo de acesso.', 'danger')
                        conn.close()
                        return redirect(url_for('main.treinamentos_realizados'))

                    participantes_ids_processados.add(participante_id)

                # 🔒 VALIDA INSTRUTOR INTERNO
                if tipo_instrutor == 'interno':

                    if perfil in ['administrador', 'avancado']:

                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                              AND pode_ser_instrutor = 1
                        """, (instrutor_usuario_id,))

                    else:

                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                              AND pode_ser_instrutor = 1
                              AND centro_custos_id = %s
                        """, (instrutor_usuario_id, centro_custo_id))

                    instrutor_valido = cursor.fetchone()

                    if not instrutor_valido:
                        flash('Instrutor inválido para seu escopo.', 'danger')
                        conn.close()
                        return redirect(url_for('main.treinamentos_realizados'))

                hora_inicio_dt = datetime.strptime(hora_inicio, '%H:%M')
                hora_fim_dt = datetime.strptime(hora_fim, '%H:%M')

                if hora_fim_dt <= hora_inicio_dt:
                    flash('A hora de fim deve ser maior que a hora de início.', 'danger')
                    conn.close()
                    return redirect(url_for('main.treinamentos_realizados'))

                cursor.execute("""
                    SELECT 
                        pr.id,
                        p.validade_dias
                    FROM procedimento_revisoes pr
                    JOIN procedimentos p
                        ON p.id = pr.procedimento_id
                    WHERE pr.id = %s
                      AND pr.procedimento_id = %s
                      AND p.ativo = 1
                """, (procedimento_revisao_id, procedimento_id))

                revisao_valida = cursor.fetchone()

                if not revisao_valida:
                    flash('A revisão selecionada não pertence ao procedimento informado.', 'danger')
                    conn.close()
                    return redirect(url_for('main.treinamentos_realizados'))

                validade_dias = revisao_valida.get('validade_dias')
                data_validade = None

                if validade_dias is not None and data_treinamento:
                    data_validade = (
                        datetime.strptime(data_treinamento, '%Y-%m-%d').date()
                        + timedelta(days=int(validade_dias))
                    )

                nome_arquivo_evidencia = None

                if evidencia_arquivo and evidencia_arquivo.filename:
                    nome_arquivo_evidencia = salvar_evidencia_treinamento(evidencia_arquivo)

                cursor.execute("""
                    INSERT INTO treinamentos_realizados (
                        data_treinamento,
                        data_validade,
                        hora_inicio,
                        hora_fim,
                        local_treinamento,
                        tipo_instrutor,
                        instrutor_usuario_id,
                        instrutor_externo_id,
                        procedimento_id,
                        procedimento_revisao_id,
                        observacoes,
                        evidencia_arquivo,
                        criado_por
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    data_treinamento,
                    data_validade,
                    hora_inicio,
                    hora_fim,
                    local_treinamento,
                    tipo_instrutor,
                    instrutor_usuario_id if tipo_instrutor == 'interno' else None,
                    instrutor_externo_id if tipo_instrutor == 'externo' else None,
                    procedimento_id,
                    procedimento_revisao_id,
                    observacoes,
                    nome_arquivo_evidencia,
                    usuario_id
                ))

                treinamento_realizado_id = cursor.lastrowid

                participantes_ids_processados = set()

                for participante_id in participante_ids:

                    participante_id = (participante_id or '').strip()

                    if not participante_id:
                        continue

                    if participante_id in participantes_ids_processados:
                        continue

                    cursor.execute("""
                        INSERT INTO treinamentos_realizados_participantes (
                            treinamento_realizado_id,
                            usuario_id,
                            presenca,
                            aprovado,
                            observacoes
                        )
                        VALUES (%s, %s, 1, 1, NULL)
                    """, (treinamento_realizado_id, participante_id))

                    participantes_ids_processados.add(participante_id)

                conn.commit()

                flash('Treinamento realizado cadastrado com sucesso!', 'success')

            except ValueError as e:

                conn.rollback()
                flash(str(e), 'danger')

            except Exception as e:

                conn.rollback()
                flash(f'Erro ao cadastrar treinamento realizado: {e}', 'danger')

            finally:

                conn.close()

            return redirect(url_for('main.treinamentos_realizados'))

        # =====================================================
        # GET
        # =====================================================

        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            WHERE p.ativo = 1
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos = cursor.fetchall()

        cursor.execute("""
            SELECT
                pr.id,
                pr.procedimento_id,
                pr.numero_revisao,
                DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
            FROM procedimento_revisoes pr
            JOIN procedimentos p
                ON p.id = pr.procedimento_id
            WHERE p.ativo = 1
            ORDER BY pr.procedimento_id, pr.data_revisao DESC, pr.id DESC
        """)
        revisoes = cursor.fetchall()

        revisoes_por_procedimento = {}

        for r in revisoes:

            procedimento_id = str(r['procedimento_id'])

            revisoes_por_procedimento.setdefault(procedimento_id, []).append({
                'id': r['id'],
                'numero_revisao': r['numero_revisao'],
                'data_revisao': r['data_revisao']
            })

        # 🔒 INSTRUTORES
        if perfil in ['administrador', 'avancado']:

            cursor.execute("""
                SELECT
                    u.id,
                    u.nome
                FROM usuarios u
                WHERE u.ativo = 1
                  AND u.pode_ser_instrutor = 1
                ORDER BY u.nome
            """)

        else:

            cursor.execute("""
                SELECT
                    u.id,
                    u.nome
                FROM usuarios u
                WHERE u.ativo = 1
                  AND u.pode_ser_instrutor = 1
                  AND u.centro_custos_id = %s
                ORDER BY u.nome
            """, (centro_custo_id,))

        instrutores_internos = cursor.fetchall()

        cursor.execute("""
            SELECT
                id,
                nome,
                empresa
            FROM instrutores_externos
            WHERE ativo = 1
            ORDER BY nome
        """)
        instrutores_externos = cursor.fetchall()

        # 🔒 PARTICIPANTES
        if perfil in ['administrador', 'avancado']:

            filtro_participantes = ""

            params_participantes = []

        else:

            filtro_participantes = "AND u.centro_custos_id = %s"

            params_participantes = [centro_custo_id]

        cursor.execute(f"""
            SELECT
                u.id,
                u.nome,
                u.matricula,
                c.nome AS nome_cargo
            FROM usuarios u
            LEFT JOIN cargos c
                ON c.id = u.cargo_id
            WHERE u.ativo = 1
            {filtro_participantes}
            ORDER BY u.nome
        """, params_participantes)

        participantes_nome = cursor.fetchall()

        cursor.execute(f"""
            SELECT
                u.id,
                u.nome,
                u.matricula,
                c.nome AS nome_cargo
            FROM usuarios u
            LEFT JOIN cargos c
                ON c.id = u.cargo_id
            WHERE u.ativo = 1
            {filtro_participantes}
            ORDER BY u.matricula, u.nome
        """, params_participantes)

        participantes_matricula = cursor.fetchall()

        conn.close()

        return render_template(
            'treinamentos_realizados.html',
            procedimentos=procedimentos,
            revisoes_por_procedimento=revisoes_por_procedimento,
            instrutores_internos=instrutores_internos,
            instrutores_externos=instrutores_externos,
            participantes_nome=participantes_nome,
            participantes_matricula=participantes_matricula
        )

    @blueprint.route('/listar_treinamentos', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def listar_treinamentos():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil_logado = session.get('perfil')
        centro_custos_logado = session.get('centro_custos_id')

        procedimento_id = request.args.get('procedimento_id', '').strip()
        data_inicio = request.args.get('data_inicio', '').strip()
        data_fim = request.args.get('data_fim', '').strip()
        sort = request.args.get('sort', 'data_treinamento').strip()
        order = request.args.get('order', 'desc').strip()

        page = request.args.get("page", 1, type=int)
        per_page = 30

        if page < 1:
            page = 1

        offset = (page - 1) * per_page

        colunas_validas = {
            'data_treinamento': 'tr.data_treinamento',
            'hora_inicio': 'tr.hora_inicio',
            'hora_fim': 'tr.hora_fim',
            'procedimento': 'p.numero_documento',
            'numero_revisao': 'pr.numero_revisao',
            'nome_instrutor': """
                CASE
                    WHEN tr.tipo_instrutor = 'interno' THEN ui.nome
                    WHEN tr.tipo_instrutor = 'externo' THEN ie.nome
                    ELSE ''
                END
            """
        }

        coluna_sort = colunas_validas.get(sort, 'tr.data_treinamento')
        direcao = 'ASC' if order == 'asc' else 'DESC'

        base_where = """
            FROM treinamentos_realizados tr
            JOIN procedimentos p
                ON p.id = tr.procedimento_id
            JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            JOIN procedimento_revisoes pr
                ON pr.id = tr.procedimento_revisao_id
            LEFT JOIN usuarios ui
                ON ui.id = tr.instrutor_usuario_id
            LEFT JOIN instrutores_externos ie
                ON ie.id = tr.instrutor_externo_id
            WHERE tr.ativo = 1
        """

        params = []

        if procedimento_id:
            base_where += " AND tr.procedimento_id = %s"
            params.append(procedimento_id)

        if data_inicio:
            base_where += " AND tr.data_treinamento >= %s"
            params.append(data_inicio)

        if data_fim:
            base_where += " AND tr.data_treinamento <= %s"
            params.append(data_fim)

        # 🔒 PERMISSIONAMENTO
        if perfil_logado == 'basico':
            base_where += """
                AND EXISTS (
                    SELECT 1
                    FROM treinamentos_realizados_participantes trp
                    WHERE trp.treinamento_realizado_id = tr.id
                      AND trp.usuario_id = %s
                )
            """
            params.append(usuario_id)

        elif perfil_logado == 'intermediario':
            base_where += """
                AND EXISTS (
                    SELECT 1
                    FROM treinamentos_realizados_participantes trp
                    JOIN usuarios up
                        ON up.id = trp.usuario_id
                    WHERE trp.treinamento_realizado_id = tr.id
                      AND up.centro_custos_id = %s
                )
            """
            params.append(centro_custos_logado)

        # avançado e administrador veem tudo

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            {base_where}
        """, params)
        total_registros = cursor.fetchone()['total']

        total_paginas = (total_registros + per_page - 1) // per_page

        if total_paginas > 0 and page > total_paginas:
            page = total_paginas
            offset = (page - 1) * per_page

        cursor.execute(f"""
            SELECT tr.id
            {base_where}
            ORDER BY {coluna_sort} {direcao}, tr.id DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])

        ids = [row['id'] for row in cursor.fetchall()]

        if not ids:
            treinamentos = []
        else:
            placeholders = ','.join(['%s'] * len(ids))

            cursor.execute(f"""
                SELECT
                    tr.id,
                    tr.procedimento_id,
                    tr.procedimento_revisao_id,
                    DATE_FORMAT(tr.data_treinamento, '%d/%m/%Y') AS data_treinamento,
                    DATE_FORMAT(tr.data_treinamento, '%Y-%m-%d') AS data_treinamento_iso,
                    TIME_FORMAT(tr.hora_inicio, '%H:%i') AS hora_inicio,
                    TIME_FORMAT(tr.hora_fim, '%H:%i') AS hora_fim,
                    tr.local_treinamento,
                    tr.tipo_instrutor,
                    tr.instrutor_usuario_id,
                    tr.instrutor_externo_id,
                    tr.observacoes,
                    tr.ativo,
                    tr.evidencia_arquivo,

                    CONCAT(td.sigla, ' ', p.numero_documento, ' - ', p.titulo) AS procedimento_descricao,
                    pr.numero_revisao,

                    CASE
                        WHEN tr.tipo_instrutor = 'interno' THEN ui.nome
                        WHEN tr.tipo_instrutor = 'externo' THEN ie.nome
                        ELSE '-'
                    END AS nome_instrutor

                FROM treinamentos_realizados tr
                JOIN procedimentos p
                    ON p.id = tr.procedimento_id
                JOIN tipos_documento td
                    ON td.id = p.tipo_documento_id
                JOIN procedimento_revisoes pr
                    ON pr.id = tr.procedimento_revisao_id
                LEFT JOIN usuarios ui
                    ON ui.id = tr.instrutor_usuario_id
                LEFT JOIN instrutores_externos ie
                    ON ie.id = tr.instrutor_externo_id
                WHERE tr.id IN ({placeholders})
                ORDER BY {coluna_sort} {direcao}, tr.id DESC
            """, ids)

            treinamentos = cursor.fetchall()

        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            WHERE p.ativo = 1
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos = cursor.fetchall()

        cursor.execute("""
            SELECT
                pr.id,
                pr.procedimento_id,
                pr.numero_revisao,
                DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
            FROM procedimento_revisoes pr
            JOIN procedimentos p
                ON p.id = pr.procedimento_id
            WHERE p.ativo = 1
            ORDER BY pr.procedimento_id, pr.data_revisao DESC, pr.id DESC
        """)
        revisoes = cursor.fetchall()

        revisoes_por_procedimento = {}
        for r in revisoes:
            procedimento_key = str(r['procedimento_id'])
            revisoes_por_procedimento.setdefault(procedimento_key, []).append({
                'id': r['id'],
                'numero_revisao': r['numero_revisao'],
                'data_revisao': r['data_revisao']
            })

        if perfil_logado in ['administrador', 'avancado']:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = 1
                ORDER BY nome
            """)
        else:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = 1
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (centro_custos_logado,))

        instrutores_internos = cursor.fetchall()

        cursor.execute("""
            SELECT id, nome, empresa
            FROM instrutores_externos
            WHERE ativo = 1
            ORDER BY nome
        """)
        instrutores_externos = cursor.fetchall()

        conn.close()

        filtros = {
            'procedimento_id': procedimento_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'sort': sort,
            'order': order
        }

        return render_template(
            'listar_treinamentos.html',
            treinamentos=treinamentos,
            procedimentos=procedimentos,
            revisoes_por_procedimento=revisoes_por_procedimento,
            instrutores_internos=instrutores_internos,
            instrutores_externos=instrutores_externos,
            filtros=filtros,
            page=page,
            per_page=per_page,
            total_paginas=total_paginas,
            total_registros=total_registros
        )

    @blueprint.route('/editar_treinamento/<int:id>', methods=['GET', 'POST'])
    @login_required
    @module_required('acesso_treinamentos')
    @perfil_required('avancado')
    def editar_treinamento(id):
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_logado_id = session.get('usuario_id')
        perfil_logado = session.get('perfil')
        centro_custos_logado = session.get('centro_custos_id')

        cursor.execute("""
            SELECT tr.*
            FROM treinamentos_realizados tr
            WHERE tr.id = %s
              AND tr.ativo = 1
        """, (id,))
        treinamento = cursor.fetchone()

        if not treinamento:
            conn.close()
            flash('Treinamento não encontrado.', 'warning')
            return redirect(url_for('main.listar_treinamentos'))

        # 🔒 PERMISSIONAMENTO
        if perfil_logado == 'basico':
            cursor.execute("""
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                WHERE trp.treinamento_realizado_id = %s
                  AND trp.usuario_id = %s
                LIMIT 1
            """, (id, usuario_logado_id))
            permitido = cursor.fetchone()

            if not permitido:
                conn.close()
                flash('Você não tem permissão para editar este treinamento.', 'danger')
                return redirect(url_for('main.listar_treinamentos'))

        elif perfil_logado == 'intermediario':
            cursor.execute("""
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                JOIN usuarios u ON u.id = trp.usuario_id
                WHERE trp.treinamento_realizado_id = %s
                  AND u.centro_custos_id = %s
                LIMIT 1
            """, (id, centro_custos_logado))
            permitido = cursor.fetchone()

            if not permitido:
                conn.close()
                flash('Você não tem permissão para editar este treinamento.', 'danger')
                return redirect(url_for('main.listar_treinamentos'))

        # avançado e administrador veem tudo

        if request.method == 'POST':
            next_url = request.form.get('next') or url_for('main.listar_treinamentos')

            procedimento_id = request.form.get('procedimento_id') or None
            procedimento_revisao_id = request.form.get('procedimento_revisao_id') or None
            data_treinamento = request.form.get('data_treinamento') or None
            hora_inicio = request.form.get('hora_inicio') or None
            hora_fim = request.form.get('hora_fim') or None
            local_treinamento = (request.form.get('local_treinamento') or '').strip()
            observacoes = (request.form.get('observacoes') or '').strip() or None
            participante_ids = request.form.getlist('participante_ids[]')
            evidencia_arquivo = request.files.get('evidencia_arquivo')

            tipo_instrutor = (request.form.get('tipo_instrutor') or treinamento.get('tipo_instrutor') or '').strip()
            instrutor_usuario_id = request.form.get('instrutor_usuario_id') or treinamento.get('instrutor_usuario_id')
            instrutor_externo_id = request.form.get('instrutor_externo_id') or treinamento.get('instrutor_externo_id')

            if not procedimento_id:
                flash('Selecione o procedimento.', 'danger')
                conn.close()
                return redirect(next_url)

            if not procedimento_revisao_id:
                flash('Selecione a revisão do procedimento.', 'danger')
                conn.close()
                return redirect(next_url)

            if not data_treinamento:
                flash('Informe a data do treinamento.', 'danger')
                conn.close()
                return redirect(next_url)

            if not hora_inicio or not hora_fim:
                flash('Informe a hora de início e a hora de fim.', 'danger')
                conn.close()
                return redirect(next_url)

            if not local_treinamento:
                flash('Informe o local do treinamento.', 'danger')
                conn.close()
                return redirect(next_url)

            if tipo_instrutor not in ('interno', 'externo'):
                flash('Selecione um tipo de instrutor válido.', 'danger')
                conn.close()
                return redirect(next_url)

            if tipo_instrutor == 'interno' and not instrutor_usuario_id:
                flash('Selecione o instrutor interno.', 'danger')
                conn.close()
                return redirect(next_url)

            if tipo_instrutor == 'externo' and not instrutor_externo_id:
                flash('Selecione o instrutor externo.', 'danger')
                conn.close()
                return redirect(next_url)

            try:
                formato_inicio = '%H:%M:%S' if len(str(hora_inicio)) == 8 else '%H:%M'
                formato_fim = '%H:%M:%S' if len(str(hora_fim)) == 8 else '%H:%M'

                hora_inicio_dt = datetime.strptime(str(hora_inicio), formato_inicio)
                hora_fim_dt = datetime.strptime(str(hora_fim), formato_fim)

                if hora_fim_dt <= hora_inicio_dt:
                    flash('A hora de fim deve ser maior que a hora de início.', 'danger')
                    conn.close()
                    return redirect(next_url)

                cursor.execute("""
                    SELECT 
                        pr.id,
                        p.validade_dias
                    FROM procedimento_revisoes pr
                    JOIN procedimentos p ON p.id = pr.procedimento_id
                    WHERE pr.id = %s
                      AND pr.procedimento_id = %s
                      AND p.ativo = 1
                """, (procedimento_revisao_id, procedimento_id))
                revisao_valida = cursor.fetchone()

                if not revisao_valida:
                    flash('A revisão selecionada não pertence ao procedimento informado.', 'danger')
                    conn.close()
                    return redirect(next_url)

                # 🔒 valida instrutor interno
                if tipo_instrutor == 'interno':
                    if perfil_logado in ['administrador', 'avancado']:
                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                              AND pode_ser_instrutor = 1
                        """, (instrutor_usuario_id,))
                    else:
                        cursor.execute("""
                            SELECT id
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                              AND pode_ser_instrutor = 1
                              AND centro_custos_id = %s
                        """, (instrutor_usuario_id, centro_custos_logado))

                    instrutor_valido = cursor.fetchone()

                    if not instrutor_valido:
                        flash('Instrutor inválido para seu escopo.', 'danger')
                        conn.close()
                        return redirect(next_url)

                validade_dias = revisao_valida.get('validade_dias')
                data_validade = None

                if validade_dias is not None and data_treinamento:
                    data_validade = datetime.strptime(data_treinamento, '%Y-%m-%d').date() + timedelta(days=int(validade_dias))

                nome_arquivo_evidencia = treinamento.get('evidencia_arquivo')

                if evidencia_arquivo and evidencia_arquivo.filename:
                    nome_arquivo_evidencia = salvar_evidencia_treinamento(evidencia_arquivo)

                cursor.execute("""
                    UPDATE treinamentos_realizados
                    SET procedimento_id = %s,
                        procedimento_revisao_id = %s,
                        data_treinamento = %s,
                        data_validade = %s,
                        hora_inicio = %s,
                        hora_fim = %s,
                        local_treinamento = %s,
                        tipo_instrutor = %s,
                        instrutor_usuario_id = %s,
                        instrutor_externo_id = %s,
                        observacoes = %s,
                        evidencia_arquivo = %s
                    WHERE id = %s
                      AND ativo = 1
                """, (
                    procedimento_id,
                    procedimento_revisao_id,
                    data_treinamento,
                    data_validade,
                    hora_inicio,
                    hora_fim,
                    local_treinamento,
                    tipo_instrutor,
                    instrutor_usuario_id if tipo_instrutor == 'interno' else None,
                    instrutor_externo_id if tipo_instrutor == 'externo' else None,
                    observacoes,
                    nome_arquivo_evidencia,
                    id
                ))

                if participante_ids:
                    cursor.execute("""
                        DELETE FROM treinamentos_realizados_participantes
                        WHERE treinamento_realizado_id = %s
                    """, (id,))

                    participantes_ids_processados = set()

                    for participante_id in participante_ids:
                        participante_id = (participante_id or '').strip()

                        if not participante_id or participante_id in participantes_ids_processados:
                            continue

                        if perfil_logado in ['administrador', 'avancado']:
                            cursor.execute("""
                                SELECT 1
                                FROM usuarios
                                WHERE id = %s
                                  AND ativo = 1
                                LIMIT 1
                            """, (participante_id,))
                        else:
                            cursor.execute("""
                                SELECT 1
                                FROM usuarios
                                WHERE id = %s
                                  AND centro_custos_id = %s
                                  AND ativo = 1
                                LIMIT 1
                            """, (participante_id, centro_custos_logado))

                        participante_permitido = cursor.fetchone()

                        if not participante_permitido:
                            continue

                        cursor.execute("""
                            INSERT INTO treinamentos_realizados_participantes (
                                treinamento_realizado_id,
                                usuario_id,
                                presenca,
                                aprovado,
                                observacoes
                            )
                            VALUES (%s, %s, 1, 1, NULL)
                        """, (id, participante_id))

                        participantes_ids_processados.add(participante_id)

                conn.commit()
                flash('Treinamento atualizado com sucesso!', 'success')

            except Exception as e:
                conn.rollback()
                flash(f'Erro ao atualizar treinamento: {e}', 'danger')

            finally:
                conn.close()

            return redirect(next_url)

        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            JOIN tipos_documento td ON td.id = p.tipo_documento_id
            WHERE p.ativo = 1
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos = cursor.fetchall()

        cursor.execute("""
            SELECT
                pr.id,
                pr.procedimento_id,
                pr.numero_revisao,
                DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
            FROM procedimento_revisoes pr
            JOIN procedimentos p ON p.id = pr.procedimento_id
            WHERE p.ativo = 1
              AND pr.procedimento_id = %s
            ORDER BY pr.data_revisao DESC, pr.id DESC
        """, (treinamento['procedimento_id'],))
        revisoes = cursor.fetchall()

        if perfil_logado in ['administrador', 'avancado']:
            filtro_participantes = ""
            params_participantes = []
        else:
            filtro_participantes = "AND u.centro_custos_id = %s"
            params_participantes = [centro_custos_logado]

        cursor.execute(f"""
            SELECT
                u.id,
                u.nome,
                u.matricula,
                c.nome AS nome_cargo
            FROM usuarios u
            LEFT JOIN cargos c ON c.id = u.cargo_id
            WHERE u.ativo = 1
            {filtro_participantes}
            ORDER BY u.nome
        """, params_participantes)
        participantes_nome = cursor.fetchall()

        cursor.execute(f"""
            SELECT
                u.id,
                u.nome,
                u.matricula,
                c.nome AS nome_cargo
            FROM usuarios u
            LEFT JOIN cargos c ON c.id = u.cargo_id
            WHERE u.ativo = 1
            {filtro_participantes}
            ORDER BY u.matricula, u.nome
        """, params_participantes)
        participantes_matricula = cursor.fetchall()

        cursor.execute("""
            SELECT
                u.id AS usuario_id,
                u.nome,
                u.matricula,
                c.nome AS nome_cargo
            FROM treinamentos_realizados_participantes trp
            JOIN usuarios u ON u.id = trp.usuario_id
            LEFT JOIN cargos c ON c.id = u.cargo_id
            WHERE trp.treinamento_realizado_id = %s
            ORDER BY u.nome
        """, (id,))
        participantes_selecionados = cursor.fetchall()

        conn.close()

        return render_template(
            'editar_treinamento.html',
            treinamento=treinamento,
            procedimentos=procedimentos,
            revisoes=revisoes,
            participantes_nome=participantes_nome,
            participantes_matricula=participantes_matricula,
            participantes_selecionados=participantes_selecionados
        )

    @blueprint.route('/excluir_treinamento/<int:id>', methods=['POST'])
    @login_required
    @module_required('acesso_treinamentos')
    @perfil_required('avancado')
    def excluir_treinamento(id):
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil_logado = session.get('perfil')
        centro_custos_logado = session.get('centro_custos_id')

        try:
            cursor.execute("""
                SELECT id, ativo
                FROM treinamentos_realizados
                WHERE id = %s
                  AND ativo = 1
            """, (id,))
            treinamento = cursor.fetchone()

            if not treinamento:
                conn.close()
                flash('Treinamento não encontrado ou já excluído.', 'warning')
                return redirect(url_for('main.listar_treinamentos'))

            if perfil_logado == 'basico':
                cursor.execute("""
                    SELECT 1
                    FROM treinamentos_realizados_participantes trp
                    WHERE trp.treinamento_realizado_id = %s
                      AND trp.usuario_id = %s
                    LIMIT 1
                """, (id, usuario_id))
                permitido = cursor.fetchone()

                if not permitido:
                    conn.close()
                    flash('Você não tem permissão para excluir este treinamento.', 'danger')
                    return redirect(url_for('main.listar_treinamentos'))

            elif perfil_logado == 'intermediario':
                cursor.execute("""
                    SELECT 1
                    FROM treinamentos_realizados_participantes trp
                    JOIN usuarios u
                        ON u.id = trp.usuario_id
                    WHERE trp.treinamento_realizado_id = %s
                      AND u.centro_custos_id = %s
                    LIMIT 1
                """, (id, centro_custos_logado))
                permitido = cursor.fetchone()

                if not permitido:
                    conn.close()
                    flash('Você não tem permissão para excluir este treinamento.', 'danger')
                    return redirect(url_for('main.listar_treinamentos'))

            # avançado e administrador veem/excluem tudo

            cursor.execute("""
                UPDATE treinamentos_realizados
                SET ativo = 0
                WHERE id = %s
            """, (id,))

            conn.commit()
            flash('Treinamento excluído logicamente com sucesso!', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao excluir treinamento: {e}', 'danger')

        finally:
            conn.close()

        return redirect(url_for('main.listar_treinamentos'))

    def _rt_table_exists(cursor, table_name):
        try:
            cursor.execute("SHOW TABLES LIKE %s", (table_name,))
            return cursor.fetchone() is not None
        except Exception:
            return False

    def _rt_get_columns(cursor, table_name):
        try:
            cursor.execute(f"SHOW COLUMNS FROM {table_name}")
            return [col["Field"] for col in cursor.fetchall()]
        except Exception:
            return []

    def _rt_first_existing(columns, options):
        for option in options:
            if option in columns:
                return option
        return None

    def _rt_to_date(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def _rt_normalize_ids(values):
        ids = []
        for value in values:
            try:
                ids.append(int(value))
            except (TypeError, ValueError):
                pass
        return list(dict.fromkeys(ids))

    def _rt_resolver_campos_procedimento(cursor):
        colunas_procedimentos = _rt_get_columns(cursor, "procedimentos")
        colunas_tipos_documento = _rt_get_columns(cursor, "tipos_documento")

        if "titulo" in colunas_procedimentos:
            campo_titulo = "p.titulo"
        elif "titulo_documento" in colunas_procedimentos:
            campo_titulo = "p.titulo_documento"
        elif "nome" in colunas_procedimentos:
            campo_titulo = "p.nome"
        else:
            campo_titulo = "''"

        if "numero_documento" in colunas_procedimentos:
            campo_numero = "p.numero_documento"
        else:
            campo_numero = "''"

        if "revisao" in colunas_procedimentos:
            campo_revisao = "p.revisao"
        elif "versao" in colunas_procedimentos:
            campo_revisao = "p.versao"
        else:
            campo_revisao = "''"

        join_tipo_documento = ""
        if "sigla" in colunas_procedimentos:
            campo_sigla = "p.sigla"
        elif "tipo_documento_id" in colunas_procedimentos and "sigla" in colunas_tipos_documento:
            join_tipo_documento = "LEFT JOIN tipos_documento td ON td.id = p.tipo_documento_id"
            campo_sigla = "td.sigla"
        else:
            campo_sigla = "''"

        filtro_ativo_procedimento = ""
        if "ativo" in colunas_procedimentos:
            filtro_ativo_procedimento = "AND p.ativo = 1"

        return {
            "campo_titulo": campo_titulo,
            "campo_numero": campo_numero,
            "campo_revisao": campo_revisao,
            "campo_sigla": campo_sigla,
            "join_tipo_documento": join_tipo_documento,
            "filtro_ativo_procedimento": filtro_ativo_procedimento
        }

    def _rt_buscar_cargos(cursor):
        cargos = []

        if _rt_table_exists(cursor, "cargos"):
            colunas_cargos = _rt_get_columns(cursor, "cargos")

            nome_expr = "CAST(c.id AS CHAR)"
            if "nome" in colunas_cargos:
                nome_expr = "c.nome"
            elif "descricao" in colunas_cargos:
                nome_expr = "c.descricao"

            filtro_ativo = "WHERE c.ativo = 1" if "ativo" in colunas_cargos else ""

            cursor.execute(f"""
                SELECT
                    c.id,
                    {nome_expr} AS nome
                FROM cargos c
                {filtro_ativo}
                ORDER BY nome
            """)
            cargos = cursor.fetchall()
        else:
            cursor.execute("""
                SELECT DISTINCT
                    u.cargo_id AS id,
                    CAST(u.cargo_id AS CHAR) AS nome
                FROM usuarios u
                WHERE u.ativo = 1
                  AND u.cargo_id IS NOT NULL
                ORDER BY nome
            """)
            cargos = cursor.fetchall()

        return cargos

    def _rt_buscar_usuarios_ativos(
        cursor,
        cargo_id=None,
        centro_custos_id=None,
        usuario_id=None,
    ):
        if _rt_table_exists(cursor, "cargos"):
            colunas_cargos = _rt_get_columns(cursor, "cargos")
            cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
            if "nome" in colunas_cargos:
                cargo_nome_expr = "c.nome"
            elif "descricao" in colunas_cargos:
                cargo_nome_expr = "c.descricao"

            query = f"""
                SELECT
                    u.id,
                    u.matricula,
                    u.nome,
                    u.cargo_id,
                    {cargo_nome_expr} AS cargo_nome
                FROM usuarios u
                LEFT JOIN cargos c ON c.id = u.cargo_id
                WHERE u.ativo = 1
            """
        else:
            query = """
                SELECT
                    u.id,
                    u.matricula,
                    u.nome,
                    u.cargo_id,
                    CAST(u.cargo_id AS CHAR) AS cargo_nome
                FROM usuarios u
                WHERE u.ativo = 1
            """

        params = []

        if cargo_id:
            query += " AND u.cargo_id = %s"
            params.append(cargo_id)

        if centro_custos_id is not None:
            query += " AND u.centro_custos_id = %s"
            params.append(centro_custos_id)

        if usuario_id is not None:
            query += " AND u.id = %s"
            params.append(usuario_id)

        query += " ORDER BY u.nome"
        cursor.execute(query, tuple(params))
        return cursor.fetchall()


    def _rt_buscar_usuarios_permitidos(cursor, cargo_id=None):
        filtros_escopo = resolver_escopo_usuarios_treinamento(
            perfil=session.get('perfil'),
            centro_custos_id=session.get('centro_custos_id'),
            usuario_id=session.get('usuario_id'),
        )

        return _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            **filtros_escopo,
        )

    def _rt_resolver_url_evidencia(valor):
        if not valor:
            return None

        valor = str(valor).strip()
        if not valor:
            return None

        valor_normalizado = valor.replace("\\", "/")

        # URL completa
        if valor_normalizado.startswith("http://") or valor_normalizado.startswith("https://"):
            return valor_normalizado

        # Já é caminho static absoluto
        if valor_normalizado.startswith("/static/"):
            return valor_normalizado

        # Caminho static relativo
        if valor_normalizado.startswith("static/"):
            return f"/{valor_normalizado}"

        # Se já vier com a pasta correta
        if valor_normalizado.startswith("evidencias_treinamentos/"):
            return url_for("static", filename=valor_normalizado)

        # Se vier com outra pasta interna de static
        if valor_normalizado.startswith("uploads/"):
            return f"/static/{valor_normalizado}"

        # Só nome do arquivo: assume a pasta correta
        nome_arquivo = os.path.basename(valor_normalizado)
        if nome_arquivo:
            return url_for("static", filename=f"evidencias_treinamentos/{nome_arquivo}")

        return None

    def _rt_buscar_treinamentos_realizados(cursor, usuario_ids=None, cargo_id=None, data_inicio=None, data_fim=None):
        if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
            return []

        if not _rt_table_exists(cursor, "treinamentos_realizados"):
            return []

        cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
        cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
        cols_usuarios = _rt_get_columns(cursor, "usuarios")
        cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

        usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id", "participante_id", "colaborador_id"])
        trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id", "treinamentos_realizados_id", "realizado_id"])
        trr_pk = _rt_first_existing(cols_trr, ["id"])

        if not usuario_col_trp or not trp_to_trr or not trr_pk:
            return []

        cfg_proc = _rt_resolver_campos_procedimento(cursor)

        cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
        join_cargos = ""
        if cols_cargos:
            if "nome" in cols_cargos:
                cargo_nome_expr = "c.nome"
            elif "descricao" in cols_cargos:
                cargo_nome_expr = "c.descricao"
            join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

        data_realizacao_col = _rt_first_existing(cols_trr, [
            "data_treinamento", "data_realizacao", "realizado_em", "data", "created_at"
        ])

        validade_col = _rt_first_existing(cols_trr, [
            "data_validade", "validade_ate", "data_vencimento", "vencimento"
        ])

        hora_inicio_col = _rt_first_existing(cols_trr, ["hora_inicio"])
        hora_fim_col = _rt_first_existing(cols_trr, ["hora_fim"])
        revisao_col = _rt_first_existing(cols_trr, ["procedimento_revisao_id"])

        evidencia_col = _rt_first_existing(cols_trr, [
            "evidencia_arquivo",
            "evidencia",
            "arquivo_evidencia",
            "anexo",
            "caminho_evidencia",
            "nome_arquivo"
        ])

        filtros = ["u.ativo = 1"]
        params = []

        if "ativo" in cols_trp:
            filtros.append("trp.ativo = 1")
        if "ativo" in cols_trr:
            filtros.append("trr.ativo = 1")
        if "ativo" in cols_usuarios:
            filtros.append("u.ativo = 1")

        if usuario_ids:
            placeholders = ", ".join(["%s"] * len(usuario_ids))
            filtros.append(f"u.id IN ({placeholders})")
            params.extend(usuario_ids)

        if cargo_id:
            filtros.append("u.cargo_id = %s")
            params.append(cargo_id)

        if data_realizacao_col:
            if data_inicio and data_fim:
                filtros.append(f"DATE(trr.{data_realizacao_col}) BETWEEN %s AND %s")
                params.append(data_inicio)
                params.append(data_fim)
            elif data_inicio:
                filtros.append(f"DATE(trr.{data_realizacao_col}) >= %s")
                params.append(data_inicio)
            elif data_fim:
                filtros.append(f"DATE(trr.{data_realizacao_col}) <= %s")
                params.append(data_fim)

        select_data = f"trr.{data_realizacao_col}" if data_realizacao_col else "NULL"
        select_validade = f"trr.{validade_col}" if validade_col else "NULL"
        select_hora_inicio = f"trr.{hora_inicio_col}" if hora_inicio_col else "NULL"
        select_hora_fim = f"trr.{hora_fim_col}" if hora_fim_col else "NULL"
        select_revisao = f"trr.{revisao_col}" if revisao_col else "NULL"
        select_evidencia = f"trr.{evidencia_col}" if evidencia_col else "NULL"

        query = f"""
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                {cargo_nome_expr} AS cargo_nome,
                p.id AS procedimento_id,
                {cfg_proc['campo_sigla']} AS sigla,
                {cfg_proc['campo_numero']} AS numero_documento,
                {cfg_proc['campo_titulo']} AS titulo,
                {select_revisao} AS revisao,
                {select_data} AS data_realizacao,
                {select_validade} AS validade,
                {select_hora_inicio} AS hora_inicio,
                {select_hora_fim} AS hora_fim,
                {select_evidencia} AS evidencia
            FROM treinamentos_realizados_participantes trp
            INNER JOIN treinamentos_realizados trr
                ON trr.{trr_pk} = trp.{trp_to_trr}
            INNER JOIN usuarios u
                ON u.id = trp.{usuario_col_trp}
            {join_cargos}
            LEFT JOIN procedimentos p
                ON p.id = trr.procedimento_id
            {cfg_proc['join_tipo_documento']}
            WHERE {" AND ".join(filtros)}
            ORDER BY u.nome, data_realizacao DESC, titulo
        """

        cursor.execute(query, tuple(params))
        dados = cursor.fetchall()

        for item in dados:
            item["status"] = "Realizado"
            item["evidencia_url"] = _rt_resolver_url_evidencia(item.get("evidencia"))

            data_realizacao = _rt_to_date(item.get("data_realizacao"))
            validade = _rt_to_date(item.get("validade"))

            item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""
            item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""

            carga_horaria_fmt = ""
            hora_inicio = item.get("hora_inicio")
            hora_fim = item.get("hora_fim")

            if hora_inicio and hora_fim:
                try:
                    h1 = datetime.strptime(str(hora_inicio), "%H:%M:%S")
                    h2 = datetime.strptime(str(hora_fim), "%H:%M:%S")
                    diferenca = h2 - h1
                    horas = diferenca.total_seconds() / 3600
                    if horas.is_integer():
                        carga_horaria_fmt = str(int(horas))
                    else:
                        carga_horaria_fmt = f"{horas:.2f}".replace(".", ",")
                except Exception:
                    try:
                        h1 = datetime.strptime(str(hora_inicio), "%H:%M")
                        h2 = datetime.strptime(str(hora_fim), "%H:%M")
                        diferenca = h2 - h1
                        horas = diferenca.total_seconds() / 3600
                        if horas.is_integer():
                            carga_horaria_fmt = str(int(horas))
                        else:
                            carga_horaria_fmt = f"{horas:.2f}".replace(".", ",")
                    except Exception:
                        carga_horaria_fmt = ""

            item["carga_horaria"] = carga_horaria_fmt

        return dados

    def _rt_exportar_excel_relatorio_treinamentos(
        registros,
        nome_aba,
        headers,
        row_builder,
        larguras,
        colunas_centralizadas=None
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = nome_aba

        ws.append(headers)

        fill_header = PatternFill(fill_type="solid", fgColor="EA6A23")
        font_header = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center

        for item in registros:
            ws.append(row_builder(item))

        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        colunas_centralizadas = colunas_centralizadas or []

        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                if idx in colunas_centralizadas:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @blueprint.route('/relatorio_treinamentos_realizados', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def relatorio_treinamentos_realizados():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        cargos = _rt_buscar_cargos(cursor)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_realizados(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        conn.close()

        return render_template(
            'relatorio_treinamentos_realizados.html',
            cargos=cargos,
            cargo_id=cargo_id,
            usuarios=usuarios,
            usuario_ids=usuario_ids,
            selecionar_todos=selecionar_todos,
            registros=registros,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

    @blueprint.route('/exportar_treinamentos_realizados_excel', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def exportar_treinamentos_realizados_excel():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_realizados(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        conn.close()

        arquivo = _rt_exportar_excel_relatorio_treinamentos(
            registros=registros,
            nome_aba="Treinamentos Realizados",
            headers=[
                "Matrícula",
                "Nome",
                "Cargo",
                "Sigla",
                "Nº Doc",
                "Título",
                "Revisão",
                "Data",
                "C.H.",
                "Validade",
                "Evidência"
            ],
            row_builder=lambda item: [
                item.get("matricula") or "",
                item.get("nome") or "",
                item.get("cargo_nome") or "",
                item.get("sigla") or "",
                item.get("numero_documento") or "",
                item.get("titulo") or "",
                item.get("revisao") or "",
                item.get("data_realizacao_fmt") or "",
                item.get("carga_horaria") or "",
                item.get("validade_fmt") or "",
                item.get("evidencia_url") or ""
            ],
            larguras={
                "A": 12,
                "B": 28,
                "C": 22,
                "D": 8,
                "E": 10,
                "F": 40,
                "G": 10,
                "H": 12,
                "I": 8,
                "J": 12,
                "K": 20
            },
            colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9, 10]
        )

        return send_file(
            arquivo,
            as_attachment=True,
            download_name="relatorio_treinamentos_realizados.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def _rt_buscar_treinamentos_a_vencer(cursor, usuario_ids=None, cargo_id=None, data_inicio=None, data_fim=None):

        if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
            return []

        if not _rt_table_exists(cursor, "treinamentos_realizados"):
            return []

        cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
        cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
        cols_usuarios = _rt_get_columns(cursor, "usuarios")
        cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

        usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id"])
        trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id"])
        trr_pk = _rt_first_existing(cols_trr, ["id"])

        if not usuario_col_trp or not trp_to_trr or not trr_pk:
            return []

        cfg_proc = _rt_resolver_campos_procedimento(cursor)

        cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
        join_cargos = ""
        if cols_cargos:
            cargo_nome_expr = "c.nome"
            join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

        hoje = date.today()
        limite = hoje + timedelta(days=30)

        filtros = [
            "u.ativo = 1",
            "trr.data_validade IS NOT NULL",
            "DATE(trr.data_validade) >= %s",
            "DATE(trr.data_validade) <= %s"
        ]

        params = [hoje, limite]

        if usuario_ids:
            placeholders = ", ".join(["%s"] * len(usuario_ids))
            filtros.append(f"u.id IN ({placeholders})")
            params.extend(usuario_ids)

        if cargo_id:
            filtros.append("u.cargo_id = %s")
            params.append(cargo_id)

        if data_inicio:
            filtros.append("DATE(trr.data_treinamento) >= %s")
            params.append(data_inicio)

        if data_fim:
            filtros.append("DATE(trr.data_treinamento) <= %s")
            params.append(data_fim)

        query = f"""
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                {cargo_nome_expr} AS cargo_nome,
                p.id AS procedimento_id,
                {cfg_proc['campo_sigla']} AS sigla,
                {cfg_proc['campo_numero']} AS numero_documento,
                {cfg_proc['campo_titulo']} AS titulo,
                trr.procedimento_revisao_id AS revisao,
                trr.data_treinamento AS data_realizacao,
                trr.data_validade AS validade
            FROM treinamentos_realizados_participantes trp
            INNER JOIN treinamentos_realizados trr
                ON trr.{trr_pk} = trp.{trp_to_trr}
            INNER JOIN usuarios u
                ON u.id = trp.{usuario_col_trp}
            {join_cargos}
            LEFT JOIN procedimentos p
                ON p.id = trr.procedimento_id
            {cfg_proc['join_tipo_documento']}
            WHERE {" AND ".join(filtros)}
            ORDER BY trr.data_validade ASC, u.nome
        """

        cursor.execute(query, tuple(params))
        dados = cursor.fetchall()

        for item in dados:
            validade = _rt_to_date(item.get("validade"))
            data_realizacao = _rt_to_date(item.get("data_realizacao"))

            item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""
            item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""

            if validade:
                dias = (validade - hoje).days
                item["dias_para_vencer"] = dias
            else:
                item["dias_para_vencer"] = None

        return dados

    @blueprint.route('/relatorio_treinamentos_a_vencer', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def relatorio_treinamentos_a_vencer():

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_ids_raw = request.args.getlist('usuario_ids')
        cargo_id = request.args.get('cargo_id') or None
        data_inicio = request.args.get('data_inicio') or None
        data_fim = request.args.get('data_fim') or None

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = [
                    int(u)
                    for u in usuario_ids_raw
                    if str(u).isdigit()
                ]

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_a_vencer(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        cursor.execute("""
            SELECT id, nome
            FROM cargos
            WHERE ativo = 1
            ORDER BY nome
        """)
        cargos = cursor.fetchall()

        conn.close()

        return render_template(
            'relatorio_treinamentos_a_vencer.html',
            registros=registros,
            usuarios=usuarios,
            cargos=cargos,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            selecionar_todos=selecionar_todos
        )

    @blueprint.route('/exportar_treinamentos_a_vencer_excel', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def exportar_treinamentos_a_vencer_excel():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_a_vencer(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id
        )

        conn.close()

        arquivo = _rt_exportar_excel_relatorio_treinamentos(
            registros=registros,
            nome_aba="Treinamentos a Vencer",
            headers=[
                "Matrícula",
                "Nome",
                "Cargo",
                "Sigla",
                "Nº Doc",
                "Título",
                "Revisão",
                "Data",
                "Validade",
                "Dias"
            ],
            row_builder=lambda item: [
                item.get("matricula") or "",
                item.get("nome") or "",
                item.get("cargo_nome") or "",
                item.get("sigla") or "",
                item.get("numero_documento") or "",
                item.get("titulo") or "",
                item.get("revisao") or "",
                item.get("data_realizacao_fmt") or "",
                item.get("validade_fmt") or "",
                item.get("dias_para_vencer") if item.get("dias_para_vencer") is not None else ""
            ],
            larguras={
                "A": 12,
                "B": 28,
                "C": 22,
                "D": 10,
                "E": 12,
                "F": 40,
                "G": 10,
                "H": 12,
                "I": 12,
                "J": 10
            },
            colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9]
        )

        return send_file(
            arquivo,
            as_attachment=True,
            download_name="relatorio_treinamentos_a_vencer.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def _rt_buscar_treinamentos_vencidos(cursor, usuario_ids=None, cargo_id=None):

        if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
            return []

        if not _rt_table_exists(cursor, "treinamentos_realizados"):
            return []

        cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
        cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
        cols_usuarios = _rt_get_columns(cursor, "usuarios")
        cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

        usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id"])
        trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id"])
        trr_pk = _rt_first_existing(cols_trr, ["id"])

        if not usuario_col_trp or not trp_to_trr or not trr_pk:
            return []

        cfg_proc = _rt_resolver_campos_procedimento(cursor)

        cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
        join_cargos = ""
        if cols_cargos:
            cargo_nome_expr = "c.nome"
            join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

        hoje = date.today()

        filtros = [
            "u.ativo = 1",
            "trr.data_validade IS NOT NULL",
            "DATE(trr.data_validade) < %s"
        ]

        params = [hoje]

        if usuario_ids:
            placeholders = ", ".join(["%s"] * len(usuario_ids))
            filtros.append(f"u.id IN ({placeholders})")
            params.extend(usuario_ids)

        if cargo_id:
            filtros.append("u.cargo_id = %s")
            params.append(cargo_id)

        query = f"""
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                {cargo_nome_expr} AS cargo_nome,
                p.id AS procedimento_id,
                {cfg_proc['campo_sigla']} AS sigla,
                {cfg_proc['campo_numero']} AS numero_documento,
                {cfg_proc['campo_titulo']} AS titulo,
                trr.procedimento_revisao_id AS revisao,
                trr.data_treinamento AS data_realizacao,
                trr.data_validade AS validade
            FROM treinamentos_realizados_participantes trp
            INNER JOIN treinamentos_realizados trr
                ON trr.{trr_pk} = trp.{trp_to_trr}
            INNER JOIN usuarios u
                ON u.id = trp.{usuario_col_trp}
            {join_cargos}
            LEFT JOIN procedimentos p
                ON p.id = trr.procedimento_id
            {cfg_proc['join_tipo_documento']}
            WHERE {" AND ".join(filtros)}
            ORDER BY trr.data_validade ASC, u.nome
        """

        cursor.execute(query, tuple(params))
        dados = cursor.fetchall()

        for item in dados:
            validade = _rt_to_date(item.get("validade"))
            data_realizacao = _rt_to_date(item.get("data_realizacao"))

            item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""
            item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""

            if validade:
                item["dias_vencido"] = (hoje - validade).days
            else:
                item["dias_vencido"] = None

        return dados

    @blueprint.route('/relatorio_treinamentos_vencidos', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def relatorio_treinamentos_vencidos():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_vencidos(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id
        )

        cursor.execute("""
            SELECT id, nome
            FROM cargos
            WHERE ativo = 1
            ORDER BY nome
        """)
        cargos = cursor.fetchall()

        conn.close()

        return render_template(
            'relatorio_treinamentos_vencidos.html',
            registros=registros,
            usuarios=usuarios,
            cargos=cargos,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            selecionar_todos=selecionar_todos
        )

    @blueprint.route('/exportar_treinamentos_vencidos_excel', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def exportar_treinamentos_vencidos_excel():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_vencidos(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id
        )

        conn.close()

        arquivo = _rt_exportar_excel_relatorio_treinamentos(
            registros=registros,
            nome_aba="Treinamentos Vencidos",
            headers=[
                "Matrícula",
                "Nome",
                "Cargo",
                "Sigla",
                "Nº Doc",
                "Título",
                "Revisão",
                "Data",
                "Validade",
                "Dias Vencido"
            ],
            row_builder=lambda item: [
                item.get("matricula") or "",
                item.get("nome") or "",
                item.get("cargo_nome") or "",
                item.get("sigla") or "",
                item.get("numero_documento") or "",
                item.get("titulo") or "",
                item.get("revisao") or "",
                item.get("data_realizacao_fmt") or "",
                item.get("validade_fmt") or "",
                item.get("dias_vencido") if item.get("dias_vencido") is not None else ""
            ],
            larguras={
                "A": 12,
                "B": 28,
                "C": 22,
                "D": 10,
                "E": 12,
                "F": 40,
                "G": 10,
                "H": 12,
                "I": 12,
                "J": 12
            },
            colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9]
        )

        return send_file(
            arquivo,
            as_attachment=True,
            download_name="relatorio_treinamentos_vencidos.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def _rt_buscar_treinamentos_pendentes(cursor, usuario_ids=None, cargo_id=None):
        hoje = date.today()

        filtros_usuarios = ["u.ativo = 1"]
        params = []

        if usuario_ids:
            placeholders = ", ".join(["%s"] * len(usuario_ids))
            filtros_usuarios.append(f"u.id IN ({placeholders})")
            params.extend(usuario_ids)

        if cargo_id:
            filtros_usuarios.append("u.cargo_id = %s")
            params.append(cargo_id)

        query = f"""
            SELECT DISTINCT
                base.usuario_id,
                base.matricula,
                base.nome,
                base.cargo_nome,
                base.procedimento_id,
                td.sigla,
                p.numero_documento,
                p.titulo,
                rev_vigente.numero_revisao AS revisao
            FROM (

                /* CAMADA CARGO */
                SELECT
                    u.id AS usuario_id,
                    u.matricula,
                    u.nome,
                    c.nome AS cargo_nome,
                    mcp.procedimento_id
                FROM usuarios u
                LEFT JOIN cargos c
                    ON c.id = u.cargo_id
                JOIN matriz_cargo_procedimentos mcp
                    ON mcp.cargo_id = u.cargo_id
                   AND mcp.ativo = 1
                   AND mcp.obrigatorio = 1
                WHERE {" AND ".join(filtros_usuarios)}

                UNION

                /* CAMADA FUNÇÃO */
                SELECT
                    u.id AS usuario_id,
                    u.matricula,
                    u.nome,
                    c.nome AS cargo_nome,
                    mfp.procedimento_id
                FROM usuarios u
                LEFT JOIN cargos c
                    ON c.id = u.cargo_id
                JOIN usuario_funcoes_setores ufs
                    ON ufs.usuario_id = u.id
                   AND ufs.ativo = 1
                JOIN matriz_funcao_procedimentos mfp
                    ON mfp.cargo_id = u.cargo_id
                   AND mfp.funcao_id = ufs.funcao_id
                   AND mfp.ativo = 1
                   AND mfp.obrigatorio = 1
                JOIN matriz_cargo_funcoes mcf
                    ON mcf.cargo_id = mfp.cargo_id
                   AND mcf.funcao_id = mfp.funcao_id
                   AND mcf.ativo = 1
                WHERE {" AND ".join(filtros_usuarios)}

                UNION

                /* CAMADA SETOR */
                SELECT
                    u.id AS usuario_id,
                    u.matricula,
                    u.nome,
                    c.nome AS cargo_nome,
                    msp.procedimento_id
                FROM usuarios u
                LEFT JOIN cargos c
                    ON c.id = u.cargo_id
                JOIN usuario_funcoes_setores ufs
                    ON ufs.usuario_id = u.id
                   AND ufs.ativo = 1
                JOIN matriz_setor_procedimentos msp
                    ON msp.cargo_id = u.cargo_id
                   AND msp.setor_id = ufs.setor_id
                   AND msp.ativo = 1
                   AND msp.obrigatorio = 1
                JOIN matriz_cargo_setores mcs
                    ON mcs.cargo_id = msp.cargo_id
                   AND mcs.setor_id = msp.setor_id
                   AND mcs.ativo = 1
                WHERE {" AND ".join(filtros_usuarios)}

            ) AS base

            JOIN procedimentos p
                ON p.id = base.procedimento_id
               AND p.ativo = 1

            JOIN tipos_documento td
                ON td.id = p.tipo_documento_id

            /* revisão vigente que exige treinamento */
            JOIN procedimento_revisoes rev_vigente
                ON rev_vigente.id = (
                    SELECT pr2.id
                    FROM procedimento_revisoes pr2
                    WHERE pr2.procedimento_id = p.id
                      AND pr2.vigente = 1
                      AND pr2.requer_treinamento = 1
                    ORDER BY pr2.data_revisao DESC, pr2.id DESC
                    LIMIT 1
                )

            /* excluir quem já tem treinamento válido da revisão vigente */
            WHERE NOT EXISTS (
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                JOIN treinamentos_realizados tr
                    ON tr.id = trp.treinamento_realizado_id
                WHERE trp.usuario_id = base.usuario_id
                  AND tr.procedimento_id = base.procedimento_id
                  AND tr.procedimento_revisao_id = rev_vigente.id
                  AND trp.presenca = 1
                  AND trp.aprovado = 1
                  AND tr.ativo = 1
                  AND tr.data_validade IS NOT NULL
                  AND DATE(tr.data_validade) >= %s
            )

            ORDER BY base.nome, td.sigla, p.numero_documento, p.titulo
        """

        params_query = params + params + params + [hoje]
        cursor.execute(query, tuple(params_query))
        return cursor.fetchall()

    @blueprint.route('/relatorio_treinamentos_pendentes', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def relatorio_treinamentos_pendentes():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_pendentes(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id
        )

        cursor.execute("""
            SELECT id, nome
            FROM cargos
            WHERE ativo = 1
            ORDER BY nome
        """)
        cargos = cursor.fetchall()

        conn.close()

        return render_template(
            'relatorio_treinamentos_pendentes.html',
            registros=registros,
            usuarios=usuarios,
            cargos=cargos,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id,
            selecionar_todos=selecionar_todos
        )

    @blueprint.route('/exportar_treinamentos_pendentes_excel', methods=['GET'])
    @login_required
    @module_required('acesso_treinamentos')
    def exportar_treinamentos_pendentes_excel():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cargo_id = request.args.get('cargo_id', type=int)
        usuario_ids_raw = request.args.getlist('usuario_ids')

        selecionar_todos = False
        usuario_ids = []

        if usuario_ids_raw:
            if "__TODOS__" in usuario_ids_raw:
                selecionar_todos = True
            else:
                usuario_ids = _rt_normalize_ids(usuario_ids_raw)

        usuarios = _rt_buscar_usuarios_permitidos(
            cursor,
            cargo_id=cargo_id,
        )
        usuario_ids = filtrar_usuario_ids_permitidos(
            usuario_ids_solicitados=usuario_ids,
            usuario_ids_permitidos=[u["id"] for u in usuarios],
            selecionar_todos=selecionar_todos,
        )

        registros = _rt_buscar_treinamentos_pendentes(
            cursor,
            usuario_ids=usuario_ids,
            cargo_id=cargo_id
        )

        conn.close()

        arquivo = _rt_exportar_excel_relatorio_treinamentos(
            registros=registros,
            nome_aba="Treinamentos Pendentes",
            headers=[
                "Matrícula",
                "Nome",
                "Cargo",
                "Sigla",
                "Nº Doc",
                "Título"
            ],
            row_builder=lambda item: [
                item.get("matricula") or "",
                item.get("nome") or "",
                item.get("cargo_nome") or "",
                item.get("sigla") or "",
                item.get("numero_documento") or "",
                item.get("titulo") or ""
            ],
            larguras={
                "A": 12,
                "B": 28,
                "C": 22,
                "D": 10,
                "E": 12,
                "F": 40
            },
            colunas_centralizadas=[0, 3, 4]
        )

        return send_file(
            arquivo,
            as_attachment=True,
            download_name="relatorio_treinamentos_pendentes.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
