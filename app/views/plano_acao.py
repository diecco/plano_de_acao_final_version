import os
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import current_app, flash, redirect, render_template, request, send_file, session, url_for
from flask_mail import Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.utils import secure_filename

from app import mail
from app.decorators import login_required, module_required, pode_acessar_acao
from app.upload_security import UploadService, UploadValidationError
from app.utils.db import get_db_connection


ALLOWED_EVIDENCE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'docx', 'xlsx'}


def allowed_evidence_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EVIDENCE_EXTENSIONS


def register_plano_acao_routes(blueprint):
    @blueprint.route('/dashboard')
    @login_required
    @module_required('acesso_plano_acao')
    def dashboard():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        from datetime import date, timedelta

        # Atualiza status para "Vencida" quando ultrapassar o prazo
        cursor.execute("""
            UPDATE acoes
            SET status = 'Vencida'
            WHERE status NOT IN ('Concluída', 'Cancelada')
              AND prazo < %s
              AND ativo = 1
        """, (date.today(),))
        conn.commit()

        usuario_id = session['usuario_id']
        perfil = session.get('perfil')

        # =========================
        # PERSISTÊNCIA DOS FILTROS
        # =========================
        if request.args.get('limpar') == '1':
            session.pop('dashboard_filtros', None)
            conn.close()
            return redirect(url_for('main.dashboard'))

        if request.args:
            filtros_sessao = {
                'responsavel': request.args.get('responsavel', ''),
                'superintendencia': request.args.get('superintendencia', ''),
                'centro_custos': request.args.get('centro_custos', ''),
                'origem': request.args.get('origem', ''),
                'status': request.args.get('status', ''),
                'data_inicio': request.args.get('data_inicio', ''),
                'data_fim': request.args.get('data_fim', '')
            }
            session['dashboard_filtros'] = filtros_sessao
        else:
            filtros_sessao = session.get('dashboard_filtros', {
                'responsavel': '',
                'superintendencia': '',
                'centro_custos': '',
                'origem': '',
                'status': '',
                'data_inicio': '',
                'data_fim': ''
            })

        responsavel_id = filtros_sessao.get('responsavel')
        superintendencia_id = filtros_sessao.get('superintendencia')
        centro_custos_id = filtros_sessao.get('centro_custos')
        origem_id = filtros_sessao.get('origem')
        status = filtros_sessao.get('status')
        data_inicio = filtros_sessao.get('data_inicio')
        data_fim = filtros_sessao.get('data_fim')

        # =========================
        # MONTAGEM DOS FILTROS
        # =========================
        filtros = []
        valores = []

        filtros.append("a.ativo = 1")

        if responsavel_id:
            filtros.append("a.responsavel_id = %s")
            valores.append(responsavel_id)

        if superintendencia_id:
            filtros.append("u.superintendencia_id = %s")
            valores.append(superintendencia_id)

        if centro_custos_id:
            filtros.append("u.centro_custos_id = %s")
            valores.append(centro_custos_id)

        if origem_id:
            filtros.append("a.origem_id = %s")
            valores.append(origem_id)

        if status:
            filtros.append("a.status = %s")
            valores.append(status)

        if data_inicio:
            filtros.append("a.prazo >= %s")
            valores.append(data_inicio)

        if data_fim:
            filtros.append("a.prazo <= %s")
            valores.append(data_fim)

        # =========================
        # CONTROLE POR PERFIL (CORRIGIDO)
        # =========================
        if perfil == 'basico':
            filtros.append("a.responsavel_id = %s")
            valores.append(usuario_id)

        elif perfil == 'intermediario':
            filtros.append("u.centro_custos_id = %s")
            valores.append(session.get('centro_custos_id'))

        elif perfil == 'avancado':
            pass

        elif perfil == 'administrador':
            pass

        where_clause = " AND ".join(filtros)
        if where_clause:
            where_clause = "WHERE " + where_clause

        hoje = date.today()
        sete_dias = hoje + timedelta(days=7)
        primeiro_dia_mes = hoje.replace(day=1)

        # =========================
        # KPIs
        # =========================
        query_kpis = f"""
            SELECT
                SUM(CASE WHEN a.status NOT IN ('Concluída', 'Cancelada') THEN 1 ELSE 0 END) AS kpi_ativas,
                SUM(CASE WHEN a.status = 'Vencida' THEN 1 ELSE 0 END) AS kpi_vencidas,
                SUM(CASE
                        WHEN a.status NOT IN ('Concluída', 'Cancelada', 'Vencida')
                         AND a.prazo BETWEEN %s AND %s
                        THEN 1 ELSE 0
                    END) AS kpi_a_vencer_7_dias,
                SUM(CASE
                        WHEN a.status = 'Concluída'
                         AND a.prazo >= %s
                        THEN 1 ELSE 0
                    END) AS kpi_concluidas_mes,
                0 AS kpi_criadas_mes
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            {where_clause}
        """
        cursor.execute(query_kpis, [hoje, sete_dias, primeiro_dia_mes] + valores)
        kpis = cursor.fetchone() or {}

        kpi_ativas = kpis.get('kpi_ativas') or 0
        kpi_vencidas = kpis.get('kpi_vencidas') or 0
        kpi_a_vencer_7_dias = kpis.get('kpi_a_vencer_7_dias') or 0
        kpi_concluidas_mes = kpis.get('kpi_concluidas_mes') or 0
        kpi_criadas_mes = kpis.get('kpi_criadas_mes') or 0

        # =========================
        # GRÁFICO POR STATUS
        # =========================
        query_status = f"""
            SELECT a.status, COUNT(*) AS total
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            {where_clause}
            GROUP BY a.status
        """
        cursor.execute(query_status, valores)
        status_data = cursor.fetchall()
        labels_status = [d['status'] for d in status_data]
        dados_status = [d['total'] for d in status_data]

        # =========================
        # GRÁFICO POR ORIGEM
        # =========================
        query_origem = f"""
            SELECT o.descricao AS origem, COUNT(*) AS total
            FROM acoes a
            JOIN origens o ON a.origem_id = o.id
            JOIN usuarios u ON a.responsavel_id = u.id
            {where_clause}
            GROUP BY o.descricao
        """
        cursor.execute(query_origem, valores)
        origem_data = cursor.fetchall()
        labels_origem = [d['origem'] for d in origem_data]
        dados_origem = [d['total'] for d in origem_data]

        # =========================
        # ATENÇÃO IMEDIATA - VENCIDAS
        # =========================
        query_vencidas_criticas = f"""
            SELECT
                a.id,
                a.descricao,
                a.prazo,
                o.descricao AS descricao_origem,
                DATEDIFF(%s, a.prazo) AS dias_atraso
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN origens o ON a.origem_id = o.id
            {where_clause}
              AND a.status = 'Vencida'
            ORDER BY a.prazo ASC
            LIMIT 5
        """
        cursor.execute(query_vencidas_criticas, [hoje] + valores)
        acoes_vencidas_criticas = cursor.fetchall()

        # =========================
        # ATENÇÃO IMEDIATA - PRÓXIMAS
        # =========================
        query_proximas = f"""
            SELECT
                a.id,
                a.descricao,
                a.prazo,
                o.descricao AS descricao_origem,
                DATEDIFF(a.prazo, %s) AS dias_restantes
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN origens o ON a.origem_id = o.id
            {where_clause}
              AND a.status NOT IN ('Concluída', 'Cancelada', 'Vencida')
              AND a.prazo >= %s
            ORDER BY a.prazo ASC
            LIMIT 5
        """
        cursor.execute(query_proximas, [hoje] + valores + [hoje])
        acoes_proximas_vencimento = cursor.fetchall()

        conn.close()

        return render_template(
            'dashboard.html',
            kpi_ativas=kpi_ativas,
            kpi_vencidas=kpi_vencidas,
            kpi_a_vencer_7_dias=kpi_a_vencer_7_dias,
            kpi_concluidas_mes=kpi_concluidas_mes,
            kpi_criadas_mes=kpi_criadas_mes,
            labels_status=labels_status,
            dados_status=dados_status,
            labels_origem=labels_origem,
            dados_origem=dados_origem,
            acoes_vencidas_criticas=acoes_vencidas_criticas,
            acoes_proximas_vencimento=acoes_proximas_vencimento,
            filtros=filtros_sessao
        )


    @blueprint.route('/cadastrar_acao', methods=['GET', 'POST'])
    @login_required
    @module_required('acesso_plano_acao')
    def cadastrar_acao():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')

        usuario_logado = {
            'id': session.get('usuario_id'),
            'superintendencia_id': session.get('superintendencia_id'),
            'centro_custos_id': session.get('centro_custos_id')
        }

        if not usuario_logado.get('id'):
            conn.close()
            flash('Usuário logado não encontrado.', 'danger')
            return redirect('/login')

        if request.method == 'POST':
            origem_id = request.form['origem_id']
            responsavel_id = request.form['responsavel_id']
            descricao = request.form['descricao']
            prazo = request.form['prazo']
            status = request.form['status']

            criado_por = usuario_id

            # Validação da origem conforme perfil
            if perfil == 'administrador':
                cursor.execute("""
                    SELECT id
                    FROM origens
                    WHERE id = %s
                      AND ativo = 1
                """, (origem_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM origens
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (origem_id, usuario_logado['centro_custos_id']))

            origem_valida = cursor.fetchone()

            if not origem_valida:
                conn.close()
                flash('Origem inválida para o seu perfil de acesso.', 'danger')
                return redirect('/cadastrar_acao')

            # Validação do responsável conforme perfil
            if perfil == 'basico':
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                      AND centro_custos_id = %s
                """, (responsavel_id, usuario_logado['centro_custos_id']))

            elif perfil == 'intermediario':
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                      AND centro_custos_id = %s
                """, (responsavel_id, usuario_logado['centro_custos_id']))

            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                """, (responsavel_id,))

            responsavel_valido = cursor.fetchone()

            if not responsavel_valido:
                conn.close()
                flash('Responsável inválido para o seu perfil de acesso.', 'danger')
                return redirect('/cadastrar_acao')

            cursor.execute("""
                INSERT INTO acoes (origem_id, responsavel_id, descricao, prazo, status, criado_por)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (origem_id, responsavel_id, descricao, prazo, status, criado_por))

            conn.commit()
            acao_id = cursor.lastrowid

            cursor.execute("""
                SELECT nome, email
                FROM usuarios
                WHERE id = %s
            """, (responsavel_id,))
            responsavel = cursor.fetchone()

            prazo_formatado = datetime.strptime(prazo, '%Y-%m-%d').strftime('%d/%m/%Y')
            link_edicao = url_for('main.editar_acao', id=acao_id, _external=True)

            try:
                if responsavel and responsavel.get('email'):
                    msg = Message(
                        subject="Nova Ação Atribuída a Você - TrackPlan",
                        recipients=[responsavel['email']]
                    )
                    msg.html = f"""
                    <div style="font-family: Arial, sans-serif; font-size: 15px;">
                        <div style="text-align: center;">
                            <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height: 50px; margin-bottom: 20px;">
                        </div>
                        <p>Olá <strong>{responsavel['nome']}</strong>,</p>

                        <p>Uma nova ação foi atribuída a você no sistema TrackPlan.</p>

                        <p><strong>Descrição:</strong> {descricao}<br>
                        <strong>Prazo:</strong> {prazo_formatado}<br>
                        <strong>Status:</strong> {status}</p>

                        <p>Acesse o sistema para mais detalhes:</p>

                        <p>
                            <a href="{link_edicao}" style="display: inline-block; background-color: #ea6a23; color: white; padding: 10px 18px; text-decoration: none; border-radius: 5px;">
                                Editar Ação
                            </a>
                        </p>

                        <br>
                        <p style="font-size: 13px; color: #666;">Equipe TrackPlan</p>
                    </div>
                    """
                    mail.send(msg)

            except Exception:
                current_app.logger.exception(
                    "Falha ao enviar notificação da nova ação %s.",
                    acao_id,
                )
                flash('Ação criada. Falha ao enviar o e-mail de notificação.', 'warning')

            conn.close()
            flash('Ação cadastrada com sucesso!', 'success')
            return redirect('/dashboard')

        # GET: carregar origens conforme perfil
        if perfil == 'administrador':
            cursor.execute("""
                SELECT id, descricao
                FROM origens
                WHERE ativo = 1
                ORDER BY descricao
            """)
        else:
            cursor.execute("""
                SELECT id, descricao
                FROM origens
                WHERE ativo = 1
                  AND centro_custos_id = %s
                ORDER BY descricao
            """, (usuario_logado['centro_custos_id'],))

        origens = cursor.fetchall()

        # GET: carregar responsáveis conforme perfil
        if perfil == 'basico':
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (usuario_logado['centro_custos_id'],))

        elif perfil == 'intermediario':
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (usuario_logado['centro_custos_id'],))

        else:
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                ORDER BY nome
            """)

        usuarios = cursor.fetchall()
        conn.close()

        return render_template(
            'cadastrar_acao.html',
            origens=origens,
            usuarios=usuarios
        )


    @blueprint.route('/editar_acao/<int:id>', methods=['GET', 'POST'])
    @login_required
    @module_required('acesso_plano_acao')
    def editar_acao(id):
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        acao = pode_acessar_acao(cursor, id)

        if not acao:
            conn.close()
            flash('Ação não encontrada.', 'warning')
            return redirect(url_for('main.dashboard'))

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')
        pode_editar_completo = acao.get('criado_por') == usuario_id

        cursor.execute("""
            SELECT id, nome, superintendencia_id, centro_custos_id
            FROM usuarios
            WHERE id = %s
        """, (usuario_id,))
        usuario_logado = cursor.fetchone()

        if not usuario_logado:
            conn.close()
            flash('Usuário logado não encontrado.', 'danger')
            return redirect('/login')

        if request.method == 'POST':
            next_url = (request.form.get('next') or '').strip()

            responsavel_antigo = acao.get('responsavel_id')
            prazo_antigo = acao.get('prazo')
            status_antigo = acao.get('status')
            descricao_antiga = (acao.get('descricao') or '').strip()

            if pode_editar_completo:
                origem_id = request.form.get('origem_id')
                responsavel_id = request.form.get('responsavel_id')
                descricao = request.form.get('descricao')
                prazo = request.form.get('prazo')

                if perfil == 'administrador':
                    cursor.execute("""
                        SELECT id
                        FROM origens
                        WHERE id = %s
                          AND ativo = 1
                    """, (origem_id,))
                else:
                    cursor.execute("""
                        SELECT id
                        FROM origens
                        WHERE id = %s
                          AND ativo = 1
                          AND centro_custos_id = %s
                    """, (origem_id, usuario_logado['centro_custos_id']))

                origem_valida = cursor.fetchone()

                if not origem_valida:
                    conn.close()
                    flash('Origem inválida para o seu perfil de acesso.', 'danger')
                    return redirect(next_url or url_for('main.editar_acao', id=id))

                if perfil == 'basico':
                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = TRUE
                          AND centro_custos_id = %s
                    """, (responsavel_id, usuario_logado['centro_custos_id']))

                elif perfil == 'intermediario':
                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = TRUE
                          AND superintendencia_id = %s
                    """, (responsavel_id, usuario_logado['superintendencia_id']))

                else:
                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = TRUE
                    """, (responsavel_id,))

                responsavel_valido = cursor.fetchone()

                if not responsavel_valido:
                    conn.close()
                    flash('Responsável inválido para o seu perfil de acesso.', 'danger')
                    return redirect(next_url or url_for('main.editar_acao', id=id))

            else:
                origem_id = acao['origem_id']
                responsavel_id = acao['responsavel_id']
                descricao = acao['descricao']
                prazo = acao['prazo']

            status = request.form.get('status')
            observacoes = (request.form.get('observacoes') or '').strip()
            data_conclusao_str = (request.form.get('data_conclusao') or '').strip()

            data_conclusao = None
            hoje = date.today()

            if data_conclusao_str:
                try:
                    data_conclusao = datetime.strptime(data_conclusao_str, '%Y-%m-%d').date()
                except ValueError:
                    flash('Data de conclusão inválida.', 'warning')
                    conn.close()
                    return redirect(next_url or url_for('main.editar_acao', id=id))

                if data_conclusao > hoje:
                    flash('A data de conclusão não pode ser superior à data de hoje.', 'warning')
                    conn.close()
                    return redirect(next_url or url_for('main.editar_acao', id=id))

                status = 'Concluída'
            else:
                if status == 'Concluída':
                    flash('Para concluir a ação, preencha a data de conclusão.', 'warning')
                    conn.close()
                    return redirect(next_url or url_for('main.editar_acao', id=id))

            if status == 'Cancelada' and not observacoes:
                flash('Ao cancelar a ação, o campo Observações é obrigatório.', 'warning')
                conn.close()
                return redirect(next_url or url_for('main.editar_acao', id=id))

            arquivo_evidencia = acao.get('arquivo_evidencia')
            arquivo = request.files.get('arquivo_evidencia')

            if arquivo and arquivo.filename:
                extensoes_permitidas = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx'}
                nome_original = secure_filename(arquivo.filename)
                extensao = nome_original.rsplit('.', 1)[-1].lower() if '.' in nome_original else ''

                if extensao not in extensoes_permitidas:
                    conn.close()
                    flash('Tipo de arquivo não permitido para evidência.', 'warning')
                    return redirect(next_url or url_for('main.editar_acao', id=id))

                try:
                    nome_arquivo = UploadService.salvar(
                        arquivo,
                        extensoes_permitidas,
                        prefixo=f"evidencia_acao_{id}",
                        diretorio=os.path.join('static', 'evidencias'),
                    )
                except UploadValidationError as exc:
                    conn.close()
                    flash(str(exc), 'warning')
                    return redirect(next_url or url_for('main.editar_acao', id=id))

                arquivo_evidencia = nome_arquivo

            cursor.execute("""
                UPDATE acoes
                SET origem_id = %s,
                    responsavel_id = %s,
                    descricao = %s,
                    prazo = %s,
                    status = %s,
                    observacoes = %s,
                    data_conclusao = %s,
                    arquivo_evidencia = %s
                WHERE id = %s
            """, (
                origem_id,
                responsavel_id,
                descricao,
                prazo,
                status,
                observacoes if observacoes else None,
                data_conclusao,
                arquivo_evidencia,
                id
            ))

            conn.commit()

            houve_mudanca_importante = (
                str(responsavel_id) != str(responsavel_antigo) or
                str(prazo) != str(prazo_antigo) or
                str(status) != str(status_antigo) or
                (descricao or '').strip() != descricao_antiga
            )

            if houve_mudanca_importante:
                try:
                    cursor.execute("SELECT nome, email FROM usuarios WHERE id = %s", (responsavel_id,))
                    responsavel_atual = cursor.fetchone()

                    if responsavel_atual and responsavel_atual.get('email'):
                        alteracoes = []

                        if str(responsavel_id) != str(responsavel_antigo):
                            if responsavel_antigo:
                                cursor.execute("SELECT nome FROM usuarios WHERE id = %s", (responsavel_antigo,))
                                resp_antigo = cursor.fetchone()
                                nome_resp_antigo = resp_antigo['nome'] if resp_antigo else 'Não identificado'
                            else:
                                nome_resp_antigo = 'Não identificado'

                            alteracoes.append(
                                f"<li><strong>Responsável:</strong> {nome_resp_antigo} → {responsavel_atual['nome']}</li>"
                            )

                        if str(prazo) != str(prazo_antigo):
                            prazo_antigo_fmt = prazo_antigo.strftime('%d/%m/%Y') if hasattr(prazo_antigo, 'strftime') else str(prazo_antigo or '-')
                            try:
                                prazo_novo_fmt = datetime.strptime(str(prazo), '%Y-%m-%d').strftime('%d/%m/%Y')
                            except Exception:
                                prazo_novo_fmt = str(prazo or '-')

                            alteracoes.append(
                                f"<li><strong>Prazo:</strong> {prazo_antigo_fmt} → {prazo_novo_fmt}</li>"
                            )

                        if str(status) != str(status_antigo):
                            alteracoes.append(
                                f"<li><strong>Status:</strong> {status_antigo or '-'} → {status or '-'}</li>"
                            )

                        if (descricao or '').strip() != descricao_antiga:
                            alteracoes.append(
                                f"<li><strong>Descrição da ação:</strong> foi atualizada.</li>"
                            )

                        link_edicao = url_for('main.editar_acao', id=id, _external=True)

                        msg = Message(
                            subject="Atualização em Ação sob sua Responsabilidade - TrackPlan",
                            recipients=[responsavel_atual['email']]
                        )

                        msg.html = f"""
                        <div style="font-family: Arial, sans-serif; font-size: 15px; color:#333;">
                            <div style="text-align:center; margin-bottom:18px;">
                                <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height:50px;">
                            </div>

                            <p>Olá <strong>{responsavel_atual['nome']}</strong>,</p>

                            <p>Uma ação sob sua responsabilidade foi atualizada no <strong>TrackPlan</strong>.</p>

                            <div style="background:#f7f7f7; border:1px solid #eee; border-radius:6px; padding:12px 14px; margin:12px 0;">
                                <p style="margin:6px 0;"><strong>ID da ação:</strong> {id}</p>
                                <p style="margin:6px 0;"><strong>Descrição atual:</strong> {descricao}</p>
                            </div>

                            <p><strong>Alterações realizadas:</strong></p>
                            <ul>
                                {''.join(alteracoes)}
                            </ul>

                            <p style="margin:22px 0; text-align:center;">
                                <a href="{link_edicao}" target="_blank"
                                   style="display:inline-block; background:#ea6a23; color:#fff; text-decoration:none; padding:10px 18px; border-radius:5px;">
                                    Acessar ação
                                </a>
                            </p>

                            <p style="font-size:13px; color:#666;">Equipe TrackPlan</p>
                        </div>
                        """

                        mail.send(msg)

                except Exception:
                    current_app.logger.exception(
                        "Falha ao enviar notificação da atualização da ação %s.",
                        id,
                    )
                    flash('Ação atualizada, mas não foi possível enviar o e-mail de notificação.', 'warning')

            conn.close()

            flash('Ação atualizada com sucesso!', 'success')

            if next_url:
                return redirect(next_url)

            return redirect(url_for('main.dashboard'))

        if perfil == 'administrador':
            cursor.execute("""
                SELECT id, descricao
                FROM origens
                WHERE ativo = 1
                ORDER BY descricao
            """)
        else:
            cursor.execute("""
                SELECT id, descricao
                FROM origens
                WHERE ativo = 1
                  AND centro_custos_id = %s
                ORDER BY descricao
            """, (usuario_logado['centro_custos_id'],))

        origens = cursor.fetchall()

        if perfil == 'basico':
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (usuario_logado['centro_custos_id'],))
        elif perfil == 'intermediario':
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                  AND superintendencia_id = %s
                ORDER BY nome
            """, (usuario_logado['superintendencia_id'],))
        else:
            cursor.execute("""
                SELECT id, nome, matricula
                FROM usuarios
                WHERE ativo = TRUE
                ORDER BY nome
            """)

        usuarios = cursor.fetchall()

        conn.close()

        return render_template(
            'editar_acao.html',
            acao=acao,
            origens=origens,
            usuarios=usuarios,
            pode_editar_completo=pode_editar_completo
        )


    @blueprint.route('/excluir_acao/<int:id>')
    @login_required
    @module_required('acesso_plano_acao')
    def excluir_acao(id):
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        acao = pode_acessar_acao(cursor, id)

        if not acao:
            conn.close()
            flash('Ação não encontrada ou você não possui permissão para excluí-la.', 'warning')
            return redirect(url_for('main.dashboard'))

        # 🔒 Regra adicional (RECOMENDADO)
        perfil = session.get('perfil')
        usuario_id = session.get('usuario_id')

        if perfil == 'basico':
            if acao.get('criado_por') != usuario_id:
                conn.close()
                flash('Você não tem permissão para excluir esta ação.', 'danger')
                return redirect(url_for('main.dashboard'))

        # 🔥 Exclusão lógica
        cursor.execute("""
            UPDATE acoes
            SET ativo = 0
            WHERE id = %s
        """, (id,))

        conn.commit()
        conn.close()

        flash('Ação excluída com sucesso!', 'success')
        return redirect(url_for('main.dashboard'))


    @blueprint.route('/anexar_evidencia/<int:acao_id>', methods=['GET', 'POST'])
    @login_required
    @module_required('acesso_plano_acao')
    def anexar_evidencia(acao_id):
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # 🔒 Validação de permissão sobre a ação
        acao = pode_acessar_acao(cursor, acao_id)

        if not acao:
            flash("Ação não encontrada ou você não possui permissão para acessá-la.", "warning")
            conn.close()
            return redirect(url_for('main.dashboard'))

        if request.method == 'POST':
            if 'arquivo' not in request.files:
                flash("Nenhum arquivo enviado.", "danger")
                conn.close()
                return redirect(request.url)

            arquivo = request.files['arquivo']

            if arquivo.filename == '':
                flash("Nenhum arquivo selecionado.", "warning")
                conn.close()
                return redirect(request.url)

            if arquivo and allowed_evidence_file(arquivo.filename):
                try:
                    filename = UploadService.salvar(
                        arquivo,
                        ALLOWED_EVIDENCE_EXTENSIONS,
                        prefixo=f"evidencia_acao_{acao_id}",
                    )
                except UploadValidationError as exc:
                    flash(str(exc), "danger")
                    conn.close()
                    return redirect(request.url)

                cursor.execute("""
                    UPDATE acoes
                    SET arquivo_evidencia = %s
                    WHERE id = %s
                """, (filename, acao_id))

                conn.commit()
                conn.close()

                flash("Evidência anexada com sucesso!", "success")
                return redirect(url_for('main.dashboard'))

            flash("Tipo de arquivo não permitido.", "danger")
            conn.close()
            return redirect(request.url)

        conn.close()
        return render_template('anexar_evidencia.html', acao=acao)


    @blueprint.route('/minhas_acoes')
    @login_required
    @module_required('acesso_plano_acao')
    def minhas_acoes():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')
        centro_custos_id = session.get('centro_custos_id')

        if request.args.get('limpar') == '1':
            session.pop('minhas_acoes_filtros', None)
            conn.close()
            return redirect(url_for('main.minhas_acoes'))

        filtros_padrao = {
            'responsavel_id': '',
            'origem_id': '',
            'status': '',
            'data_inicio': '',
            'data_fim': '',
            'sort': 'prazo',
            'order': 'asc'
        }

        filtros_sessao = session.get('minhas_acoes_filtros', filtros_padrao.copy())

        if request.args:
            for campo in filtros_padrao.keys():
                if campo in request.args:
                    filtros_sessao[campo] = request.args.get(campo, '')

            filtros_sessao['sort'] = filtros_sessao.get('sort') or 'prazo'
            filtros_sessao['order'] = filtros_sessao.get('order') or 'asc'

            session['minhas_acoes_filtros'] = filtros_sessao

        responsavel_id = filtros_sessao.get('responsavel_id', '')
        origem_id = filtros_sessao.get('origem_id', '')
        status = filtros_sessao.get('status', '')
        data_inicio = filtros_sessao.get('data_inicio', '')
        data_fim = filtros_sessao.get('data_fim', '')
        sort = filtros_sessao.get('sort', 'prazo')
        order = filtros_sessao.get('order', 'asc').lower()

        page = request.args.get('page', 1, type=int)
        per_page = 30

        if page < 1:
            page = 1

        offset = (page - 1) * per_page

        cursor.execute("""
            UPDATE acoes
            SET status = 'Vencida'
            WHERE status NOT IN ('Concluída', 'Cancelada')
              AND prazo < %s
              AND ativo = 1
        """, (date.today(),))
        conn.commit()

        def montar_filtros_sql(ignorar=None):
            if ignorar is None:
                ignorar = []

            filtros = ["a.ativo = 1"]
            valores = []

            if perfil == 'basico':
                filtros.append("a.responsavel_id = %s")
                valores.append(usuario_id)
            elif perfil == 'intermediario':
                filtros.append("ur.centro_custos_id = %s")
                valores.append(centro_custos_id)

            if (
                'responsavel_id' not in ignorar
                and responsavel_id
                and perfil != 'basico'
            ):
                filtros.append("a.responsavel_id = %s")
                valores.append(responsavel_id)

            if 'origem_id' not in ignorar and origem_id:
                filtros.append("a.origem_id = %s")
                valores.append(origem_id)

            if 'status' not in ignorar and status:
                filtros.append("a.status = %s")
                valores.append(status)

            if 'data_inicio' not in ignorar and data_inicio:
                filtros.append("a.prazo >= %s")
                valores.append(data_inicio)

            if 'data_fim' not in ignorar and data_fim:
                filtros.append("a.prazo <= %s")
                valores.append(data_fim)

            return filtros, valores

        filtros_sql, valores = montar_filtros_sql()
        where_clause = "WHERE " + " AND ".join(filtros_sql)

        campos_ordenacao = {
            'id': 'a.id',
            'descricao': 'a.descricao',
            'nome_responsavel': 'ur.nome',
            'nome_criador': 'uc.nome',
            'descricao_origem': 'o.descricao',
            'prazo': 'a.prazo',
            'status': 'a.status'
        }

        campo_sort = campos_ordenacao.get(sort, 'a.prazo')
        direcao = 'DESC' if order == 'desc' else 'ASC'

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM acoes a
            JOIN origens o ON a.origem_id = o.id
            JOIN usuarios ur ON a.responsavel_id = ur.id
            LEFT JOIN usuarios uc ON a.criado_por = uc.id
            {where_clause}
        """, valores)
        total_registros = cursor.fetchone()['total']

        total_paginas = (total_registros + per_page - 1) // per_page

        if total_paginas > 0 and page > total_paginas:
            page = total_paginas
            offset = (page - 1) * per_page

        query = f"""
            SELECT
                a.*,
                o.descricao AS descricao_origem,
                ur.nome AS nome_responsavel,
                uc.nome AS nome_criador
            FROM acoes a
            JOIN origens o ON a.origem_id = o.id
            JOIN usuarios ur ON a.responsavel_id = ur.id
            LEFT JOIN usuarios uc ON a.criado_por = uc.id
            {where_clause}
            ORDER BY {campo_sort} {direcao}, a.id DESC
            LIMIT %s OFFSET %s
        """

        cursor.execute(query, valores + [per_page, offset])
        acoes = cursor.fetchall()

        filtros_origem, valores_origem = montar_filtros_sql(ignorar=['origem_id'])
        where_origem = "WHERE " + " AND ".join(filtros_origem)

        cursor.execute(f"""
            SELECT DISTINCT o.id, o.descricao
            FROM acoes a
            JOIN origens o ON a.origem_id = o.id
            JOIN usuarios ur ON a.responsavel_id = ur.id
            {where_origem}
            ORDER BY o.descricao
        """, valores_origem)
        origens = cursor.fetchall()

        filtros_status, valores_status = montar_filtros_sql(ignorar=['status'])
        where_status = "WHERE " + " AND ".join(filtros_status)

        cursor.execute(f"""
            SELECT DISTINCT a.status
            FROM acoes a
            JOIN usuarios ur ON a.responsavel_id = ur.id
            {where_status}
            ORDER BY FIELD(a.status,
                'Não iniciada',
                'Em andamento',
                'Vencida',
                'Concluída',
                'Cancelada'
            ), a.status
        """, valores_status)
        status_opcoes = cursor.fetchall()

        if perfil in ['administrador', 'avancado']:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = 1
                ORDER BY nome
            """)
        else:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = 1
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (centro_custos_id,))

        usuarios = cursor.fetchall()

        conn.close()

        return render_template(
            'minhas_acoes.html',
            acoes=acoes,
            origens=origens,
            status_opcoes=status_opcoes,
            usuarios=usuarios,
            filtros=filtros_sessao,
            page=page,
            per_page=per_page,
            total_registros=total_registros,
            total_paginas=total_paginas
        )

    @blueprint.route('/acoes_criadas')
    @login_required
    @module_required('acesso_plano_acao')
    def acoes_criadas():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')
        centro_custos_usuario_id = session.get('centro_custos_id')

        if request.args.get('limpar') == '1':
            session.pop('acoes_criadas_filtros', None)
            conn.close()
            return redirect(url_for('main.acoes_criadas'))

        filtros_padrao = {
            'superintendencia_id': '',
            'centro_custos_id': '',
            'origem_id': '',
            'status': '',
            'data_inicio': '',
            'data_fim': '',
            'sort': 'prazo',
            'order': 'asc'
        }

        filtros_sessao = session.get('acoes_criadas_filtros', filtros_padrao.copy())

        if request.args:
            for campo in filtros_padrao.keys():
                if campo in request.args:
                    filtros_sessao[campo] = request.args.get(campo, '')

            filtros_sessao['sort'] = filtros_sessao.get('sort') or 'prazo'
            filtros_sessao['order'] = filtros_sessao.get('order') or 'asc'

            session['acoes_criadas_filtros'] = filtros_sessao

        superintendencia_id = filtros_sessao.get('superintendencia_id', '')
        centro_custos_id = filtros_sessao.get('centro_custos_id', '')
        origem_id = filtros_sessao.get('origem_id', '')
        status = filtros_sessao.get('status', '')
        data_inicio = filtros_sessao.get('data_inicio', '')
        data_fim = filtros_sessao.get('data_fim', '')
        sort = filtros_sessao.get('sort', 'prazo')
        order = filtros_sessao.get('order', 'asc').lower()

        page = request.args.get('page', 1, type=int)
        per_page = 30

        if page < 1:
            page = 1

        offset = (page - 1) * per_page

        cursor.execute("""
            UPDATE acoes
            SET status = 'Vencida'
            WHERE status NOT IN ('Concluída', 'Cancelada')
              AND prazo < %s
              AND ativo = 1
        """, (date.today(),))
        conn.commit()

        def montar_filtros_sql(ignorar=None):
            if ignorar is None:
                ignorar = []

            filtros = ["a.ativo = 1", "a.criado_por = %s"]
            valores = [usuario_id]

            if 'superintendencia_id' not in ignorar and superintendencia_id:
                filtros.append("u.superintendencia_id = %s")
                valores.append(superintendencia_id)

            if 'centro_custos_id' not in ignorar and centro_custos_id:
                filtros.append("u.centro_custos_id = %s")
                valores.append(centro_custos_id)

            if 'origem_id' not in ignorar and origem_id:
                filtros.append("a.origem_id = %s")
                valores.append(origem_id)

            if 'status' not in ignorar and status:
                filtros.append("a.status = %s")
                valores.append(status)

            if 'data_inicio' not in ignorar and data_inicio:
                filtros.append("a.prazo >= %s")
                valores.append(data_inicio)

            if 'data_fim' not in ignorar and data_fim:
                filtros.append("a.prazo <= %s")
                valores.append(data_fim)

            return filtros, valores

        filtros_sql, valores = montar_filtros_sql()
        where_clause = "WHERE " + " AND ".join(filtros_sql)

        campos_ordenacao = {
            'id': 'a.id',
            'descricao': 'a.descricao',
            'nome_responsavel': 'u.nome',
            'descricao_origem': 'o.descricao',
            'prazo': 'a.prazo',
            'status': 'a.status'
        }

        campo_sort = campos_ordenacao.get(sort, 'a.prazo')
        direcao = 'DESC' if order == 'desc' else 'ASC'

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN origens o ON a.origem_id = o.id
            {where_clause}
        """, valores)
        total_registros = cursor.fetchone()['total']

        total_paginas = (total_registros + per_page - 1) // per_page

        if total_paginas > 0 and page > total_paginas:
            page = total_paginas
            offset = (page - 1) * per_page

        query = f"""
            SELECT
                a.*,
                u.nome AS nome_responsavel,
                o.descricao AS descricao_origem
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN origens o ON a.origem_id = o.id
            {where_clause}
            ORDER BY {campo_sort} {direcao}, a.id DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query, valores + [per_page, offset])
        acoes = cursor.fetchall()

        filtros_sup, valores_sup = montar_filtros_sql(ignorar=['superintendencia_id'])
        where_sup = "WHERE " + " AND ".join(filtros_sup)

        cursor.execute(f"""
            SELECT DISTINCT s.id, s.nome
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN superintendencias s ON u.superintendencia_id = s.id
            {where_sup}
            ORDER BY s.nome
        """, valores_sup)
        superintendencias = cursor.fetchall()

        filtros_cc, valores_cc = montar_filtros_sql(ignorar=['centro_custos_id'])
        where_cc = "WHERE " + " AND ".join(filtros_cc)

        cursor.execute(f"""
            SELECT DISTINCT cc.id, cc.codigo, cc.descricao
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN centros_custos cc ON u.centro_custos_id = cc.id
            {where_cc}
            ORDER BY cc.codigo, cc.descricao
        """, valores_cc)
        centros_custos = cursor.fetchall()

        filtros_origem, valores_origem = montar_filtros_sql(ignorar=['origem_id'])
        where_origem = "WHERE " + " AND ".join(filtros_origem)

        cursor.execute(f"""
            SELECT DISTINCT o.id, o.descricao
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            JOIN origens o ON a.origem_id = o.id
            {where_origem}
            ORDER BY o.descricao
        """, valores_origem)
        origens = cursor.fetchall()

        filtros_status, valores_status = montar_filtros_sql(ignorar=['status'])
        where_status = "WHERE " + " AND ".join(filtros_status)

        cursor.execute(f"""
            SELECT DISTINCT a.status
            FROM acoes a
            JOIN usuarios u ON a.responsavel_id = u.id
            {where_status}
            ORDER BY FIELD(a.status,
                'Não iniciada',
                'Em andamento',
                'Vencida',
                'Concluída',
                'Cancelada'
            ), a.status
        """, valores_status)
        status_opcoes = cursor.fetchall()

        if perfil in ['administrador', 'avancado']:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = TRUE
                ORDER BY nome
            """)
        else:
            cursor.execute("""
                SELECT id, nome
                FROM usuarios
                WHERE ativo = TRUE
                  AND centro_custos_id = %s
                ORDER BY nome
            """, (centro_custos_usuario_id,))

        usuarios = cursor.fetchall()

        conn.close()

        return render_template(
            'acoes_criadas.html',
            acoes=acoes,
            superintendencias=superintendencias,
            centros_custos=centros_custos,
            origens=origens,
            status_opcoes=status_opcoes,
            usuarios=usuarios,
            filtros=filtros_sessao,
            page=page,
            per_page=per_page,
            total_registros=total_registros,
            total_paginas=total_paginas
        )

    @blueprint.route('/exportar_minhas_acoes')
    @login_required
    @module_required('acesso_plano_acao')
    def exportar_minhas_acoes():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')
        perfil = session.get('perfil')
        centro_custos_id = session.get('centro_custos_id')

        responsavel_id = request.args.get('responsavel_id')
        origem_id = request.args.get('origem_id')
        status = request.args.get('status')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        query = """
            SELECT 
                a.id,
                a.descricao,
                ur.nome AS responsavel,
                uc.nome AS criado_por,
                o.descricao AS origem,
                a.prazo,
                a.status
            FROM acoes a
            JOIN usuarios ur ON a.responsavel_id = ur.id
            LEFT JOIN usuarios uc ON a.criado_por = uc.id
            LEFT JOIN origens o ON a.origem_id = o.id
            WHERE a.ativo = 1
        """

        params = []

        if perfil == 'basico':
            query += " AND a.responsavel_id = %s"
            params.append(usuario_id)
        elif perfil == 'intermediario':
            query += " AND ur.centro_custos_id = %s"
            params.append(centro_custos_id)

        if responsavel_id and perfil != 'basico':
            query += " AND a.responsavel_id = %s"
            params.append(responsavel_id)

        if origem_id:
            query += " AND a.origem_id = %s"
            params.append(origem_id)

        if status:
            query += " AND a.status = %s"
            params.append(status)

        if data_inicio:
            query += " AND a.prazo >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND a.prazo <= %s"
            params.append(data_fim)

        query += " ORDER BY a.prazo ASC, a.id ASC"

        cursor.execute(query, params)
        dados = cursor.fetchall()

        cursor.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Minhas Ações"

        cabecalhos = [
            "ID",
            "Descrição",
            "Responsável",
            "Criado por",
            "Origem",
            "Prazo",
            "Status",
        ]
        ws.append(cabecalhos)

        cor_cabecalho = PatternFill(fill_type="solid", fgColor="F26719")
        fonte_cabecalho = Font(color="FFFFFF", bold=True)
        alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")
        alinhamento_texto = Alignment(vertical="center", wrap_text=True)

        borda = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for cell in ws[1]:
            cell.fill = cor_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_cabecalho
            cell.border = borda

        for item in dados:
            prazo_formatado = ""

            if item.get("prazo"):
                try:
                    prazo_formatado = item["prazo"].strftime("%d/%m/%Y")
                except AttributeError:
                    prazo_formatado = str(item["prazo"])

            ws.append([
                item.get("id", ""),
                item.get("descricao", "") or "",
                item.get("responsavel", "") or "",
                item.get("criado_por", "") or "",
                item.get("origem", "") or "",
                prazo_formatado,
                item.get("status", "") or ""
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alinhamento_texto
                cell.border = borda

        larguras = {
            "A": 10,
            "B": 50,
            "C": 25,
            "D": 25,
            "E": 25,
            "F": 15,
            "G": 18
        }

        for coluna, largura in larguras.items():
            ws.column_dimensions[coluna].width = largura

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="minhas_acoes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @blueprint.route('/exportar_acoes_criadas')
    @login_required
    @module_required('acesso_plano_acao')
    def exportar_acoes_criadas():
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        usuario_id = session.get('usuario_id')

        superintendencia_id = request.args.get('superintendencia_id')
        centro_custos_id = request.args.get('centro_custos_id')
        origem_id = request.args.get('origem_id')
        status = request.args.get('status')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        query = """
            SELECT 
                a.id,
                a.descricao,
                u.nome AS responsavel,
                o.descricao AS origem,
                a.prazo,
                a.status
            FROM acoes a
            LEFT JOIN usuarios u ON a.responsavel_id = u.id
            LEFT JOIN origens o ON a.origem_id = o.id
            WHERE a.criado_por = %s
              AND a.ativo = 1
        """

        params = [usuario_id]

        if superintendencia_id:
            query += " AND u.superintendencia_id = %s"
            params.append(superintendencia_id)

        if centro_custos_id:
            query += " AND u.centro_custos_id = %s"
            params.append(centro_custos_id)

        if origem_id:
            query += " AND a.origem_id = %s"
            params.append(origem_id)

        if status:
            query += " AND a.status = %s"
            params.append(status)

        if data_inicio:
            query += " AND a.prazo >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND a.prazo <= %s"
            params.append(data_fim)

        query += " ORDER BY a.prazo ASC, a.id ASC"

        cursor.execute(query, params)
        dados = cursor.fetchall()

        cursor.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ações Criadas"

        cabecalhos = ["ID", "Descrição", "Responsável", "Origem", "Prazo", "Status"]
        ws.append(cabecalhos)

        cor_cabecalho = PatternFill(fill_type="solid", fgColor="F26719")
        fonte_cabecalho = Font(color="FFFFFF", bold=True)
        alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")
        alinhamento_texto = Alignment(vertical="center", wrap_text=True)

        borda = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for cell in ws[1]:
            cell.fill = cor_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_cabecalho
            cell.border = borda

        for item in dados:
            prazo_formatado = ""

            if item.get("prazo"):
                try:
                    prazo_formatado = item["prazo"].strftime("%d/%m/%Y")
                except AttributeError:
                    prazo_formatado = str(item["prazo"])

            ws.append([
                item.get("id", ""),
                item.get("descricao", "") or "",
                item.get("responsavel", "") or "",
                item.get("origem", "") or "",
                prazo_formatado,
                item.get("status", "") or ""
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alinhamento_texto
                cell.border = borda

        larguras = {
            "A": 10,
            "B": 50,
            "C": 25,
            "D": 25,
            "E": 15,
            "F": 18
        }

        for coluna, largura in larguras.items():
            ws.column_dimensions[coluna].width = largura

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="acoes_criadas.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
