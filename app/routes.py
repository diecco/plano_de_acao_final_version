from flask import Blueprint, json, render_template, request, redirect, session, flash, current_app, url_for, jsonify
from app.utils.db import get_db_connection
from werkzeug.security import generate_password_hash
from functools import wraps
from flask import session, redirect
from app.decorators import login_required, admin_required, module_required, perfil_required, pode_acessar_acao, pode_acessar_ssma
from werkzeug.utils import secure_filename
from itsdangerous import URLSafeTimedSerializer
from flask_mail import Mail, Message
from app import mail
import os
from datetime import datetime
from datetime import date
from datetime import timedelta
from flask import send_file, request, session
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from uuid import uuid4
import os
import uuid
from io import BytesIO
from flask import send_file
from openpyxl import Workbook


# Configuração do token seguro
s = URLSafeTimedSerializer('CHAVE_SECRETA_DO_SISTEMA')  # Substitua por sua chave secreta

UPLOAD_FOLDER = 'app/static/evidencias'

ALLOWED_EVIDENCE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'docx', 'xlsx'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_evidence_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EVIDENCE_EXTENSIONS

def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

main_routes = Blueprint('main', __name__)

@main_routes.route('/')
def index():
    return redirect('/login')


@main_routes.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT *
            FROM usuarios
            WHERE email = %s
              AND ativo = TRUE
              AND tem_acesso_sistema = 1
        """, (email,))
        usuario = cursor.fetchone()
        conn.close()

        from werkzeug.security import check_password_hash

        if usuario and check_password_hash(usuario['senha_hash'], senha):
            session.clear()

            # 🔹 Dados básicos
            session['usuario_id'] = usuario['id']
            session['nome'] = usuario['nome']
            session['perfil'] = usuario['perfil']

            # 🔹 Escopo (ESSENCIAL)
            session['centro_custos_id'] = usuario.get('centro_custos_id')
            session['superintendencia_id'] = usuario.get('superintendencia_id')

            # 🔹 Permissões por módulo
            session['acesso_plano_acao'] = usuario.get('acesso_plano_acao', 0)
            session['acesso_ssma'] = usuario.get('acesso_ssma', 0)
            session['acesso_melhoria'] = usuario.get('acesso_melhoria', 0)
            session['acesso_gestao_pessoas'] = usuario.get('acesso_gestao_pessoas', 0)
            session['acesso_treinamentos'] = usuario.get('acesso_treinamentos', 0)
            session['acesso_procedimentos'] = usuario.get('acesso_procedimentos', 0)
            session['acesso_pcpm'] = usuario.get('acesso_pcpm', 0)

            # 🔹 Flags adicionais (se quiser usar depois)
            session['pode_ser_instrutor'] = usuario.get('pode_ser_instrutor', 0)
            session['responsavel_revisao_padrao'] = usuario.get('responsavel_revisao_padrao', 0)

            # 🔹 Forçar troca de senha
            if usuario.get('precisa_alterar_senha'):
                return redirect('/alterar_senha')

            # 🔹 Redirecionamento
            if usuario['perfil'] == 'administrador':
                return redirect('/admin')
            else:
                return redirect('/dashboard')

        else:
            flash('Email ou senha inválidos.', 'warning')

    return render_template('login.html')

@main_routes.route('/admin')
@login_required

def admin():
    if session.get('perfil') != 'administrador':
        flash('Acesso restrito ao administrador.')
        return redirect('/')
    return render_template('admin.html')

@main_routes.route('/alterar_senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    if request.method == 'POST':
        senha_atual = request.form['senha_atual']
        nova_senha = request.form['nova_senha']
        confirmar_senha = request.form['confirmar_senha']

        if nova_senha != confirmar_senha:
            flash("A nova senha e a confirmação não coincidem.", "danger")
            return redirect('/alterar_senha')

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT senha_hash FROM usuarios WHERE id = %s", (session['usuario_id'],))
        usuario = cursor.fetchone()

        from werkzeug.security import check_password_hash, generate_password_hash

        if not usuario or not check_password_hash(usuario['senha_hash'], senha_atual):
            flash("Senha atual incorreta.", "danger")
            conn.close()
            return redirect('/alterar_senha')

        nova_hash = generate_password_hash(nova_senha)
        cursor.execute("UPDATE usuarios SET senha_hash = %s, precisa_alterar_senha = 0 WHERE id = %s", (nova_hash, session['usuario_id']))
        conn.commit()
        conn.close()

        flash("Senha alterada com sucesso!", "success")
        return redirect('/dashboard')

    return render_template('alterar_senha.html')

@main_routes.route('/logout')
def logout():
    session.clear()
    return render_template('logout.html')  # ou use redirect('/login')

@main_routes.route('/esqueci_senha', methods=['GET', 'POST'])
def esqueci_senha():
    if request.method == 'POST':
        email = request.form['email']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM usuarios WHERE email = %s AND ativo = TRUE", (email,))
        usuario = cursor.fetchone()
        conn.close()

        if usuario:
            token = s.dumps(email, salt='recuperar-senha')
            link = url_for('main.redefinir_senha', token=token, _external=True)

            # Envia o e-mail com o link
            msg = Message('Redefinição de Senha - TrackPlan',
                          sender='trackplan@trackplan.com.br',
                          recipients=[email])
            msg.body = f'Olá {usuario["nome"]},\n\nClique no link abaixo para redefinir sua senha:\n{link}\n\nEste link é válido por 30 minutos.'
            mail.send(msg)

            flash('Um link para redefinir a senha foi enviado para seu e-mail.', 'success')
        else:
            flash('E-mail não encontrado ou usuário inativo.', 'danger')

    return render_template('esqueci_senha.html')

@main_routes.route('/redefinir_senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    try:
        email = s.loads(token, salt='recuperar-senha', max_age=1800)  # 30 min
    except:
        flash('O link expirou ou é inválido.', 'danger')
        return redirect(url_for('main_routes.login'))

    if request.method == 'POST':
        nova_senha = request.form['nova_senha']
        senha_hash = generate_password_hash(nova_senha)

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE usuarios SET senha_hash = %s, precisa_alterar_senha = 0 WHERE email = %s", (senha_hash, email))
        conn.commit()
        conn.close()

        flash('Senha redefinida com sucesso!', 'success')
        return redirect(url_for('main.login'))

    return render_template('redefinir_senha.html', token=token)

@main_routes.route('/admin/testar_relatorio', methods=['POST'])
@login_required
def testar_relatorio():
    # apenas admin
    if session.get('perfil') != 'administrador':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect('/')

    try:
        # Import aqui para evitar import cycle
        from app.tasks import send_weekly_reports

        # Executa e pega o resumo
        resumo = send_weekly_reports()  # {'enviados': int, 'sem_acoes': int, 'erros': [(email,msg), ...]}
        enviados = resumo.get('enviados', 0)
        sem_acoes = resumo.get('sem_acoes', 0)
        erros = resumo.get('erros', [])

        msg = f"Relatórios enviados: {enviados}. Usuários sem ações: {sem_acoes}."
        if erros:
            # mostra só os 3 primeiros para não poluir a tela
            preview = "; ".join([f"{e[0]}: {e[1]}" for e in erros[:3]])
            msg += f" Erros ({len(erros)}): {preview}"
        flash(msg, 'success' if enviados or sem_acoes >= 0 else 'warning')

    except Exception as e:
        flash(f'Falha ao enviar relatórios agora: {e}', 'danger')

    return redirect('/admin')


# =============================== #
# CENTROS DE CUSTOS               #
# =============================== #

@main_routes.route('/centros_custos', methods=['GET', 'POST'])
@login_required
@admin_required
def centros_custos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        codigo = request.form['codigo']
        descricao = request.form['descricao']
        superintendencia_id = request.form['superintendencia_id']

        cursor.execute("""
            INSERT INTO centros_custos (codigo, descricao, superintendencia_id)
            VALUES (%s, %s, %s)
        """, (codigo, descricao, superintendencia_id))

        conn.commit()
        flash('Centro de Custo cadastrado com sucesso!')

    cursor.execute("""
        SELECT cc.*, s.nome AS nome_superintendencia
        FROM centros_custos cc
        LEFT JOIN superintendencias s ON cc.superintendencia_id = s.id
    """)
    centros = cursor.fetchall()

    cursor.execute("SELECT * FROM superintendencias WHERE ativo = TRUE")
    superintendencias = cursor.fetchall()

    conn.close()

    return render_template(
        'centros_custos.html',
        centros=centros,
        superintendencias=superintendencias
    )

@main_routes.route('/editar_centro/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_centro(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        novo_codigo = request.form['codigo']
        nova_descricao = request.form['descricao']
        nova_superintendencia_id = request.form['superintendencia_id']
        cursor.execute("""
            UPDATE centros_custos
            SET codigo = %s, descricao = %s, superintendencia_id = %s
            WHERE id = %s
        """, (novo_codigo, nova_descricao, nova_superintendencia_id, id))
        conn.commit()
        conn.close()
        flash('Centro de Custo atualizado com sucesso!')
        return redirect('/centros_custos')

    cursor.execute("SELECT * FROM centros_custos WHERE id = %s", (id,))
    centro = cursor.fetchone()

    cursor.execute("SELECT * FROM superintendencias WHERE ativo = TRUE")
    superintendencias = cursor.fetchall()

    conn.close()

    return render_template('editar_centro.html', centro=centro, superintendencias=superintendencias)

@main_routes.route('/habilitar_centrocusto/<int:id>',methods=['POST'])
@login_required
@admin_required
def habilitar_centrocusto(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE centros_custos SET ativo = TRUE WHERE id = %s", (id,))
    conn.commit()
    conn.close()   
    flash('Centro de Custos habilitado com sucesso!', 'sucess')
    return redirect('/centros_custos')

@main_routes.route('/desabilitar_centrocusto/<int:id>', methods=['POST'])
@login_required
@admin_required
def desabilitar_centrocusto(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE centros_custos SET ativo = FALSE WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Centro de Custos desabilitado com sucesso!', 'success')
    return redirect('/centros_custos')

# =============================== #
# SUPERINTENDÊNCIAS               #
# =============================== #

@main_routes.route('/superintendencias', methods=['GET', 'POST'])
@login_required
@admin_required
def superintendencias():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        nome = request.form['nome']
        cursor.execute("INSERT INTO superintendencias (nome) VALUES (%s)", (nome,))
        conn.commit()
        flash('Superintendência cadastrada com sucesso!')

    cursor.execute("SELECT * FROM superintendencias")
    superintendencias = cursor.fetchall()
    conn.close()

    return render_template('superintendencias.html', superintendencias=superintendencias)

@main_routes.route('/editar_superintendencia/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_superintendencia(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        novo_nome = request.form['nome']
        cursor.execute("UPDATE superintendencias SET nome = %s WHERE id = %s", (novo_nome, id))
        conn.commit()
        flash('Superintendência atualizada com sucesso!')
        conn.close()
        return redirect('/superintendencias')

    cursor.execute("SELECT * FROM superintendencias WHERE id = %s", (id,))
    superintendencia = cursor.fetchone()
    conn.close()

    if not superintendencia:
        flash('Superintendência não encontrada.')
        return redirect('/superintendencias')

    return render_template('editar_superintendencia.html', superintendencia=superintendencia)

@main_routes.route('/habilitar_superintendencia/<int:id>',methods=['POST'])
@login_required
@admin_required
def habilitar_superintendencia(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE superintendencias SET ativo = TRUE WHERE id = %s", (id,))
    conn.commit()
    conn.close()   
    flash('Superintendência habilitada com sucesso!', 'sucess')
    return redirect('/superintendencias')

@main_routes.route('/desabilitar_superintendencia/<int:id>', methods=['POST'])
@login_required
@admin_required
def desabilitar_superintendencia(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE superintendencias SET ativo = FALSE WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Superintendencia desabilitada com sucesso!', 'success')
    return redirect('/superintendencias')

# =============================== #
# ORIGENS DAS AÇÕES               #
# =============================== #

@main_routes.route('/origens', methods=['GET', 'POST'])
@login_required
@admin_required
def origens():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        descricao = (request.form.get('descricao') or '').strip()
        centro_custos_id = request.form.get('centro_custos_id')

        if not descricao:
            flash('Informe a descrição da origem.', 'danger')
            conn.close()
            return redirect('/origens')

        if not centro_custos_id:
            flash('Selecione o centro de custo da origem.', 'danger')
            conn.close()
            return redirect('/origens')

        try:
            # Valida se o centro de custo existe e está ativo
            cursor.execute("""
                SELECT id
                FROM centros_custos
                WHERE id = %s AND ativo = 1
            """, (centro_custos_id,))
            centro = cursor.fetchone()

            if not centro:
                flash('Centro de custo inválido ou inativo.', 'danger')
                conn.close()
                return redirect('/origens')

            # Mantém descricao e nome alinhados, e grava o centro de custo
            cursor.execute("""
                INSERT INTO origens (descricao, nome, centro_custos_id, ativo)
                VALUES (%s, %s, %s, 1)
            """, (descricao, descricao, centro_custos_id))

            conn.commit()
            flash('Origem cadastrada com sucesso!', 'success')

        except Exception as e:
            msg = str(e)
            if '1062' in msg or 'Duplicate entry' in msg:
                flash('Já existe uma origem com esse nome/descrição.', 'warning')
            else:
                flash(f'Erro ao salvar origem: {e}', 'danger')
            conn.rollback()

    cursor.execute("""
        SELECT
            o.id,
            o.descricao,
            o.nome,
            o.ativo,
            o.centro_custos_id,
            cc.codigo AS codigo_cc,
            cc.descricao AS descricao_cc
        FROM origens o
        LEFT JOIN centros_custos cc ON o.centro_custos_id = cc.id
        ORDER BY o.id DESC
    """)
    origens = cursor.fetchall()

    cursor.execute("""
        SELECT id, codigo, descricao
        FROM centros_custos
        WHERE ativo = 1
        ORDER BY codigo, descricao
    """)
    centros_custos = cursor.fetchall()

    conn.close()

    return render_template(
        'origens.html',
        origens=origens,
        centros_custos=centros_custos
    )

@main_routes.route('/editar_origem/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_origem(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        nova_descricao = (request.form.get('descricao') or '').strip()
        centro_custos_id = request.form.get('centro_custos_id')

        if not nova_descricao:
            flash('Informe a descrição.', 'danger')
            conn.close()
            return redirect(f'/editar_origem/{id}')

        if not centro_custos_id:
            flash('Selecione o centro de custo da origem.', 'danger')
            conn.close()
            return redirect(f'/editar_origem/{id}')

        try:
            # Valida centro de custo
            cursor.execute("""
                SELECT id
                FROM centros_custos
                WHERE id = %s AND ativo = 1
            """, (centro_custos_id,))
            centro = cursor.fetchone()

            if not centro:
                flash('Centro de custo inválido ou inativo.', 'danger')
                conn.close()
                return redirect(f'/editar_origem/{id}')

            # Atualiza origem com centro de custo
            cursor.execute("""
                UPDATE origens
                SET descricao = %s,
                    nome = %s,
                    centro_custos_id = %s
                WHERE id = %s
            """, (nova_descricao, nova_descricao, centro_custos_id, id))

            conn.commit()
            flash('Origem atualizada com sucesso!', 'success')
            conn.close()
            return redirect('/origens')

        except Exception as e:
            msg = str(e)
            if '1062' in msg or 'Duplicate entry' in msg:
                flash('Já existe outra origem com esse nome/descrição.', 'warning')
            else:
                flash(f'Erro ao atualizar origem: {e}', 'danger')
            conn.rollback()
            conn.close()
            return redirect(f'/editar_origem/{id}')

    # GET
    cursor.execute("""
        SELECT *
        FROM origens
        WHERE id = %s
    """, (id,))
    origem = cursor.fetchone()

    if not origem:
        conn.close()
        flash('Origem não encontrada.')
        return redirect('/origens')

    # Carrega centros de custo para o dropdown
    cursor.execute("""
        SELECT id, codigo, descricao
        FROM centros_custos
        WHERE ativo = 1
        ORDER BY codigo, descricao
    """)
    centros_custos = cursor.fetchall()

    conn.close()

    return render_template(
        'editar_origens.html',
        origem=origem,
        centros_custos=centros_custos
    )

@main_routes.route('/desabilitar_origem/<int:id>', methods=['POST'])
@login_required
@admin_required
def desabilitar_origem(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE origens SET ativo = FALSE WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Origem desabilitada com sucesso!', 'success')
    return redirect('/origens')

@main_routes.route('/habilitar_origem/<int:id>', methods=['POST'])
@login_required
@admin_required
def habilitar_origem(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE origens SET ativo = TRUE WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Origem habilitada com sucesso!', 'success')
    return redirect('/origens')

# =============================== #
# USUÁRIOS                        #
# =============================== #

@main_routes.route('/usuarios', methods=['GET', 'POST'])
@login_required
@admin_required
def usuarios():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        matricula = (request.form.get('matricula') or '').strip() or None
        uid_rfid = (request.form.get('uid_rfid') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        superintendencia_id = request.form.get('superintendencia_id') or None
        centro_custos_id = request.form.get('centro_custos_id') or None
        cargo_id = request.form.get('cargo_id') or None

        funcao_ids = request.form.getlist('funcao_ids[]')
        setor_ids = request.form.getlist('setor_ids[]')

        tem_acesso_sistema = 1 if request.form.get('tem_acesso_sistema') == '1' else 0
        pode_ser_instrutor = 1 if request.form.get('pode_ser_instrutor') == '1' else 0
        responsavel_revisao_padrao = 1 if request.form.get('responsavel_revisao_padrao') == '1' else 0

        perfil = (request.form.get('perfil') or '').strip() or None
        senha_plana = request.form.get('senha') or None
        precisa_alterar_senha = 1 if request.form.get('precisa_alterar_senha') else 0

        acesso_plano_acao = 1 if request.form.get('acesso_plano_acao') else 0
        acesso_ssma = 1 if request.form.get('acesso_ssma') else 0
        acesso_melhoria = 1 if request.form.get('acesso_melhoria') else 0
        acesso_gestao_pessoas = 1 if request.form.get('acesso_gestao_pessoas') else 0
        acesso_treinamentos = 1 if request.form.get('acesso_treinamentos') else 0
        acesso_procedimentos = 1 if request.form.get('acesso_procedimentos') else 0
        acesso_pcpm = 1 if request.form.get('acesso_pcpm') else 0

        if not nome:
            flash('Informe o nome do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if not matricula:
            flash('Informe a matrícula do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if not matricula.isdigit():
            flash('A matrícula deve conter apenas números.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if len(matricula) != 6:
            flash('A matrícula deve conter exatamente 6 dígitos.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if not centro_custos_id:
            flash('Selecione o centro de custos.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if not superintendencia_id:
            flash('Selecione um centro de custos válido para preencher a superintendência.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if not cargo_id:
            flash('Selecione o cargo do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if len(funcao_ids) != len(setor_ids):
            flash('Erro ao processar as habilitações informadas.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if responsavel_revisao_padrao == 1 and tem_acesso_sistema != 1:
            flash('O responsável por revisão de padrão precisa possuir acesso ao TrackPlan.', 'danger')
            conn.close()
            return redirect(url_for('main.usuarios'))

        if tem_acesso_sistema == 1:
            if not email:
                flash('Funcionários com acesso ao sistema devem possuir e-mail.', 'danger')
                conn.close()
                return redirect(url_for('main.usuarios'))

            if not perfil:
                flash('Funcionários com acesso ao sistema devem possuir perfil.', 'danger')
                conn.close()
                return redirect(url_for('main.usuarios'))

            if not senha_plana:
                flash('Funcionários com acesso ao sistema devem possuir senha inicial.', 'danger')
                conn.close()
                return redirect(url_for('main.usuarios'))
        else:
            perfil = None
            senha_plana = None
            precisa_alterar_senha = 0

            acesso_plano_acao = 0
            acesso_ssma = 0
            acesso_melhoria = 0
            acesso_gestao_pessoas = 0
            acesso_treinamentos = 0
            acesso_procedimentos = 0
            acesso_pcpm = 0
            responsavel_revisao_padrao = 0

        try:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE matricula = %s
            """, (matricula,))
            matricula_existente = cursor.fetchone()

            if matricula_existente:
                flash('Já existe um funcionário cadastrado com esta matrícula.', 'warning')
                conn.close()
                return redirect(url_for('main.usuarios'))

            if email:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE email = %s
                """, (email,))
                usuario_existente = cursor.fetchone()
            
            if uid_rfid:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE uid_rfid = %s
                """, (uid_rfid,))
                rfid_existente = cursor.fetchone()

                if rfid_existente:
                    flash('Já existe um funcionário cadastrado com este RFID.', 'warning')
                    conn.close()
                    return redirect(url_for('main.usuarios'))

                if usuario_existente:
                    flash('Já existe um funcionário cadastrado com este e-mail.', 'warning')
                    conn.close()
                    return redirect(url_for('main.usuarios'))

            if responsavel_revisao_padrao == 1:
                cursor.execute("""
                    SELECT id, nome
                    FROM usuarios
                    WHERE responsavel_revisao_padrao = 1
                      AND ativo = 1
                      AND centro_custos_id = %s
                    LIMIT 1
                """, (centro_custos_id,))
                responsavel_existente = cursor.fetchone()

                if responsavel_existente:
                    flash(
                        f"Já existe um responsável ativo por revisão de padrão neste centro de custo: {responsavel_existente['nome']}.",
                        "warning"
                    )
                    conn.close()
                    return redirect(url_for('main.usuarios'))

            hash_senha = None
            if senha_plana:
                hash_senha = generate_password_hash(senha_plana, method="pbkdf2:sha256")

            cursor.execute("""
                INSERT INTO usuarios (
                    nome,
                    matricula,
                    uid_rfid,
                    email,
                    superintendencia_id,
                    centro_custos_id,
                    cargo_id,
                    senha_hash,
                    perfil,
                    ativo,
                    precisa_alterar_senha,
                    tem_acesso_sistema,
                    pode_ser_instrutor,
                    responsavel_revisao_padrao,
                    acesso_plano_acao,
                    acesso_ssma,
                    acesso_melhoria,
                    acesso_gestao_pessoas,
                    acesso_treinamentos,
                    acesso_procedimentos,
                    acesso_pcpm
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                nome,
                matricula,
                uid_rfid,
                email,
                superintendencia_id,
                centro_custos_id,
                cargo_id,
                hash_senha,
                perfil,
                precisa_alterar_senha,
                tem_acesso_sistema,
                pode_ser_instrutor,
                responsavel_revisao_padrao,
                acesso_plano_acao,
                acesso_ssma,
                acesso_melhoria,
                acesso_gestao_pessoas,
                acesso_treinamentos,
                acesso_procedimentos,
                acesso_pcpm
            ))

            usuario_id = cursor.lastrowid
            combinacoes_inseridas = set()

            for funcao_id, setor_id in zip(funcao_ids, setor_ids):
                funcao_id = (funcao_id or '').strip()
                setor_id = (setor_id or '').strip()

                if not funcao_id or not setor_id:
                    continue

                chave = (funcao_id, setor_id)

                if chave in combinacoes_inseridas:
                    continue

                cursor.execute("""
                    INSERT INTO usuario_funcoes_setores (
                        usuario_id,
                        funcao_id,
                        setor_id,
                        ativo
                    )
                    VALUES (%s, %s, %s, 1)
                """, (usuario_id, funcao_id, setor_id))

                combinacoes_inseridas.add(chave)

            conn.commit()

            if tem_acesso_sistema == 1 and email and senha_plana:
                try:
                    link_login = url_for('main.login', _external=True)

                    msg = Message(
                        subject="Bem-vindo ao TrackPlan — dados de acesso",
                        recipients=[email]
                    )

                    msg.html = f"""
                    <div style="font-family: Arial, sans-serif; font-size: 15px; color:#333;">
                      <div style="text-align:center; margin-bottom:18px;">
                        <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height:50px;">
                      </div>

                      <p>Olá <strong>{nome}</strong>,</p>
                      <p>Seu acesso ao <strong>TrackPlan</strong> foi criado com sucesso. Seguem seus dados de login:</p>

                      <div style="background:#f7f7f7; border:1px solid #eee; border-radius:6px; padding:12px 14px; margin:12px 0;">
                        <p style="margin:6px 0;"><strong>Endereço:</strong> <a href="{link_login}" target="_blank">{link_login}</a></p>
                        <p style="margin:6px 0;"><strong>Usuário (e-mail):</strong> {email}</p>
                        <p style="margin:6px 0;"><strong>Senha inicial:</strong> {senha_plana}</p>
                      </div>

                      <p style="margin:22px 0; text-align:center;">
                        <a href="{link_login}" target="_blank"
                           style="display:inline-block; background:#ea6a23; color:#fff; text-decoration:none; padding:10px 18px; border-radius:5px;">
                          Acessar o TrackPlan
                        </a>
                      </p>

                      <p style="font-size:13px; color:#666;">Equipe TrackPlan</p>
                    </div>
                    """

                    mail.send(msg)
                    flash('Funcionário cadastrado e e-mail de boas-vindas enviado!', 'success')

                except Exception:
                    flash('Funcionário cadastrado, mas não foi possível enviar o e-mail de boas-vindas.', 'warning')
            else:
                flash('Funcionário cadastrado com sucesso!', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao cadastrar funcionário: {e}', 'danger')

        finally:
            conn.close()

        return redirect(url_for('main.usuarios'))

    cursor.execute("""
        SELECT cc.id, cc.codigo, cc.descricao,
               s.id AS superintendencia_id,
               s.nome AS superintendencia_nome
        FROM centros_custos cc
        LEFT JOIN superintendencias s ON s.id = cc.superintendencia_id
        WHERE cc.ativo = 1
        ORDER BY cc.codigo, cc.descricao
    """)
    centros_custos = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM setores WHERE ativo = 1 ORDER BY nome")
    setores = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM cargos WHERE ativo = 1 ORDER BY nome")
    cargos = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM funcoes WHERE ativo = 1 ORDER BY nome")
    funcoes = cursor.fetchall()

    conn.close()

    return render_template(
        'usuarios.html',
        centros_custos=centros_custos,
        setores=setores,
        cargos=cargos,
        funcoes=funcoes
    )

@main_routes.route('/editar_usuario/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            u.*,
            s.nome AS nome_superintendencia
        FROM usuarios u
        LEFT JOIN superintendencias s
            ON s.id = u.superintendencia_id
        WHERE u.id = %s
    """, (id,))
    usuario = cursor.fetchone()

    if not usuario:
        conn.close()
        flash('Funcionário não encontrado.', 'warning')
        return redirect(url_for('main.listar_usuarios'))

    tinha_acesso = 1 if usuario.get('tem_acesso_sistema') else 0

    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        matricula = (request.form.get('matricula') or '').strip() or None
        uid_rfid = (request.form.get('uid_rfid') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        superintendencia_id = request.form.get('superintendencia_id') or None
        centro_custos_id = request.form.get('centro_custos_id') or None
        cargo_id = request.form.get('cargo_id') or None

        funcao_ids = request.form.getlist('funcao_ids[]')
        setor_ids = request.form.getlist('setor_ids[]')

        tem_acesso_sistema = 1 if request.form.get('tem_acesso_sistema') == '1' else 0
        pode_ser_instrutor = 1 if request.form.get('pode_ser_instrutor') == '1' else 0
        responsavel_revisao_padrao = 1 if request.form.get('responsavel_revisao_padrao') == '1' else 0

        perfil = (request.form.get('perfil') or '').strip() or None
        nova_senha = request.form.get('nova_senha') or None
        precisa_alterar_senha = 1 if request.form.get('precisa_alterar_senha') else 0
        ativo = 1 if request.form.get('ativo', '1') == '1' else 0

        acesso_plano_acao = 1 if request.form.get('acesso_plano_acao') else 0
        acesso_ssma = 1 if request.form.get('acesso_ssma') else 0
        acesso_melhoria = 1 if request.form.get('acesso_melhoria') else 0
        acesso_gestao_pessoas = 1 if request.form.get('acesso_gestao_pessoas') else 0
        acesso_treinamentos = 1 if request.form.get('acesso_treinamentos') else 0
        acesso_procedimentos = 1 if request.form.get('acesso_procedimentos') else 0
        acesso_pcpm = 1 if request.form.get('acesso_pcpm') else 0

        if not nome:
            flash('Informe o nome do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if not matricula:
            flash('Informe a matrícula do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if not matricula.isdigit():
            flash('A matrícula deve conter apenas números.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if len(matricula) != 6:
            flash('A matrícula deve conter exatamente 6 dígitos.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if uid_rfid:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE uid_rfid = %s
                AND id <> %s
            """, (uid_rfid, id))
            rfid_existente = cursor.fetchone()

            if rfid_existente:
                flash('Já existe outro funcionário cadastrado com este RFID.', 'warning')
                conn.close()
                return redirect(url_for('main.editar_usuario', id=id))

        if not centro_custos_id:
            flash('Selecione o centro de custos.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if not superintendencia_id:
            flash('Selecione um centro de custos válido para preencher a superintendência.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if not cargo_id:
            flash('Selecione o cargo do funcionário.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if len(funcao_ids) != len(setor_ids):
            flash('Erro ao processar as habilitações informadas.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if responsavel_revisao_padrao == 1 and tem_acesso_sistema != 1:
            flash('O responsável por revisão de padrão precisa possuir acesso ao TrackPlan.', 'danger')
            conn.close()
            return redirect(url_for('main.editar_usuario', id=id))

        if tem_acesso_sistema == 1:
            if not email:
                flash('Funcionários com acesso ao sistema devem possuir e-mail.', 'danger')
                conn.close()
                return redirect(url_for('main.editar_usuario', id=id))

            if not perfil:
                flash('Funcionários com acesso ao sistema devem possuir perfil.', 'danger')
                conn.close()
                return redirect(url_for('main.editar_usuario', id=id))

            if tinha_acesso == 0 and not nova_senha:
                flash('Ao conceder acesso ao TrackPlan, informe uma senha inicial.', 'danger')
                conn.close()
                return redirect(url_for('main.editar_usuario', id=id))
        else:
            perfil = None
            nova_senha = None
            precisa_alterar_senha = 0

            acesso_plano_acao = 0
            acesso_ssma = 0
            acesso_melhoria = 0
            acesso_gestao_pessoas = 0
            acesso_treinamentos = 0
            acesso_procedimentos = 0
            acesso_pcpm = 0
            responsavel_revisao_padrao = 0

        try:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE matricula = %s
                  AND id <> %s
            """, (matricula, id))
            matricula_existente = cursor.fetchone()

            if matricula_existente:
                flash('Já existe outro funcionário cadastrado com esta matrícula.', 'warning')
                conn.close()
                return redirect(url_for('main.editar_usuario', id=id))

            if email:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE email = %s
                      AND id <> %s
                """, (email, id))
                existente = cursor.fetchone()

                if existente:
                    flash('Já existe outro funcionário cadastrado com este e-mail.', 'warning')
                    conn.close()
                    return redirect(url_for('main.editar_usuario', id=id))

            if responsavel_revisao_padrao == 1:
                cursor.execute("""
                    SELECT id, nome
                    FROM usuarios
                    WHERE responsavel_revisao_padrao = 1
                      AND ativo = 1
                      AND centro_custos_id = %s
                      AND id <> %s
                    LIMIT 1
                """, (centro_custos_id, id))
                responsavel_existente = cursor.fetchone()

                if responsavel_existente:
                    flash(
                        f"Já existe um responsável ativo por revisão de padrão neste centro de custo: {responsavel_existente['nome']}.",
                        "warning"
                    )
                    conn.close()
                    return redirect(url_for('main.editar_usuario', id=id))

            if nova_senha:
                hash_senha = generate_password_hash(nova_senha, method="pbkdf2:sha256")

                cursor.execute("""
                    UPDATE usuarios
                    SET nome = %s,
                        matricula = %s,
                        uid_rfid = %s,
                        email = %s,
                        superintendencia_id = %s,
                        centro_custos_id = %s,
                        cargo_id = %s,
                        perfil = %s,
                        ativo = %s,
                        precisa_alterar_senha = %s,
                        tem_acesso_sistema = %s,
                        pode_ser_instrutor = %s,
                        responsavel_revisao_padrao = %s,
                        acesso_plano_acao = %s,
                        acesso_ssma = %s,
                        acesso_melhoria = %s,
                        acesso_gestao_pessoas = %s,
                        acesso_treinamentos = %s,
                        acesso_procedimentos = %s,
                        acesso_pcpm = %s,
                        senha_hash = %s
                    WHERE id = %s
                """, (
                    nome,
                    matricula,
                    uid_rfid,
                    email,
                    superintendencia_id,
                    centro_custos_id,
                    cargo_id,
                    perfil,
                    ativo,
                    precisa_alterar_senha,
                    tem_acesso_sistema,
                    pode_ser_instrutor,
                    responsavel_revisao_padrao,
                    acesso_plano_acao,
                    acesso_ssma,
                    acesso_melhoria,
                    acesso_gestao_pessoas,
                    acesso_treinamentos,
                    acesso_procedimentos,
                    acesso_pcpm,
                    hash_senha,
                    id
                ))
            else:
                cursor.execute("""
                    UPDATE usuarios
                    SET nome = %s,
                        matricula = %s,
                        uid_rfid = %s,
                        email = %s,
                        superintendencia_id = %s,
                        centro_custos_id = %s,
                        cargo_id = %s,
                        perfil = %s,
                        ativo = %s,
                        precisa_alterar_senha = %s,
                        tem_acesso_sistema = %s,
                        pode_ser_instrutor = %s,
                        responsavel_revisao_padrao = %s,
                        acesso_plano_acao = %s,
                        acesso_ssma = %s,
                        acesso_melhoria = %s,
                        acesso_gestao_pessoas = %s,
                        acesso_treinamentos = %s,
                        acesso_procedimentos = %s,
                        acesso_pcpm = %s
                    WHERE id = %s
                """, (
                    nome,
                    matricula,
                    uid_rfid,
                    email,
                    superintendencia_id,
                    centro_custos_id,
                    cargo_id,
                    perfil,
                    ativo,
                    precisa_alterar_senha,
                    tem_acesso_sistema,
                    pode_ser_instrutor,
                    responsavel_revisao_padrao,
                    acesso_plano_acao,
                    acesso_ssma,
                    acesso_melhoria,
                    acesso_gestao_pessoas,
                    acesso_treinamentos,
                    acesso_procedimentos,
                    acesso_pcpm,
                    id
                ))

            cursor.execute("DELETE FROM usuario_funcoes_setores WHERE usuario_id = %s", (id,))

            combinacoes_inseridas = set()

            for funcao_id, setor_id in zip(funcao_ids, setor_ids):
                funcao_id = (funcao_id or '').strip()
                setor_id = (setor_id or '').strip()

                if not funcao_id or not setor_id:
                    continue

                chave = (funcao_id, setor_id)

                if chave in combinacoes_inseridas:
                    continue

                cursor.execute("""
                    INSERT INTO usuario_funcoes_setores (
                        usuario_id,
                        funcao_id,
                        setor_id,
                        ativo
                    )
                    VALUES (%s, %s, %s, 1)
                """, (id, funcao_id, setor_id))

                combinacoes_inseridas.add(chave)

            conn.commit()
            flash('Funcionário atualizado com sucesso!', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao atualizar funcionário: {e}', 'danger')

        finally:
            conn.close()

        return redirect(url_for('main.listar_usuarios'))

    cursor.execute("""
        SELECT cc.id, cc.codigo, cc.descricao,
               s.id AS superintendencia_id,
               s.nome AS superintendencia_nome
        FROM centros_custos cc
        LEFT JOIN superintendencias s ON s.id = cc.superintendencia_id
        WHERE cc.ativo = 1
        ORDER BY cc.codigo, cc.descricao
    """)
    centros_custos = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM setores WHERE ativo = 1 ORDER BY nome")
    setores = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM cargos WHERE ativo = 1 ORDER BY nome")
    cargos = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM funcoes WHERE ativo = 1 ORDER BY nome")
    funcoes = cursor.fetchall()

    cursor.execute("""
        SELECT
            ufs.funcao_id,
            ufs.setor_id,
            f.nome AS nome_funcao,
            s.nome AS nome_setor
        FROM usuario_funcoes_setores ufs
        JOIN funcoes f
            ON f.id = ufs.funcao_id
        JOIN setores s
            ON s.id = ufs.setor_id
        WHERE ufs.usuario_id = %s
          AND ufs.ativo = 1
        ORDER BY f.nome, s.nome
    """, (id,))
    habilitacoes = cursor.fetchall()

    conn.close()

    return render_template(
        'editar_usuario.html',
        usuario=usuario,
        centros_custos=centros_custos,
        setores=setores,
        cargos=cargos,
        funcoes=funcoes,
        habilitacoes=habilitacoes
    )

@main_routes.route('/listar_usuarios', methods=['GET'])
@login_required
@admin_required
def listar_usuarios():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    status = request.args.get('status', '').strip()
    tem_acesso_sistema = request.args.get('tem_acesso_sistema', '').strip()
    pode_ser_instrutor = request.args.get('pode_ser_instrutor', '').strip()
    centro_custos_id = request.args.get('centro_custos_id', '').strip()

    sort = request.args.get('sort', 'nome').strip()
    order = request.args.get('order', 'asc').strip()

    colunas_validas = {
        'nome': 'u.nome',
        'matricula': 'u.matricula',
        'email': 'u.email',
        'cargo': 'c.nome'
    }

    coluna_sort = colunas_validas.get(sort, 'u.nome')
    direcao = 'ASC' if order == 'asc' else 'DESC'

    query = f"""
        SELECT
            u.id,
            u.nome,
            u.matricula,
            u.email,
            u.ativo,
            u.tem_acesso_sistema,
            u.pode_ser_instrutor,
            u.perfil,
            u.centro_custos_id,
            u.superintendencia_id,
            u.cargo_id,

            u.acesso_plano_acao,
            u.acesso_ssma,
            u.acesso_melhoria,
            u.acesso_gestao_pessoas,
            u.acesso_treinamentos,
            u.acesso_procedimentos,

            c.nome AS nome_cargo,

            cc.codigo AS codigo_cc,
            cc.descricao AS descricao_cc,

            s.nome AS nome_superintendencia,

            GROUP_CONCAT(
                DISTINCT f.nome
                ORDER BY f.nome
                SEPARATOR ', '
            ) AS nomes_funcoes,

            GROUP_CONCAT(
                DISTINCT st.nome
                ORDER BY st.nome
                SEPARATOR ', '
            ) AS nomes_setores

        FROM usuarios u
        LEFT JOIN cargos c
            ON c.id = u.cargo_id
        LEFT JOIN centros_custos cc
            ON cc.id = u.centro_custos_id
        LEFT JOIN superintendencias s
            ON s.id = u.superintendencia_id
        LEFT JOIN usuario_funcoes_setores ufs
            ON ufs.usuario_id = u.id
           AND ufs.ativo = 1
        LEFT JOIN funcoes f
            ON f.id = ufs.funcao_id
        LEFT JOIN setores st
            ON st.id = ufs.setor_id
        WHERE 1=1
    """

    params = []

    if status != '':
        query += " AND u.ativo = %s"
        params.append(status)

    if tem_acesso_sistema != '':
        query += " AND u.tem_acesso_sistema = %s"
        params.append(tem_acesso_sistema)

    if pode_ser_instrutor != '':
        query += " AND u.pode_ser_instrutor = %s"
        params.append(pode_ser_instrutor)

    if centro_custos_id:
        query += " AND u.centro_custos_id = %s"
        params.append(centro_custos_id)

    query += f"""
        GROUP BY
            u.id,
            u.nome,
            u.matricula,
            u.email,
            u.ativo,
            u.tem_acesso_sistema,
            u.pode_ser_instrutor,
            u.perfil,
            u.centro_custos_id,
            u.superintendencia_id,
            u.cargo_id,
            u.acesso_plano_acao,
            u.acesso_ssma,
            u.acesso_melhoria,
            u.acesso_gestao_pessoas,
            u.acesso_treinamentos,
            u.acesso_procedimentos,
            c.nome,
            cc.codigo,
            cc.descricao,
            s.nome
        ORDER BY {coluna_sort} {direcao}, u.nome ASC
    """

    cursor.execute(query, params)
    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            cc.id,
            cc.codigo,
            cc.descricao
        FROM centros_custos cc
        WHERE cc.ativo = 1
        ORDER BY cc.codigo, cc.descricao
    """)
    centros_custos = cursor.fetchall()

    conn.close()

    filtros = {
        'status': status,
        'tem_acesso_sistema': tem_acesso_sistema,
        'pode_ser_instrutor': pode_ser_instrutor,
        'centro_custos_id': centro_custos_id,
        'sort': sort,
        'order': order
    }

    return render_template(
        'listar_usuarios.html',
        usuarios=usuarios,
        centros_custos=centros_custos,
        filtros=filtros
    )

@main_routes.route('/permissoes_usuario/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def permissoes_usuario(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        perfil = request.form.get('perfil')

        acesso_plano_acao = 1 if request.form.get('acesso_plano_acao') else 0
        acesso_ssma = 1 if request.form.get('acesso_ssma') else 0
        acesso_melhoria = 1 if request.form.get('acesso_melhoria') else 0
        acesso_gestao_pessoas = 1 if request.form.get('acesso_gestao_pessoas') else 0
        acesso_treinamentos = 1 if request.form.get('acesso_treinamentos') else 0
        acesso_procedimentos = 1 if request.form.get('acesso_procedimentos') else 0
        acesso_pcpm = 1 if request.form.get('acesso_pcpm') else 0

        cursor.execute("""
            UPDATE usuarios
            SET perfil = %s,
                acesso_plano_acao = %s,
                acesso_ssma = %s,
                acesso_melhoria = %s,
                acesso_gestao_pessoas = %s,
                acesso_treinamentos = %s,
                acesso_procedimentos = %s,
                acesso_pcpm = %s
            WHERE id = %s
        """, (
            perfil,
            acesso_plano_acao,
            acesso_ssma,
            acesso_melhoria,
            acesso_gestao_pessoas,
            acesso_treinamentos,
            acesso_procedimentos,
            acesso_pcpm,
            id
        ))

        conn.commit()
        conn.close()

        flash('Permissões atualizadas com sucesso!', 'success')
        return redirect('/usuarios')

    cursor.execute("""
        SELECT *
        FROM usuarios
        WHERE id = %s
    """, (id,))
    usuario = cursor.fetchone()

    conn.close()

    if not usuario:
        flash('Usuário não encontrado.', 'warning')
        return redirect('/usuarios')

    return render_template('permissoes_usuario.html', usuario=usuario)


# =============================== #
# PLANOS DE AÇÃO                  #
# =============================== #

@main_routes.route('/dashboard')
@login_required
@module_required('acesso_plano_acao')
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    from datetime import date, timedelta

    # Atualiza status para "Vencida" quando ultrapassar o prazo
    cursor.execute("""
        UPDATE acoes
        SET status = 'Vencida'
        WHERE status NOT IN ('Concluída', 'Cancelada')
          AND prazo < %s
          AND ativo = 1
    """, (date.today(),))
    conn.commit()

    usuario_id = session['usuario_id']
    perfil = session.get('perfil')

    # =========================
    # PERSISTÊNCIA DOS FILTROS
    # =========================
    if request.args.get('limpar') == '1':
        session.pop('dashboard_filtros', None)
        conn.close()
        return redirect(url_for('main.dashboard'))

    if request.args:
        filtros_sessao = {
            'responsavel': request.args.get('responsavel', ''),
            'superintendencia': request.args.get('superintendencia', ''),
            'centro_custos': request.args.get('centro_custos', ''),
            'origem': request.args.get('origem', ''),
            'status': request.args.get('status', ''),
            'data_inicio': request.args.get('data_inicio', ''),
            'data_fim': request.args.get('data_fim', '')
        }
        session['dashboard_filtros'] = filtros_sessao
    else:
        filtros_sessao = session.get('dashboard_filtros', {
            'responsavel': '',
            'superintendencia': '',
            'centro_custos': '',
            'origem': '',
            'status': '',
            'data_inicio': '',
            'data_fim': ''
        })

    responsavel_id = filtros_sessao.get('responsavel')
    superintendencia_id = filtros_sessao.get('superintendencia')
    centro_custos_id = filtros_sessao.get('centro_custos')
    origem_id = filtros_sessao.get('origem')
    status = filtros_sessao.get('status')
    data_inicio = filtros_sessao.get('data_inicio')
    data_fim = filtros_sessao.get('data_fim')

    # =========================
    # MONTAGEM DOS FILTROS
    # =========================
    filtros = []
    valores = []

    filtros.append("a.ativo = 1")

    if responsavel_id:
        filtros.append("a.responsavel_id = %s")
        valores.append(responsavel_id)

    if superintendencia_id:
        filtros.append("u.superintendencia_id = %s")
        valores.append(superintendencia_id)

    if centro_custos_id:
        filtros.append("u.centro_custos_id = %s")
        valores.append(centro_custos_id)

    if origem_id:
        filtros.append("a.origem_id = %s")
        valores.append(origem_id)

    if status:
        filtros.append("a.status = %s")
        valores.append(status)

    if data_inicio:
        filtros.append("a.prazo >= %s")
        valores.append(data_inicio)

    if data_fim:
        filtros.append("a.prazo <= %s")
        valores.append(data_fim)

    # =========================
    # CONTROLE POR PERFIL (CORRIGIDO)
    # =========================
    if perfil == 'basico':
        filtros.append("a.responsavel_id = %s")
        valores.append(usuario_id)

    elif perfil == 'intermediario':
        filtros.append("u.centro_custos_id = %s")
        valores.append(session.get('centro_custos_id'))

    elif perfil == 'avancado':
        pass

    elif perfil == 'administrador':
        pass

    where_clause = " AND ".join(filtros)
    if where_clause:
        where_clause = "WHERE " + where_clause

    hoje = date.today()
    sete_dias = hoje + timedelta(days=7)
    primeiro_dia_mes = hoje.replace(day=1)

    # =========================
    # KPIs
    # =========================
    query_kpis = f"""
        SELECT
            SUM(CASE WHEN a.status NOT IN ('Concluída', 'Cancelada') THEN 1 ELSE 0 END) AS kpi_ativas,
            SUM(CASE WHEN a.status = 'Vencida' THEN 1 ELSE 0 END) AS kpi_vencidas,
            SUM(CASE
                    WHEN a.status NOT IN ('Concluída', 'Cancelada', 'Vencida')
                     AND a.prazo BETWEEN %s AND %s
                    THEN 1 ELSE 0
                END) AS kpi_a_vencer_7_dias,
            SUM(CASE
                    WHEN a.status = 'Concluída'
                     AND a.prazo >= %s
                    THEN 1 ELSE 0
                END) AS kpi_concluidas_mes,
            0 AS kpi_criadas_mes
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        {where_clause}
    """
    cursor.execute(query_kpis, [hoje, sete_dias, primeiro_dia_mes] + valores)
    kpis = cursor.fetchone() or {}

    kpi_ativas = kpis.get('kpi_ativas') or 0
    kpi_vencidas = kpis.get('kpi_vencidas') or 0
    kpi_a_vencer_7_dias = kpis.get('kpi_a_vencer_7_dias') or 0
    kpi_concluidas_mes = kpis.get('kpi_concluidas_mes') or 0
    kpi_criadas_mes = kpis.get('kpi_criadas_mes') or 0

    # =========================
    # GRÁFICO POR STATUS
    # =========================
    query_status = f"""
        SELECT a.status, COUNT(*) AS total
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        {where_clause}
        GROUP BY a.status
    """
    cursor.execute(query_status, valores)
    status_data = cursor.fetchall()
    labels_status = [d['status'] for d in status_data]
    dados_status = [d['total'] for d in status_data]

    # =========================
    # GRÁFICO POR ORIGEM
    # =========================
    query_origem = f"""
        SELECT o.descricao AS origem, COUNT(*) AS total
        FROM acoes a
        JOIN origens o ON a.origem_id = o.id
        JOIN usuarios u ON a.responsavel_id = u.id
        {where_clause}
        GROUP BY o.descricao
    """
    cursor.execute(query_origem, valores)
    origem_data = cursor.fetchall()
    labels_origem = [d['origem'] for d in origem_data]
    dados_origem = [d['total'] for d in origem_data]

    # =========================
    # ATENÇÃO IMEDIATA - VENCIDAS
    # =========================
    query_vencidas_criticas = f"""
        SELECT
            a.id,
            a.descricao,
            a.prazo,
            o.descricao AS descricao_origem,
            DATEDIFF(%s, a.prazo) AS dias_atraso
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN origens o ON a.origem_id = o.id
        {where_clause}
          AND a.status = 'Vencida'
        ORDER BY a.prazo ASC
        LIMIT 5
    """
    cursor.execute(query_vencidas_criticas, [hoje] + valores)
    acoes_vencidas_criticas = cursor.fetchall()

    # =========================
    # ATENÇÃO IMEDIATA - PRÓXIMAS
    # =========================
    query_proximas = f"""
        SELECT
            a.id,
            a.descricao,
            a.prazo,
            o.descricao AS descricao_origem,
            DATEDIFF(a.prazo, %s) AS dias_restantes
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN origens o ON a.origem_id = o.id
        {where_clause}
          AND a.status NOT IN ('Concluída', 'Cancelada', 'Vencida')
          AND a.prazo >= %s
        ORDER BY a.prazo ASC
        LIMIT 5
    """
    cursor.execute(query_proximas, [hoje] + valores + [hoje])
    acoes_proximas_vencimento = cursor.fetchall()

    conn.close()

    return render_template(
        'dashboard.html',
        kpi_ativas=kpi_ativas,
        kpi_vencidas=kpi_vencidas,
        kpi_a_vencer_7_dias=kpi_a_vencer_7_dias,
        kpi_concluidas_mes=kpi_concluidas_mes,
        kpi_criadas_mes=kpi_criadas_mes,
        labels_status=labels_status,
        dados_status=dados_status,
        labels_origem=labels_origem,
        dados_origem=dados_origem,
        acoes_vencidas_criticas=acoes_vencidas_criticas,
        acoes_proximas_vencimento=acoes_proximas_vencimento,
        filtros=filtros_sessao
    )


@main_routes.route('/cadastrar_acao', methods=['GET', 'POST'])
@login_required
@module_required('acesso_plano_acao')
def cadastrar_acao():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')

    usuario_logado = {
        'id': session.get('usuario_id'),
        'superintendencia_id': session.get('superintendencia_id'),
        'centro_custos_id': session.get('centro_custos_id')
    }

    if not usuario_logado.get('id'):
        conn.close()
        flash('Usuário logado não encontrado.', 'danger')
        return redirect('/login')

    if request.method == 'POST':
        origem_id = request.form['origem_id']
        responsavel_id = request.form['responsavel_id']
        descricao = request.form['descricao']
        prazo = request.form['prazo']
        status = request.form['status']

        criado_por = usuario_id

        # Validação da origem conforme perfil
        if perfil == 'administrador':
            cursor.execute("""
                SELECT id
                FROM origens
                WHERE id = %s
                  AND ativo = 1
            """, (origem_id,))
        else:
            cursor.execute("""
                SELECT id
                FROM origens
                WHERE id = %s
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (origem_id, usuario_logado['centro_custos_id']))

        origem_valida = cursor.fetchone()

        if not origem_valida:
            conn.close()
            flash('Origem inválida para o seu perfil de acesso.', 'danger')
            return redirect('/cadastrar_acao')

        # Validação do responsável conforme perfil
        if perfil == 'basico':
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = TRUE
                  AND centro_custos_id = %s
            """, (responsavel_id, usuario_logado['centro_custos_id']))

        elif perfil == 'intermediario':
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = TRUE
                  AND centro_custos_id = %s
            """, (responsavel_id, usuario_logado['centro_custos_id']))

        else:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = TRUE
            """, (responsavel_id,))

        responsavel_valido = cursor.fetchone()

        if not responsavel_valido:
            conn.close()
            flash('Responsável inválido para o seu perfil de acesso.', 'danger')
            return redirect('/cadastrar_acao')

        cursor.execute("""
            INSERT INTO acoes (origem_id, responsavel_id, descricao, prazo, status, criado_por)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (origem_id, responsavel_id, descricao, prazo, status, criado_por))

        conn.commit()
        acao_id = cursor.lastrowid

        cursor.execute("""
            SELECT nome, email
            FROM usuarios
            WHERE id = %s
        """, (responsavel_id,))
        responsavel = cursor.fetchone()

        prazo_formatado = datetime.strptime(prazo, '%Y-%m-%d').strftime('%d/%m/%Y')
        link_edicao = url_for('main.editar_acao', id=acao_id, _external=True)

        try:
            if responsavel and responsavel.get('email'):
                msg = Message(
                    subject="Nova Ação Atribuída a Você - TrackPlan",
                    recipients=[responsavel['email']]
                )
                msg.html = f"""
                <div style="font-family: Arial, sans-serif; font-size: 15px;">
                    <div style="text-align: center;">
                        <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height: 50px; margin-bottom: 20px;">
                    </div>
                    <p>Olá <strong>{responsavel['nome']}</strong>,</p>

                    <p>Uma nova ação foi atribuída a você no sistema TrackPlan.</p>

                    <p><strong>Descrição:</strong> {descricao}<br>
                    <strong>Prazo:</strong> {prazo_formatado}<br>
                    <strong>Status:</strong> {status}</p>

                    <p>Acesse o sistema para mais detalhes:</p>

                    <p>
                        <a href="{link_edicao}" style="display: inline-block; background-color: #ea6a23; color: white; padding: 10px 18px; text-decoration: none; border-radius: 5px;">
                            Editar Ação
                        </a>
                    </p>

                    <br>
                    <p style="font-size: 13px; color: #666;">Equipe TrackPlan</p>
                </div>
                """
                mail.send(msg)

        except Exception:
            flash('Ação criada. Falha ao enviar o e-mail de notificação.', 'warning')

        conn.close()
        flash('Ação cadastrada com sucesso!', 'success')
        return redirect('/dashboard')

    # GET: carregar origens conforme perfil
    if perfil == 'administrador':
        cursor.execute("""
            SELECT id, descricao
            FROM origens
            WHERE ativo = 1
            ORDER BY descricao
        """)
    else:
        cursor.execute("""
            SELECT id, descricao
            FROM origens
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY descricao
        """, (usuario_logado['centro_custos_id'],))

    origens = cursor.fetchall()

    # GET: carregar responsáveis conforme perfil
    if perfil == 'basico':
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
              AND centro_custos_id = %s
            ORDER BY nome
        """, (usuario_logado['centro_custos_id'],))

    elif perfil == 'intermediario':
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
              AND centro_custos_id = %s
            ORDER BY nome
        """, (usuario_logado['centro_custos_id'],))

    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
            ORDER BY nome
        """)

    usuarios = cursor.fetchall()
    conn.close()

    return render_template(
        'cadastrar_acao.html',
        origens=origens,
        usuarios=usuarios
    )


@main_routes.route('/editar_acao/<int:id>', methods=['GET', 'POST'])
@login_required
@module_required('acesso_plano_acao')
def editar_acao(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    acao = pode_acessar_acao(cursor, id)

    if not acao:
        conn.close()
        flash('Ação não encontrada.', 'warning')
        return redirect(url_for('main.dashboard'))

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    pode_editar_completo = acao.get('criado_por') == usuario_id

    cursor.execute("""
        SELECT id, nome, superintendencia_id, centro_custos_id
        FROM usuarios
        WHERE id = %s
    """, (usuario_id,))
    usuario_logado = cursor.fetchone()

    if not usuario_logado:
        conn.close()
        flash('Usuário logado não encontrado.', 'danger')
        return redirect('/login')

    if request.method == 'POST':
        next_url = (request.form.get('next') or '').strip()

        responsavel_antigo = acao.get('responsavel_id')
        prazo_antigo = acao.get('prazo')
        status_antigo = acao.get('status')
        descricao_antiga = (acao.get('descricao') or '').strip()

        if pode_editar_completo:
            origem_id = request.form.get('origem_id')
            responsavel_id = request.form.get('responsavel_id')
            descricao = request.form.get('descricao')
            prazo = request.form.get('prazo')

            if perfil == 'administrador':
                cursor.execute("""
                    SELECT id
                    FROM origens
                    WHERE id = %s
                      AND ativo = 1
                """, (origem_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM origens
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (origem_id, usuario_logado['centro_custos_id']))

            origem_valida = cursor.fetchone()

            if not origem_valida:
                conn.close()
                flash('Origem inválida para o seu perfil de acesso.', 'danger')
                return redirect(next_url or url_for('main.editar_acao', id=id))

            if perfil == 'basico':
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                      AND centro_custos_id = %s
                """, (responsavel_id, usuario_logado['centro_custos_id']))

            elif perfil == 'intermediario':
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                      AND superintendencia_id = %s
                """, (responsavel_id, usuario_logado['superintendencia_id']))

            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = TRUE
                """, (responsavel_id,))

            responsavel_valido = cursor.fetchone()

            if not responsavel_valido:
                conn.close()
                flash('Responsável inválido para o seu perfil de acesso.', 'danger')
                return redirect(next_url or url_for('main.editar_acao', id=id))

        else:
            origem_id = acao['origem_id']
            responsavel_id = acao['responsavel_id']
            descricao = acao['descricao']
            prazo = acao['prazo']

        status = request.form.get('status')
        observacoes = (request.form.get('observacoes') or '').strip()
        data_conclusao_str = (request.form.get('data_conclusao') or '').strip()

        data_conclusao = None
        hoje = date.today()

        if data_conclusao_str:
            try:
                data_conclusao = datetime.strptime(data_conclusao_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Data de conclusão inválida.', 'warning')
                conn.close()
                return redirect(next_url or url_for('main.editar_acao', id=id))

            if data_conclusao > hoje:
                flash('A data de conclusão não pode ser superior à data de hoje.', 'warning')
                conn.close()
                return redirect(next_url or url_for('main.editar_acao', id=id))

            status = 'Concluída'
        else:
            if status == 'Concluída':
                flash('Para concluir a ação, preencha a data de conclusão.', 'warning')
                conn.close()
                return redirect(next_url or url_for('main.editar_acao', id=id))

        if status == 'Cancelada' and not observacoes:
            flash('Ao cancelar a ação, o campo Observações é obrigatório.', 'warning')
            conn.close()
            return redirect(next_url or url_for('main.editar_acao', id=id))

        arquivo_evidencia = acao.get('arquivo_evidencia')
        arquivo = request.files.get('arquivo_evidencia')

        if arquivo and arquivo.filename:
            extensoes_permitidas = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx'}
            nome_original = secure_filename(arquivo.filename)
            extensao = nome_original.rsplit('.', 1)[-1].lower() if '.' in nome_original else ''

            if extensao not in extensoes_permitidas:
                conn.close()
                flash('Tipo de arquivo não permitido para evidência.', 'warning')
                return redirect(next_url or url_for('main.editar_acao', id=id))

            nome_arquivo = f"evidencia_acao_{id}_{uuid.uuid4().hex}.{extensao}"

            pasta_evidencias = os.path.join(
                current_app.root_path,
                'static',
                'evidencias'
            )

            os.makedirs(pasta_evidencias, exist_ok=True)

            caminho_arquivo = os.path.join(pasta_evidencias, nome_arquivo)
            arquivo.save(caminho_arquivo)

            arquivo_evidencia = nome_arquivo

        cursor.execute("""
            UPDATE acoes
            SET origem_id = %s,
                responsavel_id = %s,
                descricao = %s,
                prazo = %s,
                status = %s,
                observacoes = %s,
                data_conclusao = %s,
                arquivo_evidencia = %s
            WHERE id = %s
        """, (
            origem_id,
            responsavel_id,
            descricao,
            prazo,
            status,
            observacoes if observacoes else None,
            data_conclusao,
            arquivo_evidencia,
            id
        ))

        conn.commit()

        houve_mudanca_importante = (
            str(responsavel_id) != str(responsavel_antigo) or
            str(prazo) != str(prazo_antigo) or
            str(status) != str(status_antigo) or
            (descricao or '').strip() != descricao_antiga
        )

        if houve_mudanca_importante:
            try:
                cursor.execute("SELECT nome, email FROM usuarios WHERE id = %s", (responsavel_id,))
                responsavel_atual = cursor.fetchone()

                if responsavel_atual and responsavel_atual.get('email'):
                    alteracoes = []

                    if str(responsavel_id) != str(responsavel_antigo):
                        if responsavel_antigo:
                            cursor.execute("SELECT nome FROM usuarios WHERE id = %s", (responsavel_antigo,))
                            resp_antigo = cursor.fetchone()
                            nome_resp_antigo = resp_antigo['nome'] if resp_antigo else 'Não identificado'
                        else:
                            nome_resp_antigo = 'Não identificado'

                        alteracoes.append(
                            f"<li><strong>Responsável:</strong> {nome_resp_antigo} → {responsavel_atual['nome']}</li>"
                        )

                    if str(prazo) != str(prazo_antigo):
                        prazo_antigo_fmt = prazo_antigo.strftime('%d/%m/%Y') if hasattr(prazo_antigo, 'strftime') else str(prazo_antigo or '-')
                        try:
                            prazo_novo_fmt = datetime.strptime(str(prazo), '%Y-%m-%d').strftime('%d/%m/%Y')
                        except Exception:
                            prazo_novo_fmt = str(prazo or '-')

                        alteracoes.append(
                            f"<li><strong>Prazo:</strong> {prazo_antigo_fmt} → {prazo_novo_fmt}</li>"
                        )

                    if str(status) != str(status_antigo):
                        alteracoes.append(
                            f"<li><strong>Status:</strong> {status_antigo or '-'} → {status or '-'}</li>"
                        )

                    if (descricao or '').strip() != descricao_antiga:
                        alteracoes.append(
                            f"<li><strong>Descrição da ação:</strong> foi atualizada.</li>"
                        )

                    link_edicao = url_for('main.editar_acao', id=id, _external=True)

                    msg = Message(
                        subject="Atualização em Ação sob sua Responsabilidade - TrackPlan",
                        recipients=[responsavel_atual['email']]
                    )

                    msg.html = f"""
                    <div style="font-family: Arial, sans-serif; font-size: 15px; color:#333;">
                        <div style="text-align:center; margin-bottom:18px;">
                            <img src="https://www.trackplan.com.br/imagens/barra_email.png" alt="TrackPlan" style="height:50px;">
                        </div>

                        <p>Olá <strong>{responsavel_atual['nome']}</strong>,</p>

                        <p>Uma ação sob sua responsabilidade foi atualizada no <strong>TrackPlan</strong>.</p>

                        <div style="background:#f7f7f7; border:1px solid #eee; border-radius:6px; padding:12px 14px; margin:12px 0;">
                            <p style="margin:6px 0;"><strong>ID da ação:</strong> {id}</p>
                            <p style="margin:6px 0;"><strong>Descrição atual:</strong> {descricao}</p>
                        </div>

                        <p><strong>Alterações realizadas:</strong></p>
                        <ul>
                            {''.join(alteracoes)}
                        </ul>

                        <p style="margin:22px 0; text-align:center;">
                            <a href="{link_edicao}" target="_blank"
                               style="display:inline-block; background:#ea6a23; color:#fff; text-decoration:none; padding:10px 18px; border-radius:5px;">
                                Acessar ação
                            </a>
                        </p>

                        <p style="font-size:13px; color:#666;">Equipe TrackPlan</p>
                    </div>
                    """

                    mail.send(msg)

            except Exception:
                flash('Ação atualizada, mas não foi possível enviar o e-mail de notificação.', 'warning')

        conn.close()

        flash('Ação atualizada com sucesso!', 'success')

        if next_url:
            return redirect(next_url)

        return redirect(url_for('main.dashboard'))

    if perfil == 'administrador':
        cursor.execute("""
            SELECT id, descricao
            FROM origens
            WHERE ativo = 1
            ORDER BY descricao
        """)
    else:
        cursor.execute("""
            SELECT id, descricao
            FROM origens
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY descricao
        """, (usuario_logado['centro_custos_id'],))

    origens = cursor.fetchall()

    if perfil == 'basico':
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
              AND centro_custos_id = %s
            ORDER BY nome
        """, (usuario_logado['centro_custos_id'],))
    elif perfil == 'intermediario':
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
              AND superintendencia_id = %s
            ORDER BY nome
        """, (usuario_logado['superintendencia_id'],))
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
            ORDER BY nome
        """)

    usuarios = cursor.fetchall()

    conn.close()

    return render_template(
        'editar_acao.html',
        acao=acao,
        origens=origens,
        usuarios=usuarios,
        pode_editar_completo=pode_editar_completo
    )


@main_routes.route('/excluir_acao/<int:id>')
@login_required
@module_required('acesso_plano_acao')
def excluir_acao(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    acao = pode_acessar_acao(cursor, id)

    if not acao:
        conn.close()
        flash('Ação não encontrada ou você não possui permissão para excluí-la.', 'warning')
        return redirect(url_for('main.dashboard'))

    # 🔒 Regra adicional (RECOMENDADO)
    perfil = session.get('perfil')
    usuario_id = session.get('usuario_id')

    if perfil == 'basico':
        if acao.get('criado_por') != usuario_id:
            conn.close()
            flash('Você não tem permissão para excluir esta ação.', 'danger')
            return redirect(url_for('main.dashboard'))

    # 🔥 Exclusão lógica
    cursor.execute("""
        UPDATE acoes
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Ação excluída com sucesso!', 'success')
    return redirect(url_for('main.dashboard'))


@main_routes.route('/anexar_evidencia/<int:acao_id>', methods=['GET', 'POST'])
@login_required
@module_required('acesso_plano_acao')
def anexar_evidencia(acao_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 🔒 Validação de permissão sobre a ação
    acao = pode_acessar_acao(cursor, acao_id)

    if not acao:
        flash("Ação não encontrada ou você não possui permissão para acessá-la.", "warning")
        conn.close()
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash("Nenhum arquivo enviado.", "danger")
            conn.close()
            return redirect(request.url)

        arquivo = request.files['arquivo']

        if arquivo.filename == '':
            flash("Nenhum arquivo selecionado.", "warning")
            conn.close()
            return redirect(request.url)

        if arquivo and allowed_evidence_file(arquivo.filename):
            filename = secure_filename(arquivo.filename)
            caminho = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            arquivo.save(caminho)

            cursor.execute("""
                UPDATE acoes
                SET arquivo_evidencia = %s
                WHERE id = %s
            """, (filename, acao_id))

            conn.commit()
            conn.close()

            flash("Evidência anexada com sucesso!", "success")
            return redirect(url_for('main.dashboard'))

        flash("Tipo de arquivo não permitido.", "danger")
        conn.close()
        return redirect(request.url)

    conn.close()
    return render_template('anexar_evidencia.html', acao=acao)


@main_routes.route('/minhas_acoes')
@login_required
@module_required('acesso_plano_acao')
def minhas_acoes():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if request.args.get('limpar') == '1':
        session.pop('minhas_acoes_filtros', None)
        conn.close()
        return redirect(url_for('main.minhas_acoes'))

    filtros_padrao = {
        'origem_id': '',
        'status': '',
        'data_inicio': '',
        'data_fim': '',
        'sort': 'prazo',
        'order': 'asc'
    }

    filtros_sessao = session.get('minhas_acoes_filtros', filtros_padrao.copy())

    if request.args:
        for campo in filtros_padrao.keys():
            if campo in request.args:
                filtros_sessao[campo] = request.args.get(campo, '')

        filtros_sessao['sort'] = filtros_sessao.get('sort') or 'prazo'
        filtros_sessao['order'] = filtros_sessao.get('order') or 'asc'

        session['minhas_acoes_filtros'] = filtros_sessao

    origem_id = filtros_sessao.get('origem_id', '')
    status = filtros_sessao.get('status', '')
    data_inicio = filtros_sessao.get('data_inicio', '')
    data_fim = filtros_sessao.get('data_fim', '')
    sort = filtros_sessao.get('sort', 'prazo')
    order = filtros_sessao.get('order', 'asc').lower()

    page = request.args.get('page', 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    cursor.execute("""
        UPDATE acoes
        SET status = 'Vencida'
        WHERE status NOT IN ('Concluída', 'Cancelada')
          AND prazo < %s
          AND ativo = 1
    """, (date.today(),))
    conn.commit()

    def montar_filtros_sql(ignorar=None):
        if ignorar is None:
            ignorar = []

        filtros = ["a.ativo = 1", "a.responsavel_id = %s"]
        valores = [usuario_id]

        if 'origem_id' not in ignorar and origem_id:
            filtros.append("a.origem_id = %s")
            valores.append(origem_id)

        if 'status' not in ignorar and status:
            filtros.append("a.status = %s")
            valores.append(status)

        if 'data_inicio' not in ignorar and data_inicio:
            filtros.append("a.prazo >= %s")
            valores.append(data_inicio)

        if 'data_fim' not in ignorar and data_fim:
            filtros.append("a.prazo <= %s")
            valores.append(data_fim)

        return filtros, valores

    filtros_sql, valores = montar_filtros_sql()
    where_clause = "WHERE " + " AND ".join(filtros_sql)

    campos_ordenacao = {
        'id': 'a.id',
        'descricao': 'a.descricao',
        'nome_criador': 'uc.nome',
        'descricao_origem': 'o.descricao',
        'prazo': 'a.prazo',
        'status': 'a.status'
    }

    campo_sort = campos_ordenacao.get(sort, 'a.prazo')
    direcao = 'DESC' if order == 'desc' else 'ASC'

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        JOIN origens o ON a.origem_id = o.id
        LEFT JOIN usuarios uc ON a.criado_por = uc.id
        {where_clause}
    """, valores)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    query = f"""
        SELECT
            a.*,
            o.descricao AS descricao_origem,
            uc.nome AS nome_criador
        FROM acoes a
        JOIN origens o ON a.origem_id = o.id
        LEFT JOIN usuarios uc ON a.criado_por = uc.id
        {where_clause}
        ORDER BY {campo_sort} {direcao}, a.id DESC
        LIMIT %s OFFSET %s
    """

    cursor.execute(query, valores + [per_page, offset])
    acoes = cursor.fetchall()

    filtros_origem, valores_origem = montar_filtros_sql(ignorar=['origem_id'])
    where_origem = "WHERE " + " AND ".join(filtros_origem)

    cursor.execute(f"""
        SELECT DISTINCT o.id, o.descricao
        FROM acoes a
        JOIN origens o ON a.origem_id = o.id
        {where_origem}
        ORDER BY o.descricao
    """, valores_origem)
    origens = cursor.fetchall()

    filtros_status, valores_status = montar_filtros_sql(ignorar=['status'])
    where_status = "WHERE " + " AND ".join(filtros_status)

    cursor.execute(f"""
        SELECT DISTINCT a.status
        FROM acoes a
        {where_status}
        ORDER BY FIELD(a.status,
            'Não iniciada',
            'Em andamento',
            'Vencida',
            'Concluída',
            'Cancelada'
        ), a.status
    """, valores_status)
    status_opcoes = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    conn.close()

    return render_template(
        'minhas_acoes.html',
        acoes=acoes,
        origens=origens,
        status_opcoes=status_opcoes,
        usuarios=usuarios,
        filtros=filtros_sessao,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

@main_routes.route('/acoes_criadas')
@login_required
@module_required('acesso_plano_acao')
def acoes_criadas():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_usuario_id = session.get('centro_custos_id')

    if request.args.get('limpar') == '1':
        session.pop('acoes_criadas_filtros', None)
        conn.close()
        return redirect(url_for('main.acoes_criadas'))

    filtros_padrao = {
        'superintendencia_id': '',
        'centro_custos_id': '',
        'origem_id': '',
        'status': '',
        'data_inicio': '',
        'data_fim': '',
        'sort': 'prazo',
        'order': 'asc'
    }

    filtros_sessao = session.get('acoes_criadas_filtros', filtros_padrao.copy())

    if request.args:
        for campo in filtros_padrao.keys():
            if campo in request.args:
                filtros_sessao[campo] = request.args.get(campo, '')

        filtros_sessao['sort'] = filtros_sessao.get('sort') or 'prazo'
        filtros_sessao['order'] = filtros_sessao.get('order') or 'asc'

        session['acoes_criadas_filtros'] = filtros_sessao

    superintendencia_id = filtros_sessao.get('superintendencia_id', '')
    centro_custos_id = filtros_sessao.get('centro_custos_id', '')
    origem_id = filtros_sessao.get('origem_id', '')
    status = filtros_sessao.get('status', '')
    data_inicio = filtros_sessao.get('data_inicio', '')
    data_fim = filtros_sessao.get('data_fim', '')
    sort = filtros_sessao.get('sort', 'prazo')
    order = filtros_sessao.get('order', 'asc').lower()

    page = request.args.get('page', 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    cursor.execute("""
        UPDATE acoes
        SET status = 'Vencida'
        WHERE status NOT IN ('Concluída', 'Cancelada')
          AND prazo < %s
          AND ativo = 1
    """, (date.today(),))
    conn.commit()

    def montar_filtros_sql(ignorar=None):
        if ignorar is None:
            ignorar = []

        filtros = ["a.ativo = 1", "a.criado_por = %s"]
        valores = [usuario_id]

        if 'superintendencia_id' not in ignorar and superintendencia_id:
            filtros.append("u.superintendencia_id = %s")
            valores.append(superintendencia_id)

        if 'centro_custos_id' not in ignorar and centro_custos_id:
            filtros.append("u.centro_custos_id = %s")
            valores.append(centro_custos_id)

        if 'origem_id' not in ignorar and origem_id:
            filtros.append("a.origem_id = %s")
            valores.append(origem_id)

        if 'status' not in ignorar and status:
            filtros.append("a.status = %s")
            valores.append(status)

        if 'data_inicio' not in ignorar and data_inicio:
            filtros.append("a.prazo >= %s")
            valores.append(data_inicio)

        if 'data_fim' not in ignorar and data_fim:
            filtros.append("a.prazo <= %s")
            valores.append(data_fim)

        return filtros, valores

    filtros_sql, valores = montar_filtros_sql()
    where_clause = "WHERE " + " AND ".join(filtros_sql)

    campos_ordenacao = {
        'id': 'a.id',
        'descricao': 'a.descricao',
        'nome_responsavel': 'u.nome',
        'descricao_origem': 'o.descricao',
        'prazo': 'a.prazo',
        'status': 'a.status'
    }

    campo_sort = campos_ordenacao.get(sort, 'a.prazo')
    direcao = 'DESC' if order == 'desc' else 'ASC'

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN origens o ON a.origem_id = o.id
        {where_clause}
    """, valores)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    query = f"""
        SELECT
            a.*,
            u.nome AS nome_responsavel,
            o.descricao AS descricao_origem
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN origens o ON a.origem_id = o.id
        {where_clause}
        ORDER BY {campo_sort} {direcao}, a.id DESC
        LIMIT %s OFFSET %s
    """
    cursor.execute(query, valores + [per_page, offset])
    acoes = cursor.fetchall()

    filtros_sup, valores_sup = montar_filtros_sql(ignorar=['superintendencia_id'])
    where_sup = "WHERE " + " AND ".join(filtros_sup)

    cursor.execute(f"""
        SELECT DISTINCT s.id, s.nome
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN superintendencias s ON u.superintendencia_id = s.id
        {where_sup}
        ORDER BY s.nome
    """, valores_sup)
    superintendencias = cursor.fetchall()

    filtros_cc, valores_cc = montar_filtros_sql(ignorar=['centro_custos_id'])
    where_cc = "WHERE " + " AND ".join(filtros_cc)

    cursor.execute(f"""
        SELECT DISTINCT cc.id, cc.codigo, cc.descricao
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN centros_custos cc ON u.centro_custos_id = cc.id
        {where_cc}
        ORDER BY cc.codigo, cc.descricao
    """, valores_cc)
    centros_custos = cursor.fetchall()

    filtros_origem, valores_origem = montar_filtros_sql(ignorar=['origem_id'])
    where_origem = "WHERE " + " AND ".join(filtros_origem)

    cursor.execute(f"""
        SELECT DISTINCT o.id, o.descricao
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        JOIN origens o ON a.origem_id = o.id
        {where_origem}
        ORDER BY o.descricao
    """, valores_origem)
    origens = cursor.fetchall()

    filtros_status, valores_status = montar_filtros_sql(ignorar=['status'])
    where_status = "WHERE " + " AND ".join(filtros_status)

    cursor.execute(f"""
        SELECT DISTINCT a.status
        FROM acoes a
        JOIN usuarios u ON a.responsavel_id = u.id
        {where_status}
        ORDER BY FIELD(a.status,
            'Não iniciada',
            'Em andamento',
            'Vencida',
            'Concluída',
            'Cancelada'
        ), a.status
    """, valores_status)
    status_opcoes = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = TRUE
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_usuario_id,))

    usuarios = cursor.fetchall()

    conn.close()

    return render_template(
        'acoes_criadas.html',
        acoes=acoes,
        superintendencias=superintendencias,
        centros_custos=centros_custos,
        origens=origens,
        status_opcoes=status_opcoes,
        usuarios=usuarios,
        filtros=filtros_sessao,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

@main_routes.route('/exportar_minhas_acoes')
@login_required
@module_required('acesso_plano_acao')
def exportar_minhas_acoes():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')

    origem_id = request.args.get('origem_id')
    status = request.args.get('status')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    query = """
        SELECT 
            a.id,
            a.descricao,
            u.nome AS criado_por,
            o.descricao AS origem,
            a.prazo,
            a.status
        FROM acoes a
        LEFT JOIN usuarios u ON a.criado_por = u.id
        LEFT JOIN origens o ON a.origem_id = o.id
        WHERE a.responsavel_id = %s
          AND a.ativo = 1
    """

    params = [usuario_id]

    if origem_id:
        query += " AND a.origem_id = %s"
        params.append(origem_id)

    if status:
        query += " AND a.status = %s"
        params.append(status)

    if data_inicio:
        query += " AND a.prazo >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND a.prazo <= %s"
        params.append(data_fim)

    query += " ORDER BY a.prazo ASC, a.id ASC"

    cursor.execute(query, params)
    dados = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Minhas Ações"

    cabecalhos = ["ID", "Descrição", "Criado por", "Origem", "Prazo", "Status"]
    ws.append(cabecalhos)

    cor_cabecalho = PatternFill(fill_type="solid", fgColor="F26719")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")
    alinhamento_texto = Alignment(vertical="center", wrap_text=True)

    borda = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for cell in ws[1]:
        cell.fill = cor_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda

    for item in dados:
        prazo_formatado = ""

        if item.get("prazo"):
            try:
                prazo_formatado = item["prazo"].strftime("%d/%m/%Y")
            except AttributeError:
                prazo_formatado = str(item["prazo"])

        ws.append([
            item.get("id", ""),
            item.get("descricao", "") or "",
            item.get("criado_por", "") or "",
            item.get("origem", "") or "",
            prazo_formatado,
            item.get("status", "") or ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento_texto
            cell.border = borda

    larguras = {
        "A": 10,
        "B": 50,
        "C": 25,
        "D": 25,
        "E": 15,
        "F": 18
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="minhas_acoes.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@main_routes.route('/exportar_acoes_criadas')
@login_required
@module_required('acesso_plano_acao')
def exportar_acoes_criadas():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')

    superintendencia_id = request.args.get('superintendencia_id')
    centro_custos_id = request.args.get('centro_custos_id')
    origem_id = request.args.get('origem_id')
    status = request.args.get('status')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    query = """
        SELECT 
            a.id,
            a.descricao,
            u.nome AS responsavel,
            o.descricao AS origem,
            a.prazo,
            a.status
        FROM acoes a
        LEFT JOIN usuarios u ON a.responsavel_id = u.id
        LEFT JOIN origens o ON a.origem_id = o.id
        WHERE a.criado_por = %s
          AND a.ativo = 1
    """

    params = [usuario_id]

    if superintendencia_id:
        query += " AND u.superintendencia_id = %s"
        params.append(superintendencia_id)

    if centro_custos_id:
        query += " AND u.centro_custos_id = %s"
        params.append(centro_custos_id)

    if origem_id:
        query += " AND a.origem_id = %s"
        params.append(origem_id)

    if status:
        query += " AND a.status = %s"
        params.append(status)

    if data_inicio:
        query += " AND a.prazo >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND a.prazo <= %s"
        params.append(data_fim)

    query += " ORDER BY a.prazo ASC, a.id ASC"

    cursor.execute(query, params)
    dados = cursor.fetchall()

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ações Criadas"

    cabecalhos = ["ID", "Descrição", "Responsável", "Origem", "Prazo", "Status"]
    ws.append(cabecalhos)

    cor_cabecalho = PatternFill(fill_type="solid", fgColor="F26719")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")
    alinhamento_texto = Alignment(vertical="center", wrap_text=True)

    borda = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for cell in ws[1]:
        cell.fill = cor_cabecalho
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_cabecalho
        cell.border = borda

    for item in dados:
        prazo_formatado = ""

        if item.get("prazo"):
            try:
                prazo_formatado = item["prazo"].strftime("%d/%m/%Y")
            except AttributeError:
                prazo_formatado = str(item["prazo"])

        ws.append([
            item.get("id", ""),
            item.get("descricao", "") or "",
            item.get("responsavel", "") or "",
            item.get("origem", "") or "",
            prazo_formatado,
            item.get("status", "") or ""
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alinhamento_texto
            cell.border = borda

    larguras = {
        "A": 10,
        "B": 50,
        "C": 25,
        "D": 25,
        "E": 15,
        "F": 18
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="acoes_criadas.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =============================== #
# HORA DE SEGURANÇA               #
# =============================== #

@main_routes.route("/cadastrar_hs")
@login_required
@admin_required
def cadastrar_hs():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Busca todos os temas
    cursor.execute("SELECT * FROM hs_temas ORDER BY id DESC")
    temas = cursor.fetchall()

    # Se nenhum tema ainda foi cadastrado
    tema_atual = None
    itens_tema = []

    # Se o usuário passou um id_tema na URL, foca nele
    id_tema = request.args.get("id_tema")
    if id_tema:
        cursor.execute("SELECT * FROM hs_temas WHERE id=%s", (id_tema,))
        tema_atual = cursor.fetchone()

        if tema_atual:
            cursor.execute("""
                SELECT i.*
                FROM hs_itens_verificacao i
                WHERE i.id_tema = %s
                ORDER BY i.ordem
            """, (id_tema,))
            itens_tema = cursor.fetchall()

    conn.close()
    return render_template("cadastrar_hs.html", temas=temas, tema_atual=tema_atual, itens_tema=itens_tema)


@main_routes.route("/cadastrar_hs/tema", methods=["POST"])
@login_required
@admin_required
def cadastrar_hs_tema():
    nome_tema = request.form.get("nome_tema")

    if not nome_tema:
        flash("O nome do tema é obrigatório.", "danger")
        return redirect(url_for("main.cadastrar_hs"))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO hs_temas (nome, status) VALUES (%s, 1)", (nome_tema,))
    id_tema = cursor.lastrowid
    conn.commit()
    conn.close()

    flash("Tema cadastrado com sucesso!", "success")
    return redirect(url_for("main.cadastrar_hs", id_tema=id_tema))


@main_routes.route("/cadastrar_hs/item", methods=["POST"])
@login_required
@admin_required
def cadastrar_hs_item():
    id_tema = request.form.get("id_tema")
    texto = request.form.get("texto")
    ordem = request.form.get("ordem") or None

    if not id_tema or not texto:
        flash("Preencha todos os campos obrigatórios.", "danger")
        return redirect(url_for("main.cadastrar_hs"))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO hs_itens_verificacao (id_tema, texto, ordem, status)
        VALUES (%s, %s, %s, 1)
    """, (id_tema, texto, ordem))
    conn.commit()
    conn.close()

    flash("Item cadastrado com sucesso!", "success")
    return redirect(url_for("main.cadastrar_hs", id_tema=id_tema))


@main_routes.route("/editar_hs/item/<int:id>", methods=["GET", "POST"])
@login_required
@admin_required
def editar_hs_item(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        texto = request.form.get("texto")
        ordem = request.form.get("ordem") or None
        if not texto:
            flash("A descrição do item é obrigatória.", "danger")
            return redirect(url_for("main.editar_hs_item", id=id))

        cursor.execute(
            "UPDATE hs_itens_verificacao SET texto=%s, ordem=%s WHERE id=%s",
            (texto, ordem, id)
        )
        conn.commit()
        cursor.execute("SELECT id_tema FROM hs_itens_verificacao WHERE id=%s", (id,))
        item = cursor.fetchone()
        conn.close()

        flash("Item atualizado com sucesso!", "success")
        return redirect(url_for("main.cadastrar_hs", id_tema=item["id_tema"]))

    cursor.execute("SELECT * FROM hs_itens_verificacao WHERE id=%s", (id,))
    item = cursor.fetchone()
    conn.close()
    return render_template("editar_hs_item.html", item=item)


@main_routes.route("/habilitar_hs/tema/<int:id>")
@login_required
@admin_required
def habilitar_hs_tema(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE hs_temas SET status=1 WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Tema habilitado!", "success")
    return redirect(url_for("main.cadastrar_hs", id_tema=id))


@main_routes.route("/desabilitar_hs/tema/<int:id>")
@login_required
@admin_required
def desabilitar_hs_tema(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE hs_temas SET status=0 WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Tema desabilitado!", "warning")
    return redirect(url_for("main.cadastrar_hs", id_tema=id))


@main_routes.route("/habilitar_hs/item/<int:id>")
@login_required
@admin_required
def habilitar_hs_item(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id_tema FROM hs_itens_verificacao WHERE id=%s", (id,))
    item = cursor.fetchone()

    cursor.execute("UPDATE hs_itens_verificacao SET status=1 WHERE id=%s", (id,))
    conn.commit()
    conn.close()

    flash("Item habilitado!", "success")
    return redirect(url_for("main.cadastrar_hs", id_tema=item["id_tema"]))


@main_routes.route("/desabilitar_hs/item/<int:id>")
@login_required
@admin_required
def desabilitar_hs_item(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id_tema FROM hs_itens_verificacao WHERE id=%s", (id,))
    item = cursor.fetchone()

    cursor.execute("UPDATE hs_itens_verificacao SET status=0 WHERE id=%s", (id,))
    conn.commit()
    conn.close()

    flash("Item desabilitado!", "warning")
    return redirect(url_for("main.cadastrar_hs", id_tema=item["id_tema"]))


@main_routes.route("/lancar_hs", methods=["GET", "POST"])
@login_required
@module_required('acesso_ssma')
def lancar_hs():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get('perfil')
    usuario_id = session.get('usuario_id')
    centro_custos_id = session.get('centro_custos_id')

    def calcular_turno(hora_str):
        if not hora_str:
            return None

        try:
            hora_int = int(hora_str.split(":")[0])
        except (ValueError, IndexError):
            return None

        if 0 <= hora_int < 12:
            return "Manhã"
        elif 12 <= hora_int < 18:
            return "Tarde"
        else:
            return "Noite"

    if request.method == "POST":
        data = request.form.get("data")
        hora = request.form.get("hora")
        turno = calcular_turno(hora)
        local = request.form.get("local")
        id_tema = request.form.get("id_tema")

        # Auditor sempre será o usuário logado
        id_auditor = usuario_id

        participantes = request.form.getlist("participantes")

        if not (data and hora and id_tema and id_auditor and turno):
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(url_for("main.lancar_hs"))

        cursor.execute("SELECT id FROM origens WHERE descricao = %s", ("Hora de Segurança",))
        origem = cursor.fetchone()

        if origem:
            origem_hs = origem["id"]
        else:
            cursor.execute("""
                INSERT INTO origens (descricao, ativo)
                VALUES (%s, 1)
            """, ("Hora de Segurança",))
            origem_hs = cursor.lastrowid

        cursor.execute("""
            INSERT INTO hs_registros 
                (data, hora, turno, local, id_tema, id_auditor, participantes)
            VALUES 
                (%s, %s, %s, %s, %s, %s, %s)
        """, (
            data,
            hora,
            turno,
            local,
            id_tema,
            id_auditor,
            ",".join(participantes)
        ))

        id_registro = cursor.lastrowid

        cursor.execute("""
            SELECT *
            FROM hs_itens_verificacao
            WHERE id_tema = %s
              AND status = 1
        """, (id_tema,))

        itens = cursor.fetchall()

        for item in itens:
            resultado = request.form.get(f"resultado_{item['id']}")
            desvio = (request.form.get(f"desvio_{item['id']}") or "").strip()
            acao = (request.form.get(f"acao_{item['id']}") or "").strip()
            prazo = request.form.get(f"prazo_{item['id']}")

            id_acao_gerada = None

            if resultado == "NC":
                if not desvio or not acao or not prazo:
                    flash(
                        f"O item '{item['texto']}' foi marcado como NC, mas falta preencher desvio, ação ou prazo.",
                        "danger"
                    )
                    conn.rollback()
                    conn.close()
                    return redirect(url_for("main.lancar_hs", id_tema=id_tema))
            else:
                desvio = ""
                acao = ""
                prazo = None

            if resultado == "NC":
                cursor.execute("""
                    INSERT INTO acoes 
                        (origem_id, responsavel_id, descricao, prazo, status, criado_por)
                    VALUES 
                        (%s, %s, %s, %s, %s, %s)
                """, (
                    origem_hs,
                    id_auditor,
                    acao,
                    prazo,
                    "Não iniciada",
                    id_auditor
                ))

                id_acao_gerada = cursor.lastrowid

            cursor.execute("""
                INSERT INTO hs_respostas 
                    (id_registro, id_item, resultado, descricao_desvio, descricao_acao, id_acao_gerada)
                VALUES 
                    (%s, %s, %s, %s, %s, %s)
            """, (
                id_registro,
                item["id"],
                resultado,
                desvio,
                acao,
                id_acao_gerada
            ))

        conn.commit()
        conn.close()

        flash("Hora de Segurança registrada com sucesso!", "success")
        return redirect(url_for("main.lancar_hs"))

    # ---------- GET ----------
    cursor.execute("""
        SELECT *
        FROM hs_temas
        WHERE status = 1
        ORDER BY nome
    """)
    temas = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    id_tema = request.args.get("id_tema")
    itens = []

    if id_tema:
        cursor.execute("""
            SELECT *
            FROM hs_itens_verificacao
            WHERE id_tema = %s
              AND status = 1
            ORDER BY ordem
        """, (id_tema,))
        itens = cursor.fetchall()

    query_registros = """
        SELECT 
            r.*, 
            t.nome AS nome_tema, 
            u.nome AS nome_auditor
        FROM hs_registros r
        JOIN hs_temas t ON r.id_tema = t.id
        JOIN usuarios u ON r.id_auditor = u.id
        WHERE 1 = 1
    """

    params_registros = []

    if perfil == 'basico':
        query_registros += " AND r.id_auditor = %s"
        params_registros.append(usuario_id)

    elif perfil == 'intermediario':
        query_registros += " AND u.centro_custos_id = %s"
        params_registros.append(centro_custos_id)

    query_registros += """
        ORDER BY r.data DESC, r.hora DESC
        LIMIT 20
    """

    cursor.execute(query_registros, params_registros)
    registros = cursor.fetchall()

    for r in registros:
        if isinstance(r["hora"], timedelta):
            total_seconds = r["hora"].seconds
            horas = total_seconds // 3600
            minutos = (total_seconds % 3600) // 60
            r["hora"] = f"{horas:02d}:{minutos:02d}"
        elif hasattr(r["hora"], "strftime"):
            r["hora"] = r["hora"].strftime("%H:%M")

        if hasattr(r["data"], "strftime"):
            r["data"] = r["data"].strftime("%d/%m/%Y")

    conn.close()

    return render_template(
        "lancar_hs.html",
        hoje=datetime.today().strftime("%Y-%m-%d"),
        agora=datetime.now().strftime("%H:%M"),
        temas=temas,
        usuarios=usuarios,
        usuarios_json=json.dumps(
            [{"id": u["id"], "nome": u["nome"]} for u in usuarios],
            ensure_ascii=False
        ),
        itens=itens,
        registros=registros
    )

@main_routes.route("/editar_hs/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_ssma')
def editar_hs(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    def calcular_turno(hora_str):
        if not hora_str:
            return None

        try:
            hora_int = int(hora_str.split(":")[0])
        except (ValueError, IndexError):
            return None

        if 0 <= hora_int < 12:
            return "Manhã"
        elif 12 <= hora_int < 18:
            return "Tarde"
        else:
            return "Noite"

    # 🔒 NOVO: validação centralizada
    registro = pode_acessar_ssma(cursor, 'hs', id)

    if not registro:
        conn.close()
        flash("Hora de Segurança não encontrada ou você não possui permissão para acessá-la.", "warning")
        return redirect(url_for("main.listar_hs"))

    # 🔒 REGRA DE NEGÓCIO MANTIDA (não alterei)
    if session.get("perfil") != "administrador" and registro["id_auditor"] != session.get("usuario_id"):
        conn.close()
        flash("Você não tem permissão para editar esta Hora de Segurança.", "warning")
        return redirect(url_for("main.listar_hs"))

    if request.method == "POST":
        next_url = request.form.get("next") or url_for("main.listar_hs")

        data = request.form.get("data")
        hora = request.form.get("hora")
        turno = calcular_turno(hora)
        local = request.form.get("local")
        id_tema = request.form.get("id_tema")
        id_auditor = request.form.get("id_auditor")
        participantes = request.form.getlist("participantes")

        if not (data and hora and id_tema and id_auditor and turno):
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(next_url)

        if session.get("perfil") != "administrador":
            id_auditor = registro["id_auditor"]

        try:
            cursor.execute("""
                UPDATE hs_registros
                SET data=%s,
                    hora=%s,
                    turno=%s,
                    local=%s,
                    id_tema=%s,
                    id_auditor=%s,
                    participantes=%s
                WHERE id=%s
            """, (
                data,
                hora,
                turno,
                local,
                id_tema,
                id_auditor,
                ",".join(participantes),
                id
            ))

            cursor.execute("""
                SELECT *
                FROM hs_itens_verificacao
                WHERE id_tema=%s
                  AND status=1
            """, (id_tema,))
            itens = cursor.fetchall()

            for item in itens:
                resultado = request.form.get(f"resultado_{item['id']}")
                desvio = (request.form.get(f"desvio_{item['id']}") or "").strip()
                acao = (request.form.get(f"acao_{item['id']}") or "").strip()
                prazo = request.form.get(f"prazo_{item['id']}")

                cursor.execute("""
                    SELECT id, id_acao_gerada
                    FROM hs_respostas
                    WHERE id_registro=%s
                      AND id_item=%s
                """, (id, item["id"]))
                resposta_existente = cursor.fetchone()

                id_acao_gerada = resposta_existente["id_acao_gerada"] if resposta_existente else None

                if resultado == "NC":
                    if not desvio or not acao or not prazo:
                        flash(
                            f"O item '{item['texto']}' foi marcado como NC, mas falta preencher desvio, ação ou prazo.",
                            "danger"
                        )
                        conn.rollback()
                        conn.close()
                        return redirect(next_url)
                else:
                    desvio = ""
                    acao = ""
                    prazo = None

                if resultado == "NC":
                    if id_acao_gerada:
                        cursor.execute("""
                            UPDATE acoes
                            SET descricao=%s,
                                prazo=%s,
                                responsavel_id=%s
                            WHERE id=%s
                        """, (acao, prazo, id_auditor, id_acao_gerada))
                    else:
                        cursor.execute("SELECT id FROM origens WHERE descricao=%s", ("Hora de Segurança",))
                        origem = cursor.fetchone()

                        if origem:
                            origem_hs = origem["id"]
                        else:
                            cursor.execute(
                                "INSERT INTO origens (descricao, ativo) VALUES (%s, 1)",
                                ("Hora de Segurança",)
                            )
                            origem_hs = cursor.lastrowid

                        cursor.execute("""
                            INSERT INTO acoes (
                                origem_id,
                                responsavel_id,
                                descricao,
                                prazo,
                                status,
                                criado_por
                            )
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            origem_hs,
                            id_auditor,
                            acao,
                            prazo,
                            "Não iniciada",
                            id_auditor
                        ))
                        id_acao_gerada = cursor.lastrowid

                else:
                    if id_acao_gerada:
                        cursor.execute(
                            "UPDATE hs_respostas SET id_acao_gerada=NULL WHERE id=%s",
                            (resposta_existente["id"],)
                        )
                        cursor.execute("DELETE FROM acoes WHERE id=%s", (id_acao_gerada,))
                        id_acao_gerada = None

                if resposta_existente:
                    cursor.execute("""
                        UPDATE hs_respostas
                        SET resultado=%s,
                            descricao_desvio=%s,
                            descricao_acao=%s,
                            id_acao_gerada=%s
                        WHERE id=%s
                    """, (
                        resultado,
                        desvio,
                        acao,
                        id_acao_gerada,
                        resposta_existente["id"]
                    ))
                else:
                    cursor.execute("""
                        INSERT INTO hs_respostas (
                            id_registro,
                            id_item,
                            resultado,
                            descricao_desvio,
                            descricao_acao,
                            id_acao_gerada
                        )
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        id,
                        item["id"],
                        resultado,
                        desvio,
                        acao,
                        id_acao_gerada
                    ))

            conn.commit()
            flash("Hora de Segurança atualizada com sucesso!", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar Hora de Segurança: {e}", "danger")

        finally:
            conn.close()

        return redirect(next_url)

    cursor.execute("SELECT * FROM hs_temas WHERE status=1 ORDER BY nome")
    temas = cursor.fetchall()

    cursor.execute("SELECT id, nome FROM usuarios WHERE ativo=1 ORDER BY nome")
    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            i.*,
            r.resultado,
            r.descricao_desvio,
            r.descricao_acao,
            r.id_acao_gerada,
            a.prazo AS prazo_acao
        FROM hs_itens_verificacao i
        LEFT JOIN hs_respostas r
            ON i.id = r.id_item
           AND r.id_registro = %s
        LEFT JOIN acoes a
            ON a.id = r.id_acao_gerada
        WHERE i.id_tema = %s
        ORDER BY i.ordem
    """, (id, registro["id_tema"]))
    itens = cursor.fetchall()

    conn.close()

    return render_template(
        "editar_hs.html",
        registro=registro,
        temas=temas,
        usuarios=usuarios,
        itens=itens
    )

@main_routes.route("/excluir_hs/<int:id>", methods=["POST"])
@login_required
@module_required('acesso_ssma')
def excluir_hs(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    registro = pode_acessar_ssma(cursor, 'hs', id)

    if not registro:
        conn.close()
        flash("Hora de Segurança não encontrada ou você não possui permissão para excluí-la.", "warning")
        return redirect(url_for("main.listar_hs"))

    # Regra adicional: somente administrador ou auditor que lançou pode excluir
    if session.get("perfil") != "administrador" and registro["id_auditor"] != session.get("usuario_id"):
        conn.close()
        flash("Você não tem permissão para excluir esta Hora de Segurança.", "warning")
        return redirect(url_for("main.listar_hs"))

    cursor.execute("""
        SELECT id_acao_gerada
        FROM hs_respostas
        WHERE id_registro = %s
          AND id_acao_gerada IS NOT NULL
    """, (id,))
    acoes = cursor.fetchall()

    cursor.execute("DELETE FROM hs_respostas WHERE id_registro = %s", (id,))

    for ac in acoes:
        cursor.execute("DELETE FROM acoes WHERE id = %s", (ac["id_acao_gerada"],))

    cursor.execute("DELETE FROM hs_registros WHERE id = %s", (id,))

    conn.commit()
    conn.close()

    flash("Hora de Segurança e ações vinculadas excluídas com sucesso!", "success")
    return redirect(url_for("main.listar_hs"))


@main_routes.route('/listar_hs', methods=['GET'])
@login_required
@module_required('acesso_ssma')
def listar_hs():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.listar_hs'))

    auditor_id = request.args.get('auditor_id', '')
    tema_id = request.args.get('tema_id', '')
    turno = request.args.get('turno', '')
    local = request.args.get('local', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    resultado_nc = request.args.get('resultado_nc', '')

    sort = request.args.get('sort', 'data')
    order = request.args.get('order', 'desc')

    page = request.args.get('page', 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        'id': 'r.id',
        'data': 'r.data',
        'hora': 'r.hora',
        'turno': 'r.turno',
        'tema': 't.nome',
        'auditor': 'u.nome',
        'local': 'r.local',
        'possui_nc': 'possui_nc'
    }

    coluna_sort = colunas_validas.get(sort, 'r.data')
    direcao = 'ASC' if order == 'asc' else 'DESC'

    filtros_where = """
        FROM hs_registros r
        JOIN hs_temas t ON r.id_tema = t.id
        JOIN usuarios u ON r.id_auditor = u.id
        LEFT JOIN hs_respostas resp ON resp.id_registro = r.id
        WHERE 1=1
    """

    params = []

    # CONTROLE DE ESCOPO POR PERFIL
    if perfil == 'basico':
        filtros_where += " AND r.id_auditor = %s"
        params.append(usuario_id)

    elif perfil == 'intermediario':
        filtros_where += " AND u.centro_custos_id = %s"
        params.append(centro_custos_id)

    # avancado e administrador veem tudo

    if auditor_id:
        filtros_where += " AND r.id_auditor = %s"
        params.append(auditor_id)

    if tema_id:
        filtros_where += " AND r.id_tema = %s"
        params.append(tema_id)

    if turno:
        filtros_where += " AND r.turno = %s"
        params.append(turno)

    if local:
        filtros_where += " AND r.local LIKE %s"
        params.append(f"%{local}%")

    if data_inicio:
        filtros_where += " AND r.data >= %s"
        params.append(data_inicio)

    if data_fim:
        filtros_where += " AND r.data <= %s"
        params.append(data_fim)

    group_by = """
        GROUP BY
            r.id,
            r.id_auditor,
            r.id_tema,
            r.data,
            r.hora,
            r.turno,
            r.local,
            r.participantes,
            t.nome,
            u.nome
    """

    having_clause = ""
    if resultado_nc == 'sim':
        having_clause = " HAVING possui_nc = 1"
    elif resultado_nc == 'nao':
        having_clause = " HAVING possui_nc = 0"

    count_query = f"""
        SELECT COUNT(*) AS total
        FROM (
            SELECT
                r.id,
                MAX(CASE WHEN resp.resultado = 'NC' THEN 1 ELSE 0 END) AS possui_nc
            {filtros_where}
            GROUP BY r.id
            {having_clause}
        ) AS subquery
    """

    cursor.execute(count_query, params)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    query = f"""
        SELECT
            r.id,
            r.id_auditor,
            r.id_tema,
            r.data,
            r.hora,
            r.turno,
            r.local,
            r.participantes,
            t.nome AS nome_tema,
            u.nome AS nome_auditor,
            MAX(CASE WHEN resp.resultado = 'NC' THEN 1 ELSE 0 END) AS possui_nc
        {filtros_where}
        {group_by}
        {having_clause}
        ORDER BY {coluna_sort} {direcao}
        LIMIT %s OFFSET %s
    """

    cursor.execute(query, params + [per_page, offset])
    registros = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()
    mapa_usuarios = {str(u['id']): u['nome'] for u in usuarios}

    for r in registros:
        if r.get('participantes'):
            ids = [p.strip() for p in r['participantes'].split(',') if p.strip()]
            nomes = [mapa_usuarios.get(pid, f'ID {pid}') for pid in ids]
            r['nomes_participantes'] = ', '.join(nomes)
        else:
            r['nomes_participantes'] = ''

        if r.get('hora') and hasattr(r['hora'], 'strftime'):
            r['hora'] = r['hora'].strftime('%H:%M')

        if r.get('data') and hasattr(r['data'], 'strftime'):
            r['data_iso'] = r['data'].strftime('%Y-%m-%d')
        else:
            r['data_iso'] = r.get('data') or ''

    cursor.execute("SELECT id, nome FROM hs_temas WHERE status = 1 ORDER BY nome")
    temas = cursor.fetchall()

    ids_registros = [r["id"] for r in registros]
    itens_por_registro = {}

    if ids_registros:
        placeholders = ", ".join(["%s"] * len(ids_registros))

        cursor.execute(f"""
            SELECT
                r.id AS id_registro,
                i.id AS id_item,
                i.texto,
                i.ordem,
                resp.resultado,
                resp.descricao_desvio,
                resp.descricao_acao,
                resp.id_acao_gerada,
                a.prazo AS prazo_acao
            FROM hs_registros r
            JOIN hs_itens_verificacao i
                ON i.id_tema = r.id_tema
               AND i.status = 1
            LEFT JOIN hs_respostas resp
                ON resp.id_item = i.id
               AND resp.id_registro = r.id
            LEFT JOIN acoes a
                ON a.id = resp.id_acao_gerada
            WHERE r.id IN ({placeholders})
            ORDER BY r.id, i.ordem
        """, ids_registros)

        itens_modal = cursor.fetchall()

        for item in itens_modal:
            registro_id = str(item["id_registro"])

            if item.get("prazo_acao") and hasattr(item["prazo_acao"], "strftime"):
                item["prazo_acao"] = item["prazo_acao"].strftime("%Y-%m-%d")

            itens_por_registro.setdefault(registro_id, []).append(item)

    conn.close()

    filtros = {
        'auditor_id': auditor_id,
        'tema_id': tema_id,
        'turno': turno,
        'local': local,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'resultado_nc': resultado_nc,
        'sort': sort,
        'order': order
    }

    return render_template(
        'listar_hs.html',
        registros=registros,
        usuarios=usuarios,
        temas=temas,
        filtros=filtros,
        itens_por_registro=itens_por_registro,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

# =============================== #
# RECUSA À TAREFA                 #
# =============================== #


@main_routes.route("/causas_recusa", methods=["GET", "POST"])
@login_required
@admin_required
def causas_recusa():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        descricao = request.form.get("descricao")

        if not descricao:
            flash("Informe a descrição da causa.", "danger")
            conn.close()
            return redirect(url_for("main.causas_recusa"))

        cursor.execute(
            "INSERT INTO causas_recusa (descricao, ativo) VALUES (%s, 1)",
            (descricao,)
        )
        conn.commit()
        flash("Causa de recusa cadastrada com sucesso!", "success")

    # Lista todas as causas já cadastradas
    cursor.execute("SELECT * FROM causas_recusa ORDER BY descricao")
    causas = cursor.fetchall()

    conn.close()
    return render_template("causas_recusa.html", causas=causas)


@main_routes.route("/editar_causa/<int:id>", methods=["GET", "POST"])
@login_required
@admin_required
def editar_causa(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        descricao = request.form.get("descricao")

        if not descricao:
            flash("Informe a descrição da causa.", "danger")
            conn.close()
            return redirect(url_for("main.editar_causa", id=id))

        cursor.execute(
            "UPDATE causas_recusa SET descricao=%s WHERE id=%s",
            (descricao, id)
        )
        conn.commit()
        flash("Causa de recusa atualizada com sucesso!", "success")
        conn.close()
        return redirect(url_for("main.causas_recusa"))

    cursor.execute("SELECT * FROM causas_recusa WHERE id=%s", (id,))
    causa = cursor.fetchone()
    conn.close()

    if not causa:
        flash("Causa não encontrada.", "danger")
        return redirect(url_for("main.causas_recusa"))

    return render_template("editar_causa.html", causa=causa)



@main_routes.route("/desativar_causa/<int:id>", methods=["GET"])
@login_required
@admin_required
def desativar_causa(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE causas_recusa SET ativo=0 WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Causa desativada com sucesso!", "success")
    return redirect(url_for("main.causas_recusa"))


@main_routes.route("/ativar_causa/<int:id>", methods=["GET"])
@login_required
@admin_required
def ativar_causa(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE causas_recusa SET ativo=1 WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash("Causa ativada com sucesso!", "success")
    return redirect(url_for("main.causas_recusa"))


@main_routes.route("/lancar_recusa", methods=["GET", "POST"])
@login_required
@module_required('acesso_ssma')
def lancar_recusa():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custos_id = session.get("centro_custos_id")

    if not usuario_logado_id:
        flash("Usuário logado não encontrado.", "danger")
        conn.close()
        return redirect(url_for("main.login"))

    if request.method == "POST":
        data_recusa = request.form.get("data")
        hora_recusa = request.form.get("hora")
        id_usuario = request.form.get("usuario_id")
        local = request.form.get("local")
        classificacao = request.form.get("classificacao")
        descricao = request.form.get("descricao")
        potencial_severidade = request.form.get("potencial")
        id_causa = request.form.get("causa_id")

        criado_por = usuario_logado_id

        if not (data_recusa and hora_recusa and id_usuario and classificacao and descricao and potencial_severidade and id_causa):
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(url_for("main.lancar_recusa"))

        if perfil in ["administrador", "avancado"]:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
            """, (id_usuario,))
        else:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (id_usuario, centro_custos_id))

        usuario_valido = cursor.fetchone()

        if not usuario_valido:
            flash("Usuário selecionado não pertence ao seu centro de custo.", "danger")
            conn.close()
            return redirect(url_for("main.lancar_recusa"))

        cursor.execute("""
            INSERT INTO recusa_tarefa
                (data_recusa, hora_recusa, id_usuario, local, classificacao, descricao, potencial_severidade, id_causa, criado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data_recusa,
            hora_recusa,
            id_usuario,
            local,
            classificacao,
            descricao,
            potencial_severidade,
            id_causa,
            criado_por
        ))

        conn.commit()
        flash("Recusa registrada com sucesso!", "success")
        conn.close()
        return redirect(url_for("main.lancar_recusa"))

    if perfil in ["administrador", "avancado"]:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT id, descricao
        FROM causas_recusa
        WHERE ativo = 1
        ORDER BY descricao
    """)
    causas = cursor.fetchall()

    conn.close()

    return render_template(
        "lancar_recusa.html",
        hoje=datetime.today().strftime("%Y-%m-%d"),
        agora=datetime.now().strftime("%H:%M"),
        usuarios=usuarios,
        causas=causas
    )


@main_routes.route("/editar_recusa/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_ssma')
def editar_recusa(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get("perfil")
    centro_custos_id = session.get("centro_custos_id")

    recusa = pode_acessar_ssma(cursor, 'recusa', id)

    if not recusa:
        conn.close()
        flash("Recusa não encontrada ou você não possui permissão para editá-la.", "warning")
        return redirect(url_for("main.listar_recusa"))

    if request.method == "POST":
        data_recusa = request.form.get("data")
        hora_recusa = request.form.get("hora")
        id_usuario = request.form.get("usuario_id")
        local = request.form.get("local")
        classificacao = request.form.get("classificacao")
        descricao = request.form.get("descricao")
        potencial_severidade = request.form.get("potencial")
        id_causa = request.form.get("causa_id")

        if not (data_recusa and hora_recusa and id_usuario and classificacao and descricao and potencial_severidade and id_causa):
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(url_for("main.editar_recusa", id=id))

        if perfil in ["administrador", "avancado"]:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
            """, (id_usuario,))
        else:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (id_usuario, centro_custos_id))

        usuario_valido = cursor.fetchone()

        if not usuario_valido:
            conn.close()
            flash("Usuário selecionado não pertence ao seu centro de custo.", "danger")
            return redirect(url_for("main.editar_recusa", id=id))

        cursor.execute("""
            UPDATE recusa_tarefa
            SET data_recusa=%s, hora_recusa=%s, id_usuario=%s, local=%s,
                classificacao=%s, descricao=%s, potencial_severidade=%s, id_causa=%s
            WHERE id=%s
        """, (
            data_recusa, hora_recusa, id_usuario, local,
            classificacao, descricao, potencial_severidade, id_causa, id
        ))

        conn.commit()
        conn.close()

        flash("Recusa atualizada com sucesso!", "success")
        return redirect(url_for("main.listar_recusa"))

    cursor.execute("""
        SELECT r.*, u.nome AS usuario_nome, c.descricao AS causa_nome
        FROM recusa_tarefa r
        JOIN usuarios u ON r.id_usuario = u.id
        JOIN causas_recusa c ON r.id_causa = c.id
        WHERE r.id = %s
    """, (id,))
    recusa = cursor.fetchone()

    if recusa:
        if recusa["data_recusa"]:
            if hasattr(recusa["data_recusa"], "strftime"):
                recusa["data_recusa"] = recusa["data_recusa"].strftime("%Y-%m-%d")
            else:
                try:
                    recusa["data_recusa"] = datetime.strptime(str(recusa["data_recusa"]), "%Y-%m-%d").strftime("%Y-%m-%d")
                except:
                    recusa["data_recusa"] = ""

        if recusa["hora_recusa"]:
            if hasattr(recusa["hora_recusa"], "strftime"):
                recusa["hora_recusa"] = recusa["hora_recusa"].strftime("%H:%M")
            else:
                try:
                    recusa["hora_recusa"] = datetime.strptime(str(recusa["hora_recusa"]), "%H:%M:%S").strftime("%H:%M")
                except:
                    recusa["hora_recusa"] = ""

    if perfil in ["administrador", "avancado"]:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT id, descricao
        FROM causas_recusa
        WHERE ativo = 1
        ORDER BY descricao
    """)
    causas = cursor.fetchall()

    conn.close()

    return render_template(
        "editar_recusa.html",
        recusa=recusa,
        usuarios=usuarios,
        causas=causas
    )

@main_routes.route("/excluir_recusa/<int:id>", methods=["POST"])
@login_required
@module_required('acesso_ssma')
def excluir_recusa(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    recusa = pode_acessar_ssma(cursor, 'recusa', id)

    if not recusa:
        conn.close()
        flash("Recusa não encontrada ou você não possui permissão para excluí-la.", "warning")
        return redirect(url_for("main.listar_recusa"))

    cursor.execute("DELETE FROM recusa_tarefa WHERE id = %s", (id,))
    conn.commit()
    conn.close()

    flash("Recusa excluída com sucesso!", "success")
    return redirect(url_for("main.listar_recusa"))

@main_routes.route("/listar_recusa", methods=["GET"])
@login_required
def listar_recusa():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id") or session.get("user_id") or session.get("id")

    cursor.execute("""
        SELECT id, perfil, centro_custos_id
        FROM usuarios
        WHERE id = %s
    """, (usuario_logado_id,))
    usuario_logado = cursor.fetchone()

    if not usuario_logado:
        conn.close()
        flash("Usuário logado não encontrado.", "danger")
        return redirect(url_for("main.login"))

    if request.args.get("limpar"):
        conn.close()
        return redirect(url_for("main.listar_recusa"))

    usuario_id = request.args.get("usuario_id", "")
    classificacao = request.args.get("classificacao", "")
    potencial = request.args.get("potencial", "")
    causa_id = request.args.get("causa_id", "")
    local = request.args.get("local", "")
    criado_por_mim = request.args.get("criado_por_mim", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    sort = request.args.get("sort", "data")
    order = request.args.get("order", "desc")

    page = request.args.get("page", 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        "id": "r.id",
        "data": "r.data_recusa",
        "usuario": "u.nome",
        "classificacao": "r.classificacao",
        "potencial": "r.potencial_severidade"
    }

    coluna_sort = colunas_validas.get(sort, "r.data_recusa")
    direcao = "ASC" if order == "asc" else "DESC"

    base_from = """
        FROM recusa_tarefa r
        JOIN usuarios u ON r.id_usuario = u.id
        JOIN causas_recusa c ON r.id_causa = c.id
        WHERE 1=1
    """

    params = []

    perfil = usuario_logado["perfil"]

    if perfil == "basico":
        base_from += " AND r.criado_por = %s"
        params.append(usuario_logado_id)

    elif perfil == "intermediario":
        base_from += " AND u.centro_custos_id = %s"
        params.append(usuario_logado["centro_custos_id"])

    # avançado e administrador veem tudo

    if usuario_id:
        base_from += " AND r.id_usuario = %s"
        params.append(usuario_id)

    if classificacao:
        base_from += " AND r.classificacao = %s"
        params.append(classificacao)

    if potencial:
        base_from += " AND r.potencial_severidade = %s"
        params.append(potencial)

    if causa_id:
        base_from += " AND r.id_causa = %s"
        params.append(causa_id)

    if local:
        base_from += " AND r.local LIKE %s"
        params.append(f"%{local}%")

    if criado_por_mim == "sim":
        base_from += " AND r.criado_por = %s"
        params.append(usuario_logado_id)
    elif criado_por_mim == "nao":
        base_from += " AND r.criado_por <> %s"
        params.append(usuario_logado_id)

    if data_inicio:
        base_from += " AND r.data_recusa >= %s"
        params.append(data_inicio)

    if data_fim:
        base_from += " AND r.data_recusa <= %s"
        params.append(data_fim)

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """, params)
    total_registros = cursor.fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cursor.execute(f"""
        SELECT
            r.*,
            u.nome AS usuario_nome,
            c.descricao AS causa_nome
        {base_from}
        ORDER BY {coluna_sort} {direcao}, r.hora_recusa DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    registros = cursor.fetchall()

    if usuario_logado["perfil"] == "administrador":
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (usuario_logado["centro_custos_id"],))

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT id, descricao
        FROM causas_recusa
        WHERE ativo = 1
        ORDER BY descricao
    """)
    causas = cursor.fetchall()

    conn.close()

    filtros = {
        "usuario_id": usuario_id,
        "classificacao": classificacao,
        "potencial": potencial,
        "causa_id": causa_id,
        "local": local,
        "criado_por_mim": criado_por_mim,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "sort": sort,
        "order": order
    }

    return render_template(
        "listar_recusa.html",
        registros=registros,
        usuarios=usuarios,
        causas=causas,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

# =============================== #
# RECONHECIMENTOS                 #
# =============================== #

@main_routes.route("/lancar_reconhecimento", methods=["GET", "POST"])
@login_required
@module_required('acesso_gestao_pessoas')
def lancar_reconhecimento():
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id_logado = session.get("centro_custos_id")

    if not usuario_logado_id or not centro_custo_id_logado:
        cur.close()
        conn.close()
        flash("Usuário logado inválido.", "danger")
        return redirect(url_for("main.dashboard"))

    if perfil in ["administrador", "avancado"]:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id_logado,))

    usuarios = cur.fetchall()

    if request.method == "POST":
        apoiador_id = request.form.get("apoiador_id")
        id_reconhecido = request.form.get("id_reconhecido")
        data_reconhecimento = request.form.get("data_reconhecimento")
        reconhecimento = request.form.get("reconhecimento")
        criado_por = usuario_logado_id

        if not (apoiador_id and id_reconhecido and data_reconhecimento and reconhecimento and criado_por):
            flash("Preencha todos os campos obrigatórios.", "warning")
            cur.close()
            conn.close()

            return render_template(
                "lancar_reconhecimento.html",
                usuarios=usuarios,
                current_date=date.today().isoformat()
            )

        if perfil in ["administrador", "avancado"]:
            cur.execute("""
                SELECT id
                FROM usuarios
                WHERE id IN (%s, %s)
                  AND ativo = 1
            """, (apoiador_id, id_reconhecido))
        else:
            cur.execute("""
                SELECT id
                FROM usuarios
                WHERE id IN (%s, %s)
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (apoiador_id, id_reconhecido, centro_custo_id_logado))

        usuarios_validos = cur.fetchall()

        if len(usuarios_validos) != 2:
            flash("Apoiador e reconhecido devem pertencer ao seu escopo de acesso.", "warning")
            cur.close()
            conn.close()

            return render_template(
                "lancar_reconhecimento.html",
                usuarios=usuarios,
                current_date=date.today().isoformat()
            )

        cur.execute("""
            INSERT INTO reconhecimentos
                (apoiador_id, id_reconhecido, data_reconhecimento, reconhecimento, criado_por)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            apoiador_id,
            id_reconhecido,
            data_reconhecimento,
            reconhecimento,
            criado_por
        ))

        conn.commit()

        flash("Reconhecimento cadastrado com sucesso!", "success")

        cur.close()
        conn.close()

        return redirect(url_for("main.lancar_reconhecimento"))

    cur.close()
    conn.close()

    return render_template(
        "lancar_reconhecimento.html",
        usuarios=usuarios,
        current_date=date.today().isoformat()
    )


@main_routes.route("/listar_reconhecimento", methods=["GET"])
@login_required
@module_required('acesso_gestao_pessoas')
def listar_reconhecimento():
    usuario_logado_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id_logado = session.get("centro_custos_id")

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    if not usuario_logado_id or not centro_custo_id_logado:
        cur.close()
        conn.close()
        flash("Usuário logado inválido.", "danger")
        return redirect(url_for("main.dashboard"))

    if request.args.get("limpar"):
        cur.close()
        conn.close()
        return redirect(url_for("main.listar_reconhecimento"))

    apoiador_id = request.args.get("apoiador_id", "")
    id_reconhecido = request.args.get("id_reconhecido", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")
    sort = request.args.get("sort", "data_reconhecimento")
    order = request.args.get("order", "desc")

    page = request.args.get("page", 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    base_from = """
        FROM reconhecimentos r
        JOIN usuarios u1 ON r.apoiador_id = u1.id
        JOIN usuarios u2 ON r.id_reconhecido = u2.id
        WHERE 1 = 1
    """

    params = []

    if perfil == "basico":
        base_from += " AND r.criado_por = %s"
        params.append(usuario_logado_id)

    elif perfil == "intermediario":
        base_from += """
            AND u1.centro_custos_id = %s
            AND u2.centro_custos_id = %s
        """
        params.extend([centro_custo_id_logado, centro_custo_id_logado])

    # avançado e administrador veem tudo

    if apoiador_id:
        base_from += " AND r.apoiador_id = %s"
        params.append(apoiador_id)

    if id_reconhecido:
        base_from += " AND r.id_reconhecido = %s"
        params.append(id_reconhecido)

    if data_inicio:
        base_from += " AND r.data_reconhecimento >= %s"
        params.append(data_inicio)

    if data_fim:
        base_from += " AND r.data_reconhecimento <= %s"
        params.append(data_fim)

    allowed_sorts = {
        "id": "r.id",
        "apoiador": "u1.nome",
        "reconhecido": "u2.nome",
        "data_reconhecimento": "r.data_reconhecimento",
        "reconhecimento": "r.reconhecimento"
    }

    coluna_sort = allowed_sorts.get(sort, "r.data_reconhecimento")
    direcao = "ASC" if order == "asc" else "DESC"

    cur.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """, params)
    total_registros = cur.fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cur.execute(f"""
        SELECT
            r.id,
            r.apoiador_id,
            r.id_reconhecido,
            r.data_reconhecimento,
            r.reconhecimento,
            r.criado_por,
            u1.nome AS nome_apoiador,
            u2.nome AS nome_reconhecido
        {base_from}
        ORDER BY {coluna_sort} {direcao}, r.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    reconhecimentos = cur.fetchall()

    if perfil in ["administrador", "avancado"]:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome ASC
        """)
    else:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome ASC
        """, (centro_custo_id_logado,))

    usuarios = cur.fetchall()

    filtros = {
        "apoiador_id": apoiador_id,
        "id_reconhecido": id_reconhecido,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "sort": sort,
        "order": order
    }

    cur.close()
    conn.close()

    return render_template(
        "listar_reconhecimento.html",
        reconhecimentos=reconhecimentos,
        usuarios=usuarios,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )


@main_routes.route("/editar_reconhecimento/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_gestao_pessoas')
def editar_reconhecimento(id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    usuario_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id = session.get("centro_custos_id")

    if not usuario_id or not centro_custo_id:
        cur.close()
        conn.close()
        flash("Usuário logado inválido.", "danger")
        return redirect(url_for("main.dashboard"))

    cur.execute("""
        SELECT *
        FROM reconhecimentos
        WHERE id = %s
    """, (id,))
    reconhecimento = cur.fetchone()

    if not reconhecimento:
        cur.close()
        conn.close()
        flash("Reconhecimento não encontrado.", "danger")
        return redirect(url_for("main.listar_reconhecimento"))

    # 🔒 PERMISSIONAMENTO
    if perfil not in ["administrador", "avancado"]:
        if perfil == "intermediario":
            # precisa garantir que ambos pertencem ao mesmo CC
            cur.execute("""
                SELECT 1
                FROM usuarios u1
                JOIN usuarios u2 ON u2.id = %s
                WHERE u1.id = %s
                  AND u1.centro_custos_id = %s
                  AND u2.centro_custos_id = %s
            """, (
                reconhecimento["id_reconhecido"],
                reconhecimento["apoiador_id"],
                centro_custo_id,
                centro_custo_id
            ))
            if not cur.fetchone():
                cur.close()
                conn.close()
                flash("Você não tem permissão para editar este reconhecimento.", "danger")
                return redirect(url_for("main.listar_reconhecimento"))

        elif perfil == "basico":
            if reconhecimento.get("criado_por") != usuario_id:
                cur.close()
                conn.close()
                flash("Você não tem permissão para editar este reconhecimento.", "danger")
                return redirect(url_for("main.listar_reconhecimento"))

    # ======================================================
    # POST
    # ======================================================
    if request.method == "POST":
        apoiador_id = request.form.get("apoiador_id")
        id_reconhecido = request.form.get("id_reconhecido")
        data_reconhecimento = request.form.get("data_reconhecimento")
        reconhecimento_texto = request.form.get("reconhecimento")

        if not (apoiador_id and id_reconhecido and data_reconhecimento and reconhecimento_texto):
            flash("Preencha todos os campos obrigatórios.", "warning")
            return redirect(url_for("main.editar_reconhecimento", id=id))

        # 🔒 valida usuários conforme perfil
        if perfil in ["administrador", "avancado"]:
            cur.execute("""
                SELECT id
                FROM usuarios
                WHERE id IN (%s, %s)
                  AND ativo = 1
            """, (apoiador_id, id_reconhecido))
        else:
            cur.execute("""
                SELECT id
                FROM usuarios
                WHERE id IN (%s, %s)
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (apoiador_id, id_reconhecido, centro_custo_id))

        usuarios_validos = cur.fetchall()

        if len(usuarios_validos) != 2:
            flash("Apoiador e reconhecido devem pertencer ao seu escopo.", "warning")
            return redirect(url_for("main.editar_reconhecimento", id=id))

        cur.execute("""
            UPDATE reconhecimentos
            SET apoiador_id = %s,
                id_reconhecido = %s,
                data_reconhecimento = %s,
                reconhecimento = %s
            WHERE id = %s
        """, (
            apoiador_id,
            id_reconhecido,
            data_reconhecimento,
            reconhecimento_texto,
            id
        ))

        conn.commit()

        cur.close()
        conn.close()

        flash("Reconhecimento atualizado com sucesso!", "success")
        return redirect(url_for("main.listar_reconhecimento"))

    # ======================================================
    # GET
    # ======================================================
    if perfil in ["administrador", "avancado"]:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cur.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id,))

    usuarios = cur.fetchall()

    if reconhecimento.get("data_reconhecimento") and hasattr(reconhecimento["data_reconhecimento"], "strftime"):
        reconhecimento["data_reconhecimento"] = reconhecimento["data_reconhecimento"].strftime("%Y-%m-%d")

    cur.close()
    conn.close()

    return render_template(
        "editar_reconhecimento.html",
        reconhecimento=reconhecimento,
        usuarios=usuarios
    )


@main_routes.route("/excluir_reconhecimento/<int:id>", methods=["POST"])
@login_required
@module_required('acesso_gestao_pessoas')
def excluir_reconhecimento(id):
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    usuario_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id = session.get("centro_custos_id")

    cur.execute("""
        SELECT
            r.id,
            r.criado_por,
            u1.centro_custos_id AS centro_custo_apoiador,
            u2.centro_custos_id AS centro_custo_reconhecido
        FROM reconhecimentos r
        JOIN usuarios u1 ON u1.id = r.apoiador_id
        JOIN usuarios u2 ON u2.id = r.id_reconhecido
        WHERE r.id = %s
    """, (id,))
    reconhecimento = cur.fetchone()

    if not reconhecimento:
        cur.close()
        conn.close()
        flash("Reconhecimento não encontrado.", "danger")
        return redirect(url_for("main.listar_reconhecimento"))

    if perfil not in ["administrador", "avancado"]:
        if perfil == "intermediario":
            if (
                reconhecimento.get("centro_custo_apoiador") != centro_custo_id or
                reconhecimento.get("centro_custo_reconhecido") != centro_custo_id
            ):
                cur.close()
                conn.close()
                flash("Você não tem permissão para excluir este reconhecimento.", "danger")
                return redirect(url_for("main.listar_reconhecimento"))

        elif perfil == "basico":
            if reconhecimento.get("criado_por") != usuario_id:
                cur.close()
                conn.close()
                flash("Você não tem permissão para excluir este reconhecimento.", "danger")
                return redirect(url_for("main.listar_reconhecimento"))

    cur.execute("DELETE FROM reconhecimentos WHERE id = %s", (id,))
    conn.commit()

    cur.close()
    conn.close()

    flash("Reconhecimento excluído com sucesso!", "success")
    return redirect(url_for("main.listar_reconhecimento"))


# =============================== #
# MELHORIAS                       #
# =============================== #

def _get_upload_folder():
    # usa a pasta do config; se não existir, cria
    folder = current_app.config.get('UPLOAD_FOLDER', os.path.join('app', 'static', 'evidencias'))
    os.makedirs(folder, exist_ok=True)
    return folder

def _save_image_if_present(field_name: str, prefix: str):
    file = request.files.get(field_name)
    if file and allowed_image_file(file.filename):
        folder = _get_upload_folder()
        base = secure_filename(file.filename)
        newname = f"{prefix}_{uuid4().hex}_{base}"
        file.save(os.path.join(folder, newname))
        return newname
    return None


@main_routes.route("/listar_melhoria", methods=["GET"])
@login_required
@module_required('acesso_melhoria')
def listar_melhorias():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custos_id = session.get("centro_custos_id")

    if request.args.get("limpar"):
        cursor.close()
        conn.close()
        return redirect(url_for("main.listar_melhorias"))

    executante_id = request.args.get("executante", "").strip()
    centro_custo_id = request.args.get("centro_custo", "").strip()
    criado_por_mim = request.args.get("criado_por_mim", "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()

    sort = request.args.get("sort", "data")
    order = request.args.get("order", "desc")

    page = request.args.get("page", 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        "data": "m.data",
        "executante": "u.nome",
        "centro_custo": "cc.codigo"
    }

    coluna_sort = colunas_validas.get(sort, "m.data")
    direcao = "ASC" if order == "asc" else "DESC"

    filtros_sql = []
    valores = []

    # CONTROLE DE ESCOPO
    if perfil == "basico":
        filtros_sql.append("m.criado_por = %s")
        valores.append(usuario_id)

    elif perfil == "intermediario":
        filtros_sql.append("m.centro_custo_id = %s")
        valores.append(centro_custos_id)

    # avançado e administrador veem tudo

    if executante_id:
        filtros_sql.append("m.executante_id = %s")
        valores.append(int(executante_id))

    if centro_custo_id:
        filtros_sql.append("m.centro_custo_id = %s")
        valores.append(int(centro_custo_id))

    if criado_por_mim == "sim":
        filtros_sql.append("m.criado_por = %s")
        valores.append(usuario_id)
    elif criado_por_mim == "nao":
        filtros_sql.append("m.criado_por <> %s")
        valores.append(usuario_id)

    if data_inicio and data_fim:
        filtros_sql.append("m.data BETWEEN %s AND %s")
        valores.extend([data_inicio, data_fim])
    elif data_inicio:
        filtros_sql.append("m.data >= %s")
        valores.append(data_inicio)
    elif data_fim:
        filtros_sql.append("m.data <= %s")
        valores.append(data_fim)

    where_clause = ("WHERE " + " AND ".join(filtros_sql)) if filtros_sql else ""

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        FROM melhorias m
        JOIN usuarios u ON m.executante_id = u.id
        JOIN centros_custos cc ON m.centro_custo_id = cc.id
        {where_clause}
    """, valores)
    total_registros = cursor.fetchone()["total"]

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cursor.execute(f"""
        SELECT 
            m.*,
            u.nome AS nome_executante,
            cc.codigo AS codigo_cc,
            cc.descricao AS descricao_cc
        FROM melhorias m
        JOIN usuarios u ON m.executante_id = u.id
        JOIN centros_custos cc ON m.centro_custo_id = cc.id
        {where_clause}
        ORDER BY {coluna_sort} {direcao}, m.id DESC
        LIMIT %s OFFSET %s
    """, valores + [per_page, offset])
    melhorias = cursor.fetchall()

    if perfil in ["administrador", "avancado"]:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
        usuarios = cursor.fetchall()

        cursor.execute("""
            SELECT id, codigo, descricao
            FROM centros_custos
            ORDER BY codigo
        """)
        centros_custos = cursor.fetchall()
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))
        usuarios = cursor.fetchall()

        cursor.execute("""
            SELECT id, codigo, descricao
            FROM centros_custos
            WHERE id = %s
            ORDER BY codigo
        """, (centro_custos_id,))
        centros_custos = cursor.fetchall()

    filtros = {
        "executante": executante_id,
        "centro_custo": centro_custo_id,
        "criado_por_mim": criado_por_mim,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "sort": sort,
        "order": order
    }

    cursor.close()
    conn.close()

    return render_template(
        "listar_melhoria.html",
        melhorias=melhorias,
        usuarios=usuarios,
        centros_custos=centros_custos,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )


@main_routes.route("/lancar_melhoria", methods=["GET", "POST"])
@login_required
@module_required('acesso_melhoria')
def lancar_melhoria():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id_logado = session.get("centro_custos_id")

    if not usuario_logado_id or not centro_custo_id_logado:
        cursor.close()
        conn.close()
        flash("Usuário logado inválido.", "danger")
        return redirect(url_for("main.dashboard"))

    cursor.execute("""
        SELECT id, codigo, descricao
        FROM centros_custos
        WHERE id = %s
    """, (centro_custo_id_logado,))
    centro_custo_logado = cursor.fetchone()

    if not centro_custo_logado:
        cursor.close()
        conn.close()
        flash("Centro de custo do usuário logado não encontrado.", "danger")
        return redirect(url_for("main.dashboard"))

    if request.method == "POST":
        try:
            data = request.form.get("data")
            executante_id = request.form.get("executante")
            titulo = request.form.get("titulo")
            descricao_antes = request.form.get("descricao_antes")
            acao_realizada = request.form.get("acao_realizada")
            descricao_depois = request.form.get("descricao_depois")
            resultados_alcancados = request.form.get("resultados_alcancados")
            economia_estimada = request.form.get("economia_estimada") or None
            observacoes = request.form.get("observacoes")

            tipo_ganho = request.form.getlist("tipo_ganho")
            tipo_ganho_str = ",".join(tipo_ganho)

            if not all([
                data,
                executante_id,
                titulo,
                descricao_antes,
                acao_realizada,
                descricao_depois,
                resultados_alcancados
            ]):
                flash("Preencha todos os campos obrigatórios.", "danger")
                return redirect(url_for("main.lancar_melhoria"))

            if perfil in ["administrador", "avancado"]:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                """, (executante_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (executante_id, centro_custo_id_logado))

            if not cursor.fetchone():
                flash("Executante inválido para seu escopo.", "danger")
                return redirect(url_for("main.lancar_melhoria"))

            foto_antes = _save_image_if_present("foto_antes", "antes")
            foto_depois = _save_image_if_present("foto_depois", "depois")

            cursor.execute("""
                INSERT INTO melhorias 
                (data, executante_id, centro_custo_id, titulo, tipo_ganho,
                 descricao_antes, acao_realizada, descricao_depois, resultados_alcancados,
                 foto_antes, foto_depois, economia_estimada, observacoes, criado_por)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data,
                executante_id,
                centro_custo_id_logado,
                titulo,
                tipo_ganho_str,
                descricao_antes,
                acao_realizada,
                descricao_depois,
                resultados_alcancados,
                foto_antes,
                foto_depois,
                economia_estimada,
                observacoes,
                usuario_logado_id
            ))

            conn.commit()
            flash("Melhoria lançada com sucesso!", "success")
            return redirect(url_for("main.lancar_melhoria"))

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao lançar melhoria: {e}", "danger")
            return redirect(url_for("main.lancar_melhoria"))

        finally:
            cursor.close()
            conn.close()

    if perfil in ["administrador", "avancado"]:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id_logado,))

    usuarios = cursor.fetchall()

    centros_custos = [{
        "id": centro_custo_logado["id"],
        "codigo": centro_custo_logado["codigo"],
        "descricao": centro_custo_logado["descricao"]
    }]

    centro_custo_usuario = centros_custos[0]

    cursor.close()
    conn.close()

    return render_template(
        "lancar_melhoria.html",
        usuarios=usuarios,
        centros_custos=centros_custos,
        centro_custo_usuario=centro_custo_usuario,
        melhoria=None
    )


@main_routes.route("/editar_melhoria/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_melhoria')
def editar_melhoria(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custo_id = session.get("centro_custos_id")

    cursor.execute("""
        SELECT * 
        FROM melhorias
        WHERE id = %s
    """, (id,))
    melhoria = cursor.fetchone()

    if not melhoria:
        conn.close()
        flash("Melhoria não encontrada.", "danger")
        return redirect(url_for("main.listar_melhorias"))

    # 🔒 PERMISSIONAMENTO CORRETO
    if perfil not in ["administrador", "avancado"]:
        if perfil == "intermediario":
            if melhoria.get("centro_custo_id") != centro_custo_id:
                conn.close()
                flash("Você não tem permissão para editar esta melhoria.", "danger")
                return redirect(url_for("main.listar_melhorias"))

        elif perfil == "basico":
            if melhoria.get("criado_por") != usuario_id:
                conn.close()
                flash("Você não tem permissão para editar esta melhoria.", "danger")
                return redirect(url_for("main.listar_melhorias"))

    # ======================================================
    # POST
    # ======================================================
    if request.method == "POST":
        try:
            data = request.form.get("data")
            executante_id = request.form.get("executante")
            titulo = request.form.get("titulo")
            descricao_antes = request.form.get("descricao_antes")
            acao_realizada = request.form.get("acao_realizada")
            descricao_depois = request.form.get("descricao_depois")
            resultados_alcancados = request.form.get("resultados_alcancados")
            economia_estimada = request.form.get("economia_estimada") or None
            observacoes = request.form.get("observacoes")

            tipo_ganho = request.form.getlist("tipo_ganho")
            tipo_ganho_str = ",".join(tipo_ganho)

            if not all([
                data,
                executante_id,
                titulo,
                descricao_antes,
                acao_realizada,
                descricao_depois,
                resultados_alcancados
            ]):
                flash("Preencha todos os campos obrigatórios.", "danger")
                return redirect(url_for("main.editar_melhoria", id=id))

            # 🔒 valida executante
            if perfil in ["administrador", "avancado"]:
                cursor.execute("""
                    SELECT id FROM usuarios
                    WHERE id = %s AND ativo = 1
                """, (executante_id,))
            else:
                cursor.execute("""
                    SELECT id FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (executante_id, centro_custo_id))

            if not cursor.fetchone():
                flash("Executante inválido para seu escopo.", "danger")
                return redirect(url_for("main.editar_melhoria", id=id))

            foto_antes = _save_image_if_present("foto_antes", "antes")
            foto_depois = _save_image_if_present("foto_depois", "depois")

            cursor.execute("""
                UPDATE melhorias 
                SET data=%s,
                    executante_id=%s,
                    centro_custo_id=%s,
                    titulo=%s,
                    tipo_ganho=%s,
                    descricao_antes=%s,
                    acao_realizada=%s,
                    descricao_depois=%s,
                    resultados_alcancados=%s,
                    economia_estimada=%s,
                    observacoes=%s,
                    foto_antes = COALESCE(%s, foto_antes),
                    foto_depois = COALESCE(%s, foto_depois)
                WHERE id=%s
            """, (
                data,
                executante_id,
                centro_custo_id,
                titulo,
                tipo_ganho_str,
                descricao_antes,
                acao_realizada,
                descricao_depois,
                resultados_alcancados,
                economia_estimada,
                observacoes,
                foto_antes,
                foto_depois,
                id
            ))

            conn.commit()
            flash("Melhoria atualizada com sucesso!", "success")
            return redirect(url_for("main.listar_melhorias"))

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao atualizar melhoria: {e}", "danger")
            return redirect(url_for("main.editar_melhoria", id=id))

        finally:
            cursor.close()
            conn.close()

    # ======================================================
    # GET
    # ======================================================

    if melhoria.get("tipo_ganho"):
        if isinstance(melhoria["tipo_ganho"], str):
            melhoria["ganhos_list"] = [
                g.strip()
                for g in melhoria["tipo_ganho"].split(",")
                if g.strip()
            ]
        elif isinstance(melhoria["tipo_ganho"], set):
            melhoria["ganhos_list"] = list(melhoria["tipo_ganho"])
        else:
            melhoria["ganhos_list"] = []
    else:
        melhoria["ganhos_list"] = []

    if melhoria.get("data") and hasattr(melhoria["data"], "strftime"):
        melhoria["data"] = melhoria["data"].strftime("%Y-%m-%d")

    if perfil in ["administrador", "avancado"]:
        cursor.execute("SELECT id, nome FROM usuarios WHERE ativo = 1 ORDER BY nome")
    else:
        cursor.execute("""
            SELECT id, nome FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id,))

    usuarios = cursor.fetchall()

    centros_custos = [{
        "id": centro_custo_id,
        "codigo": session.get("codigo_cc"),
        "descricao": session.get("descricao_cc")
    }]

    conn.close()

    return render_template(
        "lancar_melhoria.html",
        usuarios=usuarios,
        centros_custos=centros_custos,
        centro_custo_usuario=centros_custos[0],
        melhoria=melhoria
    )


@main_routes.route("/excluir_melhoria/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_melhoria')
def excluir_melhoria(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get("usuario_id")
    perfil = session.get("perfil")
    centro_custos_id = session.get("centro_custos_id")

    try:
        cursor.execute("""
            SELECT id, criado_por, centro_custo_id
            FROM melhorias
            WHERE id = %s
        """, (id,))
        melhoria = cursor.fetchone()

        if not melhoria:
            flash("Melhoria não encontrada.", "warning")
            return redirect(url_for("main.listar_melhorias"))

        if perfil not in ["administrador", "avancado"]:
            if perfil == "intermediario" and melhoria.get("centro_custo_id") != centro_custos_id:
                flash("Você não tem permissão para excluir esta melhoria.", "danger")
                return redirect(url_for("main.listar_melhorias"))

            if perfil == "basico" and melhoria.get("criado_por") != usuario_id:
                flash("Você não tem permissão para excluir esta melhoria.", "danger")
                return redirect(url_for("main.listar_melhorias"))

        cursor.execute("""
            DELETE FROM melhorias
            WHERE id = %s
        """, (id,))

        conn.commit()
        flash("Melhoria excluída com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir melhoria: {e}", "danger")

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("main.listar_melhorias"))


# =============================== #
# TREINAMENTOS                    #
# =============================== #

@main_routes.route("/tipos_documento", methods=["GET", "POST"])
@login_required
@admin_required
def tipos_documento():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        tipo_id = request.form.get("id")
        sigla = request.form.get("sigla", "").strip().upper()
        descricao = request.form.get("descricao", "").strip()
        nivel = request.form.get("nivel")

        if not sigla or not descricao or not nivel:
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(url_for("main.tipos_documento"))

        try:
            if tipo_id:
                cursor.execute("""
                    SELECT id
                    FROM tipos_documento
                    WHERE sigla = %s
                      AND id <> %s
                """, (sigla, tipo_id))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe outro tipo de documento com essa sigla.", "warning")
                    conn.close()
                    return redirect(url_for("main.tipos_documento", editar_id=tipo_id))

                cursor.execute("""
                    UPDATE tipos_documento
                    SET sigla = %s,
                        descricao = %s,
                        nivel = %s
                    WHERE id = %s
                """, (sigla, descricao, nivel, tipo_id))

                flash("Tipo de documento atualizado com sucesso!", "success")

            else:
                cursor.execute("""
                    SELECT id
                    FROM tipos_documento
                    WHERE sigla = %s
                """, (sigla,))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe um tipo de documento com essa sigla.", "warning")
                    conn.close()
                    return redirect(url_for("main.tipos_documento"))

                cursor.execute("""
                    INSERT INTO tipos_documento (sigla, descricao, nivel)
                    VALUES (%s, %s, %s)
                """, (sigla, descricao, nivel))

                flash("Tipo de documento cadastrado com sucesso!", "success")

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao salvar tipo de documento: {e}", "danger")

        finally:
            conn.close()

        return redirect(url_for("main.tipos_documento"))

    editar_id = request.args.get("editar_id", "")
    filtro_sigla = request.args.get("filtro_sigla", "").strip()
    filtro_descricao = request.args.get("filtro_descricao", "").strip()
    sort = request.args.get("sort", "nivel")
    order = request.args.get("order", "asc")

    colunas_validas = {
        "id": "id",
        "sigla": "sigla",
        "descricao": "descricao",
        "nivel": "nivel"
    }

    coluna_sort = colunas_validas.get(sort, "nivel")
    direcao = "ASC" if order == "asc" else "DESC"

    query = "SELECT * FROM tipos_documento WHERE 1=1"
    params = []

    if filtro_sigla:
        query += " AND sigla LIKE %s"
        params.append(f"%{filtro_sigla}%")

    if filtro_descricao:
        query += " AND descricao LIKE %s"
        params.append(f"%{filtro_descricao}%")

    query += f" ORDER BY {coluna_sort} {direcao}, id ASC"

    cursor.execute(query, params)
    tipos_documento = cursor.fetchall()

    tipo_edicao = None
    if editar_id:
        cursor.execute("SELECT * FROM tipos_documento WHERE id = %s", (editar_id,))
        tipo_edicao = cursor.fetchone()

    filtros = {
        "filtro_sigla": filtro_sigla,
        "filtro_descricao": filtro_descricao,
        "sort": sort,
        "order": order
    }

    conn.close()

    return render_template(
        "tipos_documento.html",
        tipos_documento=tipos_documento,
        tipo_edicao=tipo_edicao,
        filtros=filtros
    )


@main_routes.route("/desativar_tipo_documento/<int:id>", methods=["POST"])
@login_required
@admin_required
def desativar_tipo_documento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE tipos_documento
            SET ativo = 0
            WHERE id = %s
        """, (id,))
        conn.commit()

        flash("Tipo de documento desativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao desativar tipo de documento: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.tipos_documento"))


@main_routes.route("/reativar_tipo_documento/<int:id>", methods=["POST"])
@login_required
@admin_required
def reativar_tipo_documento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE tipos_documento
            SET ativo = 1
            WHERE id = %s
        """, (id,))
        conn.commit()

        flash("Tipo de documento reativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao reativar tipo de documento: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.tipos_documento"))


ALLOWED_PDF_EXTENSIONS = {"pdf"}


def allowed_pdf_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_PDF_EXTENSIONS


def _save_pdf_if_present(field_name: str):
    file = request.files.get(field_name)

    if not file or file.filename == "":
        return None

    if not allowed_pdf_file(file.filename):
        raise ValueError("Apenas arquivos PDF são permitidos.")

    folder = _get_upload_folder()
    newname = f"proc_{uuid4().hex}_{secure_filename(file.filename)}"
    file.save(os.path.join(folder, newname))
    return newname


@main_routes.route("/procedimentos", methods=["GET", "POST"])
@login_required
@module_required('acesso_procedimentos')
def procedimentos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id") or session.get("user_id") or session.get("id")

    if request.method == "POST":
        tipo_documento_id = request.form.get("tipo_documento_id")
        numero_documento = (request.form.get("numero_documento") or "").strip()
        titulo = (request.form.get("titulo") or "").strip()
        validade_dias = (request.form.get("validade_dias") or "").strip()
        niveis_aplicacao = request.form.getlist("niveis_aplicacao")

        numero_revisao = request.form.get("numero_revisao")
        data_revisao = request.form.get("data_revisao")
        elaborado_por = (request.form.get("elaborado_por") or "").strip() or None
        aprovado_por = (request.form.get("aprovado_por") or "").strip() or None
        observacoes = (request.form.get("observacoes") or "").strip() or None
        requer_treinamento = 1 if request.form.get("requer_treinamento") else 0

        niveis_validos = {"cargo", "funcao", "setor"}
        niveis_aplicacao = [n for n in niveis_aplicacao if n in niveis_validos]

        if not tipo_documento_id or not numero_documento or not titulo:
            flash("Preencha os campos obrigatórios do documento.", "danger")
            conn.close()
            return redirect(url_for("main.procedimentos"))

        if not niveis_aplicacao:
            flash("Selecione pelo menos um nível de aplicação do documento.", "warning")
            conn.close()
            return redirect(url_for("main.procedimentos"))

        try:
            import re

            if not re.match(r"^\d{3}$", numero_documento):
                flash("O número do documento deve conter exatamente 3 dígitos.", "warning")
                conn.close()
                return redirect(url_for("main.procedimentos"))

            if validade_dias == "":
                validade_dias_int = None
            else:
                validade_dias_int = int(validade_dias)
                if validade_dias_int < 0:
                    flash("A validade em dias não pode ser negativa.", "warning")
                    conn.close()
                    return redirect(url_for("main.procedimentos"))

            cursor.execute("""
                SELECT id
                FROM procedimentos
                WHERE tipo_documento_id = %s
                  AND numero_documento = %s
            """, (tipo_documento_id, numero_documento))
            existente = cursor.fetchone()

            if existente:
                flash("Já existe um documento com esse tipo e número.", "warning")
                conn.close()
                return redirect(url_for("main.procedimentos"))

            if numero_revisao in (None, "") or not data_revisao:
                flash("Para cadastrar um novo procedimento, informe a revisão inicial e a data da revisão.", "danger")
                conn.close()
                return redirect(url_for("main.procedimentos"))

            try:
                numero_revisao_int = int(numero_revisao)
            except ValueError:
                flash("O número da revisão deve ser numérico.", "warning")
                conn.close()
                return redirect(url_for("main.procedimentos"))

            novo_pdf = _save_pdf_if_present("arquivo_pdf")

            cursor.execute("""
                INSERT INTO procedimentos (
                    tipo_documento_id,
                    numero_documento,
                    titulo,
                    validade_dias,
                    ativo,
                    criado_por
                )
                VALUES (%s, %s, %s, %s, 1, %s)
            """, (
                tipo_documento_id,
                numero_documento,
                titulo,
                validade_dias_int,
                usuario_logado_id
            ))
            novo_procedimento_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO procedimento_revisoes (
                    procedimento_id,
                    numero_revisao,
                    data_revisao,
                    elaborado_por,
                    aprovado_por,
                    arquivo_pdf,
                    observacoes,
                    requer_treinamento,
                    vigente,
                    criado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, %s)
            """, (
                novo_procedimento_id,
                numero_revisao_int,
                data_revisao,
                elaborado_por,
                aprovado_por,
                novo_pdf,
                observacoes,
                requer_treinamento,
                usuario_logado_id
            ))

            for nivel in niveis_aplicacao:
                cursor.execute("""
                    INSERT INTO procedimento_niveis_aplicacao (
                        procedimento_id,
                        nivel_aplicacao,
                        criado_por,
                        ativo
                    )
                    VALUES (%s, %s, %s, 1)
                """, (
                    novo_procedimento_id,
                    nivel,
                    usuario_logado_id
                ))

            conn.commit()
            flash("Procedimento cadastrado com sucesso!", "success")

        except ValueError:
            conn.rollback()
            flash("Informe um valor numérico válido para a validade em dias.", "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao salvar procedimento: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("main.listar_procedimentos"))

    cursor.execute("""
        SELECT id, sigla, descricao
        FROM tipos_documento
        WHERE ativo = 1
        ORDER BY nivel ASC, sigla ASC
    """)
    tipos_documento = cursor.fetchall()

    conn.close()

    return render_template(
        "procedimentos.html",
        tipos_documento=tipos_documento
    )


@main_routes.route("/editar_procedimento/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_procedimentos')
def editar_procedimento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id") or session.get("user_id") or session.get("id")

    cursor.execute("""
        SELECT *
        FROM procedimentos
        WHERE id = %s
    """, (id,))
    procedimento = cursor.fetchone()

    if not procedimento:
        conn.close()
        flash("Procedimento não encontrado.", "warning")
        return redirect(url_for("main.listar_procedimentos"))

    if request.method == "POST":
        tipo_documento_id = request.form.get("tipo_documento_id")
        numero_documento = (request.form.get("numero_documento") or "").strip()
        titulo = (request.form.get("titulo") or "").strip()
        validade_dias = (request.form.get("validade_dias") or "").strip()
        niveis_aplicacao = request.form.getlist("niveis_aplicacao")

        niveis_validos = {"cargo", "funcao", "setor"}
        niveis_aplicacao = [n for n in niveis_aplicacao if n in niveis_validos]

        if not tipo_documento_id or not numero_documento or not titulo:
            flash("Preencha todos os campos obrigatórios.", "danger")
            conn.close()
            return redirect(url_for("main.editar_procedimento", id=id))

        if not niveis_aplicacao:
            flash("Selecione pelo menos um nível de aplicação do documento.", "warning")
            conn.close()
            return redirect(url_for("main.editar_procedimento", id=id))

        try:
            import re

            if not re.match(r"^\d{3}$", numero_documento):
                flash("O número do documento deve conter exatamente 3 dígitos.", "warning")
                conn.close()
                return redirect(url_for("main.editar_procedimento", id=id))

            if validade_dias == "":
                validade_dias_int = None
            else:
                validade_dias_int = int(validade_dias)
                if validade_dias_int < 0:
                    flash("A validade em dias não pode ser negativa.", "warning")
                    conn.close()
                    return redirect(url_for("main.editar_procedimento", id=id))

            cursor.execute("""
                SELECT id
                FROM procedimentos
                WHERE tipo_documento_id = %s
                  AND numero_documento = %s
                  AND id <> %s
            """, (tipo_documento_id, numero_documento, id))
            existente = cursor.fetchone()

            if existente:
                flash("Já existe um documento com esse tipo e número.", "warning")
                conn.close()
                return redirect(url_for("main.editar_procedimento", id=id))

            cursor.execute("""
                UPDATE procedimentos
                SET tipo_documento_id = %s,
                    numero_documento = %s,
                    titulo = %s,
                    validade_dias = %s
                WHERE id = %s
            """, (
                tipo_documento_id,
                numero_documento,
                titulo,
                validade_dias_int,
                id
            ))

            cursor.execute("""
                UPDATE procedimento_niveis_aplicacao
                SET ativo = 0
                WHERE procedimento_id = %s
            """, (id,))

            for nivel in niveis_aplicacao:
                cursor.execute("""
                    SELECT id
                    FROM procedimento_niveis_aplicacao
                    WHERE procedimento_id = %s
                      AND nivel_aplicacao = %s
                    LIMIT 1
                """, (id, nivel))
                registro_existente = cursor.fetchone()

                if registro_existente:
                    cursor.execute("""
                        UPDATE procedimento_niveis_aplicacao
                        SET ativo = 1
                        WHERE id = %s
                    """, (registro_existente["id"],))
                else:
                    cursor.execute("""
                        INSERT INTO procedimento_niveis_aplicacao (
                            procedimento_id,
                            nivel_aplicacao,
                            criado_por,
                            ativo
                        )
                        VALUES (%s, %s, %s, 1)
                    """, (
                        id,
                        nivel,
                        usuario_logado_id
                    ))

            conn.commit()
            flash("Procedimento atualizado com sucesso!", "success")
            conn.close()
            return redirect(url_for("main.listar_procedimentos"))

        except ValueError:
            conn.rollback()
            conn.close()
            flash("Informe um valor numérico válido para a validade em dias.", "danger")
            return redirect(url_for("main.editar_procedimento", id=id))
        except Exception as e:
            conn.rollback()
            conn.close()
            flash(f"Erro ao atualizar procedimento: {e}", "danger")
            return redirect(url_for("main.editar_procedimento", id=id))

    cursor.execute("""
        SELECT id, sigla, descricao
        FROM tipos_documento
        WHERE ativo = 1
        ORDER BY nivel ASC, sigla ASC
    """)
    tipos_documento = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM procedimento_revisoes
        WHERE procedimento_id = %s
          AND vigente = 1
        LIMIT 1
    """, (id,))
    revisao_vigente = cursor.fetchone()

    cursor.execute("""
        SELECT nivel_aplicacao
        FROM procedimento_niveis_aplicacao
        WHERE procedimento_id = %s
          AND ativo = 1
    """, (id,))
    niveis_aplicacao_salvos = [row["nivel_aplicacao"] for row in cursor.fetchall()]

    conn.close()

    return render_template(
        "editar_procedimento.html",
        procedimento=procedimento,
        tipos_documento=tipos_documento,
        revisao_vigente=revisao_vigente,
        niveis_aplicacao_salvos=niveis_aplicacao_salvos
    )


@main_routes.route("/nova_revisao_procedimento/<int:id>", methods=["GET", "POST"])
@login_required
@module_required('acesso_procedimentos')
def nova_revisao_procedimento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get("usuario_id") or session.get("user_id") or session.get("id")

    cursor.execute("""
        SELECT
            p.id,
            p.tipo_documento_id,
            p.numero_documento,
            p.titulo,
            td.sigla,
            td.descricao AS descricao_tipo
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        WHERE p.id = %s
    """, (id,))
    procedimento = cursor.fetchone()

    if not procedimento:
        conn.close()
        flash("Procedimento não encontrado.", "warning")
        return redirect(url_for("main.listar_procedimentos"))

    if request.method == "POST":
        numero_revisao = request.form.get("numero_revisao")
        data_revisao = request.form.get("data_revisao")
        elaborado_por = (request.form.get("elaborado_por") or "").strip() or None
        aprovado_por = (request.form.get("aprovado_por") or "").strip() or None
        observacoes = (request.form.get("observacoes") or "").strip() or None
        requer_treinamento = 1 if request.form.get("requer_treinamento") else 0

        try:
            if numero_revisao in (None, "") or not data_revisao:
                flash("Informe o número da revisão e a data da revisão.", "danger")
                conn.close()
                return redirect(url_for("main.nova_revisao_procedimento", id=id))

            try:
                numero_revisao_int = int(numero_revisao)
            except ValueError:
                flash("O número da revisão deve ser numérico.", "warning")
                conn.close()
                return redirect(url_for("main.nova_revisao_procedimento", id=id))

            cursor.execute("""
                SELECT id
                FROM procedimento_revisoes
                WHERE procedimento_id = %s
                  AND numero_revisao = %s
            """, (id, numero_revisao_int))
            revisao_existente = cursor.fetchone()

            if revisao_existente:
                flash("Já existe uma revisão com esse número para este procedimento.", "warning")
                conn.close()
                return redirect(url_for("main.nova_revisao_procedimento", id=id))

            novo_pdf = _save_pdf_if_present("arquivo_pdf")

            cursor.execute("""
                UPDATE procedimento_revisoes
                SET vigente = 0
                WHERE procedimento_id = %s
                  AND vigente = 1
            """, (id,))

            cursor.execute("""
                INSERT INTO procedimento_revisoes (
                    procedimento_id,
                    numero_revisao,
                    data_revisao,
                    elaborado_por,
                    aprovado_por,
                    arquivo_pdf,
                    observacoes,
                    requer_treinamento,
                    vigente,
                    criado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, %s)
            """, (
                id,
                numero_revisao_int,
                data_revisao,
                elaborado_por,
                aprovado_por,
                novo_pdf,
                observacoes,
                requer_treinamento,
                usuario_logado_id
            ))

            conn.commit()
            flash("Nova revisão cadastrada com sucesso!", "success")
            conn.close()
            return redirect(url_for("main.listar_procedimentos"))

        except ValueError as e:
            conn.rollback()
            conn.close()
            flash(str(e), "danger")
            return redirect(url_for("main.nova_revisao_procedimento", id=id))
        except Exception as e:
            conn.rollback()
            conn.close()
            flash(f"Erro ao cadastrar nova revisão: {e}", "danger")
            return redirect(url_for("main.nova_revisao_procedimento", id=id))

    cursor.execute("""
        SELECT *
        FROM procedimento_revisoes
        WHERE procedimento_id = %s
          AND vigente = 1
        LIMIT 1
    """, (id,))
    revisao_vigente = cursor.fetchone()

    proxima_revisao = 0
    if revisao_vigente and revisao_vigente.get("numero_revisao") is not None:
        proxima_revisao = int(revisao_vigente["numero_revisao"]) + 1

    conn.close()

    return render_template(
        "nova_revisao_procedimento.html",
        procedimento=procedimento,
        revisao_vigente=revisao_vigente,
        proxima_revisao=proxima_revisao
    )


@main_routes.route("/historico_revisoes_procedimento/<int:procedimento_id>", methods=["GET"])
@login_required
@module_required('acesso_procedimentos')
def historico_revisoes_procedimento(procedimento_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            p.id,
            p.numero_documento,
            p.titulo,
            td.sigla
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        WHERE p.id = %s
    """, (procedimento_id,))
    procedimento = cursor.fetchone()

    if not procedimento:
        conn.close()
        flash("Procedimento não encontrado.", "warning")
        return redirect(url_for("main.listar_procedimentos"))

    cursor.execute("""
        SELECT
            pr.*
        FROM procedimento_revisoes pr
        WHERE pr.procedimento_id = %s
        ORDER BY pr.numero_revisao DESC, pr.id DESC
    """, (procedimento_id,))
    revisoes = cursor.fetchall()

    conn.close()

    return render_template(
        "historico_revisoes_procedimento.html",
        procedimento=procedimento,
        revisoes=revisoes
    )


@main_routes.route("/desativar_procedimento/<int:id>", methods=["POST"])
@login_required
@admin_required
def desativar_procedimento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE procedimentos
            SET ativo = 0
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Procedimento desativado com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao desativar procedimento: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("main.listar_procedimentos"))


@main_routes.route("/reativar_procedimento/<int:id>", methods=["POST"])
@login_required
@admin_required
def reativar_procedimento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE procedimentos
            SET ativo = 1
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Procedimento reativado com sucesso!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao reativar procedimento: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("main.listar_procedimentos"))

@main_routes.route("/listar_procedimentos", methods=["GET"])
@login_required
@module_required('acesso_procedimentos')
def listar_procedimentos():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    tipo_documento_id = request.args.get("tipo_documento_id", "")
    status = request.args.get("status", "")

    sort = request.args.get("sort", "sigla")
    order = request.args.get("order", "asc")

    colunas_validas = {
        "sigla": "td.sigla",
        "numero_documento": "p.numero_documento",
        "titulo": "p.titulo",
        "numero_revisao": "pr.numero_revisao",
        "data_revisao": "pr.data_revisao"
    }

    coluna_sort = colunas_validas.get(sort, "td.sigla")
    direcao = "ASC" if order == "asc" else "DESC"

    query = """
        SELECT
            p.id,
            p.tipo_documento_id,
            p.numero_documento,
            p.titulo,
            p.ativo,
            td.sigla,
            pr.numero_revisao,
            pr.data_revisao,
            pr.arquivo_pdf,
            pr.requer_treinamento
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        LEFT JOIN procedimento_revisoes pr
            ON pr.procedimento_id = p.id
           AND pr.vigente = 1
        WHERE 1=1
    """

    params = []

    if tipo_documento_id:
        query += " AND p.tipo_documento_id = %s"
        params.append(tipo_documento_id)

    if status == "ativo":
        query += " AND p.ativo = 1"
    elif status == "inativo":
        query += " AND p.ativo = 0"

    query += f" ORDER BY {coluna_sort} {direcao}, p.numero_documento ASC"

    cursor.execute(query, params)
    procedimentos = cursor.fetchall()

    cursor.execute("""
        SELECT id, sigla, descricao
        FROM tipos_documento
        WHERE ativo = 1
        ORDER BY nivel ASC, sigla ASC
    """)
    tipos_documento = cursor.fetchall()

    conn.close()

    filtros = {
        "tipo_documento_id": tipo_documento_id,
        "status": status,
        "sort": sort,
        "order": order
    }

    return render_template(
        "listar_procedimentos.html",
        procedimentos=procedimentos,
        tipos_documento=tipos_documento,
        filtros=filtros
    )

def allowed_pdf(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'pdf'

def salvar_evidencia_treinamento(arquivo):
    if not arquivo or not arquivo.filename:
        return None

    if not allowed_pdf(arquivo.filename):
        raise ValueError('O arquivo de evidência deve estar em PDF.')

    nome_original = secure_filename(arquivo.filename)
    extensao = nome_original.rsplit('.', 1)[1].lower()
    nome_final = f"treinamento_{uuid.uuid4().hex}.{extensao}"

    pasta_destino = os.path.join(current_app.root_path, 'static', 'evidencias_treinamentos')
    os.makedirs(pasta_destino, exist_ok=True)

    caminho_arquivo = os.path.join(pasta_destino, nome_final)
    arquivo.save(caminho_arquivo)

    return nome_final

@main_routes.route('/treinamentos_realizados', methods=['GET', 'POST'])
@login_required
@module_required('acesso_treinamentos')
def treinamentos_realizados():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custo_id = session.get('centro_custos_id')

    if not usuario_id or not centro_custo_id:
        conn.close()
        flash('Usuário logado inválido.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':

        procedimento_id = request.form.get('procedimento_id') or None
        procedimento_revisao_id = request.form.get('procedimento_revisao_id') or None
        data_treinamento = request.form.get('data_treinamento') or None
        hora_inicio = request.form.get('hora_inicio') or None
        hora_fim = request.form.get('hora_fim') or None
        local_treinamento = (request.form.get('local_treinamento') or '').strip()
        tipo_instrutor = (request.form.get('tipo_instrutor') or '').strip()
        instrutor_usuario_id = request.form.get('instrutor_usuario_id') or None
        instrutor_externo_id = request.form.get('instrutor_externo_id') or None
        observacoes = (request.form.get('observacoes') or '').strip() or None
        participante_ids = request.form.getlist('participante_ids[]')
        evidencia_arquivo = request.files.get('evidencia_arquivo')

        if not procedimento_id:
            flash('Selecione o procedimento.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if not procedimento_revisao_id:
            flash('Selecione a revisão do procedimento.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if not data_treinamento:
            flash('Informe a data do treinamento.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if not hora_inicio or not hora_fim:
            flash('Informe a hora de início e a hora de fim.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if not local_treinamento:
            flash('Informe o local do treinamento.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if tipo_instrutor not in ('interno', 'externo'):
            flash('Selecione um tipo de instrutor válido.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if tipo_instrutor == 'interno' and not instrutor_usuario_id:
            flash('Selecione o instrutor interno.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if tipo_instrutor == 'externo' and not instrutor_externo_id:
            flash('Selecione o instrutor externo.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        if not participante_ids:
            flash('Adicione pelo menos um participante.', 'danger')
            conn.close()
            return redirect(url_for('main.treinamentos_realizados'))

        try:

            # 🔒 VALIDA PARTICIPANTES CONFORME ESCOPO
            participantes_ids_processados = set()

            for participante_id in participante_ids:

                participante_id = (participante_id or '').strip()

                if not participante_id:
                    continue

                if participante_id in participantes_ids_processados:
                    continue

                if perfil in ['administrador', 'avancado']:

                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                    """, (participante_id,))

                else:

                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                          AND centro_custos_id = %s
                    """, (participante_id, centro_custo_id))

                participante_valido = cursor.fetchone()

                if not participante_valido:
                    flash('Existe participante fora do seu escopo de acesso.', 'danger')
                    conn.close()
                    return redirect(url_for('main.treinamentos_realizados'))

                participantes_ids_processados.add(participante_id)

            # 🔒 VALIDA INSTRUTOR INTERNO
            if tipo_instrutor == 'interno':

                if perfil in ['administrador', 'avancado']:

                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                          AND pode_ser_instrutor = 1
                    """, (instrutor_usuario_id,))

                else:

                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                          AND pode_ser_instrutor = 1
                          AND centro_custos_id = %s
                    """, (instrutor_usuario_id, centro_custo_id))

                instrutor_valido = cursor.fetchone()

                if not instrutor_valido:
                    flash('Instrutor inválido para seu escopo.', 'danger')
                    conn.close()
                    return redirect(url_for('main.treinamentos_realizados'))

            hora_inicio_dt = datetime.strptime(hora_inicio, '%H:%M')
            hora_fim_dt = datetime.strptime(hora_fim, '%H:%M')

            if hora_fim_dt <= hora_inicio_dt:
                flash('A hora de fim deve ser maior que a hora de início.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            cursor.execute("""
                SELECT 
                    pr.id,
                    p.validade_dias
                FROM procedimento_revisoes pr
                JOIN procedimentos p
                    ON p.id = pr.procedimento_id
                WHERE pr.id = %s
                  AND pr.procedimento_id = %s
                  AND p.ativo = 1
            """, (procedimento_revisao_id, procedimento_id))

            revisao_valida = cursor.fetchone()

            if not revisao_valida:
                flash('A revisão selecionada não pertence ao procedimento informado.', 'danger')
                conn.close()
                return redirect(url_for('main.treinamentos_realizados'))

            validade_dias = revisao_valida.get('validade_dias')
            data_validade = None

            if validade_dias is not None and data_treinamento:
                data_validade = (
                    datetime.strptime(data_treinamento, '%Y-%m-%d').date()
                    + timedelta(days=int(validade_dias))
                )

            nome_arquivo_evidencia = None

            if evidencia_arquivo and evidencia_arquivo.filename:
                nome_arquivo_evidencia = salvar_evidencia_treinamento(evidencia_arquivo)

            cursor.execute("""
                INSERT INTO treinamentos_realizados (
                    data_treinamento,
                    data_validade,
                    hora_inicio,
                    hora_fim,
                    local_treinamento,
                    tipo_instrutor,
                    instrutor_usuario_id,
                    instrutor_externo_id,
                    procedimento_id,
                    procedimento_revisao_id,
                    observacoes,
                    evidencia_arquivo,
                    criado_por
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data_treinamento,
                data_validade,
                hora_inicio,
                hora_fim,
                local_treinamento,
                tipo_instrutor,
                instrutor_usuario_id if tipo_instrutor == 'interno' else None,
                instrutor_externo_id if tipo_instrutor == 'externo' else None,
                procedimento_id,
                procedimento_revisao_id,
                observacoes,
                nome_arquivo_evidencia,
                usuario_id
            ))

            treinamento_realizado_id = cursor.lastrowid

            participantes_ids_processados = set()

            for participante_id in participante_ids:

                participante_id = (participante_id or '').strip()

                if not participante_id:
                    continue

                if participante_id in participantes_ids_processados:
                    continue

                cursor.execute("""
                    INSERT INTO treinamentos_realizados_participantes (
                        treinamento_realizado_id,
                        usuario_id,
                        presenca,
                        aprovado,
                        observacoes
                    )
                    VALUES (%s, %s, 1, 1, NULL)
                """, (treinamento_realizado_id, participante_id))

                participantes_ids_processados.add(participante_id)

            conn.commit()

            flash('Treinamento realizado cadastrado com sucesso!', 'success')

        except ValueError as e:

            conn.rollback()
            flash(str(e), 'danger')

        except Exception as e:

            conn.rollback()
            flash(f'Erro ao cadastrar treinamento realizado: {e}', 'danger')

        finally:

            conn.close()

        return redirect(url_for('main.treinamentos_realizados'))

    # =====================================================
    # GET
    # =====================================================

    cursor.execute("""
        SELECT
            p.id,
            td.sigla,
            p.numero_documento,
            p.titulo
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        WHERE p.ativo = 1
        ORDER BY td.sigla, p.numero_documento, p.titulo
    """)
    procedimentos = cursor.fetchall()

    cursor.execute("""
        SELECT
            pr.id,
            pr.procedimento_id,
            pr.numero_revisao,
            DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
        FROM procedimento_revisoes pr
        JOIN procedimentos p
            ON p.id = pr.procedimento_id
        WHERE p.ativo = 1
        ORDER BY pr.procedimento_id, pr.data_revisao DESC, pr.id DESC
    """)
    revisoes = cursor.fetchall()

    revisoes_por_procedimento = {}

    for r in revisoes:

        procedimento_id = str(r['procedimento_id'])

        revisoes_por_procedimento.setdefault(procedimento_id, []).append({
            'id': r['id'],
            'numero_revisao': r['numero_revisao'],
            'data_revisao': r['data_revisao']
        })

    # 🔒 INSTRUTORES
    if perfil in ['administrador', 'avancado']:

        cursor.execute("""
            SELECT
                u.id,
                u.nome
            FROM usuarios u
            WHERE u.ativo = 1
              AND u.pode_ser_instrutor = 1
            ORDER BY u.nome
        """)

    else:

        cursor.execute("""
            SELECT
                u.id,
                u.nome
            FROM usuarios u
            WHERE u.ativo = 1
              AND u.pode_ser_instrutor = 1
              AND u.centro_custos_id = %s
            ORDER BY u.nome
        """, (centro_custo_id,))

    instrutores_internos = cursor.fetchall()

    cursor.execute("""
        SELECT
            id,
            nome,
            empresa
        FROM instrutores_externos
        WHERE ativo = 1
        ORDER BY nome
    """)
    instrutores_externos = cursor.fetchall()

    # 🔒 PARTICIPANTES
    if perfil in ['administrador', 'avancado']:

        filtro_participantes = ""

        params_participantes = []

    else:

        filtro_participantes = "AND u.centro_custos_id = %s"

        params_participantes = [centro_custo_id]

    cursor.execute(f"""
        SELECT
            u.id,
            u.nome,
            u.matricula,
            c.nome AS nome_cargo
        FROM usuarios u
        LEFT JOIN cargos c
            ON c.id = u.cargo_id
        WHERE u.ativo = 1
        {filtro_participantes}
        ORDER BY u.nome
    """, params_participantes)

    participantes_nome = cursor.fetchall()

    cursor.execute(f"""
        SELECT
            u.id,
            u.nome,
            u.matricula,
            c.nome AS nome_cargo
        FROM usuarios u
        LEFT JOIN cargos c
            ON c.id = u.cargo_id
        WHERE u.ativo = 1
        {filtro_participantes}
        ORDER BY u.matricula, u.nome
    """, params_participantes)

    participantes_matricula = cursor.fetchall()

    conn.close()

    return render_template(
        'treinamentos_realizados.html',
        procedimentos=procedimentos,
        revisoes_por_procedimento=revisoes_por_procedimento,
        instrutores_internos=instrutores_internos,
        instrutores_externos=instrutores_externos,
        participantes_nome=participantes_nome,
        participantes_matricula=participantes_matricula
    )

@main_routes.route('/listar_treinamentos', methods=['GET'])
@login_required
@module_required('acesso_treinamentos')
def listar_treinamentos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil_logado = session.get('perfil')
    centro_custos_logado = session.get('centro_custos_id')

    procedimento_id = request.args.get('procedimento_id', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    sort = request.args.get('sort', 'data_treinamento').strip()
    order = request.args.get('order', 'desc').strip()

    page = request.args.get("page", 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        'data_treinamento': 'tr.data_treinamento',
        'hora_inicio': 'tr.hora_inicio',
        'hora_fim': 'tr.hora_fim',
        'procedimento': 'p.numero_documento',
        'numero_revisao': 'pr.numero_revisao',
        'nome_instrutor': """
            CASE
                WHEN tr.tipo_instrutor = 'interno' THEN ui.nome
                WHEN tr.tipo_instrutor = 'externo' THEN ie.nome
                ELSE ''
            END
        """
    }

    coluna_sort = colunas_validas.get(sort, 'tr.data_treinamento')
    direcao = 'ASC' if order == 'asc' else 'DESC'

    base_where = """
        FROM treinamentos_realizados tr
        JOIN procedimentos p
            ON p.id = tr.procedimento_id
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        JOIN procedimento_revisoes pr
            ON pr.id = tr.procedimento_revisao_id
        LEFT JOIN usuarios ui
            ON ui.id = tr.instrutor_usuario_id
        LEFT JOIN instrutores_externos ie
            ON ie.id = tr.instrutor_externo_id
        WHERE tr.ativo = 1
    """

    params = []

    if procedimento_id:
        base_where += " AND tr.procedimento_id = %s"
        params.append(procedimento_id)

    if data_inicio:
        base_where += " AND tr.data_treinamento >= %s"
        params.append(data_inicio)

    if data_fim:
        base_where += " AND tr.data_treinamento <= %s"
        params.append(data_fim)

    # 🔒 PERMISSIONAMENTO
    if perfil_logado == 'basico':
        base_where += """
            AND EXISTS (
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                WHERE trp.treinamento_realizado_id = tr.id
                  AND trp.usuario_id = %s
            )
        """
        params.append(usuario_id)

    elif perfil_logado == 'intermediario':
        base_where += """
            AND EXISTS (
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                JOIN usuarios up
                    ON up.id = trp.usuario_id
                WHERE trp.treinamento_realizado_id = tr.id
                  AND up.centro_custos_id = %s
            )
        """
        params.append(centro_custos_logado)

    # avançado e administrador veem tudo

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        {base_where}
    """, params)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cursor.execute(f"""
        SELECT tr.id
        {base_where}
        ORDER BY {coluna_sort} {direcao}, tr.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    ids = [row['id'] for row in cursor.fetchall()]

    if not ids:
        treinamentos = []
    else:
        placeholders = ','.join(['%s'] * len(ids))

        cursor.execute(f"""
            SELECT
                tr.id,
                tr.procedimento_id,
                tr.procedimento_revisao_id,
                DATE_FORMAT(tr.data_treinamento, '%d/%m/%Y') AS data_treinamento,
                DATE_FORMAT(tr.data_treinamento, '%Y-%m-%d') AS data_treinamento_iso,
                TIME_FORMAT(tr.hora_inicio, '%H:%i') AS hora_inicio,
                TIME_FORMAT(tr.hora_fim, '%H:%i') AS hora_fim,
                tr.local_treinamento,
                tr.tipo_instrutor,
                tr.instrutor_usuario_id,
                tr.instrutor_externo_id,
                tr.observacoes,
                tr.ativo,
                tr.evidencia_arquivo,

                CONCAT(td.sigla, ' ', p.numero_documento, ' - ', p.titulo) AS procedimento_descricao,
                pr.numero_revisao,

                CASE
                    WHEN tr.tipo_instrutor = 'interno' THEN ui.nome
                    WHEN tr.tipo_instrutor = 'externo' THEN ie.nome
                    ELSE '-'
                END AS nome_instrutor

            FROM treinamentos_realizados tr
            JOIN procedimentos p
                ON p.id = tr.procedimento_id
            JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            JOIN procedimento_revisoes pr
                ON pr.id = tr.procedimento_revisao_id
            LEFT JOIN usuarios ui
                ON ui.id = tr.instrutor_usuario_id
            LEFT JOIN instrutores_externos ie
                ON ie.id = tr.instrutor_externo_id
            WHERE tr.id IN ({placeholders})
            ORDER BY {coluna_sort} {direcao}, tr.id DESC
        """, ids)

        treinamentos = cursor.fetchall()

    cursor.execute("""
        SELECT
            p.id,
            td.sigla,
            p.numero_documento,
            p.titulo
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        WHERE p.ativo = 1
        ORDER BY td.sigla, p.numero_documento, p.titulo
    """)
    procedimentos = cursor.fetchall()

    cursor.execute("""
        SELECT
            pr.id,
            pr.procedimento_id,
            pr.numero_revisao,
            DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
        FROM procedimento_revisoes pr
        JOIN procedimentos p
            ON p.id = pr.procedimento_id
        WHERE p.ativo = 1
        ORDER BY pr.procedimento_id, pr.data_revisao DESC, pr.id DESC
    """)
    revisoes = cursor.fetchall()

    revisoes_por_procedimento = {}
    for r in revisoes:
        procedimento_key = str(r['procedimento_id'])
        revisoes_por_procedimento.setdefault(procedimento_key, []).append({
            'id': r['id'],
            'numero_revisao': r['numero_revisao'],
            'data_revisao': r['data_revisao']
        })

    if perfil_logado in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_logado,))

    instrutores_internos = cursor.fetchall()

    cursor.execute("""
        SELECT id, nome, empresa
        FROM instrutores_externos
        WHERE ativo = 1
        ORDER BY nome
    """)
    instrutores_externos = cursor.fetchall()

    conn.close()

    filtros = {
        'procedimento_id': procedimento_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'sort': sort,
        'order': order
    }

    return render_template(
        'listar_treinamentos.html',
        treinamentos=treinamentos,
        procedimentos=procedimentos,
        revisoes_por_procedimento=revisoes_por_procedimento,
        instrutores_internos=instrutores_internos,
        instrutores_externos=instrutores_externos,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_paginas=total_paginas,
        total_registros=total_registros
    )

@main_routes.route('/editar_treinamento/<int:id>', methods=['GET', 'POST'])
@login_required
@module_required('acesso_treinamentos')
def editar_treinamento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_logado_id = session.get('usuario_id')
    perfil_logado = session.get('perfil')
    centro_custos_logado = session.get('centro_custos_id')

    cursor.execute("""
        SELECT tr.*
        FROM treinamentos_realizados tr
        WHERE tr.id = %s
          AND tr.ativo = 1
    """, (id,))
    treinamento = cursor.fetchone()

    if not treinamento:
        conn.close()
        flash('Treinamento não encontrado.', 'warning')
        return redirect(url_for('main.listar_treinamentos'))

    # 🔒 PERMISSIONAMENTO
    if perfil_logado == 'basico':
        cursor.execute("""
            SELECT 1
            FROM treinamentos_realizados_participantes trp
            WHERE trp.treinamento_realizado_id = %s
              AND trp.usuario_id = %s
            LIMIT 1
        """, (id, usuario_logado_id))
        permitido = cursor.fetchone()

        if not permitido:
            conn.close()
            flash('Você não tem permissão para editar este treinamento.', 'danger')
            return redirect(url_for('main.listar_treinamentos'))

    elif perfil_logado == 'intermediario':
        cursor.execute("""
            SELECT 1
            FROM treinamentos_realizados_participantes trp
            JOIN usuarios u ON u.id = trp.usuario_id
            WHERE trp.treinamento_realizado_id = %s
              AND u.centro_custos_id = %s
            LIMIT 1
        """, (id, centro_custos_logado))
        permitido = cursor.fetchone()

        if not permitido:
            conn.close()
            flash('Você não tem permissão para editar este treinamento.', 'danger')
            return redirect(url_for('main.listar_treinamentos'))

    # avançado e administrador veem tudo

    if request.method == 'POST':
        next_url = request.form.get('next') or url_for('main.listar_treinamentos')

        procedimento_id = request.form.get('procedimento_id') or None
        procedimento_revisao_id = request.form.get('procedimento_revisao_id') or None
        data_treinamento = request.form.get('data_treinamento') or None
        hora_inicio = request.form.get('hora_inicio') or None
        hora_fim = request.form.get('hora_fim') or None
        local_treinamento = (request.form.get('local_treinamento') or '').strip()
        observacoes = (request.form.get('observacoes') or '').strip() or None
        participante_ids = request.form.getlist('participante_ids[]')
        evidencia_arquivo = request.files.get('evidencia_arquivo')

        tipo_instrutor = (request.form.get('tipo_instrutor') or treinamento.get('tipo_instrutor') or '').strip()
        instrutor_usuario_id = request.form.get('instrutor_usuario_id') or treinamento.get('instrutor_usuario_id')
        instrutor_externo_id = request.form.get('instrutor_externo_id') or treinamento.get('instrutor_externo_id')

        if not procedimento_id:
            flash('Selecione o procedimento.', 'danger')
            conn.close()
            return redirect(next_url)

        if not procedimento_revisao_id:
            flash('Selecione a revisão do procedimento.', 'danger')
            conn.close()
            return redirect(next_url)

        if not data_treinamento:
            flash('Informe a data do treinamento.', 'danger')
            conn.close()
            return redirect(next_url)

        if not hora_inicio or not hora_fim:
            flash('Informe a hora de início e a hora de fim.', 'danger')
            conn.close()
            return redirect(next_url)

        if not local_treinamento:
            flash('Informe o local do treinamento.', 'danger')
            conn.close()
            return redirect(next_url)

        if tipo_instrutor not in ('interno', 'externo'):
            flash('Selecione um tipo de instrutor válido.', 'danger')
            conn.close()
            return redirect(next_url)

        if tipo_instrutor == 'interno' and not instrutor_usuario_id:
            flash('Selecione o instrutor interno.', 'danger')
            conn.close()
            return redirect(next_url)

        if tipo_instrutor == 'externo' and not instrutor_externo_id:
            flash('Selecione o instrutor externo.', 'danger')
            conn.close()
            return redirect(next_url)

        try:
            formato_inicio = '%H:%M:%S' if len(str(hora_inicio)) == 8 else '%H:%M'
            formato_fim = '%H:%M:%S' if len(str(hora_fim)) == 8 else '%H:%M'

            hora_inicio_dt = datetime.strptime(str(hora_inicio), formato_inicio)
            hora_fim_dt = datetime.strptime(str(hora_fim), formato_fim)

            if hora_fim_dt <= hora_inicio_dt:
                flash('A hora de fim deve ser maior que a hora de início.', 'danger')
                conn.close()
                return redirect(next_url)

            cursor.execute("""
                SELECT 
                    pr.id,
                    p.validade_dias
                FROM procedimento_revisoes pr
                JOIN procedimentos p ON p.id = pr.procedimento_id
                WHERE pr.id = %s
                  AND pr.procedimento_id = %s
                  AND p.ativo = 1
            """, (procedimento_revisao_id, procedimento_id))
            revisao_valida = cursor.fetchone()

            if not revisao_valida:
                flash('A revisão selecionada não pertence ao procedimento informado.', 'danger')
                conn.close()
                return redirect(next_url)

            # 🔒 valida instrutor interno
            if tipo_instrutor == 'interno':
                if perfil_logado in ['administrador', 'avancado']:
                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                          AND pode_ser_instrutor = 1
                    """, (instrutor_usuario_id,))
                else:
                    cursor.execute("""
                        SELECT id
                        FROM usuarios
                        WHERE id = %s
                          AND ativo = 1
                          AND pode_ser_instrutor = 1
                          AND centro_custos_id = %s
                    """, (instrutor_usuario_id, centro_custos_logado))

                instrutor_valido = cursor.fetchone()

                if not instrutor_valido:
                    flash('Instrutor inválido para seu escopo.', 'danger')
                    conn.close()
                    return redirect(next_url)

            validade_dias = revisao_valida.get('validade_dias')
            data_validade = None

            if validade_dias is not None and data_treinamento:
                data_validade = datetime.strptime(data_treinamento, '%Y-%m-%d').date() + timedelta(days=int(validade_dias))

            nome_arquivo_evidencia = treinamento.get('evidencia_arquivo')

            if evidencia_arquivo and evidencia_arquivo.filename:
                nome_arquivo_evidencia = salvar_evidencia_treinamento(evidencia_arquivo)

            cursor.execute("""
                UPDATE treinamentos_realizados
                SET procedimento_id = %s,
                    procedimento_revisao_id = %s,
                    data_treinamento = %s,
                    data_validade = %s,
                    hora_inicio = %s,
                    hora_fim = %s,
                    local_treinamento = %s,
                    tipo_instrutor = %s,
                    instrutor_usuario_id = %s,
                    instrutor_externo_id = %s,
                    observacoes = %s,
                    evidencia_arquivo = %s
                WHERE id = %s
                  AND ativo = 1
            """, (
                procedimento_id,
                procedimento_revisao_id,
                data_treinamento,
                data_validade,
                hora_inicio,
                hora_fim,
                local_treinamento,
                tipo_instrutor,
                instrutor_usuario_id if tipo_instrutor == 'interno' else None,
                instrutor_externo_id if tipo_instrutor == 'externo' else None,
                observacoes,
                nome_arquivo_evidencia,
                id
            ))

            if participante_ids:
                cursor.execute("""
                    DELETE FROM treinamentos_realizados_participantes
                    WHERE treinamento_realizado_id = %s
                """, (id,))

                participantes_ids_processados = set()

                for participante_id in participante_ids:
                    participante_id = (participante_id or '').strip()

                    if not participante_id or participante_id in participantes_ids_processados:
                        continue

                    if perfil_logado in ['administrador', 'avancado']:
                        cursor.execute("""
                            SELECT 1
                            FROM usuarios
                            WHERE id = %s
                              AND ativo = 1
                            LIMIT 1
                        """, (participante_id,))
                    else:
                        cursor.execute("""
                            SELECT 1
                            FROM usuarios
                            WHERE id = %s
                              AND centro_custos_id = %s
                              AND ativo = 1
                            LIMIT 1
                        """, (participante_id, centro_custos_logado))

                    participante_permitido = cursor.fetchone()

                    if not participante_permitido:
                        continue

                    cursor.execute("""
                        INSERT INTO treinamentos_realizados_participantes (
                            treinamento_realizado_id,
                            usuario_id,
                            presenca,
                            aprovado,
                            observacoes
                        )
                        VALUES (%s, %s, 1, 1, NULL)
                    """, (id, participante_id))

                    participantes_ids_processados.add(participante_id)

            conn.commit()
            flash('Treinamento atualizado com sucesso!', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao atualizar treinamento: {e}', 'danger')

        finally:
            conn.close()

        return redirect(next_url)

    cursor.execute("""
        SELECT
            p.id,
            td.sigla,
            p.numero_documento,
            p.titulo
        FROM procedimentos p
        JOIN tipos_documento td ON td.id = p.tipo_documento_id
        WHERE p.ativo = 1
        ORDER BY td.sigla, p.numero_documento, p.titulo
    """)
    procedimentos = cursor.fetchall()

    cursor.execute("""
        SELECT
            pr.id,
            pr.procedimento_id,
            pr.numero_revisao,
            DATE_FORMAT(pr.data_revisao, '%d/%m/%Y') AS data_revisao
        FROM procedimento_revisoes pr
        JOIN procedimentos p ON p.id = pr.procedimento_id
        WHERE p.ativo = 1
          AND pr.procedimento_id = %s
        ORDER BY pr.data_revisao DESC, pr.id DESC
    """, (treinamento['procedimento_id'],))
    revisoes = cursor.fetchall()

    if perfil_logado in ['administrador', 'avancado']:
        filtro_participantes = ""
        params_participantes = []
    else:
        filtro_participantes = "AND u.centro_custos_id = %s"
        params_participantes = [centro_custos_logado]

    cursor.execute(f"""
        SELECT
            u.id,
            u.nome,
            u.matricula,
            c.nome AS nome_cargo
        FROM usuarios u
        LEFT JOIN cargos c ON c.id = u.cargo_id
        WHERE u.ativo = 1
        {filtro_participantes}
        ORDER BY u.nome
    """, params_participantes)
    participantes_nome = cursor.fetchall()

    cursor.execute(f"""
        SELECT
            u.id,
            u.nome,
            u.matricula,
            c.nome AS nome_cargo
        FROM usuarios u
        LEFT JOIN cargos c ON c.id = u.cargo_id
        WHERE u.ativo = 1
        {filtro_participantes}
        ORDER BY u.matricula, u.nome
    """, params_participantes)
    participantes_matricula = cursor.fetchall()

    cursor.execute("""
        SELECT
            u.id AS usuario_id,
            u.nome,
            u.matricula,
            c.nome AS nome_cargo
        FROM treinamentos_realizados_participantes trp
        JOIN usuarios u ON u.id = trp.usuario_id
        LEFT JOIN cargos c ON c.id = u.cargo_id
        WHERE trp.treinamento_realizado_id = %s
        ORDER BY u.nome
    """, (id,))
    participantes_selecionados = cursor.fetchall()

    conn.close()

    return render_template(
        'editar_treinamento.html',
        treinamento=treinamento,
        procedimentos=procedimentos,
        revisoes=revisoes,
        participantes_nome=participantes_nome,
        participantes_matricula=participantes_matricula,
        participantes_selecionados=participantes_selecionados
    )

@main_routes.route('/excluir_treinamento/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_treinamentos')
def excluir_treinamento(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil_logado = session.get('perfil')
    centro_custos_logado = session.get('centro_custos_id')

    try:
        cursor.execute("""
            SELECT id, ativo
            FROM treinamentos_realizados
            WHERE id = %s
              AND ativo = 1
        """, (id,))
        treinamento = cursor.fetchone()

        if not treinamento:
            conn.close()
            flash('Treinamento não encontrado ou já excluído.', 'warning')
            return redirect(url_for('main.listar_treinamentos'))

        if perfil_logado == 'basico':
            cursor.execute("""
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                WHERE trp.treinamento_realizado_id = %s
                  AND trp.usuario_id = %s
                LIMIT 1
            """, (id, usuario_id))
            permitido = cursor.fetchone()

            if not permitido:
                conn.close()
                flash('Você não tem permissão para excluir este treinamento.', 'danger')
                return redirect(url_for('main.listar_treinamentos'))

        elif perfil_logado == 'intermediario':
            cursor.execute("""
                SELECT 1
                FROM treinamentos_realizados_participantes trp
                JOIN usuarios u
                    ON u.id = trp.usuario_id
                WHERE trp.treinamento_realizado_id = %s
                  AND u.centro_custos_id = %s
                LIMIT 1
            """, (id, centro_custos_logado))
            permitido = cursor.fetchone()

            if not permitido:
                conn.close()
                flash('Você não tem permissão para excluir este treinamento.', 'danger')
                return redirect(url_for('main.listar_treinamentos'))

        # avançado e administrador veem/excluem tudo

        cursor.execute("""
            UPDATE treinamentos_realizados
            SET ativo = 0
            WHERE id = %s
        """, (id,))

        conn.commit()
        flash('Treinamento excluído logicamente com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao excluir treinamento: {e}', 'danger')

    finally:
        conn.close()

    return redirect(url_for('main.listar_treinamentos'))

def _rt_table_exists(cursor, table_name):
    try:
        cursor.execute("SHOW TABLES LIKE %s", (table_name,))
        return cursor.fetchone() is not None
    except Exception:
        return False

def _rt_get_columns(cursor, table_name):
    try:
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        return [col["Field"] for col in cursor.fetchall()]
    except Exception:
        return []

def _rt_first_existing(columns, options):
    for option in options:
        if option in columns:
            return option
    return None

def _rt_to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None

def _rt_normalize_ids(values):
    ids = []
    for value in values:
        try:
            ids.append(int(value))
        except (TypeError, ValueError):
            pass
    return list(dict.fromkeys(ids))

def _rt_resolver_campos_procedimento(cursor):
    colunas_procedimentos = _rt_get_columns(cursor, "procedimentos")
    colunas_tipos_documento = _rt_get_columns(cursor, "tipos_documento")

    if "titulo" in colunas_procedimentos:
        campo_titulo = "p.titulo"
    elif "titulo_documento" in colunas_procedimentos:
        campo_titulo = "p.titulo_documento"
    elif "nome" in colunas_procedimentos:
        campo_titulo = "p.nome"
    else:
        campo_titulo = "''"

    if "numero_documento" in colunas_procedimentos:
        campo_numero = "p.numero_documento"
    else:
        campo_numero = "''"

    if "revisao" in colunas_procedimentos:
        campo_revisao = "p.revisao"
    elif "versao" in colunas_procedimentos:
        campo_revisao = "p.versao"
    else:
        campo_revisao = "''"

    join_tipo_documento = ""
    if "sigla" in colunas_procedimentos:
        campo_sigla = "p.sigla"
    elif "tipo_documento_id" in colunas_procedimentos and "sigla" in colunas_tipos_documento:
        join_tipo_documento = "LEFT JOIN tipos_documento td ON td.id = p.tipo_documento_id"
        campo_sigla = "td.sigla"
    else:
        campo_sigla = "''"

    filtro_ativo_procedimento = ""
    if "ativo" in colunas_procedimentos:
        filtro_ativo_procedimento = "AND p.ativo = 1"

    return {
        "campo_titulo": campo_titulo,
        "campo_numero": campo_numero,
        "campo_revisao": campo_revisao,
        "campo_sigla": campo_sigla,
        "join_tipo_documento": join_tipo_documento,
        "filtro_ativo_procedimento": filtro_ativo_procedimento
    }

def _rt_buscar_cargos(cursor):
    cargos = []

    if _rt_table_exists(cursor, "cargos"):
        colunas_cargos = _rt_get_columns(cursor, "cargos")

        nome_expr = "CAST(c.id AS CHAR)"
        if "nome" in colunas_cargos:
            nome_expr = "c.nome"
        elif "descricao" in colunas_cargos:
            nome_expr = "c.descricao"

        filtro_ativo = "WHERE c.ativo = 1" if "ativo" in colunas_cargos else ""

        cursor.execute(f"""
            SELECT
                c.id,
                {nome_expr} AS nome
            FROM cargos c
            {filtro_ativo}
            ORDER BY nome
        """)
        cargos = cursor.fetchall()
    else:
        cursor.execute("""
            SELECT DISTINCT
                u.cargo_id AS id,
                CAST(u.cargo_id AS CHAR) AS nome
            FROM usuarios u
            WHERE u.ativo = 1
              AND u.cargo_id IS NOT NULL
            ORDER BY nome
        """)
        cargos = cursor.fetchall()

    return cargos

def _rt_buscar_usuarios_ativos(cursor, cargo_id=None):
    if _rt_table_exists(cursor, "cargos"):
        colunas_cargos = _rt_get_columns(cursor, "cargos")
        cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
        if "nome" in colunas_cargos:
            cargo_nome_expr = "c.nome"
        elif "descricao" in colunas_cargos:
            cargo_nome_expr = "c.descricao"

        query = f"""
            SELECT
                u.id,
                u.matricula,
                u.nome,
                u.cargo_id,
                {cargo_nome_expr} AS cargo_nome
            FROM usuarios u
            LEFT JOIN cargos c ON c.id = u.cargo_id
            WHERE u.ativo = 1
        """
    else:
        query = """
            SELECT
                u.id,
                u.matricula,
                u.nome,
                u.cargo_id,
                CAST(u.cargo_id AS CHAR) AS cargo_nome
            FROM usuarios u
            WHERE u.ativo = 1
        """

    params = []

    if cargo_id:
        query += " AND u.cargo_id = %s"
        params.append(cargo_id)

    query += " ORDER BY u.nome"
    cursor.execute(query, tuple(params))
    return cursor.fetchall()

def _rt_resolver_url_evidencia(valor):
    if not valor:
        return None

    valor = str(valor).strip()
    if not valor:
        return None

    valor_normalizado = valor.replace("\\", "/")

    # URL completa
    if valor_normalizado.startswith("http://") or valor_normalizado.startswith("https://"):
        return valor_normalizado

    # Já é caminho static absoluto
    if valor_normalizado.startswith("/static/"):
        return valor_normalizado

    # Caminho static relativo
    if valor_normalizado.startswith("static/"):
        return f"/{valor_normalizado}"

    # Se já vier com a pasta correta
    if valor_normalizado.startswith("evidencias_treinamentos/"):
        return url_for("static", filename=valor_normalizado)

    # Se vier com outra pasta interna de static
    if valor_normalizado.startswith("uploads/"):
        return f"/static/{valor_normalizado}"

    # Só nome do arquivo: assume a pasta correta
    nome_arquivo = os.path.basename(valor_normalizado)
    if nome_arquivo:
        return url_for("static", filename=f"evidencias_treinamentos/{nome_arquivo}")

    return None

def _rt_buscar_treinamentos_realizados(cursor, usuario_ids=None, cargo_id=None, data_inicio=None, data_fim=None):
    if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
        return []

    if not _rt_table_exists(cursor, "treinamentos_realizados"):
        return []

    cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
    cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
    cols_usuarios = _rt_get_columns(cursor, "usuarios")
    cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

    usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id", "participante_id", "colaborador_id"])
    trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id", "treinamentos_realizados_id", "realizado_id"])
    trr_pk = _rt_first_existing(cols_trr, ["id"])

    if not usuario_col_trp or not trp_to_trr or not trr_pk:
        return []

    cfg_proc = _rt_resolver_campos_procedimento(cursor)

    cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
    join_cargos = ""
    if cols_cargos:
        if "nome" in cols_cargos:
            cargo_nome_expr = "c.nome"
        elif "descricao" in cols_cargos:
            cargo_nome_expr = "c.descricao"
        join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

    data_realizacao_col = _rt_first_existing(cols_trr, [
        "data_treinamento", "data_realizacao", "realizado_em", "data", "created_at"
    ])

    validade_col = _rt_first_existing(cols_trr, [
        "data_validade", "validade_ate", "data_vencimento", "vencimento"
    ])

    hora_inicio_col = _rt_first_existing(cols_trr, ["hora_inicio"])
    hora_fim_col = _rt_first_existing(cols_trr, ["hora_fim"])
    revisao_col = _rt_first_existing(cols_trr, ["procedimento_revisao_id"])

    evidencia_col = _rt_first_existing(cols_trr, [
        "evidencia_arquivo",
        "evidencia",
        "arquivo_evidencia",
        "anexo",
        "caminho_evidencia",
        "nome_arquivo"
    ])

    filtros = ["u.ativo = 1"]
    params = []

    if "ativo" in cols_trp:
        filtros.append("trp.ativo = 1")
    if "ativo" in cols_trr:
        filtros.append("trr.ativo = 1")
    if "ativo" in cols_usuarios:
        filtros.append("u.ativo = 1")

    if usuario_ids:
        placeholders = ", ".join(["%s"] * len(usuario_ids))
        filtros.append(f"u.id IN ({placeholders})")
        params.extend(usuario_ids)

    if cargo_id:
        filtros.append("u.cargo_id = %s")
        params.append(cargo_id)

    if data_realizacao_col:
        if data_inicio and data_fim:
            filtros.append(f"DATE(trr.{data_realizacao_col}) BETWEEN %s AND %s")
            params.append(data_inicio)
            params.append(data_fim)
        elif data_inicio:
            filtros.append(f"DATE(trr.{data_realizacao_col}) >= %s")
            params.append(data_inicio)
        elif data_fim:
            filtros.append(f"DATE(trr.{data_realizacao_col}) <= %s")
            params.append(data_fim)

    select_data = f"trr.{data_realizacao_col}" if data_realizacao_col else "NULL"
    select_validade = f"trr.{validade_col}" if validade_col else "NULL"
    select_hora_inicio = f"trr.{hora_inicio_col}" if hora_inicio_col else "NULL"
    select_hora_fim = f"trr.{hora_fim_col}" if hora_fim_col else "NULL"
    select_revisao = f"trr.{revisao_col}" if revisao_col else "NULL"
    select_evidencia = f"trr.{evidencia_col}" if evidencia_col else "NULL"

    query = f"""
        SELECT
            u.id AS usuario_id,
            u.matricula,
            u.nome,
            {cargo_nome_expr} AS cargo_nome,
            p.id AS procedimento_id,
            {cfg_proc['campo_sigla']} AS sigla,
            {cfg_proc['campo_numero']} AS numero_documento,
            {cfg_proc['campo_titulo']} AS titulo,
            {select_revisao} AS revisao,
            {select_data} AS data_realizacao,
            {select_validade} AS validade,
            {select_hora_inicio} AS hora_inicio,
            {select_hora_fim} AS hora_fim,
            {select_evidencia} AS evidencia
        FROM treinamentos_realizados_participantes trp
        INNER JOIN treinamentos_realizados trr
            ON trr.{trr_pk} = trp.{trp_to_trr}
        INNER JOIN usuarios u
            ON u.id = trp.{usuario_col_trp}
        {join_cargos}
        LEFT JOIN procedimentos p
            ON p.id = trr.procedimento_id
        {cfg_proc['join_tipo_documento']}
        WHERE {" AND ".join(filtros)}
        ORDER BY u.nome, data_realizacao DESC, titulo
    """

    cursor.execute(query, tuple(params))
    dados = cursor.fetchall()

    for item in dados:
        item["status"] = "Realizado"
        item["evidencia_url"] = _rt_resolver_url_evidencia(item.get("evidencia"))

        data_realizacao = _rt_to_date(item.get("data_realizacao"))
        validade = _rt_to_date(item.get("validade"))

        item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""
        item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""

        carga_horaria_fmt = ""
        hora_inicio = item.get("hora_inicio")
        hora_fim = item.get("hora_fim")

        if hora_inicio and hora_fim:
            try:
                h1 = datetime.strptime(str(hora_inicio), "%H:%M:%S")
                h2 = datetime.strptime(str(hora_fim), "%H:%M:%S")
                diferenca = h2 - h1
                horas = diferenca.total_seconds() / 3600
                if horas.is_integer():
                    carga_horaria_fmt = str(int(horas))
                else:
                    carga_horaria_fmt = f"{horas:.2f}".replace(".", ",")
            except Exception:
                try:
                    h1 = datetime.strptime(str(hora_inicio), "%H:%M")
                    h2 = datetime.strptime(str(hora_fim), "%H:%M")
                    diferenca = h2 - h1
                    horas = diferenca.total_seconds() / 3600
                    if horas.is_integer():
                        carga_horaria_fmt = str(int(horas))
                    else:
                        carga_horaria_fmt = f"{horas:.2f}".replace(".", ",")
                except Exception:
                    carga_horaria_fmt = ""

        item["carga_horaria"] = carga_horaria_fmt

    return dados

def _rt_exportar_excel_relatorio_treinamentos(
    registros,
    nome_aba,
    headers,
    row_builder,
    larguras,
    colunas_centralizadas=None
):
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    ws.append(headers)

    fill_header = PatternFill(fill_type="solid", fgColor="EA6A23")
    font_header = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    for item in registros:
        ws.append(row_builder(item))

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    colunas_centralizadas = colunas_centralizadas or []

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            if idx in colunas_centralizadas:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@main_routes.route('/relatorio_treinamentos_realizados', methods=['GET'])
@login_required
@module_required('acesso_treinamentos')
def relatorio_treinamentos_realizados():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')
    usuario_id = session.get('usuario_id')

    cargo_id = request.args.get('cargo_id', type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    cargos = _rt_buscar_cargos(cursor)

    if perfil in ['administrador', 'avancado']:
        usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)
    elif perfil == 'intermediario':
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            centro_custos_id=centro_custos_id
        )
    else:
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            usuario_id=usuario_id
        )

    usuarios_permitidos_ids = [u["id"] for u in usuarios]

    if selecionar_todos:
        usuario_ids = usuarios_permitidos_ids
    elif usuario_ids:
        usuario_ids = [
            uid for uid in usuario_ids
            if uid in usuarios_permitidos_ids
        ]

    registros = _rt_buscar_treinamentos_realizados(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else usuarios_permitidos_ids,
        cargo_id=cargo_id,
        data_inicio=data_inicio,
        data_fim=data_fim
    )

    conn.close()

    return render_template(
        'relatorio_treinamentos_realizados.html',
        cargos=cargos,
        cargo_id=cargo_id,
        usuarios=usuarios,
        usuario_ids=usuario_ids,
        selecionar_todos=selecionar_todos,
        registros=registros,
        data_inicio=data_inicio,
        data_fim=data_fim
    )

@main_routes.route('/exportar_treinamentos_realizados_excel', methods=['GET'])
@login_required
def exportar_treinamentos_realizados_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')

    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)

    if selecionar_todos:
        usuario_ids = [u["id"] for u in usuarios]

    registros = _rt_buscar_treinamentos_realizados(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else None,
        cargo_id=cargo_id,
        data_inicio=data_inicio,
        data_fim=data_fim
    )

    conn.close()

    arquivo = _rt_exportar_excel_relatorio_treinamentos(
        registros=registros,
        nome_aba="Treinamentos Realizados",
        headers=[
            "Matrícula",
            "Nome",
            "Cargo",
            "Sigla",
            "Nº Doc",
            "Título",
            "Revisão",
            "Data",
            "C.H.",
            "Validade",
            "Evidência"
        ],
        row_builder=lambda item: [
            item.get("matricula") or "",
            item.get("nome") or "",
            item.get("cargo_nome") or "",
            item.get("sigla") or "",
            item.get("numero_documento") or "",
            item.get("titulo") or "",
            item.get("revisao") or "",
            item.get("data_realizacao_fmt") or "",
            item.get("carga_horaria") or "",
            item.get("validade_fmt") or "",
            item.get("evidencia_url") or ""
        ],
        larguras={
            "A": 12,
            "B": 28,
            "C": 22,
            "D": 8,
            "E": 10,
            "F": 40,
            "G": 10,
            "H": 12,
            "I": 8,
            "J": 12,
            "K": 20
        },
        colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9, 10]
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_treinamentos_realizados.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def _rt_buscar_treinamentos_a_vencer(cursor, usuario_ids=None, cargo_id=None, data_inicio=None, data_fim=None):

    if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
        return []

    if not _rt_table_exists(cursor, "treinamentos_realizados"):
        return []

    cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
    cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
    cols_usuarios = _rt_get_columns(cursor, "usuarios")
    cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

    usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id"])
    trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id"])
    trr_pk = _rt_first_existing(cols_trr, ["id"])

    if not usuario_col_trp or not trp_to_trr or not trr_pk:
        return []

    cfg_proc = _rt_resolver_campos_procedimento(cursor)

    cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
    join_cargos = ""
    if cols_cargos:
        cargo_nome_expr = "c.nome"
        join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

    hoje = date.today()
    limite = hoje + timedelta(days=30)

    filtros = [
        "u.ativo = 1",
        "trr.data_validade IS NOT NULL",
        "DATE(trr.data_validade) >= %s",
        "DATE(trr.data_validade) <= %s"
    ]

    params = [hoje, limite]

    if usuario_ids:
        placeholders = ", ".join(["%s"] * len(usuario_ids))
        filtros.append(f"u.id IN ({placeholders})")
        params.extend(usuario_ids)

    if cargo_id:
        filtros.append("u.cargo_id = %s")
        params.append(cargo_id)

    if data_inicio:
        filtros.append("DATE(trr.data_treinamento) >= %s")
        params.append(data_inicio)

    if data_fim:
        filtros.append("DATE(trr.data_treinamento) <= %s")
        params.append(data_fim)

    query = f"""
        SELECT
            u.id AS usuario_id,
            u.matricula,
            u.nome,
            {cargo_nome_expr} AS cargo_nome,
            p.id AS procedimento_id,
            {cfg_proc['campo_sigla']} AS sigla,
            {cfg_proc['campo_numero']} AS numero_documento,
            {cfg_proc['campo_titulo']} AS titulo,
            trr.procedimento_revisao_id AS revisao,
            trr.data_treinamento AS data_realizacao,
            trr.data_validade AS validade
        FROM treinamentos_realizados_participantes trp
        INNER JOIN treinamentos_realizados trr
            ON trr.{trr_pk} = trp.{trp_to_trr}
        INNER JOIN usuarios u
            ON u.id = trp.{usuario_col_trp}
        {join_cargos}
        LEFT JOIN procedimentos p
            ON p.id = trr.procedimento_id
        {cfg_proc['join_tipo_documento']}
        WHERE {" AND ".join(filtros)}
        ORDER BY trr.data_validade ASC, u.nome
    """

    cursor.execute(query, tuple(params))
    dados = cursor.fetchall()

    for item in dados:
        validade = _rt_to_date(item.get("validade"))
        data_realizacao = _rt_to_date(item.get("data_realizacao"))

        item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""
        item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""

        if validade:
            dias = (validade - hoje).days
            item["dias_para_vencer"] = dias
        else:
            item["dias_para_vencer"] = None

    return dados

@main_routes.route('/relatorio_treinamentos_a_vencer', methods=['GET'])
@login_required
@module_required('acesso_treinamentos')
def relatorio_treinamentos_a_vencer():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')
    usuario_id_logado = session.get('usuario_id')

    usuario_ids_raw = request.args.getlist('usuario_ids')
    cargo_id = request.args.get('cargo_id') or None
    data_inicio = request.args.get('data_inicio') or None
    data_fim = request.args.get('data_fim') or None

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = [
                int(u)
                for u in usuario_ids_raw
                if str(u).isdigit()
            ]

    # 🔒 USUÁRIOS CONFORME ESCOPO
    if perfil in ['administrador', 'avancado']:

        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id
        )

    elif perfil == 'intermediario':

        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            centro_custos_id=centro_custos_id
        )

    else:

        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            usuario_id=usuario_id_logado
        )

    usuarios_permitidos_ids = [u["id"] for u in usuarios]

    if selecionar_todos:
        usuario_ids = usuarios_permitidos_ids

    elif usuario_ids:
        usuario_ids = [
            uid for uid in usuario_ids
            if uid in usuarios_permitidos_ids
        ]

    registros = _rt_buscar_treinamentos_a_vencer(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else usuarios_permitidos_ids,
        cargo_id=cargo_id,
        data_inicio=data_inicio,
        data_fim=data_fim
    )

    cursor.execute("""
        SELECT id, nome
        FROM cargos
        WHERE ativo = 1
        ORDER BY nome
    """)
    cargos = cursor.fetchall()

    conn.close()

    return render_template(
        'relatorio_treinamentos_a_vencer.html',
        registros=registros,
        usuarios=usuarios,
        cargos=cargos,
        usuario_ids=usuario_ids,
        cargo_id=cargo_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        selecionar_todos=selecionar_todos
    )

@main_routes.route('/exportar_treinamentos_a_vencer_excel', methods=['GET'])
@login_required
def exportar_treinamentos_a_vencer_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)
    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)

    if selecionar_todos:
        usuario_ids = [u["id"] for u in usuarios]

    registros = _rt_buscar_treinamentos_a_vencer(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else None,
        cargo_id=cargo_id
    )

    conn.close()

    arquivo = _rt_exportar_excel_relatorio_treinamentos(
        registros=registros,
        nome_aba="Treinamentos a Vencer",
        headers=[
            "Matrícula",
            "Nome",
            "Cargo",
            "Sigla",
            "Nº Doc",
            "Título",
            "Revisão",
            "Data",
            "Validade",
            "Dias"
        ],
        row_builder=lambda item: [
            item.get("matricula") or "",
            item.get("nome") or "",
            item.get("cargo_nome") or "",
            item.get("sigla") or "",
            item.get("numero_documento") or "",
            item.get("titulo") or "",
            item.get("revisao") or "",
            item.get("data_realizacao_fmt") or "",
            item.get("validade_fmt") or "",
            item.get("dias_para_vencer") if item.get("dias_para_vencer") is not None else ""
        ],
        larguras={
            "A": 12,
            "B": 28,
            "C": 22,
            "D": 10,
            "E": 12,
            "F": 40,
            "G": 10,
            "H": 12,
            "I": 12,
            "J": 10
        },
        colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9]
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_treinamentos_a_vencer.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def _rt_buscar_treinamentos_vencidos(cursor, usuario_ids=None, cargo_id=None):

    if not _rt_table_exists(cursor, "treinamentos_realizados_participantes"):
        return []

    if not _rt_table_exists(cursor, "treinamentos_realizados"):
        return []

    cols_trp = _rt_get_columns(cursor, "treinamentos_realizados_participantes")
    cols_trr = _rt_get_columns(cursor, "treinamentos_realizados")
    cols_usuarios = _rt_get_columns(cursor, "usuarios")
    cols_cargos = _rt_get_columns(cursor, "cargos") if _rt_table_exists(cursor, "cargos") else []

    usuario_col_trp = _rt_first_existing(cols_trp, ["usuario_id"])
    trp_to_trr = _rt_first_existing(cols_trp, ["treinamento_realizado_id"])
    trr_pk = _rt_first_existing(cols_trr, ["id"])

    if not usuario_col_trp or not trp_to_trr or not trr_pk:
        return []

    cfg_proc = _rt_resolver_campos_procedimento(cursor)

    cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
    join_cargos = ""
    if cols_cargos:
        cargo_nome_expr = "c.nome"
        join_cargos = "LEFT JOIN cargos c ON c.id = u.cargo_id"

    hoje = date.today()

    filtros = [
        "u.ativo = 1",
        "trr.data_validade IS NOT NULL",
        "DATE(trr.data_validade) < %s"
    ]

    params = [hoje]

    if usuario_ids:
        placeholders = ", ".join(["%s"] * len(usuario_ids))
        filtros.append(f"u.id IN ({placeholders})")
        params.extend(usuario_ids)

    if cargo_id:
        filtros.append("u.cargo_id = %s")
        params.append(cargo_id)

    query = f"""
        SELECT
            u.id AS usuario_id,
            u.matricula,
            u.nome,
            {cargo_nome_expr} AS cargo_nome,
            p.id AS procedimento_id,
            {cfg_proc['campo_sigla']} AS sigla,
            {cfg_proc['campo_numero']} AS numero_documento,
            {cfg_proc['campo_titulo']} AS titulo,
            trr.procedimento_revisao_id AS revisao,
            trr.data_treinamento AS data_realizacao,
            trr.data_validade AS validade
        FROM treinamentos_realizados_participantes trp
        INNER JOIN treinamentos_realizados trr
            ON trr.{trr_pk} = trp.{trp_to_trr}
        INNER JOIN usuarios u
            ON u.id = trp.{usuario_col_trp}
        {join_cargos}
        LEFT JOIN procedimentos p
            ON p.id = trr.procedimento_id
        {cfg_proc['join_tipo_documento']}
        WHERE {" AND ".join(filtros)}
        ORDER BY trr.data_validade ASC, u.nome
    """

    cursor.execute(query, tuple(params))
    dados = cursor.fetchall()

    for item in dados:
        validade = _rt_to_date(item.get("validade"))
        data_realizacao = _rt_to_date(item.get("data_realizacao"))

        item["validade_fmt"] = validade.strftime("%d/%m/%Y") if validade else ""
        item["data_realizacao_fmt"] = data_realizacao.strftime("%d/%m/%Y") if data_realizacao else ""

        if validade:
            item["dias_vencido"] = (hoje - validade).days
        else:
            item["dias_vencido"] = None

    return dados

@main_routes.route('/relatorio_treinamentos_vencidos', methods=['GET'])
@login_required
@module_required('acesso_treinamentos')
def relatorio_treinamentos_vencidos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')
    usuario_id = session.get('usuario_id')

    cargo_id = request.args.get('cargo_id', type=int)
    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    if perfil in ['administrador', 'avancado']:
        usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)
    elif perfil == 'intermediario':
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            centro_custos_id=centro_custos_id
        )
    else:
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            usuario_id=usuario_id
        )

    usuarios_permitidos_ids = [u["id"] for u in usuarios]

    if selecionar_todos:
        usuario_ids = usuarios_permitidos_ids
    elif usuario_ids:
        usuario_ids = [
            uid for uid in usuario_ids
            if uid in usuarios_permitidos_ids
        ]

    registros = _rt_buscar_treinamentos_vencidos(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else usuarios_permitidos_ids,
        cargo_id=cargo_id
    )

    cursor.execute("""
        SELECT id, nome
        FROM cargos
        WHERE ativo = 1
        ORDER BY nome
    """)
    cargos = cursor.fetchall()

    conn.close()

    return render_template(
        'relatorio_treinamentos_vencidos.html',
        registros=registros,
        usuarios=usuarios,
        cargos=cargos,
        usuario_ids=usuario_ids,
        cargo_id=cargo_id,
        selecionar_todos=selecionar_todos
    )

@main_routes.route('/exportar_treinamentos_vencidos_excel', methods=['GET'])
@login_required
def exportar_treinamentos_vencidos_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)
    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)

    if selecionar_todos:
        usuario_ids = [u["id"] for u in usuarios]

    registros = _rt_buscar_treinamentos_vencidos(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else None,
        cargo_id=cargo_id
    )

    conn.close()

    arquivo = _rt_exportar_excel_relatorio_treinamentos(
        registros=registros,
        nome_aba="Treinamentos Vencidos",
        headers=[
            "Matrícula",
            "Nome",
            "Cargo",
            "Sigla",
            "Nº Doc",
            "Título",
            "Revisão",
            "Data",
            "Validade",
            "Dias Vencido"
        ],
        row_builder=lambda item: [
            item.get("matricula") or "",
            item.get("nome") or "",
            item.get("cargo_nome") or "",
            item.get("sigla") or "",
            item.get("numero_documento") or "",
            item.get("titulo") or "",
            item.get("revisao") or "",
            item.get("data_realizacao_fmt") or "",
            item.get("validade_fmt") or "",
            item.get("dias_vencido") if item.get("dias_vencido") is not None else ""
        ],
        larguras={
            "A": 12,
            "B": 28,
            "C": 22,
            "D": 10,
            "E": 12,
            "F": 40,
            "G": 10,
            "H": 12,
            "I": 12,
            "J": 12
        },
        colunas_centralizadas=[0, 3, 4, 6, 7, 8, 9]
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_treinamentos_vencidos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def _rt_buscar_treinamentos_pendentes(cursor, usuario_ids=None, cargo_id=None):
    hoje = date.today()

    filtros_usuarios = ["u.ativo = 1"]
    params = []

    if usuario_ids:
        placeholders = ", ".join(["%s"] * len(usuario_ids))
        filtros_usuarios.append(f"u.id IN ({placeholders})")
        params.extend(usuario_ids)

    if cargo_id:
        filtros_usuarios.append("u.cargo_id = %s")
        params.append(cargo_id)

    query = f"""
        SELECT DISTINCT
            base.usuario_id,
            base.matricula,
            base.nome,
            base.cargo_nome,
            base.procedimento_id,
            td.sigla,
            p.numero_documento,
            p.titulo,
            rev_vigente.numero_revisao AS revisao
        FROM (

            /* CAMADA CARGO */
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                c.nome AS cargo_nome,
                mcp.procedimento_id
            FROM usuarios u
            LEFT JOIN cargos c
                ON c.id = u.cargo_id
            JOIN matriz_cargo_procedimentos mcp
                ON mcp.cargo_id = u.cargo_id
               AND mcp.ativo = 1
               AND mcp.obrigatorio = 1
            WHERE {" AND ".join(filtros_usuarios)}

            UNION

            /* CAMADA FUNÇÃO */
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                c.nome AS cargo_nome,
                mfp.procedimento_id
            FROM usuarios u
            LEFT JOIN cargos c
                ON c.id = u.cargo_id
            JOIN usuario_funcoes_setores ufs
                ON ufs.usuario_id = u.id
               AND ufs.ativo = 1
            JOIN matriz_funcao_procedimentos mfp
                ON mfp.cargo_id = u.cargo_id
               AND mfp.funcao_id = ufs.funcao_id
               AND mfp.ativo = 1
               AND mfp.obrigatorio = 1
            WHERE {" AND ".join(filtros_usuarios)}

            UNION

            /* CAMADA SETOR */
            SELECT
                u.id AS usuario_id,
                u.matricula,
                u.nome,
                c.nome AS cargo_nome,
                msp.procedimento_id
            FROM usuarios u
            LEFT JOIN cargos c
                ON c.id = u.cargo_id
            JOIN usuario_funcoes_setores ufs
                ON ufs.usuario_id = u.id
               AND ufs.ativo = 1
            JOIN matriz_setor_procedimentos msp
                ON msp.cargo_id = u.cargo_id
               AND msp.setor_id = ufs.setor_id
               AND msp.ativo = 1
               AND msp.obrigatorio = 1
            WHERE {" AND ".join(filtros_usuarios)}

        ) AS base

        JOIN procedimentos p
            ON p.id = base.procedimento_id
           AND p.ativo = 1

        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id

        /* revisão vigente que exige treinamento */
        JOIN procedimento_revisoes rev_vigente
            ON rev_vigente.id = (
                SELECT pr2.id
                FROM procedimento_revisoes pr2
                WHERE pr2.procedimento_id = p.id
                  AND pr2.vigente = 1
                  AND pr2.requer_treinamento = 1
                ORDER BY pr2.data_revisao DESC, pr2.id DESC
                LIMIT 1
            )

        /* excluir quem já tem treinamento válido da revisão vigente */
        WHERE NOT EXISTS (
            SELECT 1
            FROM treinamentos_realizados_participantes trp
            JOIN treinamentos_realizados tr
                ON tr.id = trp.treinamento_realizado_id
            WHERE trp.usuario_id = base.usuario_id
              AND tr.procedimento_id = base.procedimento_id
              AND tr.procedimento_revisao_id = rev_vigente.id
              AND tr.ativo = 1
              AND tr.data_validade IS NOT NULL
              AND DATE(tr.data_validade) >= %s
        )

        ORDER BY base.nome, td.sigla, p.numero_documento, p.titulo
    """

    params_query = params + params + params + [hoje]
    cursor.execute(query, tuple(params_query))
    return cursor.fetchall()

@main_routes.route('/relatorio_treinamentos_pendentes', methods=['GET'])
@login_required
@module_required('acesso_treinamentos')
def relatorio_treinamentos_pendentes():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')
    usuario_id = session.get('usuario_id')

    cargo_id = request.args.get('cargo_id', type=int)
    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    if perfil in ['administrador', 'avancado']:
        usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)
    elif perfil == 'intermediario':
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            centro_custos_id=centro_custos_id
        )
    else:
        usuarios = _rt_buscar_usuarios_ativos(
            cursor,
            cargo_id=cargo_id,
            usuario_id=usuario_id
        )

    usuarios_permitidos_ids = [u["id"] for u in usuarios]

    if selecionar_todos:
        usuario_ids = usuarios_permitidos_ids
    elif usuario_ids:
        usuario_ids = [
            uid for uid in usuario_ids
            if uid in usuarios_permitidos_ids
        ]

    registros = _rt_buscar_treinamentos_pendentes(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else usuarios_permitidos_ids,
        cargo_id=cargo_id
    )

    cursor.execute("""
        SELECT id, nome
        FROM cargos
        WHERE ativo = 1
        ORDER BY nome
    """)
    cargos = cursor.fetchall()

    conn.close()

    return render_template(
        'relatorio_treinamentos_pendentes.html',
        registros=registros,
        usuarios=usuarios,
        cargos=cargos,
        usuario_ids=usuario_ids,
        cargo_id=cargo_id,
        selecionar_todos=selecionar_todos
    )

@main_routes.route('/exportar_treinamentos_pendentes_excel', methods=['GET'])
@login_required
def exportar_treinamentos_pendentes_excel():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)
    usuario_ids_raw = request.args.getlist('usuario_ids')

    selecionar_todos = False
    usuario_ids = []

    if usuario_ids_raw:
        if "__TODOS__" in usuario_ids_raw:
            selecionar_todos = True
        else:
            usuario_ids = _rt_normalize_ids(usuario_ids_raw)

    usuarios = _rt_buscar_usuarios_ativos(cursor, cargo_id=cargo_id)

    if selecionar_todos:
        usuario_ids = [u["id"] for u in usuarios]

    registros = _rt_buscar_treinamentos_pendentes(
        cursor,
        usuario_ids=usuario_ids if usuario_ids else None,
        cargo_id=cargo_id
    )

    conn.close()

    arquivo = _rt_exportar_excel_relatorio_treinamentos(
        registros=registros,
        nome_aba="Treinamentos Pendentes",
        headers=[
            "Matrícula",
            "Nome",
            "Cargo",
            "Sigla",
            "Nº Doc",
            "Título"
        ],
        row_builder=lambda item: [
            item.get("matricula") or "",
            item.get("nome") or "",
            item.get("cargo_nome") or "",
            item.get("sigla") or "",
            item.get("numero_documento") or "",
            item.get("titulo") or ""
        ],
        larguras={
            "A": 12,
            "B": 28,
            "C": 22,
            "D": 10,
            "E": 12,
            "F": 40
        },
        colunas_centralizadas=[0, 3, 4]
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="relatorio_treinamentos_pendentes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =============================== #
# SETORES                         #
# =============================== #

@main_routes.route("/setores", methods=["GET", "POST"])
@login_required
@admin_required
def setores():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        setor_id = request.form.get("id")
        nome = (request.form.get("nome") or "").strip()

        if not nome:
            flash("Informe o nome do setor.", "danger")
            conn.close()
            return redirect(url_for("main.setores"))

        try:
            if setor_id:
                cursor.execute("""
                    SELECT id
                    FROM setores
                    WHERE nome = %s
                      AND id <> %s
                """, (nome, setor_id))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe outro setor com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.setores", editar_id=setor_id))

                cursor.execute("""
                    UPDATE setores
                    SET nome = %s
                    WHERE id = %s
                """, (nome, setor_id))

                flash("Setor atualizado com sucesso!", "success")

            else:
                cursor.execute("""
                    SELECT id
                    FROM setores
                    WHERE nome = %s
                """, (nome,))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe um setor com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.setores"))

                cursor.execute("""
                    INSERT INTO setores (nome, ativo)
                    VALUES (%s, 1)
                """, (nome,))

                flash("Setor cadastrado com sucesso!", "success")

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao salvar setor: {e}", "danger")

        finally:
            conn.close()

        return redirect(url_for("main.setores"))

    editar_id = request.args.get("editar_id")
    setor_edicao = None

    if editar_id:
        cursor.execute("""
            SELECT *
            FROM setores
            WHERE id = %s
        """, (editar_id,))
        setor_edicao = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM setores
        ORDER BY nome ASC
    """)
    setores = cursor.fetchall()

    conn.close()

    return render_template(
        "setores.html",
        setores=setores,
        setor_edicao=setor_edicao
    )


@main_routes.route("/desativar_setor/<int:id>", methods=["POST"])
@login_required
@admin_required
def desativar_setor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE setores
            SET ativo = 0
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Setor desativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao desativar setor: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.setores"))


@main_routes.route("/reativar_setor/<int:id>", methods=["POST"])
@login_required
@admin_required
def reativar_setor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE setores
            SET ativo = 1
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Setor reativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao reativar setor: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.setores"))

# =============================== #
# FUNÇÕES                         #
# =============================== #

@main_routes.route("/funcoes", methods=["GET", "POST"])
@login_required
@admin_required
def funcoes():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        funcao_id = request.form.get("id")
        nome = (request.form.get("nome") or "").strip()

        if not nome:
            flash("Informe o nome da função.", "danger")
            conn.close()
            return redirect(url_for("main.funcoes"))

        try:
            if funcao_id:
                cursor.execute("""
                    SELECT id
                    FROM funcoes
                    WHERE nome = %s
                      AND id <> %s
                """, (nome, funcao_id))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe outra função com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.funcoes", editar_id=funcao_id))

                cursor.execute("""
                    UPDATE funcoes
                    SET nome = %s
                    WHERE id = %s
                """, (nome, funcao_id))

                flash("Função atualizada com sucesso!", "success")

            else:
                cursor.execute("""
                    SELECT id
                    FROM funcoes
                    WHERE nome = %s
                """, (nome,))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe uma função com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.funcoes"))

                cursor.execute("""
                    INSERT INTO funcoes (nome, ativo)
                    VALUES (%s, 1)
                """, (nome,))

                flash("Função cadastrada com sucesso!", "success")

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao salvar função: {e}", "danger")

        finally:
            conn.close()

        return redirect(url_for("main.funcoes"))

    editar_id = request.args.get("editar_id")
    funcao_edicao = None

    if editar_id:
        cursor.execute("""
            SELECT *
            FROM funcoes
            WHERE id = %s
        """, (editar_id,))
        funcao_edicao = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM funcoes
        ORDER BY nome ASC
    """)
    funcoes = cursor.fetchall()

    conn.close()

    return render_template(
        "funcoes.html",
        funcoes=funcoes,
        funcao_edicao=funcao_edicao
    )


@main_routes.route("/desativar_funcao/<int:id>", methods=["POST"])
@login_required
@admin_required
def desativar_funcao(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE funcoes
            SET ativo = 0
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Função desativada com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao desativar função: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.funcoes"))


@main_routes.route("/reativar_funcao/<int:id>", methods=["POST"])
@login_required
@admin_required
def reativar_funcao(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE funcoes
            SET ativo = 1
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Função reativada com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao reativar função: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.funcoes"))

# =============================== #
# CARGOS                          #
# =============================== #

@main_routes.route("/cargos", methods=["GET", "POST"])
@login_required
@admin_required
def cargos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":
        cargo_id = request.form.get("id")
        nome = (request.form.get("nome") or "").strip()

        if not nome:
            flash("Informe o nome do cargo.", "danger")
            conn.close()
            return redirect(url_for("main.cargos"))

        try:
            if cargo_id:
                # Validação duplicidade (edição)
                cursor.execute("""
                    SELECT id
                    FROM cargos
                    WHERE nome = %s
                      AND id <> %s
                """, (nome, cargo_id))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe outro cargo com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.cargos", editar_id=cargo_id))

                cursor.execute("""
                    UPDATE cargos
                    SET nome = %s
                    WHERE id = %s
                """, (nome, cargo_id))

                flash("Cargo atualizado com sucesso!", "success")

            else:
                # Validação duplicidade (inclusão)
                cursor.execute("""
                    SELECT id
                    FROM cargos
                    WHERE nome = %s
                """, (nome,))
                existente = cursor.fetchone()

                if existente:
                    flash("Já existe um cargo com esse nome.", "warning")
                    conn.close()
                    return redirect(url_for("main.cargos"))

                cursor.execute("""
                    INSERT INTO cargos (nome, ativo)
                    VALUES (%s, 1)
                """, (nome,))

                flash("Cargo cadastrado com sucesso!", "success")

            conn.commit()

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao salvar cargo: {e}", "danger")

        finally:
            conn.close()

        return redirect(url_for("main.cargos"))

    # =========================
    # EDIÇÃO
    # =========================
    editar_id = request.args.get("editar_id")
    cargo_edicao = None

    if editar_id:
        cursor.execute("""
            SELECT *
            FROM cargos
            WHERE id = %s
        """, (editar_id,))
        cargo_edicao = cursor.fetchone()

    # =========================
    # LISTA
    # =========================
    cursor.execute("""
        SELECT *
        FROM cargos
        ORDER BY nome ASC
    """)
    cargos = cursor.fetchall()

    conn.close()

    return render_template(
        "cargos.html",
        cargos=cargos,
        cargo_edicao=cargo_edicao
    )


@main_routes.route("/desativar_cargo/<int:id>", methods=["POST"])
@login_required
@admin_required
def desativar_cargo(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE cargos
            SET ativo = 0
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Cargo desativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao desativar cargo: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.cargos"))


@main_routes.route("/reativar_cargo/<int:id>", methods=["POST"])
@login_required
@admin_required
def reativar_cargo(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            UPDATE cargos
            SET ativo = 1
            WHERE id = %s
        """, (id,))
        conn.commit()
        flash("Cargo reativado com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao reativar cargo: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for("main.cargos"))

# =============================== #
# INSTRUTORES                     #
# =============================== #

@main_routes.route('/instrutores', methods=['GET', 'POST'])
@login_required
@admin_required
def instrutores():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        empresa = (request.form.get('empresa') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        telefone = (request.form.get('telefone') or '').strip() or None
        observacoes = (request.form.get('observacoes') or '').strip() or None

        if not nome:
            flash('Informe o nome do instrutor.', 'danger')
            conn.close()
            return redirect(url_for('main.instrutores'))

        try:
            cursor.execute("""
                SELECT id
                FROM instrutores_externos
                WHERE nome = %s
                  AND IFNULL(email, '') = IFNULL(%s, '')
            """, (nome, email))
            existente = cursor.fetchone()

            if existente:
                flash('Já existe um instrutor externo cadastrado com esse nome/e-mail.', 'warning')
                conn.close()
                return redirect(url_for('main.instrutores'))

            cursor.execute("""
                INSERT INTO instrutores_externos (
                    nome,
                    empresa,
                    email,
                    telefone,
                    observacoes,
                    ativo
                )
                VALUES (%s, %s, %s, %s, %s, 1)
            """, (
                nome,
                empresa,
                email,
                telefone,
                observacoes
            ))

            conn.commit()
            flash('Instrutor externo cadastrado com sucesso!', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao cadastrar instrutor: {e}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('main.instrutores'))

    cursor.execute("""
        SELECT *
        FROM instrutores_externos
        ORDER BY nome ASC
    """)
    instrutores = cursor.fetchall()

    conn.close()

    return render_template(
        'instrutores.html',
        instrutores=instrutores
    )


@main_routes.route('/editar_instrutor/<int:id>', methods=['POST'])
@login_required
@admin_required
def editar_instrutor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    nome = (request.form.get('nome') or '').strip()
    empresa = (request.form.get('empresa') or '').strip() or None
    email = (request.form.get('email') or '').strip() or None
    telefone = (request.form.get('telefone') or '').strip() or None
    observacoes = (request.form.get('observacoes') or '').strip() or None

    if not nome:
        flash('Informe o nome do instrutor.', 'danger')
        conn.close()
        return redirect(url_for('main.instrutores'))

    try:
        cursor.execute("""
            SELECT id
            FROM instrutores_externos
            WHERE nome = %s
              AND IFNULL(email, '') = IFNULL(%s, '')
              AND id <> %s
        """, (nome, email, id))
        existente = cursor.fetchone()

        if existente:
            flash('Já existe outro instrutor externo cadastrado com esse nome/e-mail.', 'warning')
            conn.close()
            return redirect(url_for('main.instrutores'))

        cursor.execute("""
            UPDATE instrutores_externos
            SET nome = %s,
                empresa = %s,
                email = %s,
                telefone = %s,
                observacoes = %s
            WHERE id = %s
        """, (
            nome,
            empresa,
            email,
            telefone,
            observacoes,
            id
        ))

        conn.commit()
        flash('Instrutor externo atualizado com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar instrutor: {e}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('main.instrutores'))


@main_routes.route('/alternar_instrutor/<int:id>', methods=['POST'])
@login_required
@admin_required
def alternar_instrutor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT ativo
            FROM instrutores_externos
            WHERE id = %s
        """, (id,))
        instrutor = cursor.fetchone()

        if not instrutor:
            conn.close()
            flash('Instrutor não encontrado.', 'warning')
            return redirect(url_for('main.instrutores'))

        novo_status = 0 if instrutor['ativo'] else 1

        cursor.execute("""
            UPDATE instrutores_externos
            SET ativo = %s
            WHERE id = %s
        """, (novo_status, id))

        conn.commit()

        if novo_status == 1:
            flash('Instrutor externo ativado com sucesso!', 'success')
        else:
            flash('Instrutor externo inativado com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao alterar status do instrutor: {e}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('main.instrutores'))

# =============================== #
# MATRIZ DE CAPACITAÇÃO           #
# =============================== #

@main_routes.route('/matriz_capacitacao', methods=['GET'])
@login_required
@module_required('acesso_procedimentos')
def matriz_capacitacao():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)

    # Inicialização
    cargos = []
    procedimentos_cargo_disponiveis = []
    procedimentos_cargo = []
    funcoes_disponiveis = []
    funcoes_vinculadas = []
    setores_disponiveis = []
    setores_vinculados = []
    procedimentos_funcao_disponiveis = []
    procedimentos_setor_disponiveis = []

    # =========================
    # CARGOS
    # =========================
    cursor.execute("""
        SELECT id, nome
        FROM cargos
        WHERE ativo = 1
        ORDER BY nome
    """)
    cargos = cursor.fetchall()

    if cargo_id:

        # =========================
        # CAMADA 1 - DISPONÍVEIS (CARGO)
        # =========================
        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            INNER JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            INNER JOIN procedimento_niveis_aplicacao pna
                ON pna.procedimento_id = p.id
            WHERE p.ativo = 1
              AND pna.ativo = 1
              AND pna.nivel_aplicacao = 'cargo'
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos_cargo_disponiveis = cursor.fetchall()

        # =========================
        # CAMADA 1 - VINCULADOS
        # =========================
        cursor.execute("""
            SELECT
                mcp.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM matriz_cargo_procedimentos mcp
            INNER JOIN procedimentos p
                ON p.id = mcp.procedimento_id
            INNER JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            WHERE mcp.cargo_id = %s
              AND mcp.ativo = 1
              AND p.ativo = 1
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """, (cargo_id,))
        procedimentos_cargo = cursor.fetchall()

        # =========================
        # FUNÇÕES DISPONÍVEIS
        # =========================
        cursor.execute("""
            SELECT id, nome
            FROM funcoes
            WHERE ativo = 1
            ORDER BY nome
        """)
        funcoes_disponiveis = cursor.fetchall()

        # =========================
        # CAMADA 2 - DISPONÍVEIS (FUNÇÃO)
        # =========================
        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            INNER JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            INNER JOIN procedimento_niveis_aplicacao pna
                ON pna.procedimento_id = p.id
            WHERE p.ativo = 1
              AND pna.ativo = 1
              AND pna.nivel_aplicacao = 'funcao'
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos_funcao_disponiveis = cursor.fetchall()

        # =========================
        # FUNÇÕES VINCULADAS AO CARGO
        # =========================
        cursor.execute("""
            SELECT
                mcf.id,
                mcf.funcao_id,
                f.nome
            FROM matriz_cargo_funcoes mcf
            INNER JOIN funcoes f
                ON f.id = mcf.funcao_id
            WHERE mcf.cargo_id = %s
              AND mcf.ativo = 1
              AND f.ativo = 1
            ORDER BY f.nome
        """, (cargo_id,))
        funcoes_raw = cursor.fetchall()

        for funcao in funcoes_raw:
            cursor.execute("""
                SELECT
                    mfp.id,
                    td.sigla,
                    p.numero_documento,
                    p.titulo
                FROM matriz_funcao_procedimentos mfp
                INNER JOIN procedimentos p
                    ON p.id = mfp.procedimento_id
                INNER JOIN tipos_documento td
                    ON td.id = p.tipo_documento_id
                WHERE mfp.cargo_id = %s
                  AND mfp.funcao_id = %s
                  AND mfp.ativo = 1
                  AND p.ativo = 1
                ORDER BY td.sigla, p.numero_documento, p.titulo
            """, (cargo_id, funcao['funcao_id']))

            procedimentos = cursor.fetchall()

            funcoes_vinculadas.append({
                'id': funcao['id'],
                'funcao_id': funcao['funcao_id'],
                'nome': funcao['nome'],
                'procedimentos': procedimentos
            })

        # =========================
        # SETORES DISPONÍVEIS
        # =========================
        cursor.execute("""
            SELECT id, nome
            FROM setores
            WHERE ativo = 1
            ORDER BY nome
        """)
        setores_disponiveis = cursor.fetchall()

        # =========================
        # CAMADA 3 - DISPONÍVEIS (SETOR)
        # =========================
        cursor.execute("""
            SELECT
                p.id,
                td.sigla,
                p.numero_documento,
                p.titulo
            FROM procedimentos p
            INNER JOIN tipos_documento td
                ON td.id = p.tipo_documento_id
            INNER JOIN procedimento_niveis_aplicacao pna
                ON pna.procedimento_id = p.id
            WHERE p.ativo = 1
              AND pna.ativo = 1
              AND pna.nivel_aplicacao = 'setor'
            ORDER BY td.sigla, p.numero_documento, p.titulo
        """)
        procedimentos_setor_disponiveis = cursor.fetchall()

        # =========================
        # SETORES VINCULADOS AO CARGO
        # =========================
        cursor.execute("""
            SELECT
                mcs.id,
                mcs.setor_id,
                s.nome
            FROM matriz_cargo_setores mcs
            INNER JOIN setores s
                ON s.id = mcs.setor_id
            WHERE mcs.cargo_id = %s
              AND mcs.ativo = 1
              AND s.ativo = 1
            ORDER BY s.nome
        """, (cargo_id,))
        setores_raw = cursor.fetchall()

        for setor in setores_raw:
            cursor.execute("""
                SELECT
                    msp.id,
                    td.sigla,
                    p.numero_documento,
                    p.titulo
                FROM matriz_setor_procedimentos msp
                INNER JOIN procedimentos p
                    ON p.id = msp.procedimento_id
                INNER JOIN tipos_documento td
                    ON td.id = p.tipo_documento_id
                WHERE msp.cargo_id = %s
                  AND msp.setor_id = %s
                  AND msp.ativo = 1
                  AND p.ativo = 1
                ORDER BY td.sigla, p.numero_documento, p.titulo
            """, (cargo_id, setor['setor_id']))

            procedimentos = cursor.fetchall()

            setores_vinculados.append({
                'id': setor['id'],
                'setor_id': setor['setor_id'],
                'nome': setor['nome'],
                'procedimentos': procedimentos
            })

    conn.close()

    return render_template(
        'matriz_capacitacao.html',
        cargos=cargos,
        cargo_id=cargo_id,
        procedimentos_cargo_disponiveis=procedimentos_cargo_disponiveis,
        procedimentos_cargo=procedimentos_cargo,
        funcoes_disponiveis=funcoes_disponiveis,
        funcoes_vinculadas=funcoes_vinculadas,
        setores_disponiveis=setores_disponiveis,
        setores_vinculados=setores_vinculados,
        procedimentos_funcao_disponiveis=procedimentos_funcao_disponiveis,
        procedimentos_setor_disponiveis=procedimentos_setor_disponiveis
    )

@main_routes.route('/salvar_procedimento_cargo', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def salvar_procedimento_cargo():
    cargo_id = request.form.get('cargo_id', type=int)
    procedimento_id = request.form.get('procedimento_id', type=int)
    criado_por = session.get('usuario_id')

    if not cargo_id or not procedimento_id:
        flash('Cargo e procedimento são obrigatórios.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM matriz_cargo_procedimentos
        WHERE cargo_id = %s
          AND procedimento_id = %s
        LIMIT 1
    """, (cargo_id, procedimento_id))
    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE matriz_cargo_procedimentos
            SET ativo = 1
            WHERE id = %s
        """, (existente['id'],))
    else:
        cursor.execute("""
            INSERT INTO matriz_cargo_procedimentos
                (cargo_id, procedimento_id, criado_por, ativo)
            VALUES (%s, %s, %s, 1)
        """, (cargo_id, procedimento_id, criado_por))

    conn.commit()
    conn.close()

    flash('Procedimento vinculado ao cargo com sucesso.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/excluir_procedimento_cargo/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def excluir_procedimento_cargo(id):
    cargo_id = request.form.get('cargo_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE matriz_cargo_procedimentos
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Procedimento removido do cargo.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/adicionar_funcao_cargo', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def adicionar_funcao_cargo():
    cargo_id = request.form.get('cargo_id', type=int)
    funcao_id = request.form.get('funcao_id', type=int)
    criado_por = session.get('usuario_id')

    if not cargo_id or not funcao_id:
        flash('Cargo e função são obrigatórios.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM matriz_cargo_funcoes
        WHERE cargo_id = %s
          AND funcao_id = %s
        LIMIT 1
    """, (cargo_id, funcao_id))
    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE matriz_cargo_funcoes
            SET ativo = 1
            WHERE id = %s
        """, (existente['id'],))
    else:
        cursor.execute("""
            INSERT INTO matriz_cargo_funcoes
                (cargo_id, funcao_id, criado_por, ativo)
            VALUES (%s, %s, %s, 1)
        """, (cargo_id, funcao_id, criado_por))

    conn.commit()
    conn.close()

    flash('Função vinculada ao cargo com sucesso.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/remover_funcao_cargo/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def remover_funcao_cargo(id):
    cargo_id = request.form.get('cargo_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT cargo_id, funcao_id
        FROM matriz_cargo_funcoes
        WHERE id = %s
        LIMIT 1
    """, (id,))
    vinculo = cursor.fetchone()

    if not vinculo:
        conn.close()
        flash('Função não encontrada.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    cursor.execute("""
        UPDATE matriz_cargo_funcoes
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    cursor.execute("""
        UPDATE matriz_funcao_procedimentos
        SET ativo = 0
        WHERE cargo_id = %s
          AND funcao_id = %s
    """, (vinculo['cargo_id'], vinculo['funcao_id']))

    conn.commit()
    conn.close()

    flash('Função removida do cargo.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id or vinculo['cargo_id']))

@main_routes.route('/adicionar_procedimento_funcao', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def adicionar_procedimento_funcao():
    cargo_id = request.form.get('cargo_id', type=int)
    funcao_id = request.form.get('funcao_id', type=int)
    procedimento_id = request.form.get('procedimento_id', type=int)
    criado_por = session.get('usuario_id')

    if not cargo_id or not funcao_id or not procedimento_id:
        flash('Cargo, função e procedimento são obrigatórios.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM matriz_funcao_procedimentos
        WHERE cargo_id = %s
          AND funcao_id = %s
          AND procedimento_id = %s
        LIMIT 1
    """, (cargo_id, funcao_id, procedimento_id))
    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE matriz_funcao_procedimentos
            SET ativo = 1
            WHERE id = %s
        """, (existente['id'],))
    else:
        cursor.execute("""
            INSERT INTO matriz_funcao_procedimentos
                (cargo_id, funcao_id, procedimento_id, obrigatorio, criado_por, ativo)
            VALUES (%s, %s, %s, 1, %s, 1)
        """, (cargo_id, funcao_id, procedimento_id, criado_por))

    conn.commit()
    conn.close()

    flash('Procedimento vinculado à função com sucesso.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/remover_procedimento_funcao/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def remover_procedimento_funcao(id):
    cargo_id = request.form.get('cargo_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE matriz_funcao_procedimentos
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Procedimento removido da função.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/adicionar_setor_cargo', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def adicionar_setor_cargo():
    cargo_id = request.form.get('cargo_id', type=int)
    setor_id = request.form.get('setor_id', type=int)
    criado_por = session.get('usuario_id')

    if not cargo_id or not setor_id:
        flash('Cargo e setor são obrigatórios.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM matriz_cargo_setores
        WHERE cargo_id = %s
          AND setor_id = %s
        LIMIT 1
    """, (cargo_id, setor_id))
    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE matriz_cargo_setores
            SET ativo = 1
            WHERE id = %s
        """, (existente['id'],))
    else:
        cursor.execute("""
            INSERT INTO matriz_cargo_setores
                (cargo_id, setor_id, criado_por, ativo)
            VALUES (%s, %s, %s, 1)
        """, (cargo_id, setor_id, criado_por))

    conn.commit()
    conn.close()

    flash('Setor vinculado ao cargo com sucesso.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/remover_setor_cargo/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def remover_setor_cargo(id):
    cargo_id = request.form.get('cargo_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT cargo_id, setor_id
        FROM matriz_cargo_setores
        WHERE id = %s
        LIMIT 1
    """, (id,))
    vinculo = cursor.fetchone()

    if not vinculo:
        conn.close()
        flash('Setor não encontrado.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    cursor.execute("""
        UPDATE matriz_cargo_setores
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    cursor.execute("""
        UPDATE matriz_setor_procedimentos
        SET ativo = 0
        WHERE cargo_id = %s
          AND setor_id = %s
    """, (vinculo['cargo_id'], vinculo['setor_id']))

    conn.commit()
    conn.close()

    flash('Setor removido do cargo.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id or vinculo['cargo_id']))

@main_routes.route('/adicionar_procedimento_setor', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def adicionar_procedimento_setor():
    cargo_id = request.form.get('cargo_id', type=int)
    setor_id = request.form.get('setor_id', type=int)
    procedimento_id = request.form.get('procedimento_id', type=int)
    criado_por = session.get('usuario_id')

    if not cargo_id or not setor_id or not procedimento_id:
        flash('Cargo, setor e procedimento são obrigatórios.', 'warning')
        return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM matriz_setor_procedimentos
        WHERE cargo_id = %s
          AND setor_id = %s
          AND procedimento_id = %s
        LIMIT 1
    """, (cargo_id, setor_id, procedimento_id))
    existente = cursor.fetchone()

    if existente:
        cursor.execute("""
            UPDATE matriz_setor_procedimentos
            SET ativo = 1
            WHERE id = %s
        """, (existente['id'],))
    else:
        cursor.execute("""
            INSERT INTO matriz_setor_procedimentos
                (cargo_id, setor_id, procedimento_id, obrigatorio, criado_por, ativo)
            VALUES (%s, %s, %s, 1, %s, 1)
        """, (cargo_id, setor_id, procedimento_id, criado_por))

    conn.commit()
    conn.close()

    flash('Procedimento vinculado ao setor com sucesso.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/remover_procedimento_setor/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_procedimentos')
def remover_procedimento_setor(id):
    cargo_id = request.form.get('cargo_id', type=int)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE matriz_setor_procedimentos
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Procedimento removido do setor.', 'success')
    return redirect(url_for('main.matriz_capacitacao', cargo_id=cargo_id))

@main_routes.route('/listar_matrizes_capacitacao', methods=['GET'])
@login_required
@module_required('acesso_procedimentos')
def listar_matrizes_capacitacao():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cargo_id = request.args.get('cargo_id', type=int)

    # =========================
    # LISTA DE CARGOS (FILTRO)
    # =========================
    cursor.execute("""
        SELECT id, nome
        FROM cargos
        WHERE ativo = 1
        ORDER BY nome
    """)
    cargos = cursor.fetchall()

    # =========================
    # QUERY PRINCIPAL
    # =========================
    query = """
        SELECT
            c.id,
            c.nome,

            COUNT(DISTINCT mcp.id) AS qtd_procedimentos,
            COUNT(DISTINCT mcf.id) AS qtd_funcoes,
            COUNT(DISTINCT mcs.id) AS qtd_setores

        FROM cargos c

        LEFT JOIN matriz_cargo_procedimentos mcp
            ON mcp.cargo_id = c.id AND mcp.ativo = 1

        LEFT JOIN matriz_cargo_funcoes mcf
            ON mcf.cargo_id = c.id AND mcf.ativo = 1

        LEFT JOIN matriz_cargo_setores mcs
            ON mcs.cargo_id = c.id AND mcs.ativo = 1

        WHERE c.ativo = 1
    """

    params = []

    if cargo_id:
        query += " AND c.id = %s"
        params.append(cargo_id)

    query += """
        GROUP BY c.id, c.nome
        ORDER BY c.nome
    """

    cursor.execute(query, params)
    matrizes = cursor.fetchall()

    conn.close()

    return render_template(
        'listar_matrizes_capacitacao.html',
        matrizes=matrizes,
        cargos=cargos,
        cargo_id=cargo_id
    )

@main_routes.route('/exportar_matriz_capacitacao/<int:cargo_id>')
@login_required
@module_required('acesso_procedimentos')
def exportar_matriz_capacitacao(cargo_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT nome
        FROM cargos
        WHERE id = %s
    """, (cargo_id,))
    cargo = cursor.fetchone()

    if not cargo:
        conn.close()
        flash('Cargo não encontrado.', 'warning')
        return redirect(url_for('main.listar_matrizes_capacitacao'))

    # CAMADA 1 - CARGO
    cursor.execute("""
        SELECT
            'CARGO' AS camada,
            '' AS referencia,
            CONCAT(td.sigla, ' - ', p.numero_documento, ' - ', p.titulo) AS procedimento
        FROM matriz_cargo_procedimentos mcp
        JOIN procedimentos p ON p.id = mcp.procedimento_id
        JOIN tipos_documento td ON td.id = p.tipo_documento_id
        WHERE mcp.cargo_id = %s
          AND mcp.ativo = 1
          AND p.ativo = 1
        ORDER BY td.sigla, p.numero_documento, p.titulo
    """, (cargo_id,))
    cargo_proc = cursor.fetchall()

    # CAMADA 2 - FUNÇÃO
    cursor.execute("""
        SELECT
            'FUNÇÃO' AS camada,
            f.nome AS referencia,
            CONCAT(td.sigla, ' - ', p.numero_documento, ' - ', p.titulo) AS procedimento
        FROM matriz_funcao_procedimentos mfp
        JOIN funcoes f ON f.id = mfp.funcao_id
        JOIN procedimentos p ON p.id = mfp.procedimento_id
        JOIN tipos_documento td ON td.id = p.tipo_documento_id
        WHERE mfp.cargo_id = %s
          AND mfp.ativo = 1
          AND p.ativo = 1
        ORDER BY f.nome, td.sigla, p.numero_documento, p.titulo
    """, (cargo_id,))
    func_proc = cursor.fetchall()

    # CAMADA 3 - SETOR
    cursor.execute("""
        SELECT
            'SETOR' AS camada,
            s.nome AS referencia,
            CONCAT(td.sigla, ' - ', p.numero_documento, ' - ', p.titulo) AS procedimento
        FROM matriz_setor_procedimentos msp
        JOIN setores s ON s.id = msp.setor_id
        JOIN procedimentos p ON p.id = msp.procedimento_id
        JOIN tipos_documento td ON td.id = p.tipo_documento_id
        WHERE msp.cargo_id = %s
          AND msp.ativo = 1
          AND p.ativo = 1
        ORDER BY s.nome, td.sigla, p.numero_documento, p.titulo
    """, (cargo_id,))
    setor_proc = cursor.fetchall()

    conn.close()

    dados = cargo_proc + func_proc + setor_proc

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz"

    # Cabeçalhos
    ws.append(["Camada", "Referência", "Procedimento"])

    # Linhas
    for linha in dados:
        ws.append([
            linha["camada"],
            linha["referencia"],
            linha["procedimento"]
        ])

    # Ajuste simples de largura
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"matriz_{cargo['nome'].replace(' ', '_')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@main_routes.route('/verificar_matriz_funcionario', methods=['GET'])
@login_required
@module_required('acesso_procedimentos')
def verificar_matriz_funcionario():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = request.args.get('usuario_id', type=int)

    usuarios = []
    usuario = None
    procedimentos_cargo = []
    funcoes_com_procedimentos = []
    setores_com_procedimentos = []

    # -------------------------------------------------
    # Helpers internos
    # -------------------------------------------------
    def get_table_columns(table_name):
        try:
            cursor.execute(f"SHOW COLUMNS FROM {table_name}")
            return [col["Field"] for col in cursor.fetchall()]
        except Exception:
            return []

    def table_exists(table_name):
        try:
            cursor.execute("SHOW TABLES LIKE %s", (table_name,))
            return cursor.fetchone() is not None
        except Exception:
            return False

    def first_existing(columns, options):
        for option in options:
            if option in columns:
                return option
        return None

    def to_date(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def resolver_campos_procedimento():
        colunas_procedimentos = get_table_columns("procedimentos")
        colunas_tipos_documento = get_table_columns("tipos_documento")

        if "titulo" in colunas_procedimentos:
            campo_titulo = "p.titulo"
        elif "titulo_documento" in colunas_procedimentos:
            campo_titulo = "p.titulo_documento"
        elif "nome" in colunas_procedimentos:
            campo_titulo = "p.nome"
        else:
            campo_titulo = "''"

        if "numero_documento" in colunas_procedimentos:
            campo_numero = "p.numero_documento"
        else:
            campo_numero = "''"

        join_tipo_documento = ""
        if "sigla" in colunas_procedimentos:
            campo_sigla = "p.sigla"
        elif "tipo_documento_id" in colunas_procedimentos and "sigla" in colunas_tipos_documento:
            join_tipo_documento = "LEFT JOIN tipos_documento td ON td.id = p.tipo_documento_id"
            campo_sigla = "td.sigla"
        else:
            campo_sigla = "''"

        filtro_ativo_procedimento = ""
        if "ativo" in colunas_procedimentos:
            filtro_ativo_procedimento = "AND p.ativo = 1"

        return {
            "campo_titulo": campo_titulo,
            "campo_numero": campo_numero,
            "campo_sigla": campo_sigla,
            "join_tipo_documento": join_tipo_documento,
            "filtro_ativo_procedimento": filtro_ativo_procedimento
        }

    def obter_status_procedimento(usuario_id_local, procedimento_id):
        """
        Retorna: 'Realizado', 'Vencido' ou 'Pendente'
        """
        if not table_exists("treinamentos_realizados_participantes") or not table_exists("treinamentos_realizados"):
            return "Pendente"

        cols_trp = get_table_columns("treinamentos_realizados_participantes")
        cols_trr = get_table_columns("treinamentos_realizados")
        cols_t = get_table_columns("treinamentos") if table_exists("treinamentos") else []

        # Coluna do usuário/participante
        usuario_col_trp = first_existing(cols_trp, ["usuario_id", "participante_id", "colaborador_id"])
        if not usuario_col_trp:
            return "Pendente"

        # Chaves de ligação
        trp_to_trr = first_existing(cols_trp, [
            "treinamento_realizado_id",
            "treinamentos_realizados_id",
            "realizado_id"
        ])

        trr_pk = first_existing(cols_trr, ["id"])
        if not trp_to_trr or not trr_pk:
            return "Pendente"

        # Como relacionar o realizado ao procedimento
        join_treinamentos = ""
        filtro_procedimento = ""
        params = [usuario_id_local]

        if "procedimento_id" in cols_trr:
            filtro_procedimento = "trr.procedimento_id = %s"
            params.append(procedimento_id)
        else:
            trr_to_t = first_existing(cols_trr, ["treinamento_id"])
            t_pk = first_existing(cols_t, ["id"]) if cols_t else None
            t_procedimento = first_existing(cols_t, ["procedimento_id"])

            if trr_to_t and t_pk and t_procedimento:
                join_treinamentos = f"INNER JOIN treinamentos t ON t.{t_pk} = trr.{trr_to_t}"
                filtro_procedimento = "t.procedimento_id = %s"
                params.append(procedimento_id)
            else:
                return "Pendente"

        # Data de realização
        data_realizacao_col = first_existing(cols_trr, [
            "data_realizacao",
            "data_treinamento",
            "realizado_em",
            "data",
            "created_at"
        ])

        # Data de validade/vencimento já pronta
        validade_col = first_existing(cols_trr, [
            "data_validade",
            "validade_ate",
            "data_vencimento",
            "vencimento"
        ])

        # Dias de validade no treinamento
        validade_dias_col = first_existing(cols_t, [
            "validade_dias",
            "dias_validade",
            "prazo_validade_dias"
        ]) if cols_t else None

        # Filtros ativos
        filtro_ativo_trp = "AND trp.ativo = 1" if "ativo" in cols_trp else ""
        filtro_ativo_trr = "AND trr.ativo = 1" if "ativo" in cols_trr else ""
        filtro_ativo_t = "AND t.ativo = 1" if "ativo" in cols_t and join_treinamentos else ""

        select_data_realizacao = f"trr.{data_realizacao_col} AS data_realizacao" if data_realizacao_col else "NULL AS data_realizacao"
        select_validade = f"trr.{validade_col} AS data_validade" if validade_col else "NULL AS data_validade"
        select_validade_dias = f"t.{validade_dias_col} AS validade_dias" if validade_dias_col and join_treinamentos else "NULL AS validade_dias"

        order_by = f"ORDER BY trr.{data_realizacao_col} DESC, trr.id DESC" if data_realizacao_col else "ORDER BY trr.id DESC"

        query = f"""
            SELECT
                trr.id,
                {select_data_realizacao},
                {select_validade},
                {select_validade_dias}
            FROM treinamentos_realizados_participantes trp
            INNER JOIN treinamentos_realizados trr
                ON trr.{trr_pk} = trp.{trp_to_trr}
            {join_treinamentos}
            WHERE trp.{usuario_col_trp} = %s
              AND {filtro_procedimento}
              {filtro_ativo_trp}
              {filtro_ativo_trr}
              {filtro_ativo_t}
            {order_by}
            LIMIT 1
        """

        cursor.execute(query, tuple(params))
        registro = cursor.fetchone()

        if not registro:
            return "Pendente"

        data_validade = to_date(registro.get("data_validade"))
        data_realizacao = to_date(registro.get("data_realizacao"))
        validade_dias = registro.get("validade_dias")

        if not data_validade and data_realizacao and validade_dias:
            try:
                validade_dias = int(validade_dias)
                data_validade = data_realizacao + timedelta(days=validade_dias)
            except Exception:
                data_validade = None

        hoje = date.today()

        if data_validade and data_validade < hoje:
            return "Vencido"

        return "Realizado"

    def aplicar_status(lista_procedimentos):
        for item in lista_procedimentos:
            item["status"] = obter_status_procedimento(usuario_id, item["id"])
        return lista_procedimentos

    def buscar_procedimentos_cargo(cargo_id):
        if not cargo_id:
            return []

        if not table_exists("matriz_cargo_procedimentos"):
            return []

        colunas_mcp = get_table_columns("matriz_cargo_procedimentos")
        cfg = resolver_campos_procedimento()

        filtro_ativo_mcp = "AND mcp.ativo = 1" if "ativo" in colunas_mcp else ""

        query = f"""
            SELECT
                p.id,
                {cfg['campo_sigla']} AS sigla,
                {cfg['campo_numero']} AS numero_documento,
                {cfg['campo_titulo']} AS titulo,
                {'mcp.obrigatorio' if 'obrigatorio' in colunas_mcp else '1'} AS obrigatorio
            FROM matriz_cargo_procedimentos mcp
            INNER JOIN procedimentos p ON p.id = mcp.procedimento_id
            {cfg['join_tipo_documento']}
            WHERE mcp.cargo_id = %s
              {filtro_ativo_mcp}
              {cfg['filtro_ativo_procedimento']}
            ORDER BY {cfg['campo_numero']}, p.id
        """
        cursor.execute(query, (cargo_id,))
        return aplicar_status(cursor.fetchall())

    def buscar_procedimentos_funcao(cargo_id, funcao_id):
        if not cargo_id or not funcao_id:
            return []

        cfg = resolver_campos_procedimento()
        colunas_mfp = get_table_columns("matriz_funcao_procedimentos")
        filtro_ativo_mfp = "AND mfp.ativo = 1" if "ativo" in colunas_mfp else ""

        query = f"""
            SELECT
                p.id,
                {cfg['campo_sigla']} AS sigla,
                {cfg['campo_numero']} AS numero_documento,
                {cfg['campo_titulo']} AS titulo,
                {'mfp.obrigatorio' if 'obrigatorio' in colunas_mfp else '1'} AS obrigatorio
            FROM matriz_funcao_procedimentos mfp
            INNER JOIN procedimentos p ON p.id = mfp.procedimento_id
            {cfg['join_tipo_documento']}
            WHERE mfp.cargo_id = %s
              AND mfp.funcao_id = %s
              {filtro_ativo_mfp}
              {cfg['filtro_ativo_procedimento']}
            ORDER BY {cfg['campo_numero']}, p.id
        """
        cursor.execute(query, (cargo_id, funcao_id))
        return aplicar_status(cursor.fetchall())

    def buscar_procedimentos_setor(cargo_id, setor_id):
        if not cargo_id or not setor_id:
            return []

        if not table_exists("matriz_setor_procedimentos"):
            return []

        colunas_msp = get_table_columns("matriz_setor_procedimentos")
        if "setor_id" not in colunas_msp:
            return []

        cfg = resolver_campos_procedimento()
        filtro_ativo_msp = "AND msp.ativo = 1" if "ativo" in colunas_msp else ""

        query = f"""
            SELECT
                p.id,
                {cfg['campo_sigla']} AS sigla,
                {cfg['campo_numero']} AS numero_documento,
                {cfg['campo_titulo']} AS titulo,
                {'msp.obrigatorio' if 'obrigatorio' in colunas_msp else '1'} AS obrigatorio
            FROM matriz_setor_procedimentos msp
            INNER JOIN procedimentos p ON p.id = msp.procedimento_id
            {cfg['join_tipo_documento']}
            WHERE msp.cargo_id = %s
              AND msp.setor_id = %s
              {filtro_ativo_msp}
              {cfg['filtro_ativo_procedimento']}
            ORDER BY {cfg['campo_numero']}, p.id
        """
        cursor.execute(query, (cargo_id, setor_id))
        return aplicar_status(cursor.fetchall())

    # -------------------------------------------------
    # Lista de usuários
    # -------------------------------------------------
    cursor.execute("""
        SELECT
            u.id,
            u.matricula,
            u.nome
        FROM usuarios u
        WHERE u.ativo = 1
        ORDER BY u.nome
    """)
    usuarios = cursor.fetchall()

    # -------------------------------------------------
    # Dados do usuário selecionado
    # -------------------------------------------------
    if usuario_id:
        cargo_nome_expr = "CAST(u.cargo_id AS CHAR)"
        if table_exists("cargos"):
            colunas_cargos = get_table_columns("cargos")
            if "nome" in colunas_cargos:
                cargo_nome_expr = "c.nome"
            elif "descricao" in colunas_cargos:
                cargo_nome_expr = "c.descricao"

            cursor.execute(f"""
                SELECT
                    u.id,
                    u.nome,
                    u.matricula,
                    u.cargo_id,
                    {cargo_nome_expr} AS cargo_nome
                FROM usuarios u
                LEFT JOIN cargos c ON c.id = u.cargo_id
                WHERE u.id = %s
                  AND u.ativo = 1
                LIMIT 1
            """, (usuario_id,))
        else:
            cursor.execute("""
                SELECT
                    u.id,
                    u.nome,
                    u.matricula,
                    u.cargo_id,
                    CAST(u.cargo_id AS CHAR) AS cargo_nome
                FROM usuarios u
                WHERE u.id = %s
                  AND u.ativo = 1
                LIMIT 1
            """, (usuario_id,))

        usuario = cursor.fetchone()

        if usuario:
            procedimentos_cargo = buscar_procedimentos_cargo(usuario["cargo_id"])

            cursor.execute("""
                SELECT DISTINCT
                    f.id,
                    f.nome
                FROM usuario_funcoes_setores ufs
                INNER JOIN funcoes f ON f.id = ufs.funcao_id
                WHERE ufs.usuario_id = %s
                  AND ufs.ativo = 1
                  AND ufs.funcao_id IS NOT NULL
                ORDER BY f.nome
            """, (usuario_id,))
            funcoes_usuario = cursor.fetchall()

            for funcao in funcoes_usuario:
                funcoes_com_procedimentos.append({
                    "id": funcao["id"],
                    "nome": funcao["nome"],
                    "procedimentos": buscar_procedimentos_funcao(usuario["cargo_id"], funcao["id"])
                })

            cursor.execute("""
                SELECT DISTINCT
                    s.id,
                    s.nome
                FROM usuario_funcoes_setores ufs
                INNER JOIN setores s ON s.id = ufs.setor_id
                WHERE ufs.usuario_id = %s
                  AND ufs.ativo = 1
                  AND ufs.setor_id IS NOT NULL
                ORDER BY s.nome
            """, (usuario_id,))
            setores_usuario = cursor.fetchall()

            for setor in setores_usuario:
                setores_com_procedimentos.append({
                    "id": setor["id"],
                    "nome": setor["nome"],
                    "procedimentos": buscar_procedimentos_setor(usuario["cargo_id"], setor["id"])
                })

    conn.close()

    return render_template(
        "verificar_matriz_funcionario.html",
        usuarios=usuarios,
        usuario=usuario,
        usuario_id=usuario_id,
        procedimentos_cargo=procedimentos_cargo,
        funcoes_com_procedimentos=funcoes_com_procedimentos,
        setores_com_procedimentos=setores_com_procedimentos
    )

# =============================== #
# AUDITORIA DE PADRÃO             #
# =============================== #

@main_routes.route('/lancar_ap', methods=['GET', 'POST'])
@login_required
@module_required('acesso_ssma')
def lancar_ap():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if request.method == 'POST':
        try:
            auditor_id = usuario_id
            auditado_id = request.form.get('auditado_id')
            area_auditada = request.form.get('area_auditada')
            data_auditoria = request.form.get('data_auditoria')
            atividade_auditada = request.form.get('atividade_auditada')

            procedimento_id = request.form.get('procedimento_id')
            procedimento_revisao_id = request.form.get('procedimento_revisao_id')

            pontos_observados = request.form.get('pontos_observados')

            criado_por = usuario_id

            # =========================
            # VALIDAR AUDITADO
            # =========================
            if perfil in ['administrador', 'avancado']:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                """, (auditado_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (auditado_id, centro_custos_id))

            usuario_valido = cursor.fetchone()

            if not usuario_valido:
                flash("Usuário auditado inválido para seu escopo.", "danger")
                conn.close()
                return redirect(url_for('main.lancar_ap'))

            # =========================
            # BUSCAR/CRIAR ORIGEM AP
            # =========================
            cursor.execute("""
                SELECT id
                FROM origens
                WHERE descricao = %s
                AND centro_custos_id = %s
                AND ativo = 1
                LIMIT 1
            """, ("Auditoria de Padrão", centro_custos_id))

            origem = cursor.fetchone()

            if origem:
                origem_ap_id = origem["id"]
            else:
                cursor.execute("""
                    INSERT INTO origens (nome, descricao, centro_custos_id, ativo)
                    VALUES (%s, %s, %s, 1)
                """, (
                    "Auditoria de Padrão",
                    "Auditoria de Padrão",
                    centro_custos_id
                ))
                origem_ap_id = cursor.lastrowid

            # =========================
            # INSERIR AUDITORIA
            # =========================
            cursor.execute("""
                INSERT INTO auditorias_padrao (
                    auditor_id,
                    auditado_id,
                    area_auditada,
                    data_auditoria,
                    atividade_auditada,
                    procedimento_id,
                    procedimento_revisao_id,
                    pontos_observados,
                    criado_por
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                auditor_id,
                auditado_id,
                area_auditada,
                data_auditoria,
                atividade_auditada,
                procedimento_id,
                procedimento_revisao_id,
                pontos_observados,
                criado_por
            ))

            auditoria_id = cursor.lastrowid

            # =========================
            # SALVAR CHECKLIST
            # =========================
            for i in range(1, 11):
                item_texto = request.form.get(f'item_texto_{i}')
                resultado = request.form.get(f'resultado_{i}')

                cursor.execute("""
                    INSERT INTO auditoria_padrao_respostas (
                        auditoria_id,
                        numero_item,
                        item_verificacao,
                        resultado
                    )
                    VALUES (%s,%s,%s,%s)
                """, (auditoria_id, i, item_texto, resultado))

            # =========================
            # DESVIOS
            # =========================
            desvios_json = request.form.get('desvios_json')

            if desvios_json:
                desvios = json.loads(desvios_json)

                for d in desvios:
                    cursor.execute("""
                        INSERT INTO acoes (
                            origem_id,
                            descricao,
                            responsavel_id,
                            prazo,
                            status,
                            criado_por
                        )
                        VALUES (%s,%s,%s,%s,'Não iniciada',%s)
                    """, (
                        origem_ap_id,
                        f"Desvio AP - Item {d['numero_item']}: {d['acao_proposta']}",
                        d['responsavel_id'],
                        d['prazo'],
                        criado_por
                    ))

                    acao_id = cursor.lastrowid

                    cursor.execute("""
                        INSERT INTO auditoria_padrao_desvios (
                            auditoria_id,
                            numero_item,
                            desvio_observado,
                            acao_proposta,
                            responsavel_id,
                            prazo,
                            acao_id
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        auditoria_id,
                        d['numero_item'],
                        d['desvio_observado'],
                        d['acao_proposta'],
                        d['responsavel_id'],
                        d['prazo'],
                        acao_id
                    ))

            # =========================
            # SOLICITAÇÃO DE REVISÃO
            # =========================
            revisao_json = request.form.get('solicitacao_revisao_json')

            if revisao_json:
                revisao = json.loads(revisao_json)

                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE responsavel_revisao_padrao = 1
                      AND ativo = 1
                    LIMIT 1
                """)
                resp = cursor.fetchone()

                if resp:
                    responsavel_id = resp['id']
                    prazo = datetime.now().date() + timedelta(days=30)

                    cursor.execute("""
                        INSERT INTO acoes (
                            origem_id,
                            descricao,
                            responsavel_id,
                            prazo,
                            status,
                            criado_por
                        )
                        VALUES (%s,%s,%s,%s,'Não iniciada',%s)
                    """, (
                        origem_ap_id,
                        f"Revisão de procedimento: {revisao['sugestao_revisao']}",
                        responsavel_id,
                        prazo,
                        criado_por
                    ))

                    acao_revisao_id = cursor.lastrowid

                    cursor.execute("""
                        UPDATE auditorias_padrao
                        SET necessita_revisao_padrao = 1,
                            oportunidade_revisao = %s,
                            justificativa_revisao = %s,
                            sugestao_revisao = %s,
                            acao_revisao_id = %s
                        WHERE id = %s
                    """, (
                        revisao['oportunidade_revisao'],
                        revisao['justificativa_revisao'],
                        revisao['sugestao_revisao'],
                        acao_revisao_id,
                        auditoria_id
                    ))

            conn.commit()

            flash('Auditoria registrada com sucesso!', 'success')
            return redirect(url_for('main.listar_ap'))

        except Exception as e:
            conn.rollback()
            flash(f'Erro ao salvar auditoria: {e}', 'danger')
            return redirect(url_for('main.lancar_ap'))

        finally:
            conn.close()

    # =========================
    # GET
    # =========================

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT 
            p.id AS procedimento_id,
            pr.id AS revisao_id,
            p.numero_documento,
            p.titulo,
            td.sigla,
            pr.numero_revisao
        FROM procedimentos p
        JOIN tipos_documento td ON td.id = p.tipo_documento_id
        JOIN procedimento_revisoes pr ON pr.procedimento_id = p.id
        WHERE pr.vigente = 1
        ORDER BY td.sigla, p.numero_documento
    """)
    procedimentos = cursor.fetchall()

    conn.close()

    return render_template(
        'lancar_ap.html',
        usuarios=usuarios,
        procedimentos=procedimentos
    )

@main_routes.route('/listar_ap', methods=['GET'])
@login_required
@module_required('acesso_ssma')
def listar_ap():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.listar_ap'))

    auditor_id = request.args.get('auditor_id', '').strip()
    auditado_id = request.args.get('auditado_id', '').strip()
    procedimento_id = request.args.get('procedimento_id', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    revisao_padrao = request.args.get('revisao_padrao', '').strip()

    sort = request.args.get('sort', 'data_auditoria').strip()
    order = request.args.get('order', 'desc').strip().lower()

    page = request.args.get('page', 1, type=int)
    per_page = 30

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        'id': 'ap.id',
        'data_auditoria': 'ap.data_auditoria',
        'auditor': 'auditor.nome',
        'auditado': 'auditado.nome',
        'area_auditada': 'ap.area_auditada',
        'procedimento': 'p.numero_documento',
        'revisao_padrao': 'ap.necessita_revisao_padrao'
    }

    coluna_sort = colunas_validas.get(sort, 'ap.data_auditoria')
    direcao = 'ASC' if order == 'asc' else 'DESC'

    filtros_sql = ["ap.ativo = 1"]
    params = []

    # CONTROLE DE ESCOPO POR PERFIL
    if perfil == 'basico':
        filtros_sql.append("ap.criado_por = %s")
        params.append(usuario_id)

    elif perfil == 'intermediario':
        filtros_sql.append("auditado.centro_custos_id = %s")
        params.append(centro_custos_id)

    # avançado e administrador veem tudo

    if auditor_id:
        filtros_sql.append("ap.auditor_id = %s")
        params.append(auditor_id)

    if auditado_id:
        filtros_sql.append("ap.auditado_id = %s")
        params.append(auditado_id)

    if procedimento_id:
        filtros_sql.append("ap.procedimento_id = %s")
        params.append(procedimento_id)

    if data_inicio:
        filtros_sql.append("ap.data_auditoria >= %s")
        params.append(data_inicio)

    if data_fim:
        filtros_sql.append("ap.data_auditoria <= %s")
        params.append(data_fim)

    if revisao_padrao == 'sim':
        filtros_sql.append("ap.necessita_revisao_padrao = 1")
    elif revisao_padrao == 'nao':
        filtros_sql.append("ap.necessita_revisao_padrao = 0")

    where_clause = "WHERE " + " AND ".join(filtros_sql)

    base_from = f"""
        FROM auditorias_padrao ap
        JOIN usuarios auditor
            ON auditor.id = ap.auditor_id
        JOIN usuarios auditado
            ON auditado.id = ap.auditado_id
        JOIN procedimentos p
            ON p.id = ap.procedimento_id
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        JOIN procedimento_revisoes pr
            ON pr.id = ap.procedimento_revisao_id
        {where_clause}
    """

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """, params)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cursor.execute(f"""
        SELECT
            ap.id,
            ap.auditor_id,
            ap.auditado_id,
            ap.data_auditoria,
            ap.area_auditada,
            ap.atividade_auditada,
            ap.pontos_observados,
            ap.necessita_revisao_padrao,
            ap.criado_por,

            auditor.nome AS auditor_nome,
            auditado.nome AS auditado_nome,

            td.sigla,
            p.numero_documento,
            p.titulo,
            pr.numero_revisao

        {base_from}
        ORDER BY {coluna_sort} {direcao}, ap.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    auditorias = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            p.id,
            td.sigla,
            p.numero_documento,
            p.titulo
        FROM procedimentos p
        JOIN tipos_documento td
            ON td.id = p.tipo_documento_id
        WHERE p.ativo = 1
        ORDER BY td.sigla, p.numero_documento, p.titulo
    """)
    procedimentos = cursor.fetchall()

    conn.close()

    filtros = {
        'auditor_id': auditor_id,
        'auditado_id': auditado_id,
        'procedimento_id': procedimento_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'revisao_padrao': revisao_padrao,
        'sort': sort,
        'order': order
    }

    return render_template(
        'listar_ap.html',
        auditorias=auditorias,
        usuarios=usuarios,
        procedimentos=procedimentos,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

@main_routes.route('/excluir_ap/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def excluir_ap(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        auditoria = pode_acessar_ssma(cursor, 'auditoria_padrao', id)

        if not auditoria:
            flash('Auditoria de Padrão não encontrada ou você não possui permissão para excluí-la.', 'warning')
            conn.close()
            return redirect(url_for('main.listar_ap'))

        perfil = session.get('perfil')
        usuario_id = session.get('usuario_id')

        # Regra adicional: somente administrador ou quem criou pode excluir
        if perfil != 'administrador' and auditoria['criado_por'] != usuario_id:
            flash('Você não tem permissão para excluir esta Auditoria de Padrão.', 'warning')
            conn.close()
            return redirect(url_for('main.listar_ap'))

        cursor.execute("""
            UPDATE auditorias_padrao
            SET ativo = 0
            WHERE id = %s
        """, (id,))

        conn.commit()
        flash('Auditoria de Padrão excluída com sucesso.', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao excluir Auditoria de Padrão: {e}', 'danger')

    finally:
        conn.close()

    return redirect(url_for('main.listar_ap'))

@main_routes.route('/editar_ap/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def editar_ap(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        auditoria = pode_acessar_ssma(cursor, 'auditoria_padrao', id)

        if not auditoria:
            flash('Auditoria de Padrão não encontrada ou você não possui permissão para editá-la.', 'warning')
            conn.close()
            return redirect(url_for('main.listar_ap'))

        perfil = session.get('perfil')
        usuario_id = session.get('usuario_id')
        centro_custos_id = session.get('centro_custos_id')

        # Regra adicional: somente administrador ou quem criou pode editar
        if perfil != 'administrador' and auditoria['criado_por'] != usuario_id:
            flash('Você não tem permissão para editar esta Auditoria de Padrão.', 'warning')
            conn.close()
            return redirect(url_for('main.listar_ap'))

        auditor_id = request.form.get('auditor_id')
        auditado_id = request.form.get('auditado_id')
        data_auditoria = request.form.get('data_auditoria')
        area_auditada = (request.form.get('area_auditada') or '').strip()
        atividade_auditada = (request.form.get('atividade_auditada') or '').strip()
        pontos_observados = (request.form.get('pontos_observados') or '').strip()

        if not auditor_id or not auditado_id or not data_auditoria or not area_auditada or not atividade_auditada:
            flash('Preencha todos os campos obrigatórios da Auditoria de Padrão.', 'danger')
            conn.close()
            return redirect(request.form.get('next') or url_for('main.listar_ap'))

        if perfil in ['administrador', 'avancado']:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
            """, (auditado_id,))
        else:
            cursor.execute("""
                SELECT id
                FROM usuarios
                WHERE id = %s
                  AND ativo = 1
                  AND centro_custos_id = %s
            """, (auditado_id, centro_custos_id))

        auditado_valido = cursor.fetchone()

        if not auditado_valido:
            flash('Usuário auditado inválido para seu escopo.', 'danger')
            conn.close()
            return redirect(request.form.get('next') or url_for('main.listar_ap'))

        if perfil == 'administrador':
            auditor_id_final = auditor_id
        else:
            auditor_id_final = auditoria['auditor_id']

        cursor.execute("""
            UPDATE auditorias_padrao
            SET auditor_id = %s,
                auditado_id = %s,
                data_auditoria = %s,
                area_auditada = %s,
                atividade_auditada = %s,
                pontos_observados = %s
            WHERE id = %s
              AND ativo = 1
        """, (
            auditor_id_final,
            auditado_id,
            data_auditoria,
            area_auditada,
            atividade_auditada,
            pontos_observados,
            id
        ))

        conn.commit()
        flash('Auditoria de Padrão atualizada com sucesso.', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar Auditoria de Padrão: {e}', 'danger')

    finally:
        conn.close()

    return redirect(request.form.get('next') or url_for('main.listar_ap'))

# =============================== #
# IFS                             #
# =============================== #

@main_routes.route('/lancar_ifs', methods=['GET', 'POST'])
@login_required
@module_required('acesso_ssma')
def lancar_ifs():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custo_id = session.get('centro_custos_id')

    if not usuario_id:
        conn.close()
        flash('Usuário logado não encontrado.', 'danger')
        return redirect(url_for('main.login'))

    if not centro_custo_id:
        conn.close()
        flash('Não foi possível identificar o centro de custo do usuário logado.', 'danger')
        return redirect(url_for('main.dashboard'))

    # =========================================================
    # POST → SALVAR IFS
    # =========================================================
    if request.method == 'POST':

        profissional_sesmt_id = usuario_id
        participantes = request.form.get('participantes')
        descricao_atividade_os = request.form.get('descricao_atividade_os')
        local_inspecao = request.form.get('local_inspecao')
        data_inspecao = request.form.get('data_inspecao')
        hora_inspecao = request.form.get('hora_inspecao')

        respostas_json = request.form.get('respostas_json')
        desvios_json = request.form.get('desvios_json')

        respostas = json.loads(respostas_json) if respostas_json else []
        desvios = json.loads(desvios_json) if desvios_json else []

        total_itens_avaliados = len(respostas)
        total_nao_conformidades = sum(1 for r in respostas if r.get('resultado') == 'N')

        # Busca origem IFS do centro de custo
        cursor.execute("""
            SELECT id
            FROM origens
            WHERE nome = 'IFS - Inspeção de Frente de Serviço'
              AND centro_custos_id = %s
              AND ativo = 1
            LIMIT 1
        """, (centro_custo_id,))
        origem_ifs = cursor.fetchone()

        if not origem_ifs:
            cursor.execute("""
                INSERT INTO origens (
                    nome,
                    descricao,
                    ativo,
                    centro_custos_id
                )
                VALUES (%s, %s, 1, %s)
            """, (
                'IFS - Inspeção de Frente de Serviço',
                'IFS - Inspeção de Frente de Serviço',
                centro_custo_id
            ))

            origem_id_ifs = cursor.lastrowid
        else:
            origem_id_ifs = origem_ifs['id']

        # Inserir inspeção
        cursor.execute("""
            INSERT INTO ifs_inspecoes (
                profissional_sesmt_id,
                participantes,
                descricao_atividade_os,
                local_inspecao,
                data_inspecao,
                hora_inspecao,
                total_itens_avaliados,
                total_nao_conformidades,
                criado_por
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            profissional_sesmt_id,
            participantes,
            descricao_atividade_os,
            local_inspecao,
            data_inspecao,
            hora_inspecao,
            total_itens_avaliados,
            total_nao_conformidades,
            usuario_id
        ))

        inspecao_id = cursor.lastrowid

        # Salvar respostas
        for r in respostas:
            cursor.execute("""
                INSERT INTO ifs_respostas (
                    inspecao_id,
                    item_id,
                    resultado
                )
                VALUES (%s,%s,%s)
            """, (
                inspecao_id,
                r.get('item_id'),
                r.get('resultado')
            ))

        # Salvar desvios + criar ações
        for d in desvios:

            responsavel_id = d.get('responsavel_id')

            if perfil in ['administrador', 'avancado']:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                """, (responsavel_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (responsavel_id, centro_custo_id))

            responsavel_valido = cursor.fetchone()

            if not responsavel_valido:
                conn.rollback()
                conn.close()
                flash('Responsável inválido para seu escopo.', 'danger')
                return redirect(url_for('main.lancar_ifs'))

            descricao_acao = (
                f"IFS - Item {d.get('codigo')}: {d.get('descricao')}\n\n"
                f"Desvio observado: {d.get('desvio_observado')}\n\n"
                f"Ação proposta: {d.get('acao_proposta')}"
            )

            cursor.execute("""
                INSERT INTO acoes (
                    origem_id,
                    descricao,
                    responsavel_id,
                    prazo,
                    status,
                    criado_por
                )
                VALUES (%s,%s,%s,%s,'Não iniciada',%s)
            """, (
                origem_id_ifs,
                descricao_acao,
                responsavel_id,
                d.get('prazo'),
                usuario_id
            ))

            acao_id = cursor.lastrowid

            cursor.execute("""
                INSERT INTO ifs_desvios (
                    inspecao_id,
                    item_id,
                    desvio_observado,
                    acao_proposta,
                    responsavel_id,
                    prazo,
                    acao_id
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                inspecao_id,
                d.get('item_id'),
                d.get('desvio_observado'),
                d.get('acao_proposta'),
                responsavel_id,
                d.get('prazo'),
                acao_id
            ))

        conn.commit()
        conn.close()

        flash('IFS registrada com sucesso.', 'success')
        return redirect(url_for('main.listar_ifs'))

    # =========================================================
    # GET → CARREGAR TELA
    # =========================================================

    cursor.execute("""
        SELECT id, nome
        FROM ifs_blocos
        WHERE ativo = 1
        ORDER BY ordem
    """)
    blocos = cursor.fetchall()

    cursor.execute("""
        SELECT
            id,
            bloco_id,
            codigo,
            descricao,
            potencial
        FROM ifs_itens
        WHERE ativo = 1
        ORDER BY bloco_id, ordem
    """)
    itens = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id,))

    usuarios = cursor.fetchall()

    conn.close()

    itens_por_bloco = {}

    for bloco in blocos:
        itens_por_bloco[bloco['id']] = []

    for item in itens:
        itens_por_bloco[item['bloco_id']].append(item)

    return render_template(
        'lancar_ifs.html',
        blocos=blocos,
        itens_por_bloco=itens_por_bloco,
        usuarios=usuarios
    )

@main_routes.route('/listar_ifs', methods=['GET'])
@login_required
@module_required('acesso_ssma')
def listar_ifs():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custo_id = session.get('centro_custos_id')

    # 🔹 limpar filtros
    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.listar_ifs'))

    # 🔹 filtros
    profissional_sesmt_id = request.args.get('profissional_sesmt_id', '').strip()
    local_inspecao = request.args.get('local_inspecao', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()

    # 🔹 paginação
    page = request.args.get('page', 1, type=int)
    per_page = 20

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    # 🔹 filtros SQL
    filtros_sql = ["ifs.ativo = 1"]
    params = []

    # 🔒 CONTROLE DE ESCOPO
    if perfil == 'basico':
        filtros_sql.append("ifs.criado_por = %s")
        params.append(usuario_id)

    elif perfil == 'intermediario':
        filtros_sql.append("u.centro_custos_id = %s")
        params.append(centro_custo_id)

    # avançado e administrador veem tudo

    if profissional_sesmt_id:
        filtros_sql.append("ifs.profissional_sesmt_id = %s")
        params.append(profissional_sesmt_id)

    if local_inspecao:
        filtros_sql.append("ifs.local_inspecao LIKE %s")
        params.append(f"%{local_inspecao}%")

    if data_inicio:
        filtros_sql.append("ifs.data_inspecao >= %s")
        params.append(data_inicio)

    if data_fim:
        filtros_sql.append("ifs.data_inspecao <= %s")
        params.append(data_fim)

    where_clause = "WHERE " + " AND ".join(filtros_sql)

    # 🔹 base da query
    base_from = f"""
        FROM ifs_inspecoes ifs
        JOIN usuarios u
            ON u.id = ifs.profissional_sesmt_id
        {where_clause}
    """

    # 🔹 total de registros
    cursor.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """, params)

    total_registros = cursor.fetchone()['total']
    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    # 🔹 query principal
    cursor.execute(f"""
        SELECT
            ifs.id,
            ifs.data_inspecao,
            ifs.hora_inspecao,
            ifs.local_inspecao,
            ifs.descricao_atividade_os,
            ifs.total_itens_avaliados,
            ifs.total_nao_conformidades,
            ifs.criado_por,

            u.nome AS profissional_sesmt_nome

        {base_from}
        ORDER BY ifs.data_inspecao DESC, ifs.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    inspecoes = cursor.fetchall()

    # 🔹 usuários (para filtro)
    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id,))

    usuarios = cursor.fetchall()

    conn.close()

    # 🔹 filtros para manter estado na tela
    filtros = {
        'profissional_sesmt_id': profissional_sesmt_id,
        'local_inspecao': local_inspecao,
        'data_inicio': data_inicio,
        'data_fim': data_fim
    }

    return render_template(
        'listar_ifs.html',
        inspecoes=inspecoes,
        usuarios=usuarios,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

@main_routes.route('/excluir_ifs/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def excluir_ifs(id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    inspecao = pode_acessar_ssma(cursor, 'ifs', id)

    if not inspecao:
        conn.close()
        flash('IFS não encontrada ou você não possui permissão para excluí-la.', 'warning')
        return redirect(url_for('main.listar_ifs'))

    # Regra adicional: somente administrador ou quem criou pode excluir
    if inspecao['criado_por'] != session.get('usuario_id') and session.get('perfil') != 'administrador':
        conn.close()
        flash('Você não tem permissão para excluir esta IFS.', 'danger')
        return redirect(url_for('main.listar_ifs'))

    cursor.execute("""
        UPDATE ifs_inspecoes
        SET ativo = 0,
            atualizado_em = NOW()
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('IFS excluída com sucesso.', 'success')
    return redirect(url_for('main.listar_ifs'))

@main_routes.route('/editar_ifs/<int:id>', methods=['GET', 'POST'])
@login_required
@module_required('acesso_ssma')
def editar_ifs(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custo_id = session.get('centro_custos_id')

    inspecao = pode_acessar_ssma(cursor, 'ifs', id)

    if not inspecao:
        conn.close()
        flash('IFS não encontrada ou você não possui permissão para acessá-la.', 'warning')
        return redirect(url_for('main.listar_ifs'))

    if inspecao['criado_por'] != usuario_id and perfil != 'administrador':
        conn.close()
        flash('Você não tem permissão para editar esta IFS.', 'danger')
        return redirect(url_for('main.listar_ifs'))

    if request.method == 'POST':
        profissional_sesmt_id = usuario_id
        participantes = request.form.get('participantes')
        descricao_atividade_os = request.form.get('descricao_atividade_os')
        local_inspecao = request.form.get('local_inspecao')
        data_inspecao = request.form.get('data_inspecao')
        hora_inspecao = request.form.get('hora_inspecao')

        respostas_json = request.form.get('respostas_json')
        desvios_json = request.form.get('desvios_json')

        respostas = json.loads(respostas_json) if respostas_json else []
        desvios = json.loads(desvios_json) if desvios_json else []

        total_itens_avaliados = len(respostas)
        total_nao_conformidades = sum(1 for r in respostas if r.get('resultado') == 'N')

        cursor.execute("""
            UPDATE ifs_inspecoes
            SET profissional_sesmt_id = %s,
                participantes = %s,
                descricao_atividade_os = %s,
                local_inspecao = %s,
                data_inspecao = %s,
                hora_inspecao = %s,
                total_itens_avaliados = %s,
                total_nao_conformidades = %s,
                atualizado_em = NOW()
            WHERE id = %s
        """, (
            profissional_sesmt_id,
            participantes,
            descricao_atividade_os,
            local_inspecao,
            data_inspecao,
            hora_inspecao,
            total_itens_avaliados,
            total_nao_conformidades,
            id
        ))

        cursor.execute("DELETE FROM ifs_respostas WHERE inspecao_id = %s", (id,))

        for r in respostas:
            cursor.execute("""
                INSERT INTO ifs_respostas (inspecao_id, item_id, resultado)
                VALUES (%s, %s, %s)
            """, (id, r.get('item_id'), r.get('resultado')))

        cursor.execute("SELECT * FROM ifs_desvios WHERE inspecao_id = %s", (id,))
        desvios_existentes_lista = cursor.fetchall()

        desvios_existentes = {
            str(d['item_id']): d for d in desvios_existentes_lista
        }

        novos_item_ids = [str(d.get('item_id')) for d in desvios]

        if novos_item_ids:
            placeholders = ','.join(['%s'] * len(novos_item_ids))
            cursor.execute(f"""
                DELETE FROM ifs_desvios
                WHERE inspecao_id = %s
                  AND item_id NOT IN ({placeholders})
            """, [id] + novos_item_ids)
        else:
            cursor.execute("DELETE FROM ifs_desvios WHERE inspecao_id = %s", (id,))

        cursor.execute("""
            SELECT id
            FROM origens
            WHERE nome = 'IFS - Inspeção de Frente de Serviço'
              AND centro_custos_id = %s
              AND ativo = 1
            LIMIT 1
        """, (centro_custo_id,))
        origem_ifs = cursor.fetchone()

        if not origem_ifs:
            cursor.execute("""
                INSERT INTO origens (nome, descricao, ativo, centro_custos_id)
                VALUES (%s, %s, 1, %s)
            """, (
                'IFS - Inspeção de Frente de Serviço',
                'IFS - Inspeção de Frente de Serviço',
                centro_custo_id
            ))
            origem_id_ifs = cursor.lastrowid
        else:
            origem_id_ifs = origem_ifs['id']

        for d in desvios:
            responsavel_id = d.get('responsavel_id')

            if perfil in ['administrador', 'avancado']:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                """, (responsavel_id,))
            else:
                cursor.execute("""
                    SELECT id
                    FROM usuarios
                    WHERE id = %s
                      AND ativo = 1
                      AND centro_custos_id = %s
                """, (responsavel_id, centro_custo_id))

            if not cursor.fetchone():
                conn.rollback()
                conn.close()
                flash('Responsável inválido para seu escopo.', 'danger')
                return redirect(url_for('main.editar_ifs', id=id))

            item_id = str(d.get('item_id'))

            descricao_acao = (
                f"IFS - Item {d.get('codigo')}: {d.get('descricao')}\n\n"
                f"Desvio observado: {d.get('desvio_observado')}\n\n"
                f"Ação proposta: {d.get('acao_proposta')}"
            )

            desvio_existente = desvios_existentes.get(item_id)

            if desvio_existente:
                acao_id = desvio_existente.get('acao_id')

                cursor.execute("""
                    UPDATE ifs_desvios
                    SET desvio_observado = %s,
                        acao_proposta = %s,
                        responsavel_id = %s,
                        prazo = %s
                    WHERE id = %s
                """, (
                    d.get('desvio_observado'),
                    d.get('acao_proposta'),
                    responsavel_id,
                    d.get('prazo'),
                    desvio_existente['id']
                ))

                if acao_id:
                    cursor.execute("""
                        UPDATE acoes
                        SET origem_id = %s,
                            descricao = %s,
                            responsavel_id = %s,
                            prazo = %s
                        WHERE id = %s
                    """, (
                        origem_id_ifs,
                        descricao_acao,
                        responsavel_id,
                        d.get('prazo'),
                        acao_id
                    ))

            else:
                cursor.execute("""
                    INSERT INTO acoes (
                        origem_id,
                        descricao,
                        responsavel_id,
                        prazo,
                        status,
                        criado_por
                    )
                    VALUES (%s, %s, %s, %s, 'Não iniciada', %s)
                """, (
                    origem_id_ifs,
                    descricao_acao,
                    responsavel_id,
                    d.get('prazo'),
                    usuario_id
                ))

                acao_id = cursor.lastrowid

                cursor.execute("""
                    INSERT INTO ifs_desvios (
                        inspecao_id,
                        item_id,
                        desvio_observado,
                        acao_proposta,
                        responsavel_id,
                        prazo,
                        acao_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    id,
                    d.get('item_id'),
                    d.get('desvio_observado'),
                    d.get('acao_proposta'),
                    responsavel_id,
                    d.get('prazo'),
                    acao_id
                ))

        conn.commit()
        conn.close()

        flash('IFS atualizada com sucesso.', 'success')
        return redirect(url_for('main.listar_ifs'))

    # =========================================================
    # GET → CARREGAR TELA
    # =========================================================

    cursor.execute("""
        SELECT id, nome
        FROM ifs_blocos
        WHERE ativo = 1
        ORDER BY ordem
    """)
    blocos = cursor.fetchall()

    cursor.execute("""
        SELECT id, bloco_id, codigo, descricao, potencial
        FROM ifs_itens
        WHERE ativo = 1
        ORDER BY bloco_id, ordem
    """)
    itens = cursor.fetchall()

    cursor.execute("""
        SELECT item_id, resultado
        FROM ifs_respostas
        WHERE inspecao_id = %s
    """, (id,))
    respostas_lista = cursor.fetchall()

    respostas_por_item = {
        str(r['item_id']): r['resultado'] for r in respostas_lista
    }

    cursor.execute("""
        SELECT
            d.id,
            d.inspecao_id,
            d.item_id,
            d.desvio_observado,
            d.acao_proposta,
            d.responsavel_id,
            DATE_FORMAT(d.prazo, '%Y-%m-%d') AS prazo,
            d.acao_id,
            i.codigo,
            i.descricao,
            i.potencial,
            i.bloco_id
        FROM ifs_desvios d
        JOIN ifs_itens i ON i.id = d.item_id
        WHERE d.inspecao_id = %s
    """, (id,))
    desvios_lista = cursor.fetchall()

    desvios_por_item = {
        str(d['item_id']): d for d in desvios_lista
    }

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custo_id,))

    usuarios = cursor.fetchall()

    conn.close()

    itens_por_bloco = {}

    for bloco in blocos:
        itens_por_bloco[bloco['id']] = []

    for item in itens:
        itens_por_bloco[item['bloco_id']].append(item)

    return render_template(
        'editar_ifs.html',
        inspecao=inspecao,
        blocos=blocos,
        itens_por_bloco=itens_por_bloco,
        respostas_por_item=respostas_por_item,
        desvios_por_item=desvios_por_item,
        usuarios=usuarios
    )

#===============================#
# APR                           #
#===============================#

@main_routes.route('/listar_apr', methods=['GET'])
@login_required
@module_required('acesso_ssma')
def listar_apr():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.listar_apr'))

    emitente_id = request.args.get('emitente_id', '').strip()
    aprovador_id = request.args.get('aprovador_id', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()

    sort = request.args.get('sort', 'data_apr').strip()
    order = request.args.get('order', 'desc').strip().lower()

    page = request.args.get('page', 1, type=int)
    per_page = 20

    if page < 1:
        page = 1

    offset = (page - 1) * per_page

    colunas_validas = {
        'data_apr': 'apr.data_apr',
        'area_setor': 'apr.area_setor',
        'atividade': 'apr.atividade',
        'emitente_nome': 'emitente.nome',
        'aprovador_nome': 'aprovador.nome'
    }

    sort_sql = colunas_validas.get(sort, 'apr.data_apr')
    order_sql = 'ASC' if order == 'asc' else 'DESC'

    filtros_sql = ["apr.ativo = 1"]
    params = []

    # Controle de escopo por perfil
    if perfil == 'basico':
        filtros_sql.append("apr.criado_por = %s")
        params.append(usuario_id)

    elif perfil == 'intermediario':
        filtros_sql.append("emitente.centro_custos_id = %s")
        params.append(centro_custos_id)

    # avançado e administrador veem tudo

    if emitente_id:
        filtros_sql.append("apr.emitente_id = %s")
        params.append(emitente_id)

    if aprovador_id:
        filtros_sql.append("apr.aprovador_id = %s")
        params.append(aprovador_id)

    if data_inicio:
        filtros_sql.append("apr.data_apr >= %s")
        params.append(data_inicio)

    if data_fim:
        filtros_sql.append("apr.data_apr <= %s")
        params.append(data_fim)

    where_clause = "WHERE " + " AND ".join(filtros_sql)

    base_from = f"""
        FROM aprs apr
        JOIN usuarios emitente
            ON emitente.id = apr.emitente_id
        JOIN usuarios aprovador
            ON aprovador.id = apr.aprovador_id
        {where_clause}
    """

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        {base_from}
    """, params)
    total_registros = cursor.fetchone()['total']

    total_paginas = (total_registros + per_page - 1) // per_page

    if total_paginas > 0 and page > total_paginas:
        page = total_paginas
        offset = (page - 1) * per_page

    cursor.execute(f"""
        SELECT
            apr.id,
            apr.data_apr,
            apr.area_setor,
            apr.atividade,
            apr.emitente_id,
            apr.aprovador_id,
            apr.arquivo_pdf,
            apr.criado_por,
            apr.criado_em,

            emitente.nome AS emitente_nome,
            aprovador.nome AS aprovador_nome

        {base_from}
        ORDER BY {sort_sql} {order_sql}, apr.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    aprs = cursor.fetchall()

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, matricula
            FROM usuarios
            WHERE ativo = 1
              AND centro_custos_id = %s
            ORDER BY nome
        """, (centro_custos_id,))

    usuarios = cursor.fetchall()

    conn.close()

    filtros = {
        'emitente_id': emitente_id,
        'aprovador_id': aprovador_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'sort': sort,
        'order': order
    }

    return render_template(
        'listar_apr.html',
        aprs=aprs,
        usuarios=usuarios,
        filtros=filtros,
        page=page,
        per_page=per_page,
        total_registros=total_registros,
        total_paginas=total_paginas
    )

@main_routes.route('/cadastrar_apr', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def cadastrar_apr():
    data_apr = request.form.get('data_apr', '').strip()
    area_setor = request.form.get('area_setor', '').strip()
    atividade = request.form.get('atividade', '').strip()
    emitente_id = request.form.get('emitente_id', '').strip()
    aprovador_id = request.form.get('aprovador_id', '').strip()
    arquivo = request.files.get('arquivo_pdf')

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if not data_apr or not area_setor or not atividade or not emitente_id or not aprovador_id:
        flash('Preencha todos os campos obrigatórios.', 'warning')
        return redirect(url_for('main.listar_apr'))

    if not arquivo or arquivo.filename == '':
        flash('Anexe o arquivo PDF da APR.', 'warning')
        return redirect(url_for('main.listar_apr'))

    extensao = arquivo.filename.rsplit('.', 1)[-1].lower() if '.' in arquivo.filename else ''

    if extensao != 'pdf':
        flash('Somente arquivos PDF são permitidos.', 'danger')
        return redirect(url_for('main.listar_apr'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id
            FROM usuarios
            WHERE id IN (%s, %s)
              AND ativo = 1
        """, (emitente_id, aprovador_id))
    else:
        cursor.execute("""
            SELECT id
            FROM usuarios
            WHERE id IN (%s, %s)
              AND ativo = 1
              AND centro_custos_id = %s
        """, (emitente_id, aprovador_id, centro_custos_id))

    usuarios_validos = cursor.fetchall()

    if len(usuarios_validos) != 2:
        conn.close()
        flash('Emitente ou aprovador inválido para seu escopo.', 'danger')
        return redirect(url_for('main.listar_apr'))

    pasta_aprs = os.path.join(current_app.root_path, 'static', 'aprs')
    os.makedirs(pasta_aprs, exist_ok=True)

    nome_original = secure_filename(arquivo.filename)
    nome_arquivo = f"apr_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nome_original}"

    caminho_arquivo = os.path.join(pasta_aprs, nome_arquivo)
    arquivo.save(caminho_arquivo)

    cursor.execute("""
        INSERT INTO aprs (
            data_apr,
            area_setor,
            atividade,
            emitente_id,
            aprovador_id,
            arquivo_pdf,
            criado_por
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        data_apr,
        area_setor,
        atividade,
        emitente_id,
        aprovador_id,
        nome_arquivo,
        usuario_id
    ))

    conn.commit()
    conn.close()

    flash('APR cadastrada com sucesso.', 'success')
    return redirect(url_for('main.listar_apr'))

@main_routes.route('/editar_apr/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def editar_apr(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    apr = pode_acessar_ssma(cursor, 'apr', id)

    if not apr:
        conn.close()
        flash('APR não encontrada ou você não possui permissão para editá-la.', 'warning')
        return redirect(url_for('main.listar_apr'))

    usuario_id = session.get('usuario_id')
    perfil = session.get('perfil')
    centro_custos_id = session.get('centro_custos_id')

    if apr['criado_por'] != usuario_id and perfil != 'administrador':
        conn.close()
        flash('Você não tem permissão para editar esta APR.', 'danger')
        return redirect(url_for('main.listar_apr'))

    data_apr = request.form.get('data_apr', '').strip()
    area_setor = request.form.get('area_setor', '').strip()
    atividade = request.form.get('atividade', '').strip()
    emitente_id = request.form.get('emitente_id', '').strip()
    aprovador_id = request.form.get('aprovador_id', '').strip()
    arquivo = request.files.get('arquivo_pdf')

    if not data_apr or not area_setor or not atividade or not emitente_id or not aprovador_id:
        conn.close()
        flash('Preencha todos os campos obrigatórios.', 'warning')
        return redirect(url_for('main.listar_apr'))

    if perfil in ['administrador', 'avancado']:
        cursor.execute("""
            SELECT id
            FROM usuarios
            WHERE id IN (%s, %s)
              AND ativo = 1
        """, (emitente_id, aprovador_id))
    else:
        cursor.execute("""
            SELECT id
            FROM usuarios
            WHERE id IN (%s, %s)
              AND ativo = 1
              AND centro_custos_id = %s
        """, (emitente_id, aprovador_id, centro_custos_id))

    usuarios_validos = cursor.fetchall()

    if len(usuarios_validos) != 2:
        conn.close()
        flash('Emitente ou aprovador inválido para seu escopo.', 'danger')
        return redirect(url_for('main.listar_apr'))

    nome_arquivo = apr['arquivo_pdf']

    if arquivo and arquivo.filename:
        extensao = arquivo.filename.rsplit('.', 1)[-1].lower() if '.' in arquivo.filename else ''

        if extensao != 'pdf':
            conn.close()
            flash('Somente arquivos PDF são permitidos.', 'danger')
            return redirect(url_for('main.listar_apr'))

        pasta_aprs = os.path.join(current_app.root_path, 'static', 'aprs')
        os.makedirs(pasta_aprs, exist_ok=True)

        nome_original = secure_filename(arquivo.filename)
        novo_nome_arquivo = f"apr_{id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nome_original}"

        caminho_novo = os.path.join(pasta_aprs, novo_nome_arquivo)
        arquivo.save(caminho_novo)

        if nome_arquivo:
            caminho_antigo = os.path.join(pasta_aprs, nome_arquivo)
            if os.path.exists(caminho_antigo):
                try:
                    os.remove(caminho_antigo)
                except Exception:
                    pass

        nome_arquivo = novo_nome_arquivo

    cursor.execute("""
        UPDATE aprs
        SET data_apr = %s,
            area_setor = %s,
            atividade = %s,
            emitente_id = %s,
            aprovador_id = %s,
            arquivo_pdf = %s,
            atualizado_em = NOW()
        WHERE id = %s
          AND ativo = 1
    """, (
        data_apr,
        area_setor,
        atividade,
        emitente_id,
        aprovador_id,
        nome_arquivo,
        id
    ))

    conn.commit()
    conn.close()

    flash('APR atualizada com sucesso.', 'success')
    return redirect(url_for('main.listar_apr'))

@main_routes.route('/excluir_apr/<int:id>', methods=['POST'])
@login_required
@module_required('acesso_ssma')
def excluir_apr(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    apr = pode_acessar_ssma(cursor, 'apr', id)

    if not apr:
        conn.close()
        flash('APR não encontrada ou você não possui permissão para excluí-la.', 'warning')
        return redirect(url_for('main.listar_apr'))

    if apr['criado_por'] != session.get('usuario_id') and session.get('perfil') != 'administrador':
        conn.close()
        flash('Você não tem permissão para excluir esta APR.', 'danger')
        return redirect(url_for('main.listar_apr'))

    cursor.execute("""
        UPDATE aprs
        SET ativo = 0,
            atualizado_em = NOW()
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('APR excluída com sucesso.', 'success')
    return redirect(url_for('main.listar_apr'))

# =============================== #
# EQUIPAMENTOS                    #
# =============================== #


@main_routes.route('/tipos_equipamento', methods=['GET'])
@login_required
def tipos_equipamento():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')

    colunas_validas = {
        'id': 'id',
        'tag': 'tag',
        'nome': 'nome'
    }

    if sort not in colunas_validas:
        sort = 'nome'

    if order not in ['asc', 'desc']:
        order = 'asc'

    cursor.execute(f"""
        SELECT
            id,
            tag,
            nome,
            ativo
        FROM pcpm_tipos_equipamento
        ORDER BY {colunas_validas[sort]} {order.upper()}
    """)

    tipos_equipamento = cursor.fetchall()
    conn.close()

    return render_template(
        'tipos_equipamento.html',
        tipos_equipamento=tipos_equipamento
    )


@main_routes.route('/cadastrar_tipo_equipamento', methods=['POST'])
@login_required
def cadastrar_tipo_equipamento():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tag = (request.form.get('tag') or '').strip().upper()
    nome = (request.form.get('nome') or '').strip()

    if not tag:
        flash('Informe a TAG do tipo de equipamento.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    if not nome:
        flash('Informe o nome do tipo de equipamento.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM pcpm_tipos_equipamento
        WHERE UPPER(tag) = UPPER(%s)
           OR UPPER(nome) = UPPER(%s)
    """, (tag, nome))

    existente = cursor.fetchone()

    if existente:
        conn.close()
        flash('Já existe um tipo de equipamento com essa TAG ou nome.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    cursor.execute("""
        INSERT INTO pcpm_tipos_equipamento (
            tag,
            nome,
            ativo
        )
        VALUES (%s, %s, 1)
    """, (tag, nome))

    conn.commit()
    conn.close()

    flash('Tipo de equipamento cadastrado com sucesso!', 'success')
    return redirect(url_for('main.tipos_equipamento'))


@main_routes.route('/editar_tipo_equipamento/<int:id>', methods=['POST'])
@login_required
def editar_tipo_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tag = (request.form.get('tag') or '').strip().upper()
    nome = (request.form.get('nome') or '').strip()

    if not tag:
        flash('Informe a TAG do tipo de equipamento.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    if not nome:
        flash('Informe o nome do tipo de equipamento.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT id
        FROM pcpm_tipos_equipamento
        WHERE (UPPER(tag) = UPPER(%s)
            OR UPPER(nome) = UPPER(%s))
          AND id <> %s
    """, (tag, nome, id))

    existente = cursor.fetchone()

    if existente:
        conn.close()
        flash('Já existe outro tipo de equipamento com essa TAG ou nome.', 'warning')
        return redirect(url_for('main.tipos_equipamento'))

    cursor.execute("""
        UPDATE pcpm_tipos_equipamento
        SET tag = %s,
            nome = %s
        WHERE id = %s
    """, (tag, nome, id))

    conn.commit()
    conn.close()

    flash('Tipo de equipamento atualizado com sucesso!', 'success')
    return redirect(url_for('main.tipos_equipamento'))


@main_routes.route('/inativar_tipo_equipamento/<int:id>', methods=['POST'])
@login_required
def inativar_tipo_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_tipos_equipamento
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Tipo de equipamento inativado com sucesso!', 'success')
    return redirect(url_for('main.tipos_equipamento'))


@main_routes.route('/ativar_tipo_equipamento/<int:id>', methods=['POST'])
@login_required
def ativar_tipo_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_tipos_equipamento
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Tipo de equipamento ativado com sucesso!', 'success')
    return redirect(url_for('main.tipos_equipamento'))

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_tipos_equipamento
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Tipo de equipamento ativado com sucesso!', 'success')

    return redirect(url_for('main.tipos_equipamento'))


@main_routes.route('/equipamentos', methods=['GET'])
@login_required
def equipamentos():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.equipamentos'))

    filtros = {
        'codigo_frota': (request.args.get('codigo_frota') or '').strip().upper(),
        'tipo_equipamento_id': request.args.get('tipo_equipamento_id') or '',
        'centro_custo_id': request.args.get('centro_custo_id') or ''
    }

    where = []
    params = []

    if filtros['codigo_frota']:
        where.append("e.codigo_frota LIKE %s")
        params.append(f"%{filtros['codigo_frota']}%")

    if filtros['tipo_equipamento_id']:
        where.append("e.tipo_equipamento_id = %s")
        params.append(filtros['tipo_equipamento_id'])

    if filtros['centro_custo_id']:
        where.append("e.centro_custo_id = %s")
        params.append(filtros['centro_custo_id'])

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    cursor.execute("""
        SELECT
            id,
            tag,
            nome
        FROM pcpm_tipos_equipamento
        WHERE ativo = 1
        ORDER BY tag, nome
    """)
    tipos_equipamento = cursor.fetchall()

    cursor.execute("""
        SELECT
            id,
            codigo,
            descricao
        FROM centros_custos
        WHERE ativo = 1
        ORDER BY codigo, descricao
    """)
    centros_custos = cursor.fetchall()

    cursor.execute(f"""
        SELECT
            e.id,
            e.tipo_equipamento_id,
            e.centro_custo_id,
            e.codigo_frota,
            e.marca,
            e.modelo,
            e.ativo,
            te.tag AS tipo_tag,
            te.nome AS tipo_nome,
            cc.codigo AS centro_custo_codigo,
            cc.descricao AS centro_custo_descricao
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        JOIN centros_custos cc
            ON cc.id = e.centro_custo_id
        {where_sql}
        ORDER BY e.codigo_frota ASC
    """, params)

    equipamentos = cursor.fetchall()

    conn.close()

    return render_template(
        'equipamentos.html',
        equipamentos=equipamentos,
        tipos_equipamento=tipos_equipamento,
        centros_custos=centros_custos,
        filtros=filtros
    )


@main_routes.route('/cadastrar_equipamento', methods=['POST'])
@login_required
def cadastrar_equipamento():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tipo_equipamento_id = request.form.get('tipo_equipamento_id')
    centro_custo_id = request.form.get('centro_custo_id')

    numero_frota = (request.form.get('codigo_frota') or '').strip()

    marca = (request.form.get('marca') or '').strip()
    modelo = (request.form.get('modelo') or '').strip()

    if not tipo_equipamento_id:
        flash('Selecione o tipo de equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not centro_custo_id:
        flash('Selecione o centro de custo.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not numero_frota:
        flash('Informe o número da frota.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not numero_frota.isdigit() or len(numero_frota) != 4:
        flash('Informe os 4 dígitos numéricos da frota.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not marca:
        flash('Informe a marca do equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not modelo:
        flash('Informe o modelo do equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT
                id,
                tag
            FROM pcpm_tipos_equipamento
            WHERE id = %s
              AND ativo = 1
        """, (tipo_equipamento_id,))

        tipo = cursor.fetchone()

        if not tipo:
            flash('Tipo de equipamento inválido.', 'warning')
            conn.close()
            return redirect(url_for('main.equipamentos'))

        codigo_frota = f"{tipo['tag']}-{numero_frota}"

        cursor.execute("""
            SELECT id
            FROM pcpm_equipamentos
            WHERE UPPER(codigo_frota) = UPPER(%s)
        """, (codigo_frota,))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe um equipamento cadastrado com esta frota.', 'warning')
            conn.close()
            return redirect(url_for('main.equipamentos'))

        cursor.execute("""
            INSERT INTO pcpm_equipamentos (
                tipo_equipamento_id,
                centro_custo_id,
                codigo_frota,
                marca,
                modelo,
                ativo
            )
            VALUES (%s, %s, %s, %s, %s, 1)
        """, (
            tipo_equipamento_id,
            centro_custo_id,
            codigo_frota,
            marca,
            modelo
        ))

        conn.commit()

        flash('Equipamento cadastrado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao cadastrar equipamento: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.equipamentos'))


@main_routes.route('/editar_equipamento/<int:id>', methods=['POST'])
@login_required
def editar_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tipo_equipamento_id = request.form.get('tipo_equipamento_id')
    centro_custo_id = request.form.get('centro_custo_id')

    numero_frota = (request.form.get('codigo_frota') or '').strip()

    marca = (request.form.get('marca') or '').strip()
    modelo = (request.form.get('modelo') or '').strip()

    if not tipo_equipamento_id:
        flash('Selecione o tipo de equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not centro_custo_id:
        flash('Selecione o centro de custo.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not numero_frota:
        flash('Informe o número da frota.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not numero_frota.isdigit() or len(numero_frota) != 4:
        flash('Informe os 4 dígitos numéricos da frota.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not marca:
        flash('Informe a marca do equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    if not modelo:
        flash('Informe o modelo do equipamento.', 'warning')
        return redirect(url_for('main.equipamentos'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT
                id,
                tag
            FROM pcpm_tipos_equipamento
            WHERE id = %s
              AND ativo = 1
        """, (tipo_equipamento_id,))

        tipo = cursor.fetchone()

        if not tipo:
            flash('Tipo de equipamento inválido.', 'warning')
            conn.close()
            return redirect(url_for('main.equipamentos'))

        codigo_frota = f"{tipo['tag']}-{numero_frota}"

        cursor.execute("""
            SELECT id
            FROM pcpm_equipamentos
            WHERE UPPER(codigo_frota) = UPPER(%s)
              AND id <> %s
        """, (codigo_frota, id))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe outro equipamento cadastrado com esta frota.', 'warning')
            conn.close()
            return redirect(url_for('main.equipamentos'))

        cursor.execute("""
            UPDATE pcpm_equipamentos
            SET tipo_equipamento_id = %s,
                centro_custo_id = %s,
                codigo_frota = %s,
                marca = %s,
                modelo = %s
            WHERE id = %s
        """, (
            tipo_equipamento_id,
            centro_custo_id,
            codigo_frota,
            marca,
            modelo,
            id
        ))

        conn.commit()

        flash('Equipamento atualizado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao atualizar equipamento: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.equipamentos'))


@main_routes.route('/inativar_equipamento/<int:id>', methods=['POST'])
@login_required
def inativar_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_equipamentos
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Equipamento inativado com sucesso!', 'success')

    return redirect(url_for('main.equipamentos'))


@main_routes.route('/ativar_equipamento/<int:id>', methods=['POST'])
@login_required
def ativar_equipamento(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_equipamentos
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Equipamento ativado com sucesso!', 'success')

    return redirect(url_for('main.equipamentos'))

# =============================== #
# CLIENTES / OPERADORES           #
# =============================== #

@main_routes.route('/pcpm_pessoas', methods=['GET'])
@login_required
def pcpm_pessoas():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.pcpm_pessoas'))

    filtros = {
        'matricula': (request.args.get('matricula') or '').strip(),
        'nome': (request.args.get('nome') or '').strip(),
        'empresa_id': request.args.get('empresa_id') or '',
        'setor_area': (request.args.get('setor_area') or '').strip()
    }

    cursor.execute("""
        SELECT
            id,
            nome
        FROM pcpm_empresas
        WHERE ativo = 1
        ORDER BY nome ASC
    """)
    empresas = cursor.fetchall()

    where = []
    params = []

    if filtros['matricula']:
        where.append("p.matricula LIKE %s")
        params.append(f"%{filtros['matricula']}%")

    if filtros['nome']:
        where.append("p.nome LIKE %s")
        params.append(f"%{filtros['nome']}%")

    if filtros['empresa_id']:
        where.append("p.empresa_id = %s")
        params.append(filtros['empresa_id'])

    if filtros['setor_area']:
        where.append("p.setor_area LIKE %s")
        params.append(f"%{filtros['setor_area']}%")

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    cursor.execute(f"""
        SELECT
            p.id,
            p.rfid,
            p.matricula,
            p.nome,
            p.empresa_id,
            emp.nome AS empresa_nome,
            p.setor_area,
            p.ativo
        FROM pcpm_pessoas p
        LEFT JOIN pcpm_empresas emp
            ON emp.id = p.empresa_id
        {where_sql}
        ORDER BY p.nome ASC
    """, params)

    pessoas = cursor.fetchall()

    conn.close()

    return render_template(
        'pcpm_pessoas.html',
        pessoas=pessoas,
        empresas=empresas,
        filtros=filtros
    )


@main_routes.route('/cadastrar_pcpm_pessoa', methods=['POST'])
@login_required
def cadastrar_pcpm_pessoa():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    rfid = (request.form.get('rfid') or '').strip()
    matricula = (request.form.get('matricula') or '').strip()
    nome = (request.form.get('nome') or '').strip()
    empresa_id = request.form.get('empresa_id') or None
    setor_area = (request.form.get('setor_area') or '').strip()

    if not rfid:
        flash('Informe o RFID do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not matricula:
        flash('Informe a matrícula do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not nome:
        flash('Informe o nome do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not empresa_id:
        flash('Selecione a empresa do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not setor_area:
        flash('Informe o setor/área do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT id
            FROM pcpm_empresas
            WHERE id = %s
              AND ativo = 1
        """, (empresa_id,))
        empresa = cursor.fetchone()

        if not empresa:
            flash('Empresa inválida ou inativa.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_pessoas'))

        cursor.execute("""
            SELECT id
            FROM pcpm_pessoas
            WHERE rfid = %s
               OR matricula = %s
        """, (rfid, matricula))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe um cliente/operador cadastrado com este RFID ou matrícula.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_pessoas'))

        cursor.execute("""
            INSERT INTO pcpm_pessoas (
                rfid,
                matricula,
                nome,
                empresa_id,
                setor_area,
                ativo
            )
            VALUES (%s, %s, %s, %s, %s, 1)
        """, (
            rfid,
            matricula,
            nome,
            empresa_id,
            setor_area
        ))

        conn.commit()
        flash('Cliente/operador cadastrado com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao cadastrar cliente/operador: {e}', 'danger')

    finally:
        conn.close()

    return redirect(url_for('main.pcpm_pessoas'))


@main_routes.route('/editar_pcpm_pessoa/<int:id>', methods=['POST'])
@login_required
def editar_pcpm_pessoa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    rfid = (request.form.get('rfid') or '').strip()
    matricula = (request.form.get('matricula') or '').strip()
    nome = (request.form.get('nome') or '').strip()
    empresa_id = request.form.get('empresa_id') or None
    setor_area = (request.form.get('setor_area') or '').strip()

    if not rfid:
        flash('Informe o RFID do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not matricula:
        flash('Informe a matrícula do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not nome:
        flash('Informe o nome do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not empresa_id:
        flash('Selecione a empresa do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    if not setor_area:
        flash('Informe o setor/área do cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_pessoas'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT id
            FROM pcpm_empresas
            WHERE id = %s
              AND ativo = 1
        """, (empresa_id,))
        empresa = cursor.fetchone()

        if not empresa:
            flash('Empresa inválida ou inativa.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_pessoas'))

        cursor.execute("""
            SELECT id
            FROM pcpm_pessoas
            WHERE (rfid = %s OR matricula = %s)
              AND id <> %s
        """, (rfid, matricula, id))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe outro cliente/operador cadastrado com este RFID ou matrícula.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_pessoas'))

        cursor.execute("""
            UPDATE pcpm_pessoas
            SET rfid = %s,
                matricula = %s,
                nome = %s,
                empresa_id = %s,
                setor_area = %s
            WHERE id = %s
        """, (
            rfid,
            matricula,
            nome,
            empresa_id,
            setor_area,
            id
        ))

        conn.commit()
        flash('Cliente/operador atualizado com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar cliente/operador: {e}', 'danger')

    finally:
        conn.close()

    return redirect(url_for('main.pcpm_pessoas'))


@main_routes.route('/inativar_pcpm_pessoa/<int:id>', methods=['POST'])
@login_required
def inativar_pcpm_pessoa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_pessoas
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Cliente/operador inativado com sucesso!', 'success')
    return redirect(url_for('main.pcpm_pessoas'))


@main_routes.route('/ativar_pcpm_pessoa/<int:id>', methods=['POST'])
@login_required
def ativar_pcpm_pessoa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_pessoas
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Cliente/operador ativado com sucesso!', 'success')
    return redirect(url_for('main.pcpm_pessoas'))

# ==========================================================
# PCP-M - EMPRESAS
# ==========================================================

@main_routes.route('/pcpm_empresas', methods=['GET'])
@login_required
def pcpm_empresas():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')

    colunas_validas = {
        'id': 'id',
        'nome': 'nome'
    }

    if sort not in colunas_validas:
        sort = 'nome'

    if order not in ['asc', 'desc']:
        order = 'asc'

    cursor.execute(f"""
        SELECT
            id,
            nome,
            ativo
        FROM pcpm_empresas
        ORDER BY {colunas_validas[sort]} {order.upper()}
    """)

    empresas = cursor.fetchall()

    conn.close()

    return render_template(
        'pcpm_empresas.html',
        empresas=empresas
    )


# ==========================================================
# CADASTRAR EMPRESA
# ==========================================================

@main_routes.route('/cadastrar_pcpm_empresa', methods=['POST'])
@login_required
def cadastrar_pcpm_empresa():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    nome = (request.form.get('nome') or '').strip()

    if not nome:
        flash('Informe o nome da empresa.', 'warning')
        return redirect(url_for('main.pcpm_empresas'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_empresas
            WHERE UPPER(nome) = UPPER(%s)
        """, (nome,))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe uma empresa cadastrada com este nome.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_empresas'))

        cursor.execute("""
            INSERT INTO pcpm_empresas (
                nome,
                ativo
            )
            VALUES (%s, 1)
        """, (nome,))

        conn.commit()

        flash('Empresa cadastrada com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao cadastrar empresa: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_empresas'))


# ==========================================================
# EDITAR EMPRESA
# ==========================================================

@main_routes.route('/editar_pcpm_empresa/<int:id>', methods=['POST'])
@login_required
def editar_pcpm_empresa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    nome = (request.form.get('nome') or '').strip()

    if not nome:
        flash('Informe o nome da empresa.', 'warning')
        return redirect(url_for('main.pcpm_empresas'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_empresas
            WHERE UPPER(nome) = UPPER(%s)
              AND id <> %s
        """, (nome, id))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe outra empresa cadastrada com este nome.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_empresas'))

        cursor.execute("""
            UPDATE pcpm_empresas
            SET nome = %s
            WHERE id = %s
        """, (
            nome,
            id
        ))

        conn.commit()

        flash('Empresa atualizada com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao atualizar empresa: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_empresas'))


# ==========================================================
# INATIVAR EMPRESA
# ==========================================================

@main_routes.route('/inativar_pcpm_empresa/<int:id>', methods=['POST'])
@login_required
def inativar_pcpm_empresa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_empresas
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Empresa inativada com sucesso!', 'success')

    return redirect(url_for('main.pcpm_empresas'))


# ==========================================================
# ATIVAR EMPRESA
# ==========================================================

@main_routes.route('/ativar_pcpm_empresa/<int:id>', methods=['POST'])
@login_required
def ativar_pcpm_empresa(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_empresas
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Empresa ativada com sucesso!', 'success')

    return redirect(url_for('main.pcpm_empresas'))

# ==========================================================
# API PCP-M - LISTAR EMPRESAS
# ==========================================================

@main_routes.route('/api/pcpm/empresas', methods=['GET'])
@login_required
def api_pcpm_empresas():

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            id,
            nome
        FROM pcpm_empresas
        WHERE ativo = 1
        ORDER BY nome
    """)

    empresas = cursor.fetchall()

    conn.close()

    return jsonify({
        'sucesso': True,
        'empresas': empresas
    })

# ==========================================================
# PCP-M - MODELOS DE CHECKLIST
# ==========================================================

@main_routes.route('/pcpm_checklist_modelos', methods=['GET'])
@login_required
def pcpm_checklist_modelos():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.pcpm_checklist_modelos'))

    filtros = {
        'tipo_equipamento_id': request.args.get('tipo_equipamento_id') or '',
        'nome': (request.args.get('nome') or '').strip()
    }

    cursor.execute("""
        SELECT
            id,
            tag,
            nome
        FROM pcpm_tipos_equipamento
        WHERE ativo = 1
        ORDER BY tag, nome
    """)
    tipos_equipamento = cursor.fetchall()

    where = []
    params = []

    if filtros['tipo_equipamento_id']:
        where.append("m.tipo_equipamento_id = %s")
        params.append(filtros['tipo_equipamento_id'])

    if filtros['nome']:
        where.append("m.nome LIKE %s")
        params.append(f"%{filtros['nome']}%")

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    cursor.execute(f"""
        SELECT
            m.id,
            m.tipo_equipamento_id,
            m.nome,
            m.ativo,
            te.tag AS tipo_tag,
            te.nome AS tipo_nome
        FROM pcpm_checklist_modelos m
        JOIN pcpm_tipos_equipamento te
            ON te.id = m.tipo_equipamento_id
        {where_sql}
        ORDER BY te.nome ASC, m.nome ASC
    """, params)

    modelos = cursor.fetchall()

    conn.close()

    return render_template(
        'pcpm_checklist_modelos.html',
        modelos=modelos,
        tipos_equipamento=tipos_equipamento,
        filtros=filtros
    )


# ==========================================================
# CADASTRAR MODELO CHECKLIST
# ==========================================================

@main_routes.route('/cadastrar_pcpm_checklist_modelo', methods=['POST'])
@login_required
def cadastrar_pcpm_checklist_modelo():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tipo_equipamento_id = request.form.get('tipo_equipamento_id')
    nome = (request.form.get('nome') or '').strip()

    if not tipo_equipamento_id:
        flash('Selecione o tipo de equipamento.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    if not nome:
        flash('Informe o nome do modelo.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_tipos_equipamento
            WHERE id = %s
              AND ativo = 1
        """, (tipo_equipamento_id,))

        tipo = cursor.fetchone()

        if not tipo:
            flash('Tipo de equipamento inválido.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        cursor.execute("""
            SELECT id
            FROM pcpm_checklist_modelos
            WHERE tipo_equipamento_id = %s
              AND UPPER(nome) = UPPER(%s)
        """, (tipo_equipamento_id, nome))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe um modelo com este nome para o tipo de equipamento selecionado.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        cursor.execute("""
            INSERT INTO pcpm_checklist_modelos (
                tipo_equipamento_id,
                nome,
                ativo
            )
            VALUES (%s, %s, 1)
        """, (
            tipo_equipamento_id,
            nome
        ))

        conn.commit()

        flash('Modelo de checklist cadastrado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao cadastrar modelo de checklist: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_checklist_modelos'))


# ==========================================================
# EDITAR MODELO CHECKLIST
# ==========================================================

@main_routes.route('/editar_pcpm_checklist_modelo/<int:id>', methods=['POST'])
@login_required
def editar_pcpm_checklist_modelo(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    tipo_equipamento_id = request.form.get('tipo_equipamento_id')
    nome = (request.form.get('nome') or '').strip()

    if not tipo_equipamento_id:
        flash('Selecione o tipo de equipamento.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    if not nome:
        flash('Informe o nome do modelo.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_tipos_equipamento
            WHERE id = %s
              AND ativo = 1
        """, (tipo_equipamento_id,))

        tipo = cursor.fetchone()

        if not tipo:
            flash('Tipo de equipamento inválido.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        cursor.execute("""
            SELECT id
            FROM pcpm_checklist_modelos
            WHERE tipo_equipamento_id = %s
              AND UPPER(nome) = UPPER(%s)
              AND id <> %s
        """, (
            tipo_equipamento_id,
            nome,
            id
        ))

        existente = cursor.fetchone()

        if existente:
            flash('Já existe outro modelo com este nome para o tipo de equipamento selecionado.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        cursor.execute("""
            UPDATE pcpm_checklist_modelos
            SET tipo_equipamento_id = %s,
                nome = %s
            WHERE id = %s
        """, (
            tipo_equipamento_id,
            nome,
            id
        ))

        conn.commit()

        flash('Modelo de checklist atualizado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao atualizar modelo de checklist: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_checklist_modelos'))


# ==========================================================
# INATIVAR MODELO CHECKLIST
# ==========================================================

@main_routes.route('/inativar_pcpm_checklist_modelo/<int:id>', methods=['POST'])
@login_required
def inativar_pcpm_checklist_modelo(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_checklist_modelos
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Modelo de checklist inativado com sucesso!', 'success')

    return redirect(url_for('main.pcpm_checklist_modelos'))


# ==========================================================
# ATIVAR MODELO CHECKLIST
# ==========================================================

@main_routes.route('/ativar_pcpm_checklist_modelo/<int:id>', methods=['POST'])
@login_required
def ativar_pcpm_checklist_modelo(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE pcpm_checklist_modelos
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Modelo de checklist ativado com sucesso!', 'success')

    return redirect(url_for('main.pcpm_checklist_modelos'))

# ==========================================================
# PCP-M - ITENS DO CHECKLIST
# ==========================================================

@main_routes.route('/pcpm_checklist_itens/<int:modelo_id>', methods=['GET'])
@login_required
def pcpm_checklist_itens(modelo_id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            m.id,
            m.nome,
            m.ativo,
            te.tag AS tipo_tag,
            te.nome AS tipo_nome
        FROM pcpm_checklist_modelos m
        JOIN pcpm_tipos_equipamento te
            ON te.id = m.tipo_equipamento_id
        WHERE m.id = %s
    """, (modelo_id,))

    modelo = cursor.fetchone()

    if not modelo:
        conn.close()
        flash('Modelo de checklist não encontrado.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    cursor.execute("""
        SELECT
            id,
            modelo_id,
            ordem,
            item,
            criterio,
            exige_foto_nok,
            exige_observacao_nok,
            ativo
        FROM pcpm_checklist_itens
        WHERE modelo_id = %s
        ORDER BY ordem ASC
    """, (modelo_id,))

    itens = cursor.fetchall()

    conn.close()

    return render_template(
        'pcpm_checklist_itens.html',
        modelo=modelo,
        itens=itens
    )


# ==========================================================
# CADASTRAR ITEM CHECKLIST
# ==========================================================

@main_routes.route('/cadastrar_pcpm_checklist_item/<int:modelo_id>', methods=['POST'])
@login_required
def cadastrar_pcpm_checklist_item(modelo_id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    ordem = request.form.get('ordem')
    item = (request.form.get('item') or '').strip()
    criterio = (request.form.get('criterio') or '').strip()

    exige_foto_nok = request.form.get('exige_foto_nok', 1)
    exige_observacao_nok = request.form.get('exige_observacao_nok', 1)

    if not ordem:
        flash('Informe a ordem do item.', 'warning')
        return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

    if not item:
        flash('Informe o item do checklist.', 'warning')
        return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

    if not criterio:
        flash('Informe o critério/verificação.', 'warning')
        return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_checklist_modelos
            WHERE id = %s
        """, (modelo_id,))

        modelo = cursor.fetchone()

        if not modelo:
            flash('Modelo de checklist inválido.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        cursor.execute("""
            SELECT id
            FROM pcpm_checklist_itens
            WHERE modelo_id = %s
              AND ordem = %s
        """, (modelo_id, ordem))

        ordem_existente = cursor.fetchone()

        if ordem_existente:
            flash('Já existe um item cadastrado com esta ordem.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

        cursor.execute("""
            INSERT INTO pcpm_checklist_itens (
                modelo_id,
                ordem,
                item,
                criterio,
                exige_foto_nok,
                exige_observacao_nok,
                ativo
            )
            VALUES (%s, %s, %s, %s, %s, %s, 1)
        """, (
            modelo_id,
            ordem,
            item,
            criterio,
            exige_foto_nok,
            exige_observacao_nok
        ))

        conn.commit()

        flash('Item do checklist cadastrado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao cadastrar item do checklist: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))


# ==========================================================
# EDITAR ITEM CHECKLIST
# ==========================================================

@main_routes.route('/editar_pcpm_checklist_item/<int:id>', methods=['POST'])
@login_required
def editar_pcpm_checklist_item(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    ordem = request.form.get('ordem')
    item = (request.form.get('item') or '').strip()
    criterio = (request.form.get('criterio') or '').strip()

    exige_foto_nok = request.form.get('exige_foto_nok', 1)
    exige_observacao_nok = request.form.get('exige_observacao_nok', 1)

    if not ordem:
        flash('Informe a ordem do item.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    if not item:
        flash('Informe o item do checklist.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    if not criterio:
        flash('Informe o critério/verificação.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT
                id,
                modelo_id
            FROM pcpm_checklist_itens
            WHERE id = %s
        """, (id,))

        item_atual = cursor.fetchone()

        if not item_atual:
            flash('Item do checklist não encontrado.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_modelos'))

        modelo_id = item_atual['modelo_id']

        cursor.execute("""
            SELECT id
            FROM pcpm_checklist_itens
            WHERE modelo_id = %s
              AND ordem = %s
              AND id <> %s
        """, (
            modelo_id,
            ordem,
            id
        ))

        ordem_existente = cursor.fetchone()

        if ordem_existente:
            flash('Já existe outro item cadastrado com esta ordem.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

        cursor.execute("""
            UPDATE pcpm_checklist_itens
            SET ordem = %s,
                item = %s,
                criterio = %s,
                exige_foto_nok = %s,
                exige_observacao_nok = %s
            WHERE id = %s
        """, (
            ordem,
            item,
            criterio,
            exige_foto_nok,
            exige_observacao_nok,
            id
        ))

        conn.commit()

        flash('Item do checklist atualizado com sucesso!', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao atualizar item do checklist: {e}', 'danger')

    finally:

        conn.close()

    return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))


# ==========================================================
# INATIVAR ITEM CHECKLIST
# ==========================================================

@main_routes.route('/inativar_pcpm_checklist_item/<int:id>', methods=['POST'])
@login_required
def inativar_pcpm_checklist_item(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT modelo_id
        FROM pcpm_checklist_itens
        WHERE id = %s
    """, (id,))

    item = cursor.fetchone()

    if not item:
        conn.close()
        flash('Item do checklist não encontrado.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    modelo_id = item['modelo_id']

    cursor.execute("""
        UPDATE pcpm_checklist_itens
        SET ativo = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Item do checklist inativado com sucesso!', 'success')

    return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))


# ==========================================================
# ATIVAR ITEM CHECKLIST
# ==========================================================

@main_routes.route('/ativar_pcpm_checklist_item/<int:id>', methods=['POST'])
@login_required
def ativar_pcpm_checklist_item(id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT modelo_id
        FROM pcpm_checklist_itens
        WHERE id = %s
    """, (id,))

    item = cursor.fetchone()

    if not item:
        conn.close()
        flash('Item do checklist não encontrado.', 'warning')
        return redirect(url_for('main.pcpm_checklist_modelos'))

    modelo_id = item['modelo_id']

    cursor.execute("""
        UPDATE pcpm_checklist_itens
        SET ativo = 1
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash('Item do checklist ativado com sucesso!', 'success')

    return redirect(url_for('main.pcpm_checklist_itens', modelo_id=modelo_id))

# ==========================================================
# PCP-M - MOVIMENTAÇÃO DE MÁQUINAS
# ==========================================================

@main_routes.route('/pcpm_movimentacao', methods=['GET'])
@login_required
def pcpm_movimentacao():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    from datetime import datetime

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    equipamento_id = request.args.get('equipamento_id') or None

    cursor.execute("""
        SELECT
            e.id,
            e.codigo_frota,
            e.marca,
            e.modelo,
            e.status_localizacao,
            te.nome AS tipo_nome,
            te.tag AS tipo_tag
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        WHERE e.ativo = 1
        ORDER BY e.codigo_frota
    """)
    equipamentos = cursor.fetchall()

    cursor.execute("""
        SELECT
            id,
            matricula,
            nome
        FROM usuarios
        WHERE ativo = 1
        ORDER BY nome
    """)
    funcionarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            p.id,
            p.matricula,
            p.nome,
            p.rfid,
            emp.nome AS empresa_nome
        FROM pcpm_pessoas p
        LEFT JOIN pcpm_empresas emp
            ON emp.id = p.empresa_id
        WHERE p.ativo = 1
        ORDER BY p.nome
    """)
    pessoas = cursor.fetchall()

    equipamento_selecionado = None
    modelo_checklist = None
    itens_checklist = []
    tipo_movimentacao_permitido = None

    if equipamento_id:
        cursor.execute("""
            SELECT
                e.id,
                e.codigo_frota,
                e.marca,
                e.modelo,
                e.tipo_equipamento_id,
                e.status_localizacao,
                te.nome AS tipo_nome,
                te.tag AS tipo_tag
            FROM pcpm_equipamentos e
            JOIN pcpm_tipos_equipamento te
                ON te.id = e.tipo_equipamento_id
            WHERE e.id = %s
              AND e.ativo = 1
        """, (equipamento_id,))
        equipamento_selecionado = cursor.fetchone()

        if equipamento_selecionado:
            status = equipamento_selecionado.get('status_localizacao') or 'indefinido'

            if status == 'area':
                tipo_movimentacao_permitido = 'retirada'
            elif status == 'tradimaq':
                tipo_movimentacao_permitido = 'entrega'
            else:
                tipo_movimentacao_permitido = None

            cursor.execute("""
                SELECT
                    id,
                    nome
                FROM pcpm_checklist_modelos
                WHERE tipo_equipamento_id = %s
                  AND ativo = 1
                ORDER BY id DESC
                LIMIT 1
            """, (equipamento_selecionado['tipo_equipamento_id'],))
            modelo_checklist = cursor.fetchone()

            if modelo_checklist:
                cursor.execute("""
                    SELECT
                        id,
                        ordem,
                        item,
                        criterio,
                        exige_foto_nok,
                        exige_observacao_nok
                    FROM pcpm_checklist_itens
                    WHERE modelo_id = %s
                      AND ativo = 1
                    ORDER BY ordem
                """, (modelo_checklist['id'],))
                itens_checklist = cursor.fetchall()

    conn.close()

    return render_template(
        'pcpm_movimentacao.html',
        equipamentos=equipamentos,
        funcionarios=funcionarios,
        pessoas=pessoas,
        equipamento_selecionado=equipamento_selecionado,
        modelo_checklist=modelo_checklist,
        itens_checklist=itens_checklist,
        tipo_movimentacao_permitido=tipo_movimentacao_permitido,
        data_atual=datetime.now().strftime('%d/%m/%Y')
    )


@main_routes.route('/salvar_pcpm_movimentacao', methods=['POST'])
@login_required
def salvar_pcpm_movimentacao():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    import os
    from werkzeug.utils import secure_filename
    from datetime import datetime

    equipamento_id = request.form.get('equipamento_id')
    funcionario_id = request.form.get('funcionario_id')
    cliente_id = request.form.get('cliente_id')

    tipo_movimentacao = request.form.get('tipo_movimentacao')
    motivo_movimentacao = request.form.get('motivo_movimentacao')
    numero_os = (request.form.get('numero_os') or '').strip() or None
    horimetro = request.form.get('horimetro')
    descricao_anomalia = (request.form.get('descricao_anomalia') or '').strip()

    latitude = request.form.get('latitude')
    longitude = request.form.get('longitude')

    checklist_item_ids = request.form.getlist('checklist_item_ids[]')

    if not equipamento_id:
        flash('Selecione o equipamento.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao'))

    if not funcionario_id:
        flash('Selecione o funcionário Tradimaq.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not cliente_id:
        flash('Selecione o cliente/operador.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not tipo_movimentacao:
        flash('Selecione o tipo de movimentação.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if tipo_movimentacao not in ['retirada', 'entrega']:
        flash('Tipo de movimentação inválido.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not motivo_movimentacao:
        flash('Selecione o motivo da movimentação.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not horimetro:
        flash('Informe o horímetro.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not descricao_anomalia:
        flash('Informe o relato/condições encontradas.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not latitude or not longitude:
        flash('Não foi possível capturar a geolocalização.', 'danger')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    if not checklist_item_ids:
        flash('Não há itens de checklist cadastrados para este equipamento.', 'warning')
        return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT
                id,
                status_localizacao
            FROM pcpm_equipamentos
            WHERE id = %s
              AND ativo = 1
        """, (equipamento_id,))
        equipamento = cursor.fetchone()

        if not equipamento:
            flash('Equipamento inválido ou inativo.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_movimentacao'))

        status_atual = equipamento.get('status_localizacao') or 'indefinido'

        if status_atual == 'area' and tipo_movimentacao != 'retirada':
            flash('Este equipamento está registrado como na área. A próxima movimentação deve ser retirada.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

        if status_atual == 'tradimaq' and tipo_movimentacao != 'entrega':
            flash('Este equipamento está registrado como em posse da Tradimaq. A próxima movimentação deve ser entrega.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

        cursor.execute("""
            SELECT
                id,
                uid_rfid
            FROM usuarios
            WHERE id = %s
              AND ativo = 1
        """, (funcionario_id,))
        funcionario = cursor.fetchone()

        if not funcionario:
            flash('Funcionário Tradimaq inválido ou inativo.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

        cursor.execute("""
            SELECT
                id,
                rfid
            FROM pcpm_pessoas
            WHERE id = %s
              AND ativo = 1
        """, (cliente_id,))
        cliente = cursor.fetchone()

        if not cliente:
            flash('Cliente/operador inválido ou inativo.', 'warning')
            conn.close()
            return redirect(url_for('main.pcpm_movimentacao', equipamento_id=equipamento_id))

        rfid_funcionario = funcionario.get('uid_rfid') or ''
        rfid_cliente = cliente.get('rfid') or ''

        cursor.execute("""
            INSERT INTO pcpm_movimentacoes (
                equipamento_id,
                cliente_id,
                funcionario_id,
                tipo_movimentacao,
                motivo_movimentacao,
                numero_os,
                horimetro,
                descricao_anomalia,
                rfid_funcionario,
                rfid_cliente,
                latitude,
                longitude,
                criado_por,
                ativo
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
        """, (
            equipamento_id,
            cliente_id,
            funcionario_id,
            tipo_movimentacao,
            motivo_movimentacao,
            numero_os,
            horimetro,
            descricao_anomalia,
            rfid_funcionario,
            rfid_cliente,
            latitude,
            longitude,
            session.get('usuario_id')
        ))

        movimentacao_id = cursor.lastrowid

        pasta_upload = os.path.join(
            current_app.root_path,
            'static',
            'pcpm_movimentacoes',
            str(movimentacao_id)
        )

        os.makedirs(pasta_upload, exist_ok=True)

        for item_id in checklist_item_ids:
            resultado = request.form.get(f'resultado_{item_id}')
            observacao = (request.form.get(f'observacao_{item_id}') or '').strip()

            if not resultado:
                raise Exception('Todos os itens do checklist devem ser respondidos.')

            cursor.execute("""
                SELECT
                    id,
                    ordem,
                    item,
                    criterio,
                    exige_foto_nok,
                    exige_observacao_nok
                FROM pcpm_checklist_itens
                WHERE id = %s
            """, (item_id,))
            item_checklist = cursor.fetchone()

            if not item_checklist:
                raise Exception('Item de checklist inválido.')

            arquivo_foto = request.files.get(f'foto_{item_id}')

            if resultado == 'NOK':
                if item_checklist['exige_observacao_nok'] and not observacao:
                    raise Exception(f"O item '{item_checklist['item']}' exige observação quando NOK.")

                if item_checklist['exige_foto_nok'] and (not arquivo_foto or arquivo_foto.filename == ''):
                    raise Exception(f"O item '{item_checklist['item']}' exige foto quando NOK.")

            cursor.execute("""
                INSERT INTO pcpm_movimentacao_checklist_respostas (
                    movimentacao_id,
                    checklist_item_id,
                    resultado,
                    observacao,
                    item_snapshot,
                    criterio_snapshot,
                    ordem_snapshot
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                movimentacao_id,
                item_id,
                resultado,
                observacao,
                item_checklist['item'],
                item_checklist['criterio'],
                item_checklist['ordem']
            ))

            resposta_id = cursor.lastrowid

            if arquivo_foto and arquivo_foto.filename:
                nome_arquivo = secure_filename(arquivo_foto.filename)
                nome_final = f"checklist_{resposta_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{nome_arquivo}"

                caminho_fisico = os.path.join(pasta_upload, nome_final)
                arquivo_foto.save(caminho_fisico)

                caminho_relativo = f"pcpm_movimentacoes/{movimentacao_id}/{nome_final}"

                cursor.execute("""
                    INSERT INTO pcpm_movimentacao_fotos (
                        movimentacao_id,
                        checklist_resposta_id,
                        tipo_foto,
                        caminho_arquivo
                    )
                    VALUES (%s, %s, 'checklist', %s)
                """, (
                    movimentacao_id,
                    resposta_id,
                    caminho_relativo
                ))

        novo_status = 'tradimaq' if tipo_movimentacao == 'retirada' else 'area'

        cursor.execute("""
            UPDATE pcpm_equipamentos
            SET status_localizacao = %s
            WHERE id = %s
        """, (
            novo_status,
            equipamento_id
        ))

        conn.commit()
        flash('Movimentação registrada com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao registrar movimentação: {e}', 'danger')

    finally:
        conn.close()

    return redirect(url_for('main.pcpm_movimentacao'))

# ==========================================================
# PCP-M - PAINEL DE MOVIMENTAÇÕES
# ==========================================================

@main_routes.route('/pcpm_painel_movimentacoes', methods=['GET'])
@login_required
def pcpm_painel_movimentacoes():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    filtros = {
        'codigo_frota': (request.args.get('codigo_frota') or '').strip().upper(),
        'tipo_equipamento_id': request.args.get('tipo_equipamento_id') or '',
        'status_localizacao': request.args.get('status_localizacao') or '',
        'sort': request.args.get('sort', 'codigo_frota'),
        'order': request.args.get('order', 'asc')
    }

    sort = filtros['sort']
    order = filtros['order']

    colunas_validas = {
        'codigo_frota': 'e.codigo_frota',
        'tipo_nome': 'te.nome',
        'status_localizacao': 'e.status_localizacao',
        'ultima_movimentacao_data': 'mov.criado_em'
    }

    if sort not in colunas_validas:
        sort = 'codigo_frota'

    if order not in ['asc', 'desc']:
        order = 'asc'

    cursor.execute("""
        SELECT
            id,
            tag,
            nome
        FROM pcpm_tipos_equipamento
        WHERE ativo = 1
        ORDER BY tag, nome
    """)
    tipos_equipamento = cursor.fetchall()

    where = ["e.ativo = 1"]
    params = []

    if filtros['codigo_frota']:
        where.append("e.codigo_frota LIKE %s")
        params.append(f"%{filtros['codigo_frota']}%")

    if filtros['tipo_equipamento_id']:
        where.append("e.tipo_equipamento_id = %s")
        params.append(filtros['tipo_equipamento_id'])

    if filtros['status_localizacao']:
        where.append("e.status_localizacao = %s")
        params.append(filtros['status_localizacao'])

    where_sql = "WHERE " + " AND ".join(where)

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        {where_sql}
    """, params)

    total_registros = cursor.fetchone()['total']
    total_paginas = (total_registros + per_page - 1) // per_page

    params_paginacao = params + [per_page, offset]

    cursor.execute(f"""
        SELECT
            e.id,
            e.codigo_frota,
            e.marca,
            e.modelo,
            e.status_localizacao,

            te.nome AS tipo_nome,
            te.tag AS tipo_tag,

            mov.id AS ultima_movimentacao_id,
            mov.tipo_movimentacao,
            mov.motivo_movimentacao,

            DATE_FORMAT(
                mov.criado_em,
                '%d/%m/%Y %H:%i'
            ) AS ultima_movimentacao_data,

            u.nome AS funcionario_nome,
            p.nome AS cliente_nome

        FROM pcpm_equipamentos e

        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id

        LEFT JOIN pcpm_movimentacoes mov
            ON mov.id = (
                SELECT m2.id
                FROM pcpm_movimentacoes m2
                WHERE m2.equipamento_id = e.id
                  AND m2.ativo = 1
                ORDER BY m2.id DESC
                LIMIT 1
            )

        LEFT JOIN usuarios u
            ON u.id = mov.funcionario_id

        LEFT JOIN pcpm_pessoas p
            ON p.id = mov.cliente_id

        {where_sql}

        ORDER BY
            {colunas_validas[sort]} {order.upper()},
            e.codigo_frota ASC

        LIMIT %s OFFSET %s
    """, params_paginacao)

    equipamentos = cursor.fetchall()

    for equipamento in equipamentos:
        status = equipamento.get('status_localizacao') or 'indefinido'

        if status == 'area':
            equipamento['status_descricao'] = 'Na área'
            equipamento['proxima_acao'] = 'Retirada da área'
            equipamento['tipo_movimentacao_permitido'] = 'retirada'

        elif status == 'tradimaq':
            equipamento['status_descricao'] = 'Em posse da Tradimaq'
            equipamento['proxima_acao'] = 'Entrega na área'
            equipamento['tipo_movimentacao_permitido'] = 'entrega'

        else:
            equipamento['status_descricao'] = 'Sem histórico'
            equipamento['proxima_acao'] = 'Primeira movimentação'
            equipamento['tipo_movimentacao_permitido'] = None

    conn.close()

    return render_template(
        'pcpm_painel_movimentacoes.html',
        equipamentos=equipamentos,
        tipos_equipamento=tipos_equipamento,
        filtros=filtros,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros
    )

# ==========================================================
# PCP-M - VISUALIZAR MOVIMENTAÇÃO
# ==========================================================

@main_routes.route('/visualizar_pcpm_movimentacao/<int:movimentacao_id>', methods=['GET'])
@login_required
def visualizar_pcpm_movimentacao(movimentacao_id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            mov.id,
            mov.equipamento_id,
            mov.cliente_id,
            mov.funcionario_id,
            mov.tipo_movimentacao,
            mov.motivo_movimentacao,
            mov.numero_os,
            mov.horimetro,
            mov.descricao_anomalia,
            mov.rfid_funcionario,
            mov.rfid_cliente,
            mov.latitude,
            mov.longitude,
            DATE_FORMAT(mov.criado_em, '%d/%m/%Y %H:%i') AS criado_em_formatado,

            e.codigo_frota,
            e.marca,
            e.modelo,
            te.nome AS tipo_equipamento,
            te.tag AS tipo_tag,

            func.nome AS funcionario_nome,
            func.matricula AS funcionario_matricula,

            cli.nome AS cliente_nome,
            cli.matricula AS cliente_matricula,
            cli.setor_area AS cliente_setor_area,
            emp.nome AS cliente_empresa,

            criador.nome AS criado_por_nome

        FROM pcpm_movimentacoes mov

        JOIN pcpm_equipamentos e
            ON e.id = mov.equipamento_id

        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id

        LEFT JOIN usuarios func
            ON func.id = mov.funcionario_id

        LEFT JOIN pcpm_pessoas cli
            ON cli.id = mov.cliente_id

        LEFT JOIN pcpm_empresas emp
            ON emp.id = cli.empresa_id

        LEFT JOIN usuarios criador
            ON criador.id = mov.criado_por

        WHERE mov.id = %s
          AND mov.ativo = 1
    """, (movimentacao_id,))

    movimentacao = cursor.fetchone()

    if not movimentacao:
        conn.close()
        flash('Movimentação não encontrada.', 'warning')
        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    cursor.execute("""
        SELECT
            r.id,
            r.checklist_item_id,
            r.resultado,
            r.observacao,
            r.foto_path,
            r.item_snapshot,
            r.criterio_snapshot,
            r.ordem_snapshot

        FROM pcpm_movimentacao_checklist_respostas r

        WHERE r.movimentacao_id = %s

        ORDER BY r.ordem_snapshot ASC
    """, (movimentacao_id,))

    respostas = cursor.fetchall()

    conn.close()

    return render_template(
        'visualizar_pcpm_movimentacao.html',
        movimentacao=movimentacao,
        respostas=respostas,
    )

# ==========================================================
# PCP-M - EDITAR MOVIMENTAÇÃO
# ==========================================================

@main_routes.route('/editar_pcpm_movimentacao/<int:movimentacao_id>', methods=['GET', 'POST'])
@login_required
def editar_pcpm_movimentacao(movimentacao_id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ======================================================
    # BUSCAR MOVIMENTAÇÃO
    # ======================================================

    cursor.execute("""
        SELECT
            mov.*,
            e.codigo_frota
        FROM pcpm_movimentacoes mov
        JOIN pcpm_equipamentos e
            ON e.id = mov.equipamento_id
        WHERE mov.id = %s
          AND mov.ativo = 1
    """, (movimentacao_id,))

    movimentacao = cursor.fetchone()

    if not movimentacao:
        conn.close()
        flash('Movimentação não encontrada.', 'warning')
        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    # ======================================================
    # VALIDAR SE É A ÚLTIMA MOVIMENTAÇÃO
    # ======================================================

    cursor.execute("""
        SELECT id
        FROM pcpm_movimentacoes
        WHERE equipamento_id = %s
          AND ativo = 1
        ORDER BY id DESC
        LIMIT 1
    """, (movimentacao['equipamento_id'],))

    ultima_movimentacao = cursor.fetchone()

    if not ultima_movimentacao or ultima_movimentacao['id'] != movimentacao_id:
        conn.close()

        flash(
            'Somente a última movimentação do equipamento pode ser editada.',
            'warning'
        )

        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    # ======================================================
    # POST
    # ======================================================

    if request.method == 'POST':

        numero_os = request.form.get('numero_os')
        horimetro = request.form.get('horimetro')
        descricao_anomalia = request.form.get('descricao_anomalia')

        try:

            cursor.execute("""
                UPDATE pcpm_movimentacoes
                SET
                    numero_os = %s,
                    horimetro = %s,
                    descricao_anomalia = %s
                WHERE id = %s
            """, (
                numero_os,
                horimetro,
                descricao_anomalia,
                movimentacao_id
            ))

            conn.commit()

            flash('Movimentação atualizada com sucesso.', 'success')

            return redirect(
                url_for(
                    'main.visualizar_pcpm_movimentacao',
                    movimentacao_id=movimentacao_id
                )
            )

        except Exception as e:

            conn.rollback()

            flash(f'Erro ao atualizar movimentação: {e}', 'danger')

    conn.close()

    return render_template(
        'pcpm_editar_movimentacao.html',
        movimentacao=movimentacao
    )


# ==========================================================
# PCP-M - EXCLUIR MOVIMENTAÇÃO
# ==========================================================

@main_routes.route('/excluir_pcpm_movimentacao/<int:movimentacao_id>')
@login_required
def excluir_pcpm_movimentacao(movimentacao_id):

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ======================================================
    # BUSCAR MOVIMENTAÇÃO
    # ======================================================

    cursor.execute("""
        SELECT *
        FROM pcpm_movimentacoes
        WHERE id = %s
          AND ativo = 1
    """, (movimentacao_id,))

    movimentacao = cursor.fetchone()

    if not movimentacao:
        conn.close()

        flash('Movimentação não encontrada.', 'warning')

        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    equipamento_id = movimentacao['equipamento_id']

    # ======================================================
    # VALIDAR ÚLTIMA MOVIMENTAÇÃO
    # ======================================================

    cursor.execute("""
        SELECT id
        FROM pcpm_movimentacoes
        WHERE equipamento_id = %s
          AND ativo = 1
        ORDER BY id DESC
        LIMIT 1
    """, (equipamento_id,))

    ultima_movimentacao = cursor.fetchone()

    if not ultima_movimentacao or ultima_movimentacao['id'] != movimentacao_id:

        conn.close()

        flash(
            'Somente a última movimentação do equipamento pode ser excluída.',
            'warning'
        )

        return redirect(url_for('main.pcpm_painel_movimentacoes'))

    try:

        # ==================================================
        # EXCLUSÃO LÓGICA
        # ==================================================

        cursor.execute("""
            UPDATE pcpm_movimentacoes
            SET ativo = 0
            WHERE id = %s
        """, (movimentacao_id,))

        # ==================================================
        # BUSCAR NOVA ÚLTIMA MOVIMENTAÇÃO
        # ==================================================

        cursor.execute("""
            SELECT *
            FROM pcpm_movimentacoes
            WHERE equipamento_id = %s
              AND ativo = 1
            ORDER BY id DESC
            LIMIT 1
        """, (equipamento_id,))

        nova_ultima = cursor.fetchone()

        # ==================================================
        # RECALCULAR STATUS
        # ==================================================

        if nova_ultima:

            if nova_ultima['tipo_movimentacao'] == 'entrega':
                novo_status = 'area'
            else:
                novo_status = 'tradimaq'

        else:

            novo_status = 'indefinido'

        cursor.execute("""
            UPDATE pcpm_equipamentos
            SET status_localizacao = %s
            WHERE id = %s
        """, (
            novo_status,
            equipamento_id
        ))

        conn.commit()

        flash('Movimentação excluída com sucesso.', 'success')

    except Exception as e:

        conn.rollback()

        flash(f'Erro ao excluir movimentação: {e}', 'danger')

    conn.close()

    return redirect(url_for('main.pcpm_painel_movimentacoes'))

# ==========================================================
# PCP-M - LISTAR MOVIMENTAÇÕES
# ==========================================================

@main_routes.route('/listar_pcpm_movimentacoes', methods=['GET'])
@login_required
def listar_pcpm_movimentacoes():

    if not session.get('acesso_pcpm'):
        flash('Você não possui acesso ao módulo PCP-M.', 'danger')
        return redirect(url_for('main.dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.args.get('limpar'):
        conn.close()
        return redirect(url_for('main.listar_pcpm_movimentacoes'))

    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    filtros = {
        'codigo_frota': (request.args.get('codigo_frota') or '').strip().upper(),
        'tipo_movimentacao': request.args.get('tipo_movimentacao') or '',
        'data_inicio': request.args.get('data_inicio') or '',
        'data_fim': request.args.get('data_fim') or '',
        'sort': request.args.get('sort', 'criado_em'),
        'order': request.args.get('order', 'desc')
    }

    sort = filtros['sort']
    order = filtros['order']

    colunas_validas = {
        'criado_em': 'mov.criado_em',
        'codigo_frota': 'e.codigo_frota',
        'tipo_movimentacao': 'mov.tipo_movimentacao',
        'funcionario_nome': 'func.nome',
        'cliente_nome': 'cli.nome'
    }

    if sort not in colunas_validas:
        sort = 'criado_em'

    if order not in ['asc', 'desc']:
        order = 'desc'

    cursor.execute("""
        SELECT
            e.id,
            e.codigo_frota,
            e.marca,
            e.modelo,
            te.nome AS tipo_nome
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        WHERE e.ativo = 1
        ORDER BY e.codigo_frota
    """)
    equipamentos = cursor.fetchall()

    where = ["mov.ativo = 1"]
    params = []

    if filtros['codigo_frota']:
        where.append("e.codigo_frota LIKE %s")
        params.append(f"%{filtros['codigo_frota']}%")

    if filtros['tipo_movimentacao']:
        where.append("mov.tipo_movimentacao = %s")
        params.append(filtros['tipo_movimentacao'])

    if filtros['data_inicio']:
        where.append("DATE(mov.criado_em) >= %s")
        params.append(filtros['data_inicio'])

    if filtros['data_fim']:
        where.append("DATE(mov.criado_em) <= %s")
        params.append(filtros['data_fim'])

    where_sql = "WHERE " + " AND ".join(where)

    cursor.execute(f"""
        SELECT COUNT(*) AS total
        FROM pcpm_movimentacoes mov

        JOIN pcpm_equipamentos e
            ON e.id = mov.equipamento_id

        LEFT JOIN usuarios func
            ON func.id = mov.funcionario_id

        LEFT JOIN pcpm_pessoas cli
            ON cli.id = mov.cliente_id

        {where_sql}
    """, params)

    total_registros = cursor.fetchone()['total']
    total_paginas = (total_registros + per_page - 1) // per_page

    params_paginacao = params + [per_page, offset]

    cursor.execute(f"""
        SELECT
            mov.id,
            mov.tipo_movimentacao,

            DATE_FORMAT(mov.criado_em, '%d/%m/%Y %H:%i') AS data_movimentacao,

            e.codigo_frota,
            te.nome AS tipo_equipamento,

            func.nome AS funcionario_nome,
            cli.nome AS cliente_nome

        FROM pcpm_movimentacoes mov

        JOIN pcpm_equipamentos e
            ON e.id = mov.equipamento_id

        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id

        LEFT JOIN usuarios func
            ON func.id = mov.funcionario_id

        LEFT JOIN pcpm_pessoas cli
            ON cli.id = mov.cliente_id

        {where_sql}

        ORDER BY {colunas_validas[sort]} {order.upper()}, mov.id DESC

        LIMIT %s OFFSET %s
    """, params_paginacao)

    movimentacoes = cursor.fetchall()

    conn.close()

    return render_template(
        'listar_pcpm_movimentacoes.html',
        movimentacoes=movimentacoes,
        equipamentos=equipamentos,
        filtros=filtros,
        page=page,
        total_paginas=total_paginas,
        total_registros=total_registros
    )


# ===================================================================== #
# ROTAS PARA O APLICATIVO                                               #
# ===================================================================== #

# ==========================================================
# API PCP-M - TESTE
# ==========================================================

@main_routes.route('/api/pcpm/teste', methods=['GET'])
@login_required
def api_pcpm_teste():

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    return jsonify({
        'sucesso': True,
        'mensagem': 'API PCP-M funcionando.',
        'usuario_id': session.get('usuario_id'),
        'nome': session.get('nome')
    })

# ==========================================================
# API PCP-M - LISTAR EQUIPAMENTOS
# ==========================================================

@main_routes.route('/api/pcpm/equipamentos', methods=['GET'])
@login_required
def api_pcpm_equipamentos():

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            e.id,
            e.codigo_frota,
            e.marca,
            e.modelo,
            e.status_localizacao,
            te.tag AS tipo_tag,
            te.nome AS tipo_nome
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        WHERE e.ativo = 1
        ORDER BY e.codigo_frota
    """)

    equipamentos = cursor.fetchall()

    conn.close()

    return jsonify({
        'sucesso': True,
        'total': len(equipamentos),
        'equipamentos': equipamentos
    })

# ==========================================================
# API PCP-M - CHECKLIST DO EQUIPAMENTO
# ==========================================================

@main_routes.route('/api/pcpm/checklist/<int:equipamento_id>', methods=['GET'])
@login_required
def api_pcpm_checklist(equipamento_id):

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            e.id,
            e.codigo_frota,
            e.status_localizacao,
            e.marca,
            e.modelo,
            e.tipo_equipamento_id,
            te.tag AS tipo_tag,
            te.nome AS tipo_nome
        FROM pcpm_equipamentos e
        JOIN pcpm_tipos_equipamento te
            ON te.id = e.tipo_equipamento_id
        WHERE e.id = %s
          AND e.ativo = 1
    """, (equipamento_id,))

    equipamento = cursor.fetchone()

    if not equipamento:
        conn.close()
        return jsonify({
            'sucesso': False,
            'mensagem': 'Equipamento não encontrado.'
        }), 404

    status = equipamento.get('status_localizacao') or 'indefinido'

    if status == 'area':
        tipo_movimentacao_permitido = 'retirada'
    elif status == 'tradimaq':
        tipo_movimentacao_permitido = 'entrega'
    else:
        tipo_movimentacao_permitido = None

    cursor.execute("""
        SELECT
            id,
            nome
        FROM pcpm_checklist_modelos
        WHERE tipo_equipamento_id = %s
          AND ativo = 1
        ORDER BY id DESC
        LIMIT 1
    """, (equipamento['tipo_equipamento_id'],))

    modelo_checklist = cursor.fetchone()

    if not modelo_checklist:
        conn.close()
        return jsonify({
            'sucesso': False,
            'mensagem': 'Nenhum modelo de checklist ativo encontrado para este tipo de equipamento.',
            'equipamento': equipamento,
            'tipo_movimentacao_permitido': tipo_movimentacao_permitido,
            'checklist': []
        }), 404

    cursor.execute("""
        SELECT
            id,
            ordem,
            item,
            criterio,
            exige_foto_nok,
            exige_observacao_nok
        FROM pcpm_checklist_itens
        WHERE modelo_id = %s
          AND ativo = 1
        ORDER BY ordem
    """, (modelo_checklist['id'],))

    checklist = cursor.fetchall()

    conn.close()

    return jsonify({
        'sucesso': True,
        'equipamento': equipamento,
        'modelo_checklist': modelo_checklist,
        'tipo_movimentacao_permitido': tipo_movimentacao_permitido,
        'checklist': checklist
    })

# ==========================================================
# API PCP-M - SALVAR MOVIMENTAÇÃO
# ==========================================================

@main_routes.route('/api/pcpm/movimentacao', methods=['POST'])
@login_required
def api_pcpm_salvar_movimentacao():

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    dados = request.get_json()

    print("JSON RECEBIDO /api/pcpm/movimentacao:")
    print(dados)

    if not dados:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Nenhum JSON recebido.'
        }), 400

    offline_id = dados.get('offline_id')
    equipamento_id = dados.get('equipamento_id')
    tipo_movimentacao = dados.get('tipo_movimentacao')
    cliente_id = dados.get('cliente_id')
    horimetro = dados.get('horimetro')
    numero_os = dados.get('numero_os')
    descricao_anomalia = dados.get('descricao_anomalia')
    latitude = dados.get('latitude')
    longitude = dados.get('longitude')
    itens = dados.get('itens', [])

    rfid_funcionario = (dados.get('rfid_funcionario') or '').strip()
    rfid_cliente = (dados.get('rfid_cliente') or '').strip()

    if not equipamento_id:
        return jsonify({'sucesso': False, 'mensagem': 'Equipamento não informado.'}), 400

    if not tipo_movimentacao:
        return jsonify({'sucesso': False, 'mensagem': 'Tipo de movimentação não informado.'}), 400

    if not cliente_id and not rfid_cliente:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Cliente não informado. Informe o cliente ou leia o RFID do cliente.'
        }), 400

    if not rfid_funcionario:
        return jsonify({'sucesso': False, 'mensagem': 'RFID do funcionário não informado.'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if offline_id:
            cursor.execute("""
                SELECT id
                FROM pcpm_movimentacoes
                WHERE offline_id = %s
                LIMIT 1
            """, (offline_id,))
            movimentacao_existente = cursor.fetchone()

            if movimentacao_existente:
                conn.close()
                return jsonify({
                    'sucesso': True,
                    'mensagem': 'Movimentação offline já sincronizada anteriormente.',
                    'movimentacao_id': movimentacao_existente['id']
                })

        cursor.execute("""
            SELECT id, status_localizacao
            FROM pcpm_equipamentos
            WHERE id = %s
              AND ativo = 1
        """, (equipamento_id,))
        equipamento = cursor.fetchone()

        if not equipamento:
            conn.close()
            return jsonify({'sucesso': False, 'mensagem': 'Equipamento não encontrado.'}), 404

        status_atual = equipamento['status_localizacao']

        # Para movimentação feita online, mantém a trava de status.
        # Para movimentação offline, aceita sincronizar mesmo que o status atual já tenha mudado.
        if not offline_id:
            if status_atual == 'area' and tipo_movimentacao != 'retirada':
                conn.close()
                return jsonify({
                    'sucesso': False,
                    'mensagem': 'Equipamento está na área e só permite retirada.'
                }), 400

            if status_atual == 'tradimaq' and tipo_movimentacao != 'entrega':
                conn.close()
                return jsonify({
                    'sucesso': False,
                    'mensagem': 'Equipamento está na Tradimaq e só permite entrega.'
                }), 400

        cursor.execute("""
            SELECT id, nome, uid_rfid
            FROM usuarios
            WHERE uid_rfid = %s
              AND ativo = 1
              AND tem_acesso_sistema = 1
            LIMIT 1
        """, (rfid_funcionario,))
        funcionario = cursor.fetchone()

        if not funcionario:
            conn.close()
            return jsonify({
                'sucesso': False,
                'mensagem': 'Funcionário não localizado pelo RFID.'
            }), 400

        funcionario_id = funcionario['id']

        cliente_cadastrado = 0
        status_sync = 'pendente_cliente'

        if cliente_id:
            cursor.execute("""
                SELECT id
                FROM pcpm_pessoas
                WHERE id = %s
                  AND ativo = 1
                LIMIT 1
            """, (cliente_id,))
            cliente = cursor.fetchone()

            if not cliente:
                conn.close()
                return jsonify({
                    'sucesso': False,
                    'mensagem': 'Cliente informado não encontrado.'
                }), 400

            cliente_cadastrado = 1
            status_sync = 'sincronizado'

        else:
            cursor.execute("""
                SELECT id
                FROM pcpm_pessoas
                WHERE rfid = %s
                  AND ativo = 1
                LIMIT 1
            """, (rfid_cliente,))
            cliente = cursor.fetchone()

            if cliente:
                cliente_id = cliente['id']
                cliente_cadastrado = 1
                status_sync = 'sincronizado'

        cursor.execute("""
            INSERT INTO pcpm_movimentacoes (
                offline_id,
                equipamento_id,
                cliente_id,
                funcionario_id,
                tipo_movimentacao,
                horimetro,
                numero_os,
                descricao_anomalia,
                rfid_funcionario,
                rfid_cliente,
                latitude,
                longitude,
                status_sync,
                cliente_cadastrado,
                criado_por,
                ativo
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, 1
            )
        """, (
            offline_id,
            equipamento_id,
            cliente_id,
            funcionario_id,
            tipo_movimentacao,
            horimetro,
            numero_os,
            descricao_anomalia,
            rfid_funcionario,
            rfid_cliente,
            latitude,
            longitude,
            status_sync,
            cliente_cadastrado,
            session.get('usuario_id')
        ))

        movimentacao_id = cursor.lastrowid

        for item in itens:
            item_id = item.get('item_id')
            resultado = item.get('resultado')
            observacao = item.get('observacao') or ''
            foto_path = item.get('foto_path') or ''

            cursor.execute("""
                SELECT id, ordem, item, criterio
                FROM pcpm_checklist_itens
                WHERE id = %s
                  AND ativo = 1
            """, (item_id,))
            item_checklist = cursor.fetchone()

            if not item_checklist:
                raise Exception(f'Item de checklist inválido: {item_id}')

            cursor.execute("""
                INSERT INTO pcpm_movimentacao_checklist_respostas (
                    movimentacao_id,
                    checklist_item_id,
                    resultado,
                    observacao,
                    foto_path,
                    item_snapshot,
                    criterio_snapshot,
                    ordem_snapshot
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                movimentacao_id,
                item_id,
                resultado,
                observacao,
                foto_path,
                item_checklist['item'],
                item_checklist['criterio'],
                item_checklist['ordem']
            ))

        novo_status = 'tradimaq' if tipo_movimentacao == 'retirada' else 'area'

        cursor.execute("""
            UPDATE pcpm_equipamentos
            SET status_localizacao = %s
            WHERE id = %s
        """, (novo_status, equipamento_id))

        conn.commit()
        conn.close()

        return jsonify({
            'sucesso': True,
            'mensagem': 'Movimentação registrada com sucesso.',
            'movimentacao_id': movimentacao_id,
            'novo_status': novo_status,
            'status_sync': status_sync,
            'cliente_cadastrado': cliente_cadastrado,
            'cliente_id': cliente_id,
            'funcionario_id': funcionario_id
        })

    except Exception as e:
        conn.rollback()
        conn.close()

        print("ERRO AO SALVAR MOVIMENTAÇÃO:")
        print(str(e))

        return jsonify({
            'sucesso': False,
            'mensagem': str(e)
        }), 500

@main_routes.route('/api/pcpm/upload_foto', methods=['POST'])
@login_required
def api_pcpm_upload_foto():
    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    movimentacao_id = request.form.get('movimentacao_id')
    checklist_item_id = request.form.get('checklist_item_id')
    foto = request.files.get('foto')

    if not movimentacao_id:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Movimentação não informada.'
        }), 400

    if not checklist_item_id:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Item do checklist não informado.'
        }), 400

    if not foto:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Foto não enviada.'
        }), 400

    try:
        pasta_destino = os.path.join(
            current_app.root_path,
            'static',
            'evidencias',
            'pcpm'
        )

        os.makedirs(pasta_destino, exist_ok=True)

        extensao = os.path.splitext(foto.filename)[1].lower()

        if extensao not in ['.jpg', '.jpeg', '.png']:
            return jsonify({
                'sucesso': False,
                'mensagem': 'Formato de imagem não permitido.'
            }), 400

        nome_arquivo = f"pcpm_mov_{movimentacao_id}_item_{checklist_item_id}_{uuid.uuid4().hex}{extensao}"

        caminho_completo = os.path.join(
            pasta_destino,
            nome_arquivo
        )

        foto.save(caminho_completo)

        caminho_relativo = f"evidencias/pcpm/{nome_arquivo}"

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            UPDATE pcpm_movimentacao_checklist_respostas
            SET foto_path = %s
            WHERE movimentacao_id = %s
              AND checklist_item_id = %s
        """, (
            caminho_relativo,
            movimentacao_id,
            checklist_item_id
        ))

        conn.commit()
        conn.close()

        return jsonify({
            'sucesso': True,
            'mensagem': 'Foto enviada com sucesso.',
            'foto_path': caminho_relativo
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'mensagem': str(e)
        }), 500
    
# ==========================================================
# API PCP-M - CONSULTAR CLIENTE POR RFID
# ==========================================================

@main_routes.route('/api/pcpm/cliente/rfid/<rfid>', methods=['GET'])
@login_required
def api_pcpm_cliente_por_rfid(rfid):

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            p.id,
            p.rfid,
            p.matricula,
            p.nome,
            p.empresa_id,
            emp.nome AS empresa_nome,
            p.setor_area
        FROM pcpm_pessoas p
        LEFT JOIN pcpm_empresas emp
            ON emp.id = p.empresa_id
        WHERE p.rfid = %s
          AND p.ativo = 1
        LIMIT 1
    """, (rfid,))

    pessoa = cursor.fetchone()

    conn.close()

    if not pessoa:
        return jsonify({
            'sucesso': True,
            'encontrado': False,
            'rfid': rfid,
            'mensagem': 'Cliente não cadastrado.'
        })

    return jsonify({
        'sucesso': True,
        'encontrado': True,
        'pessoa': pessoa
    })

# ==========================================================
# API PCP-M - CADASTRAR CLIENTE PELO APP
# ==========================================================

@main_routes.route('/api/pcpm/cliente/cadastrar', methods=['POST'])
@login_required
def api_pcpm_cadastrar_cliente_app():

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    dados = request.get_json()

    if not dados:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Nenhum JSON recebido.'
        }), 400

    rfid = (dados.get('rfid') or '').strip()
    matricula = (dados.get('matricula') or '').strip()
    nome = (dados.get('nome') or '').strip()
    empresa_id = dados.get('empresa_id')
    setor_area = (dados.get('setor_area') or '').strip()

    if not rfid:
        return jsonify({
            'sucesso': False,
            'mensagem': 'RFID não informado.'
        }), 400

    if not matricula:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Matrícula/registro não informado.'
        }), 400

    if not nome:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Nome não informado.'
        }), 400

    if not setor_area:
        return jsonify({
            'sucesso': False,
            'mensagem': 'Setor/área não informado.'
        }), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:

        cursor.execute("""
            SELECT id
            FROM pcpm_pessoas
            WHERE rfid = %s
              AND ativo = 1
            LIMIT 1
        """, (rfid,))

        pessoa_existente = cursor.fetchone()

        if pessoa_existente:
            conn.close()

            return jsonify({
                'sucesso': False,
                'mensagem': 'Já existe cliente cadastrado com este RFID.',
                'pessoa_id': pessoa_existente['id']
            }), 409

        cursor.execute("""
            INSERT INTO pcpm_pessoas (
                rfid,
                matricula,
                nome,
                empresa_id,
                setor_area,
                ativo
            )
            VALUES (%s, %s, %s, %s, %s, 1)
        """, (
            rfid,
            matricula,
            nome,
            empresa_id,
            setor_area
        ))

        pessoa_id = cursor.lastrowid

        conn.commit()

        return jsonify({
            'sucesso': True,
            'mensagem': 'Cliente cadastrado com sucesso.',
            'pessoa_id': pessoa_id
        })

    except Exception as e:

        conn.rollback()

        return jsonify({
            'sucesso': False,
            'mensagem': str(e)
        }), 500

    finally:
        conn.close()

# ==========================================================
# API PCP-M - CONSULTAR FUNCIONÁRIO POR RFID
# ==========================================================

@main_routes.route('/api/pcpm/funcionario/rfid/<rfid>', methods=['GET'])
@login_required
def api_pcpm_funcionario_por_rfid(rfid):

    if not session.get('acesso_pcpm'):
        return jsonify({
            'sucesso': False,
            'mensagem': 'Você não possui acesso ao módulo PCP-M.'
        }), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            id,
            nome,
            matricula,
            uid_rfid
        FROM usuarios
        WHERE uid_rfid = %s
          AND ativo = 1
          AND tem_acesso_sistema = 1
        LIMIT 1
    """, (rfid,))

    funcionario = cursor.fetchone()

    conn.close()

    if not funcionario:
        return jsonify({
            'sucesso': True,
            'encontrado': False,
            'mensagem': 'Funcionário não localizado.'
        })

    return jsonify({
        'sucesso': True,
        'encontrado': True,
        'funcionario': funcionario
    })